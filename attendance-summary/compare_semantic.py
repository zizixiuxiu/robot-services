import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SUMMARY_DAY_START_COL = 6
SUMMARY_DAY_COLS_PER_DAY = 4
DIST_DAY_START_COL = 8


def parse_dist_time(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s


def parse_summary_day(ws, row, day):
    cin_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (day - 1)
    cout_col = cin_col + 1
    cin = ws.cell(row=row, column=cin_col).value
    cout = ws.cell(row=row, column=cout_col).value
    return (parse_dist_time(cin), parse_dist_time(cout))


def parse_dist_day(ws, row, day):
    col = DIST_DAY_START_COL + (day - 1)
    val = ws.cell(row=row, column=col).value
    return parse_dist_time(val)


def load_distribution_data(path, sheet_name=None, max_days=31):
    wb = load_workbook(path, data_only=False)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    data = defaultdict(dict)
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or name == '姓名':
            continue
        for day in range(1, max_days + 1):
            v = parse_dist_day(ws, r, day)
            if v:
                data[name][day] = v
    wb.close()
    return dict(data)


def load_summary_distribution_data(path, sheet_name=None, max_days=31):
    wb = load_workbook(path, data_only=False)
    if sheet_name is None:
        # try month-named sheet first
        for sn in wb.sheetnames:
            if '部门分发' in sn:
                sheet_name = sn
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    data = defaultdict(dict)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or name == '姓名':
            continue
        for day in range(1, max_days + 1):
            v = parse_summary_day(ws, r, day)
            if any(v):
                data[name][day] = v
    wb.close()
    return dict(data)


def compare_dicts(a, b, label):
    diffs = []
    all_names = set(a.keys()) | set(b.keys())
    for name in sorted(all_names):
        if name not in a:
            diffs.append((label, name, 'missing in generated', None, b[name]))
            continue
        if name not in b:
            diffs.append((label, name, 'missing in target', a[name], None))
            continue
        days_a = a[name]
        days_b = b[name]
        all_days = set(days_a.keys()) | set(days_b.keys())
        for day in sorted(all_days):
            va = days_a.get(day)
            vb = days_b.get(day)
            if va != vb:
                diffs.append((label, name, day, va, vb))
    return diffs


def main():
    mode = sys.argv[1]  # dist or summary
    gen = sys.argv[2]
    tgt = sys.argv[3]
    max_days = int(sys.argv[4]) if len(sys.argv) > 4 else 31
    if mode == 'dist':
        a = load_distribution_data(gen, max_days=max_days)
        b = load_distribution_data(tgt, max_days=max_days)
    else:
        a = load_summary_distribution_data(gen, max_days=max_days)
        b = load_summary_distribution_data(tgt, max_days=max_days)
    diffs = compare_dicts(a, b, mode)
    print(f'Total semantic diffs: {len(diffs)}')
    for d in diffs[:50]:
        print(d)
    if not diffs:
        print('OK: time data matches')


if __name__ == '__main__':
    main()
