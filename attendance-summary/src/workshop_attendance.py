from __future__ import annotations

import argparse
import calendar
import copy
import json
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


SHEET_RAW = "\u539f\u59cb\u8bb0\u5f55"
SHEET_STAT = "\u7edf\u8ba1"
SHEET_SUMMARY = "\u6c47\u603b"

H_PUNCH_LOCATION = "\u6253\u5361\u5730\u70b9"
H_PERSON_ID = "\u4eba\u5458ID"
H_NAME = "\u59d3\u540d"
H_DEPT = "\u90e8\u95e8"
H_DEPT1 = "\u4e00\u7ea7\u90e8\u95e8"
H_DEPT2 = "\u4e8c\u7ea7\u90e8\u95e8"
H_POSITION = "\u5c97\u4f4d"
H_DATE = "\u65e5\u671f"
H_TIME = "\u65f6\u95f4"
H_MIN_TIME = "\u6700\u5c0f\u503c\u9879:\u65f6\u95f4"
H_MAX_TIME = "\u6700\u5927\u503c\u9879:\u65f6\u95f4"
H_HOURS = "\u5c0f\u65f6\u6570"
H_DAYS = "\u5929\u6570"
H_START_DATE = "\u4e0a\u73ed\u65e5\u671f"
H_ID_CARD = "\u8eab\u4efd\u8bc1\u53f7"
H_MONTH_HOURS = "\u6708\u5408\u8ba1\u5c0f\u65f6\u6570"
H_MONTH_MINUTES = "\u6708\u5408\u8ba1\u5206\u949f\u6570"
H_MONTH_DAYS = "\u5f53\u6708\u6253\u5361\u5929\u6570"
H_REST_DAYS = "\u6b63\u5e38\u4f11\u5047\u5929\u6570"
H_ATTEND_TOTAL = "\u51fa\u52e4\u5408\u8ba1"

MANUAL_HEADERS = {
    "\u8c03\u4f11\u5929\u6570",
    "\u51fa\u5dee\u5929\u6570",
    "\u8865\u5361\u5929\u6570",
    "\u5f02\u5e38\u6253\u5361",
    "\u6f0f\u5361",
    "\u6f0f\u5361\u6b21\u6570",
    "\u8bf7\u5047\u5929\u6570",
    "\u8865\u5361\u65f6\u95f4",
    "\u51fa\u5dee\u65f6\u95f4",
    "\u8bf7\u5047\u65f6\u95f4",
    "\u672a\u6253\u5361\u65f6\u95f4",
    "\u5907\u6ce8",
}



@dataclass
class RosterEntry:
    row_no: int
    person_id: str
    name: str
    punch_location: Any = None
    dept1: Any = None
    dept2: Any = None
    position: Any = None
    start_date: Any = None
    id_card: Any = None
    manual_values: dict[str, Any] | None = None


@dataclass
class DailyStat:
    person_id: str
    name: str
    work_date: date
    first_dt: datetime
    last_dt: datetime
    hours: float
    day_count: float
    punch_count: int
    entry: RosterEntry


@dataclass
class TransformReport:
    output_file: str
    year: int
    month: int
    roster_count: int
    raw_record_count: int
    matched_record_count: int
    daily_stat_count: int
    unmatched_person_count: int
    unmatched_people: list[str]


def norm_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def norm_name(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"Unsupported datetime value: {value!r}")


def combine_date_time(date_value: Any, time_value: Any) -> datetime:
    """把拆开的 日期列 + 时间列 合并成完整 datetime。"""
    if isinstance(date_value, datetime):
        d = date_value.date()
    elif isinstance(date_value, date):
        d = date_value
    else:
        text = str(date_value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                d = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                pass
        else:
            raise ValueError(f"Unsupported date value: {date_value!r}")
    if isinstance(time_value, datetime):
        t = time_value.time()
    elif isinstance(time_value, time):
        t = time_value
    else:
        parts = str(time_value).strip().split(":")
        try:
            t = time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
        except (ValueError, IndexError):
            raise ValueError(f"Unsupported time value: {time_value!r}")
    return datetime.combine(d, t)


def read_headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        headers[str(value).strip()] = col
    return headers


def is_day_header(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    if isinstance(value, int) and 1 <= value <= 31:
        return True
    if isinstance(value, float) and value.is_integer() and 1 <= int(value) <= 31:
        return True
    text = str(value).strip() if value is not None else ""
    return bool(re.fullmatch(r"\d{1,2}", text)) and 1 <= int(text) <= 31


def find_day_columns(ws) -> list[int]:
    cols = []
    for col in range(1, ws.max_column + 1):
        if is_day_header(ws.cell(1, col).value):
            cols.append(col)
    return cols


def copy_cell_format(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.protection:
        dst.protection = copy.copy(src.protection)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    if src_row == dst_row:
        return
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        copy_cell_format(ws.cell(src_row, col), ws.cell(dst_row, col))


def copy_col_style(ws, src_col: int, dst_col: int, max_row: int) -> None:
    if src_col == dst_col:
        return
    src_letter = get_column_letter(src_col)
    dst_letter = get_column_letter(dst_col)
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    for row in range(1, max_row + 1):
        copy_cell_format(ws.cell(row, src_col), ws.cell(row, dst_col))


def ensure_rows(ws, required_rows: int, style_row: int = 2) -> None:
    if required_rows <= ws.max_row:
        return
    max_col = ws.max_column
    for row in range(ws.max_row + 1, required_rows + 1):
        copy_row_style(ws, min(style_row, ws.max_row), row, max_col)


def ensure_day_columns(ws, required_days: int) -> list[int]:
    day_cols = find_day_columns(ws)
    if not day_cols:
        raise ValueError(f"No day columns found in sheet {ws.title!r}")
    max_row_for_style = max(ws.max_row, 2)
    while len(day_cols) < required_days:
        new_col = max(day_cols) + 1
        copy_col_style(ws, day_cols[-1], new_col, max_row_for_style)
        day_cols.append(new_col)
    return day_cols[:required_days]


def clear_values(ws, min_row: int, max_row: int, min_col: int = 1, max_col: int | None = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).value = None


def cell_value_by_header(ws, headers: dict[str, int], row: int, header: str) -> Any:
    col = headers.get(header)
    return ws.cell(row, col).value if col else None


def extract_roster(wb) -> list[RosterEntry]:
    raw_ws = wb[SHEET_RAW]
    raw_headers = read_headers(raw_ws)
    stat_by_name: dict[str, dict[str, Any]] = {}
    if SHEET_STAT in wb.sheetnames:
        stat_ws = wb[SHEET_STAT]
        stat_headers = read_headers(stat_ws)
        for row in range(2, stat_ws.max_row + 1):
            name = norm_name(cell_value_by_header(stat_ws, stat_headers, row, H_NAME))
            person_id = norm_id(cell_value_by_header(stat_ws, stat_headers, row, H_PERSON_ID))
            if not name or not person_id or name in stat_by_name:
                continue
            stat_by_name[name] = {
                H_PERSON_ID: person_id,
                H_DEPT1: cell_value_by_header(stat_ws, stat_headers, row, H_DEPT1)
                or cell_value_by_header(stat_ws, stat_headers, row, H_DEPT),
                H_DEPT2: cell_value_by_header(stat_ws, stat_headers, row, H_DEPT2),
                H_POSITION: cell_value_by_header(stat_ws, stat_headers, row, H_POSITION),
                H_PUNCH_LOCATION: cell_value_by_header(stat_ws, stat_headers, row, H_PUNCH_LOCATION),
            }
    summary_ws = wb[SHEET_SUMMARY] if SHEET_SUMMARY in wb.sheetnames else None
    summary_by_name: dict[str, dict[str, Any]] = {}

    if summary_ws:
        summary_headers = read_headers(summary_ws)
        for row in range(2, summary_ws.max_row + 1):
            name = norm_name(cell_value_by_header(summary_ws, summary_headers, row, H_NAME))
            if not name:
                continue
            summary_by_name[name] = {
                header: summary_ws.cell(row, col).value
                for header, col in summary_headers.items()
                if header
            }

    entries: list[RosterEntry] = []
    seen: set[tuple[str, str]] = set()
    seen_names: set[str] = set()
    for row in range(2, raw_ws.max_row + 1):
        name = norm_name(cell_value_by_header(raw_ws, raw_headers, row, H_NAME))
        stat_meta = stat_by_name.get(name or "", {})
        person_id = norm_id(stat_meta.get(H_PERSON_ID)) or norm_id(cell_value_by_header(raw_ws, raw_headers, row, H_PERSON_ID))
        if not name or not person_id:
            continue
        key = (person_id, name)
        if key in seen:
            continue
        seen.add(key)
        seen_names.add(name)
        summary = summary_by_name.get(name, {})
        entries.append(
            RosterEntry(
                row_no=len(entries) + 2,
                person_id=person_id,
                name=name,
                punch_location=cell_value_by_header(raw_ws, raw_headers, row, H_PUNCH_LOCATION)
                or stat_meta.get(H_PUNCH_LOCATION)
                or summary.get(H_PUNCH_LOCATION),
                dept1=cell_value_by_header(raw_ws, raw_headers, row, H_DEPT1)
                or stat_meta.get(H_DEPT1)
                or summary.get(H_DEPT1)
                or summary.get(H_DEPT),
                dept2=cell_value_by_header(raw_ws, raw_headers, row, H_DEPT2)
                or stat_meta.get(H_DEPT2)
                or summary.get(H_DEPT2),
                position=cell_value_by_header(raw_ws, raw_headers, row, H_POSITION)
                or stat_meta.get(H_POSITION)
                or summary.get(H_POSITION),
                start_date=summary.get(H_START_DATE),
                id_card=summary.get(H_ID_CARD),
                manual_values={h: summary.get(h) for h in MANUAL_HEADERS if h in summary},
            )
        )
    for name, stat_meta in stat_by_name.items():
        if name in seen_names:
            continue
        person_id = norm_id(stat_meta.get(H_PERSON_ID))
        if not person_id:
            continue
        summary = summary_by_name.get(name, {})
        entries.append(
            RosterEntry(
                row_no=len(entries) + 2,
                person_id=person_id,
                name=name,
                punch_location=stat_meta.get(H_PUNCH_LOCATION) or summary.get(H_PUNCH_LOCATION),
                dept1=stat_meta.get(H_DEPT1) or summary.get(H_DEPT1) or summary.get(H_DEPT),
                dept2=stat_meta.get(H_DEPT2) or summary.get(H_DEPT2),
                position=stat_meta.get(H_POSITION) or summary.get(H_POSITION),
                start_date=summary.get(H_START_DATE),
                id_card=summary.get(H_ID_CARD),
                manual_values={h: summary.get(h) for h in MANUAL_HEADERS if h in summary},
            )
        )
    return entries


def scan_input_people(input_file: Path) -> dict[str, dict[str, int]]:
    """
    扫描原始输入，按人员ID统计各姓名拼写出现次数。
    姓名取出现次数最多的拼写（同 ID 多名字时打印 warning）。
    """
    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    ws = None
    headers: dict[str, int] = {}
    for candidate in wb.worksheets:
        header_row = next(candidate.iter_rows(min_row=1, max_row=1, values_only=True))
        candidate_headers = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}
        if all(required in candidate_headers for required in (H_PERSON_ID, H_NAME)):
            ws = candidate
            headers = candidate_headers
            break
    if ws is None:
        wb.close()
        return {}

    people: dict[str, dict[str, int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        person_id = norm_id(row[headers[H_PERSON_ID]])
        name = norm_name(row[headers[H_NAME]])
        if not person_id or not name:
            continue
        names = people.setdefault(person_id, defaultdict(int))
        names[name] += 1
    wb.close()
    return people


def merge_input_roster(roster: list[RosterEntry], people: dict[str, dict[str, int]]) -> list[RosterEntry]:
    """
    把原始输入中有、但模板名单没有的人员追加进名单，按人员ID去重。
    新补人员的一级部门/二级部门/岗位留空（后续由员工自己导入），姓名、人员ID照常。
    """
    seen_ids = {e.person_id for e in roster}
    merged = list(roster)
    default_location = roster[0].punch_location if roster else None
    for person_id in sorted(people, key=lambda p: (0, int(p)) if p.isdigit() else (1, p)):
        if person_id in seen_ids:
            continue
        seen_ids.add(person_id)
        names = people[person_id]
        name = max(names.items(), key=lambda kv: kv[1])[0]
        if len(names) > 1:
            print(
                f"[workshop] WARNING: 人员ID {person_id} 存在多个姓名拼写 {dict(names)}，采用 {name!r}",
                file=sys.stderr,
            )
        merged.append(RosterEntry(
            row_no=len(merged) + 2,
            person_id=person_id,
            name=name,
            punch_location=default_location,
            dept1=None,
            dept2=None,
            position=None,
            start_date=None,
            id_card=None,
            manual_values=None,
        ))
    merged.sort(key=lambda e: (0, int(e.person_id)) if e.person_id.isdigit() else (1, e.person_id))
    return merged


def read_raw_records(input_file: Path, roster: list[RosterEntry]) -> tuple[dict[tuple[str, date], list[datetime]], int, int, set[str]]:
    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    ws = None
    headers: dict[str, int] = {}
    for candidate in wb.worksheets:
        header_row = next(candidate.iter_rows(min_row=1, max_row=1, values_only=True))
        candidate_headers = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}
        if all(required in candidate_headers for required in (H_PERSON_ID, H_NAME, H_TIME)):
            ws = candidate
            headers = candidate_headers
            break
    if ws is None:
        raise ValueError(f"Input file is missing required columns: {H_PERSON_ID}, {H_NAME}, {H_TIME}")

    by_id = {entry.person_id: entry for entry in roster}
    # 姓名兜底仅在姓名唯一时启用，避免重名（同名不同人员ID）匹配错人
    name_counts = defaultdict(int)
    for entry in roster:
        name_counts[entry.name] += 1
    by_name = {entry.name: entry for entry in roster if name_counts[entry.name] == 1}
    grouped: dict[tuple[str, date], list[datetime]] = defaultdict(list)
    total = 0
    matched = 0
    unmatched: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        person_id = norm_id(row[headers[H_PERSON_ID]])
        name = norm_name(row[headers[H_NAME]])
        entry = by_id.get(person_id or "") or by_name.get(name or "")
        if not entry:
            if name:
                unmatched.add(name)
            continue
        value = row[headers[H_TIME]]
        if value is None:
            continue
        # 日期/时间拆成两列的新格式：合并；单列完整 datetime 的旧格式保持兼容
        if H_DATE in headers and not isinstance(value, datetime):
            dt = combine_date_time(row[headers[H_DATE]], value)
        else:
            dt = parse_datetime(value)
        grouped[(entry.person_id, dt.date())].append(dt)
        matched += 1

    return grouped, total, matched, unmatched


def infer_year_month(grouped: dict[tuple[str, date], list[datetime]], explicit: str | None = None) -> tuple[int, int]:
    if explicit:
        m = re.fullmatch(r"(\d{4})[-/](\d{1,2})", explicit.strip())
        if not m:
            raise ValueError("--month must look like 2026-06")
        return int(m.group(1)), int(m.group(2))
    months: dict[tuple[int, int], int] = defaultdict(int)
    for _, work_date in grouped:
        months[(work_date.year, work_date.month)] += 1
    if not months:
        today = date.today()
        return today.year, today.month
    return max(months.items(), key=lambda item: item[1])[0]


def build_daily_stats(
    grouped: dict[tuple[str, date], list[datetime]],
    roster: list[RosterEntry],
    year: int,
    month: int,
    excluded_dates: set[tuple[str, date]] | None = None,
    half_day_dates: set[tuple[str, date]] | None = None,
    full_day_dates: set[tuple[str, date]] | None = None,
) -> list[DailyStat]:
    excluded_dates = excluded_dates or set()
    half_day_dates = half_day_dates or set()
    full_day_dates = full_day_dates or set()
    entry_by_id = {entry.person_id: entry for entry in roster}
    stats: list[DailyStat] = []
    for entry in roster:
        person_dates = sorted(
            work_date
            for person_id, work_date in grouped
            if person_id == entry.person_id and work_date.year == year and work_date.month == month
        )
        for work_date in person_dates:
            key = (entry.person_id, work_date)
            if key in excluded_dates:
                continue
            times = grouped[key]
            first_dt = min(times)
            last_dt = max(times)
            hours = (last_dt - first_dt).total_seconds() / 3600
            if hours < 0:
                hours += 24
            # 默认按工时：超过 8 小时算 1 天，8 小时以内（含）算 0.5 天；config 手工覆盖优先
            is_half_day = hours <= 8
            if key in half_day_dates:
                is_half_day = True
            if key in full_day_dates:
                is_half_day = False
            stats.append(
                DailyStat(
                    person_id=entry.person_id,
                    name=entry.name,
                    work_date=work_date,
                    first_dt=first_dt,
                    last_dt=last_dt,
                    hours=hours,
                    day_count=0.5 if is_half_day else 1,
                    punch_count=len(times),
                    entry=entry_by_id[entry.person_id],
                )
            )
    return stats


def set_cell_by_header(ws, headers: dict[str, int], row: int, header: str, value: Any) -> None:
    col = headers.get(header)
    if col:
        ws.cell(row, col).value = value


def write_raw_sheet(wb, roster: list[RosterEntry], daily_by_id_date: dict[tuple[str, date], DailyStat], year: int, month: int) -> None:
    ws = wb[SHEET_RAW]
    days_in_month = calendar.monthrange(year, month)[1]
    day_cols = ensure_day_columns(ws, days_in_month)
    ensure_rows(ws, len(roster) + 1, style_row=2)
    clear_values(ws, 2, ws.max_row)
    headers = read_headers(ws)

    first_day_header = ws.cell(1, day_cols[0]).value
    use_date_headers = isinstance(first_day_header, datetime)
    for day, col in enumerate(day_cols, start=1):
        ws.cell(1, col).value = datetime(year, month, day) if use_date_headers else day
    for col in find_day_columns(ws)[days_in_month:]:
        ws.cell(1, col).value = None

    for i, entry in enumerate(roster, start=2):
        write_common_person_cells(ws, headers, i, entry, include_location=True)
        for day, col in enumerate(day_cols, start=1):
            stat = daily_by_id_date.get((entry.person_id, date(year, month, day)))
            if stat:
                ws.cell(i, col).value = f"{stat.first_dt:%H:%M}\n{stat.last_dt:%H:%M}"
            else:
                ws.cell(i, col).value = None

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{get_column_letter(max(day_cols))}{len(roster) + 1}"


def write_common_person_cells(ws, headers: dict[str, int], row: int, entry: RosterEntry, include_location: bool = False) -> None:
    if include_location:
        set_cell_by_header(ws, headers, row, H_PUNCH_LOCATION, entry.punch_location)
    set_cell_by_header(ws, headers, row, H_PERSON_ID, int(entry.person_id) if entry.person_id.isdigit() else entry.person_id)
    set_cell_by_header(ws, headers, row, H_NAME, entry.name)
    set_cell_by_header(ws, headers, row, H_DEPT1, entry.dept1)
    set_cell_by_header(ws, headers, row, H_DEPT, entry.dept1)
    set_cell_by_header(ws, headers, row, H_DEPT2, entry.dept2)
    set_cell_by_header(ws, headers, row, H_POSITION, entry.position)


def write_stat_sheet(wb, stats: list[DailyStat]) -> None:
    ws = wb[SHEET_STAT]
    headers = read_headers(ws)
    template_single_punch_blank_max = False
    if H_MIN_TIME in headers and H_MAX_TIME in headers:
        min_col_for_scan = headers[H_MIN_TIME]
        max_col_for_scan = headers[H_MAX_TIME]
        blank_max_count = 0
        for scan_row in range(2, min(ws.max_row, 5000) + 1):
            if ws.cell(scan_row, min_col_for_scan).value is not None and ws.cell(scan_row, max_col_for_scan).value is None:
                blank_max_count += 1
                if blank_max_count >= 5:
                    template_single_punch_blank_max = True
                    break
    ensure_rows(ws, len(stats) + 1, style_row=2)
    clear_values(ws, 2, ws.max_row)

    min_col = headers.get(H_MIN_TIME)
    max_col = headers.get(H_MAX_TIME)
    hours_col = headers.get(H_HOURS)

    for row, stat in enumerate(stats, start=2):
        entry = stat.entry
        write_common_person_cells(ws, headers, row, entry, include_location=True)
        set_cell_by_header(ws, headers, row, H_DATE, datetime(stat.work_date.year, stat.work_date.month, stat.work_date.day))
        set_cell_by_header(ws, headers, row, H_MIN_TIME, stat.first_dt.time().replace(microsecond=0))
        near_single_punch = (stat.last_dt - stat.first_dt).total_seconds() <= 20
        max_time_value = (
            None
            if template_single_punch_blank_max and (stat.punch_count <= 1 or near_single_punch)
            else stat.last_dt.time().replace(microsecond=0)
        )
        set_cell_by_header(ws, headers, row, H_MAX_TIME, max_time_value)
        set_cell_by_header(ws, headers, row, H_DAYS, stat.day_count)
        if hours_col and min_col and max_col:
            min_letter = get_column_letter(min_col)
            max_letter = get_column_letter(max_col)
            ws.cell(row, hours_col).value = f"=MOD({max_letter}{row}-{min_letter}{row},1)*24"
        else:
            set_cell_by_header(ws, headers, row, H_HOURS, stat.hours)

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{len(stats) + 1}"


def numeric_or_zero(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def write_summary_sheet(
    wb,
    roster: list[RosterEntry],
    stats: list[DailyStat],
    year: int,
    month: int,
    preserve_manual: bool = False,
) -> None:
    ws = wb[SHEET_SUMMARY]
    headers = read_headers(ws)
    ensure_rows(ws, len(roster) + 1, style_row=2)
    clear_values(ws, 2, ws.max_row)

    hours_by_id: dict[str, float] = defaultdict(float)
    days_by_id: dict[str, float] = defaultdict(float)
    for stat in stats:
        hours_by_id[stat.person_id] += stat.hours
        days_by_id[stat.person_id] += stat.day_count

    days_in_month = calendar.monthrange(year, month)[1]
    for row, entry in enumerate(roster, start=2):
        write_common_person_cells(ws, headers, row, entry, include_location=True)
        set_cell_by_header(ws, headers, row, H_START_DATE, entry.start_date)
        set_cell_by_header(ws, headers, row, H_ID_CARD, entry.id_card)
        hours = hours_by_id.get(entry.person_id, 0.0)
        punch_days = days_by_id.get(entry.person_id, 0.0)
        set_cell_by_header(ws, headers, row, H_MONTH_HOURS, hours)
        set_cell_by_header(ws, headers, row, H_MONTH_MINUTES, hours * 60)
        set_cell_by_header(ws, headers, row, H_MONTH_DAYS, punch_days)
        set_cell_by_header(ws, headers, row, H_REST_DAYS, days_in_month - punch_days)
        if preserve_manual and entry.manual_values:
            for header, value in entry.manual_values.items():
                set_cell_by_header(ws, headers, row, header, value)
        attend_col = headers.get(H_ATTEND_TOTAL)
        if attend_col:
            month_days_col = headers.get(H_MONTH_DAYS)
            manual_cols = [
                headers.get("\u8c03\u4f11\u5929\u6570"),
                headers.get("\u51fa\u5dee\u5929\u6570"),
                headers.get("\u8865\u5361\u5929\u6570"),
                headers.get("\u8bf7\u5047\u5929\u6570"),
            ]
            refs = [get_column_letter(c) + str(row) for c in [month_days_col, *manual_cols] if c]
            ws.cell(row, attend_col).value = f"=SUM({','.join(refs)})" if refs else None

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{len(roster) + 1}"


def _load_date_rules(path: Path | None) -> tuple[set[tuple[str, date]], set[tuple[str, date]], set[tuple[str, date]]]:
    if not path:
        return set(), set(), set()
    data = json.loads(path.read_text(encoding="utf-8"))
    excluded: set[tuple[str, date]] = set()
    half_day: set[tuple[str, date]] = set()
    full_day: set[tuple[str, date]] = set()
    for item in data.get("exclude_dates", []):
        person_id = norm_id(item.get("person_id"))
        day = datetime.strptime(item["date"], "%Y-%m-%d").date()
        if person_id:
            excluded.add((person_id, day))
    for item in data.get("half_day_dates", []):
        person_id = norm_id(item.get("person_id"))
        day = datetime.strptime(item["date"], "%Y-%m-%d").date()
        if person_id:
            half_day.add((person_id, day))
    for item in data.get("full_day_dates", []):
        person_id = norm_id(item.get("person_id"))
        day = datetime.strptime(item["date"], "%Y-%m-%d").date()
        if person_id:
            full_day.add((person_id, day))
    return excluded, half_day, full_day


def _setup_print(ws, title_rows: str = "1:1") -> None:
    """打印友好：横向、缩放至一页宽、顶端重复表头、页脚页码。"""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = title_rows
    ws.oddFooter.center.text = "第 &P 页 / 共 &N 页"


def finalize_workshop_output(wb) -> None:
    """车间输出展示层优化：冻结、分钟数整数显示、统计按人员分组折叠（默认展开）、打印设置。"""
    ws_raw = wb["原始记录"]
    ws_raw.freeze_panes = "A2"
    _setup_print(ws_raw)

    ws_stat = wb["统计"]
    ws_stat.freeze_panes = "A2"
    headers = read_headers(ws_stat)
    id_col = headers.get(H_PERSON_ID)
    if id_col:
        for r in range(2, ws_stat.max_row + 1):
            if norm_id(ws_stat.cell(r, id_col).value):
                ws_stat.row_dimensions[r].outline_level = 1
    _setup_print(ws_stat)

    ws_sum = wb["汇总"]
    headers = read_headers(ws_sum)
    minutes_col = headers.get("月合计分钟数")
    if minutes_col:
        for r in range(2, ws_sum.max_row + 1):
            ws_sum.cell(r, minutes_col).number_format = "0"
    _setup_print(ws_sum)


def transform_attendance(
    input_file: Path,
    template_file: Path,
    output_file: Path,
    month: str | None = None,
    config_file: Path | None = None,
    preserve_manual: bool = False,
) -> TransformReport:
    input_file = Path(input_file)
    template_file = Path(template_file)
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_file)
    for sheet_name in (SHEET_RAW, SHEET_STAT, SHEET_SUMMARY):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Template workbook is missing sheet: {sheet_name}")

    roster = extract_roster(wb)
    # 名单以原始输入为准：先并入输入人员再读原始记录，保证月份从完整打卡数据推断
    roster = merge_input_roster(roster, scan_input_people(input_file))
    grouped, raw_count, matched_count, unmatched = read_raw_records(input_file, roster)
    year, month_no = infer_year_month(grouped, month)
    # 模板名单中没有打卡记录的人不出现在输出里
    present_ids = {person_id for person_id, _ in grouped}
    roster = [entry for entry in roster if entry.person_id in present_ids]
    excluded, half_day, full_day = _load_date_rules(config_file)
    stats = build_daily_stats(
        grouped,
        roster,
        year,
        month_no,
        excluded_dates=excluded,
        half_day_dates=half_day,
        full_day_dates=full_day,
    )
    daily_by_id_date = {(stat.person_id, stat.work_date): stat for stat in stats}

    write_raw_sheet(wb, roster, daily_by_id_date, year, month_no)
    write_stat_sheet(wb, stats)
    write_summary_sheet(wb, roster, stats, year, month_no, preserve_manual=preserve_manual)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    finalize_workshop_output(wb)

    tmp = output_file.with_suffix(output_file.suffix + ".tmp")
    wb.save(tmp)
    shutil.move(str(tmp), output_file)

    return TransformReport(
        output_file=str(output_file),
        year=year,
        month=month_no,
        roster_count=len(roster),
        raw_record_count=raw_count,
        matched_record_count=matched_count,
        daily_stat_count=len(stats),
        unmatched_person_count=len(unmatched),
        unmatched_people=sorted(unmatched),
    )


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate workshop attendance workbooks from raw gate logs.")
    parser.add_argument("--input", required=True, type=Path, help="Raw attendance xlsx file.")
    parser.add_argument("--template", required=True, type=Path, help="Existing output workbook used as style/roster template.")
    parser.add_argument("--output", required=True, type=Path, help="Output xlsx path.")
    parser.add_argument("--month", help="Optional target month, e.g. 2026-06.")
    parser.add_argument("--config", type=Path, help="Optional JSON config with exclude_dates.")
    parser.add_argument("--preserve-manual", action="store_true", help="Keep manual adjustment columns from the template roster.")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    report = transform_attendance(
        input_file=args.input,
        template_file=args.template,
        output_file=args.output,
        month=args.month,
        config_file=args.config,
        preserve_manual=args.preserve_manual,
    )
    print(json.dumps(asdict(report), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
