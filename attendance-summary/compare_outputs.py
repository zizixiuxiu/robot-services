import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def load_cells(path):
    wb = load_workbook(path, data_only=False)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                cells[(cell.row, cell.column)] = cell.value
        result[sheet_name] = cells
    wb.close()
    return result


def compare_files(a_path, b_path, ignore_sheets=None, ignore_cols=None):
    ignore_sheets = set(ignore_sheets or [])
    ignore_cols = set(ignore_cols or [])
    a = load_cells(a_path)
    b = load_cells(b_path)
    diffs = []
    all_sheets = set(a.keys()) | set(b.keys())
    for sheet in all_sheets:
        if sheet in ignore_sheets:
            continue
        a_cells = a.get(sheet, {})
        b_cells = b.get(sheet, {})
        all_keys = set(a_cells.keys()) | set(b_cells.keys())
        for key in all_keys:
            r, c = key
            if c in ignore_cols:
                continue
            av = a_cells.get(key)
            bv = b_cells.get(key)
            # Normalize empty values
            if av in (None, ''):
                av = None
            if bv in (None, ''):
                bv = None
            if av != bv:
                diffs.append((sheet, r, get_column_letter(c), av, bv))
    return diffs


def main():
    a = sys.argv[1]
    b = sys.argv[2]
    ignore_sheets = sys.argv[3].split(',') if len(sys.argv) > 3 else []
    ignore_cols = [int(x) for x in sys.argv[4].split(',')] if len(sys.argv) > 4 else []
    diffs = compare_files(a, b, ignore_sheets, ignore_cols)
    if not diffs:
        print('No diffs')
        return
    print(f'Total diffs: {len(diffs)}')
    for d in diffs[:50]:
        print(d)


if __name__ == '__main__':
    main()
