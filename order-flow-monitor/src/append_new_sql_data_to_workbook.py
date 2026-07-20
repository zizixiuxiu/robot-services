import argparse
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WORKDIR = Path(r"C:\Users\Administrator\Documents\Codex\2026-05-28\sqlserver")
DATA_JSON = WORKDIR / "sql_data.json"
SOURCE_SHEETS = ("SQLData", "OrderLookup", "OrderList")


def stat_signature(path):
    stat = Path(path).stat()
    return (stat.st_size, stat.st_mtime_ns)


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def normalize_key(value):
    if value is None:
        return ""
    return str(value).strip()


def canonical_order_id(value):
    key = normalize_key(value)
    if not key:
        return ""
    key = re.sub(r"\s+", "-", key)
    key = key.replace("_", "-")
    key = re.sub(r"-{2,}", "-", key)
    return key.strip("-")


def key_variants(order_id):
    key = normalize_key(order_id)
    if not key:
        return []

    variants = []

    def add(value):
        value = normalize_key(value)
        if value and value not in variants:
            variants.append(value)

    def add_separator_variants(value):
        add(value)
        add(value.replace("_", "-"))
        add(value.replace("-", "_"))
        parts = re.split(r"[_-]+", value)
        if len(parts) >= 2:
            add(" ".join(parts))
            add(" ".join(parts[:2]) + "_" + "_".join(parts[2:]))
            add(" ".join(parts[:2]) + "-" + "-".join(parts[2:]))

    add_separator_variants(key)
    canonical = canonical_order_id(key)
    if canonical != key:
        add_separator_variants(canonical)

    j_alias = re.sub(r"([_-])J([_-])", r"\g<1>8\2", key, flags=re.IGNORECASE)
    if j_alias != key:
        add_separator_variants(j_alias)
        j_alias_canonical = canonical_order_id(j_alias)
        if j_alias_canonical != j_alias:
            add_separator_variants(j_alias_canonical)

    match = re.match(r"^(.*?)[_-](\d{2})$", key)
    if match:
        base = match.group(1)
        add_separator_variants(base)
        base_j_alias = re.sub(r"([_-])J$", r"\g<1>8", base, flags=re.IGNORECASE)
        if base_j_alias != base:
            add_separator_variants(base_j_alias)

    return variants


def get_or_create_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def ensure_header(ws, expected):
    current = [ws.cell(1, col).value for col in range(1, len(expected) + 1)]
    if ws.max_row == 1 and all(value is None for value in current):
        for col, header in enumerate(expected, start=1):
            ws.cell(1, col).value = header
        style_header(ws[1])
        return
    if current != expected:
        raise RuntimeError(f"{ws.title} header mismatch. Refusing to append because the data source layout changed.")


def clear_body(ws):
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def update_lookup_formula_ranges(wb, lookup_last_row, lookup_last_col, order_list_last_row):
    return

    lookup_col_letter = get_column_letter(lookup_last_col)
    lookup_range = f"OrderLookup!$A$2:${lookup_col_letter}${lookup_last_row}"
    order_list_range = f"OrderList!$A$2:$A${order_list_last_row}"
    lookup_pattern = re.compile(r"OrderLookup!\$A\$2:\$[A-Za-z]+\$\d+")
    order_list_pattern = re.compile(r"OrderList!\$A\$2:\$A\$\d+")

    def solid_wood_formula(row_num):
        normalized = f'SUBSTITUTE(SUBSTITUTE(UPPER(TRIM($D{row_num}&"")),"_","-")," ","-")'
        return (
            f'=IF(LEN(TRIM($D{row_num}&""))=0,"",'
            f'IF(OR(ISNUMBER(SEARCH("-J-",{normalized})),RIGHT({normalized},2)="-J"),"鈭?,""))'
        )

    order_ws = wb.worksheets[0] if wb.worksheets else None

    for ws in wb.worksheets:
        if ws.title in SOURCE_SHEETS:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if ws is order_ws and cell.column == 26:
                        cell.value = solid_wood_formula(cell.row)
                        continue
                    value = lookup_pattern.sub(lookup_range, cell.value)
                    value = order_list_pattern.sub(order_list_range, value)
                    if value != cell.value:
                        cell.value = value

        if ws is order_ws:
            for row_num in range(4, ws.max_row + 1):
                ws[f"Z{row_num}"] = solid_wood_formula(row_num)

        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                dv.formula1 = order_list_pattern.sub(order_list_range, dv.formula1)


def snapshot_visible_user_values(wb):
    values = {}
    for ws in wb.worksheets:
        if ws.title in SOURCE_SHEETS:
            continue
        sheet_values = {}
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None and not (isinstance(value, str) and value.startswith("=")):
                    sheet_values[cell.coordinate] = value
        values[ws.title] = sheet_values
    return values


def restore_visible_user_values(wb, snapshot):
    restored = 0
    for sheet_name, sheet_values in snapshot.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for coordinate, value in sheet_values.items():
            cell = ws[coordinate]
            if cell.value != value and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = value
                restored += 1
    return restored


def restore_order_sheet_values_from_backup(wb, backup_path):
    if not backup_path:
        return 0

    backup = Path(backup_path)
    if not backup.exists():
        raise FileNotFoundError(f"Visible sheet backup not found: {backup}")

    backup_wb = load_workbook(backup, data_only=False)
    source_ws = backup_wb.worksheets[0]
    target_ws = wb[source_ws.title] if source_ws.title in wb.sheetnames else wb.worksheets[0]
    restored = 0

    for row in source_ws.iter_rows():
        for source_cell in row:
            value = source_cell.value
            if isinstance(value, str) and value.startswith("="):
                continue

            target_cell = target_ws[source_cell.coordinate]
            target_value = target_cell.value
            if isinstance(target_value, str) and target_value.startswith("="):
                continue

            if target_value != value:
                target_cell.value = value
                restored += 1

    return restored


def main():
    parser = argparse.ArgumentParser(description="Append only new SQL rows to hidden workbook data-source sheets.")
    parser.add_argument("--target", default=os.environ.get("TARGET_XLSX"), help="Workbook to update.")
    parser.add_argument(
        "--visible-backup",
        default=os.environ.get("REFRESH_BACKUP_XLSX"),
        help="Workbook backup whose first visible/order sheet user-entered values should be restored before saving.",
    )
    parser.add_argument(
        "--rebuild-source",
        action="store_true",
        help="Clear and rebuild hidden SQLData, OrderLookup, and OrderList sheets from sql_data.json.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Report what would change without saving.")
    args = parser.parse_args()

    if not args.target:
        raise SystemExit("TARGET_XLSX or --target is required.")

    target = Path(args.target)
    target_signature_before_load = stat_signature(target)
    payload = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    headers = payload["headers"]
    rows = payload["rows"]

    wb = load_workbook(target)
    visible_user_values = snapshot_visible_user_values(wb)
    data_ws = get_or_create_sheet(wb, "SQLData")
    lookup_ws = get_or_create_sheet(wb, "OrderLookup")
    order_list_ws = get_or_create_sheet(wb, "OrderList")

    ensure_header(data_ws, headers)
    ensure_header(lookup_ws, ["订单号Key", *headers])
    ensure_header(order_list_ws, ["订单号"])

    if args.rebuild_source:
        clear_body(data_ws)
        clear_body(lookup_ws)
        clear_body(order_list_ws)

    existing_exact_keys = {
        canonical_order_id(lookup_ws.cell(row, 4).value)
        for row in range(2, lookup_ws.max_row + 1)
        if canonical_order_id(lookup_ws.cell(row, 4).value)
    }
    existing_data_order_ids = {
        canonical_order_id(data_ws.cell(row, 3).value)
        for row in range(2, data_ws.max_row + 1)
        if canonical_order_id(data_ws.cell(row, 3).value)
    }
    existing_lookup_keys = {
        normalize_key(lookup_ws.cell(row, 1).value)
        for row in range(2, lookup_ws.max_row + 1)
        if normalize_key(lookup_ws.cell(row, 1).value)
    }
    existing_order_list_keys = {
        normalize_key(order_list_ws.cell(row, 1).value)
        for row in range(2, order_list_ws.max_row + 1)
        if normalize_key(order_list_ws.cell(row, 1).value)
    }
    new_rows = [row for row in rows if canonical_order_id(row[2]) and canonical_order_id(row[2]) not in existing_exact_keys]

    appended_data_rows = 0
    for row in rows:
        data_key = canonical_order_id(row[2])
        if data_key in existing_data_order_ids:
            continue
        display_row = list(row)
        display_row[2] = canonical_order_id(row[2])
        display_row[3] = canonical_order_id(row[3])
        data_ws.append(display_row)
        existing_data_order_ids.add(data_key)
        appended_data_rows += 1

    appended_lookup_rows = 0
    appended_order_list_rows = 0
    for row in rows:
        display_row = list(row)
        display_row[2] = canonical_order_id(row[2])
        display_row[3] = canonical_order_id(row[3])
        for key in key_variants(row[2]):
            raw_key = normalize_key(key)
            if raw_key in existing_lookup_keys:
                continue
            existing_lookup_keys.add(raw_key)
            display_key = canonical_order_id(key)
            lookup_ws.append([raw_key, *display_row])
            appended_lookup_rows += 1
            if display_key and display_key not in existing_order_list_keys:
                existing_order_list_keys.add(display_key)
                order_list_ws.append([display_key])
                appended_order_list_rows += 1

    for ws in (data_ws, lookup_ws, order_list_ws):
        ws.sheet_state = "hidden"
        ws.freeze_panes = "A2"

    update_lookup_formula_ranges(
        wb,
        lookup_last_row=lookup_ws.max_row,
        lookup_last_col=lookup_ws.max_column,
        order_list_last_row=order_list_ws.max_row,
    )

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    restored_visible_values = restore_visible_user_values(wb, visible_user_values)
    restored_order_values_from_backup = restore_order_sheet_values_from_backup(wb, args.visible_backup)

    print(f"target={target}")
    print(f"sql_rows={len(rows)}")
    print(f"existing_rows={len(existing_exact_keys)}")
    print(f"new_rows={len(new_rows)}")
    print(f"appended_data_rows={appended_data_rows}")
    print(f"appended_lookup_rows={appended_lookup_rows}")
    print(f"appended_order_list_rows={appended_order_list_rows}")
    print(f"restored_visible_user_values={restored_visible_values}")
    print(f"restored_order_values_from_backup={restored_order_values_from_backup}")

    if args.dry_run:
        print("dry_run=true")
        return

    target_signature_before_save = stat_signature(target)
    if target_signature_before_save != target_signature_before_load:
        raise RuntimeError(
            "refresh_status=skipped reason=workbook_changed_during_refresh "
            "message='Target workbook changed while refresh was running. "
            "Skipped save to avoid overwriting user edits; monitor will retry.'"
        )

    wb.save(target)
    print("refresh_status=saved")


if __name__ == "__main__":
    main()
