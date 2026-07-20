# -*- coding: utf-8 -*-
"""
Dealer Data Analyzer integration.
Adapted from dealer-data-analyzer skill to generate 4 analysis sheets
that are appended after the account-detail sheet.
"""

import os
import json
import urllib.request
import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== Constants ==========
DARK_BLUE = "1F4E78"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D9E2F3"
VLIGHT_BLUE = "EDF7FF"
GREEN = "70AD47"
RED = "C00000"
ORANGE = "ED7D31"
GRAY = "808080"
DK_GRAY = "404040"

# Fonts
ft_title = Font(name='微软雅黑', size=18, bold=True, color=DARK_BLUE)
ft_subtitle = Font(name='微软雅黑', size=11, color=GRAY)
ft_section = Font(name='微软雅黑', size=13, bold=True, color="FFFFFF")
ft_header = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
ft_data = Font(name='微软雅黑', size=11, color=DK_GRAY)
ft_bold = Font(name='微软雅黑', size=11, bold=True, color=DK_GRAY)
ft_insight = Font(name='微软雅黑', size=11, bold=True, color=RED)
ft_positive = Font(name='微软雅黑', size=11, bold=True, color=GREEN)
ft_small = Font(name='微软雅黑', size=10, color=GRAY)
ft_card_label = Font(name='微软雅黑', size=10, color=GRAY)
ft_card_value = Font(name='微软雅黑', size=20, bold=True, color=DARK_BLUE)
ft_card_note = Font(name='微软雅黑', size=10, color=MED_BLUE)

# Fills
fill_dark = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_light = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fill_red = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
fill_orange = PatternFill(start_color="FEF0E7", end_color="FEF0E7", fill_type="solid")
fill_yellow = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
fill_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_vlight = PatternFill(start_color=VLIGHT_BLUE, end_color=VLIGHT_BLUE, fill_type="solid")

# Alignments
al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
al_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Border
bd_thin = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)


def prepare_dealer_df(df):
    """Convert the internal dealer DataFrame to the analyzer format."""
    out = pd.DataFrame()
    out['大区'] = df['区域'].astype(str).str.strip()
    out['经销商'] = df['经销商'].astype(str).str.strip()
    out['创建方案数'] = pd.to_numeric(df['创建方案数'], errors='coerce').fillna(0).astype(int)
    out['渲染方案数'] = pd.to_numeric(df['新增渲染方案数'], errors='coerce').fillna(0).astype(int)
    out['提审方案数'] = pd.to_numeric(df['提审方案数'], errors='coerce').fillna(0).astype(int)

    out['渲染率'] = np.where(out['创建方案数'] > 0,
                              (out['渲染方案数'] / out['创建方案数'] * 100).round(1), 0)
    out['提审率_创建'] = np.where(out['创建方案数'] > 0,
                                 (out['提审方案数'] / out['创建方案数'] * 100).round(1), 0)
    out['提审率_渲染'] = np.where(out['渲染方案数'] > 0,
                                 (out['提审方案数'] / out['渲染方案数'] * 100).round(1), 0)

    def classify(row):
        if row['创建方案数'] == 0:
            return '零活跃'
        if row['提审方案数'] > 0:
            if row['提审率_创建'] >= 20:
                return '高转化'
            elif row['提审率_创建'] >= 10:
                return '中转化'
            else:
                return '低转化'
        else:
            if row['渲染方案数'] > 0:
                return '有渲染无提审'
            else:
                return '无深化'

    out['转化分层'] = out.apply(classify, axis=1)
    return out


def build_region_stats(df):
    """Build region-level statistics."""
    stats = df.groupby('大区').agg({
        '经销商': 'count',
        '创建方案数': ['sum', 'mean'],
        '渲染方案数': ['sum', 'mean'],
        '提审方案数': ['sum', 'mean']
    }).round(1)

    stats.columns = ['经销商数', '创建总和', '创建均值', '渲染总和', '渲染均值', '提审总和', '提审均值']
    stats['活跃经销商数'] = df.groupby('大区').apply(lambda x: (x['创建方案数'] > 0).sum(), include_groups=False)
    stats['活跃度'] = (stats['活跃经销商数'] / stats['经销商数'] * 100).round(1)
    stats['渲染率'] = (stats['渲染总和'] / stats['创建总和'] * 100).round(1)
    stats['提审率(创建)'] = (stats['提审总和'] / stats['创建总和'] * 100).round(1)
    stats['人均创建'] = (stats['创建总和'] / stats['活跃经销商数']).round(1)
    stats['人均提审'] = (stats['提审总和'] / stats['活跃经销商数']).round(1)
    stats['效率指数'] = (stats['提审率(创建)'] * 0.6 + stats['活跃度'] * 0.4).round(1)

    def get_diagnosis(row):
        if row['活跃度'] >= 90 and row['提审率(创建)'] >= 20:
            return "★ 明星大区：高活跃高转化，经验复制的首选"
        elif row['活跃度'] >= 90 and row['提审率(创建)'] < 15:
            return "▲ 高活低效：活跃度高但转化低，需强化销售闭环培训"
        elif row['活跃度'] < 70 and row['提审率(创建)'] >= 15:
            return "● 低效高转：转化率高但活跃度低，需激活沉睡经销商"
        elif row['活跃度'] < 70 and row['提审率(创建)'] < 10:
            return "▼ 低效低转：双重困境，需要系统性诊断和扶持"
        else:
            return "◆ 潜力大区：中等水平，可针对性提升"

    stats['AI诊断'] = stats.apply(get_diagnosis, axis=1)
    return stats


def _safe_json_loads(text):
    """安全解析 AI 返回的 JSON，支持 markdown 代码块和截断修复。"""
    if not text:
        return None
    text = text.strip()
    if text.startswith('```json'):
        text = text[7:]
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()

    # 正常解析
    try:
        return json.loads(text)
    except Exception:
        pass

    # 截断修复：尝试去掉末尾不完整的对象/数组内容，找到最后一个完整结构
    # 策略：从末尾向前找最后一个 '}' 或 ']'，尝试截断并补全
    for end_char in ['}', ']']:
        idx = text.rfind(end_char)
        if idx > 0:
            truncated = text[:idx + 1]
            try:
                return json.loads(truncated)
            except Exception:
                pass

    return None


def call_deepseek(prompt: str) -> str:
    """调用 DeepSeek API，未配置 key 时返回空字符串。"""
    api_key = os.getenv('AI_API_KEY')
    base_url = os.getenv('AI_BASE_URL', 'https://api.deepseek.com/v1')
    model = os.getenv('AI_MODEL', 'deepseek-v4-flash')

    if not api_key:
        return ''

    try:
        body = json.dumps({
            'model': model,
            'messages': [{'role': 'user', 'content': prompt}],
            'temperature': 0.5,
            'max_tokens': 6000,
        }, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(
            f'{base_url}/chat/completions',
            data=body,
            headers={
                'Content-Type': 'application/json; charset=utf-8',
                'Authorization': f'Bearer {api_key}',
            },
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode('utf-8'))
        return result['choices'][0]['message']['content']
    except Exception as e:
        print(f'DeepSeek 调用失败: {e}')
        return ''


def build_ai_prompt(df, region_stats) -> str:
    """构建给 DeepSeek 的 prompt。"""
    total_create = int(df['创建方案数'].sum())
    total_render = int(df['渲染方案数'].sum())
    total_submit = int(df['提审方案数'].sum())
    total_count = len(df)
    active_count = int(len(df[df['创建方案数'] > 0]))

    lines = [
        '你是一位资深经销商运营分析师。请根据以下数据生成中文分析结论。',
        '必须严格按下方 JSON Schema 返回，不要输出 JSON 以外的任何内容（不要加 markdown 代码块）。',
        '为控制输出长度，每个字符串请控制在 80 字以内，JSON 总长度尽量不超过 2500 字：',
        '',
        '{',
        '  "核心洞察": ["1. ...", "2. ...", "3. ...", "4. ...", "5. ..."],',
        '  "大区诊断": {"华东区": "...", ...},',
        '  "行动建议": {',
        '    "P0紧急行动": [{"优先级":"P0", "行动项": "...", "目标对象": "...", "预期指标变化": "...", "预估影响": "...", "投入估算": "...", "预期ROI": "...", "负责人": "...", "验收标准": "..."}],',
        '    "P1重要行动": [{"优先级":"P1", "行动项": "...", "目标对象": "...", "预期指标变化": "...", "预估影响": "...", "投入估算": "...", "预期ROI": "...", "负责人": "...", "验收标准": "..."}],',
        '    "标杆案例": [{"排名":1, "大区":"...", "经销商":"...", "创建":0, "渲染":0, "提审":0, "提审率":"...", "成功模式":"...", "可复制要点":"..."}],',
        '    "问题清单": [{"类型":"沉睡", "大区":"...", "经销商":"...", "创建":0, "渲染":0, "提审":0, "转化率":"0%", "问题诊断":"...", "跟进建议":"..."}]',
        '  }',
        '}',
        '',
        '## 整体数据',
        f'- 经销商总数：{total_count} 家（活跃 {active_count} 家）',
        f'- 创建方案数：{total_create}',
        f'- 渲染方案数：{total_render}',
        f'- 提审方案数：{total_submit}',
        f'- 渲染率：{total_render/total_create*100:.1f}%',
        f'- 提审率（占创建）：{total_submit/total_create*100:.1f}%',
        f'- 提审率（占渲染）：{total_submit/total_render*100:.1f}%' if total_render > 0 else '- 提审率（占渲染）：N/A',
        '',
        '## 大区数据（按效率指数降序）',
    ]

    region_sorted = region_stats.sort_values('效率指数', ascending=False)
    for region, row in region_sorted.iterrows():
        lines.append(
            f'- {region}: 经销商{int(row["经销商数"])}家，活跃{int(row["活跃经销商数"])}家（活跃度{row["活跃度"]:.1f}%），'
            f'创建{int(row["创建总和"])}，渲染{int(row["渲染总和"])}，提审{int(row["提审总和"])}，'
            f'渲染率{row["渲染率"]:.1f}%，提审率{row["提审率(创建)"]:.1f}%，效率指数{row["效率指数"]}'
        )

    lines.extend(['', '## 经销商分层'])
    for layer in ['高转化', '中转化', '低转化', '有渲染无提审', '无深化', '零活跃']:
        count = len(df[df['转化分层'] == layer])
        lines.append(f'- {layer}: {count} 家')

    lines.extend(['', '## TOP5 提审经销商'])
    for _, r in df.nlargest(5, '提审方案数').iterrows():
        lines.append(
            f'- {r["大区"]} {r["经销商"]}: 创建{r["创建方案数"]}，渲染{r["渲染方案数"]}，'
            f'提审{r["提审方案数"]}，提审率{r["提审率_创建"]:.1f}%'
        )

    lines.extend(['', '## 沉睡经销商（创建≥20 且 提审=0）'])
    sleeping = df[(df['创建方案数'] >= 20) & (df['提审方案数'] == 0)].sort_values('创建方案数', ascending=False)
    if len(sleeping) == 0:
        lines.append('- 无')
    else:
        for _, r in sleeping.head(10).iterrows():
            lines.append(
                f'- {r["大区"]} {r["经销商"]}: 创建{r["创建方案数"]}，渲染{r["渲染方案数"]}，提审{r["提审方案数"]}'
            )

    lines.extend(['', '请确保分析具体、数据驱动、建议可执行。'])
    return '\n'.join(lines)


def generate_ai_insights(df, region_stats) -> dict | None:
    """调用 DeepSeek 生成 AI 洞察，失败返回 None。"""
    prompt = build_ai_prompt(df, region_stats)
    ai_text = call_deepseek(prompt)
    if not ai_text:
        print('DeepSeek 返回为空')
        return None
    parsed = _safe_json_loads(ai_text)
    if parsed is None:
        print('DeepSeek 返回 JSON 解析失败，原始内容前500字:', ai_text[:500])
    return parsed


def add_analysis_sheets(wb, df, use_ai=True):
    """Append 4 analysis sheets to the existing workbook.
    当 use_ai=True 且配置了 AI_API_KEY 时，优先使用 DeepSeek 生成诊断/建议文本。
    """
    region_stats = build_region_stats(df)
    total_create = int(df['创建方案数'].sum())
    total_render = int(df['渲染方案数'].sum())
    total_submit = int(df['提审方案数'].sum())
    active_count = int(len(df[df['创建方案数'] > 0]))
    total_count = int(len(df))

    render_rate = total_render / total_create if total_create > 0 else 0
    submit_rate_of_render = total_submit / total_render if total_render > 0 else 0
    submit_rate_of_create = total_submit / total_create if total_create > 0 else 0
    best_rate = region_stats['提审率(创建)'].max()
    potential = int(total_create * best_rate / 100) - total_submit

    # 尝试调用 AI 生成洞察
    ai_insights = generate_ai_insights(df, region_stats) if use_ai else None
    if ai_insights:
        print('DeepSeek AI 洞察已生成并应用')
    else:
        print('DeepSeek AI 洞察未生成，使用本地规则填充')

    # ========== Sheet 1: Data Overview ==========
    ws = wb.create_sheet("数据总览")
    ws.column_dimensions['A'].width = 2
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 16

    # Title
    ws.merge_cells('B2:H2')
    c = ws.cell(2, 2, "经销商数据分析报告")
    c.font = ft_title
    c.alignment = al_center
    ws.row_dimensions[2].height = 40

    ws.merge_cells('B3:H3')
    c = ws.cell(3, 2, f"{df['大区'].nunique()}大区 | {total_count}家经销商 | 从数据洞察到行动指引")
    c.font = ft_subtitle
    c.alignment = al_center

    # Key metrics section
    r = 5
    ws.merge_cells(f'B{r}:H{r}')
    c = ws.cell(r, 2, "关键指标总览")
    c.font = ft_section
    c.fill = fill_dark
    c.alignment = al_left
    for col in range(2, 9):
        ws.cell(r, col).fill = fill_dark
        ws.cell(r, col).border = bd_thin
    ws.row_dimensions[r].height = 32

    cards = [
        ('B', '经销商总数', str(total_count), f"活跃{active_count}家({active_count/total_count*100:.1f}%)"),
        ('C', '创建方案数', f"{total_create:,}", f"月均{total_create/active_count:.1f}个/经销商"),
        ('D', '渲染方案数', f"{total_render:,}", f"渲染率{render_rate*100:.1f}%"),
        ('E', '提审方案数', f"{total_submit:,}", f"提审率{submit_rate_of_create*100:.1f}%"),
        ('F', '提审/渲染', f"{submit_rate_of_render*100:.1f}%", f"流失率{(1-submit_rate_of_render)*100:.1f}%"),
        ('G', '潜在增收', f"+{potential}", f"若全部大区达最佳转化率{best_rate:.1f}%"),
    ]

    for row_idx in [6, 7, 8]:
        for col_letter, label, value, note in cards:
            col_idx = ord(col_letter) - ord('A') + 1
            if row_idx == 6:
                ws.cell(6, col_idx, label).font = ft_card_label
                ws.cell(6, col_idx).alignment = al_center
                ws.cell(6, col_idx).border = bd_thin
            elif row_idx == 7:
                ws.cell(7, col_idx, value).font = ft_card_value
                ws.cell(7, col_idx).alignment = al_center
                ws.cell(7, col_idx).border = bd_thin
            else:
                ws.cell(8, col_idx, note).font = ft_card_note
                ws.cell(8, col_idx).alignment = al_center
                ws.cell(8, col_idx).border = bd_thin
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 38
    ws.row_dimensions[8].height = 20

    # Region comparison
    r = 10
    ws.merge_cells(f'B{r}:H{r}')
    c = ws.cell(r, 2, "大区核心指标对比")
    c.font = ft_section
    c.fill = fill_dark
    c.alignment = al_left
    for col in range(2, 9):
        ws.cell(r, col).fill = fill_dark
        ws.cell(r, col).border = bd_thin
    ws.row_dimensions[r].height = 32

    headers = ['大区', '经销商数', '创建总数', '渲染总数', '提审总数', '渲染率', '提审率(创建)']
    for i, h in enumerate(headers):
        c = ws.cell(11, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws.row_dimensions[11].height = 28

    region_sorted = region_stats.sort_values('效率指数', ascending=False)
    r = 12
    for region, data in region_sorted.iterrows():
        vals = [region, int(data['经销商数']), int(data['创建总和']), int(data['渲染总和']),
                int(data['提审总和']), f"{data['渲染率']:.1f}%", f"{data['提审率(创建)']:.1f}%"]
        fills = [fill_white] * 7
        if data['活跃度'] < 60:
            fills[6] = fill_red
        elif data['活跃度'] >= 90:
            fills[6] = fill_green
        if data['提审率(创建)'] >= 20:
            fills[6] = fill_green
        elif data['提审率(创建)'] < 10:
            fills[6] = fill_red
        if data['渲染率'] < 50:
            fills[5] = fill_orange

        for i, (v, fl) in enumerate(zip(vals, fills)):
            c = ws.cell(r, 2 + i, v)
            c.font = ft_bold if i == 0 else ft_data
            c.alignment = al_center
            c.border = bd_thin
            c.fill = fl
        ws.row_dimensions[r].height = 28
        r += 1

    # Total row
    vals = ['合计', total_count, total_create, total_render, total_submit,
            f"{total_render/total_create*100:.1f}%", f"{total_submit/total_create*100:.1f}%"]
    for i, v in enumerate(vals):
        c = ws.cell(r, 2 + i, v)
        c.font = ft_bold
        c.alignment = al_center
        c.border = bd_thin
        c.fill = fill_light
    ws.row_dimensions[r].height = 28

    # Funnel analysis
    r += 2
    ws.merge_cells(f'B{r}:H{r}')
    c = ws.cell(r, 2, "转化漏斗分析")
    c.font = ft_section
    c.fill = fill_dark
    c.alignment = al_left
    for col in range(2, 9):
        ws.cell(r, col).fill = fill_dark
        ws.cell(r, col).border = bd_thin
    ws.row_dimensions[r].height = 32

    funnel_headers = ['环节', '数量', '转化率', '流失数', '流失率', '诊断']
    for i, h in enumerate(funnel_headers):
        c = ws.cell(r + 1, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws.row_dimensions[r + 1].height = 28

    # 漏斗诊断根据实际流失率动态判断
    if render_rate >= 0.7:
        render_diag = "渲染转化优秀"
    elif render_rate >= 0.5:
        render_diag = "正常水平"
    else:
        render_diag = "⚠ 渲染转化偏低"

    if submit_rate_of_render >= 0.4:
        submit_diag = "提审转化优秀"
    elif submit_rate_of_render >= 0.3:
        submit_diag = "正常水平"
    else:
        submit_diag = "⚠ 最大瓶颈！"

    funnel_rows = [
        ('创建方案', total_create, '100%', '-', '-', '入口'),
        ('渲染方案', total_render, f"{render_rate*100:.1f}%",
         total_create - total_render, f"{(1-render_rate)*100:.1f}%", render_diag),
        ('提审方案', total_submit, f"{submit_rate_of_render*100:.1f}%（占渲染）",
         total_render - total_submit, f"{(1-submit_rate_of_render)*100:.1f}%", submit_diag),
    ]

    for i, (stage, qty, rate, loss, loss_rate, diag) in enumerate(funnel_rows):
        row = r + 2 + i
        vals = [stage, qty, rate, loss, loss_rate, diag]
        for j, v in enumerate(vals):
            c = ws.cell(row, 2 + j, v)
            c.font = ft_bold if j == 0 else ft_data
            c.alignment = al_center
            c.border = bd_thin
            if '最大瓶颈' in str(v):
                c.fill = fill_red
                c.font = ft_insight
        ws.row_dimensions[row].height = 28

    # Insights
    r += 6
    ws.merge_cells(f'B{r}:H{r}')
    c = ws.cell(r, 2, "核心洞察")
    c.font = ft_section
    c.fill = fill_dark
    c.alignment = al_left
    for col in range(2, 9):
        ws.cell(r, col).fill = fill_dark
        ws.cell(r, col).border = bd_thin
    ws.row_dimensions[r].height = 32

    if ai_insights and isinstance(ai_insights.get('核心洞察'), list) and len(ai_insights['核心洞察']) >= 5:
        insights = list(ai_insights['核心洞察'])[:5]
    else:
        insights = [
            f"1. 【漏斗瓶颈】渲染→提审流失{(1-submit_rate_of_render)*100:.1f}%是当前最大痛点，{total_render - total_submit}个效果图未转化为订单",
            f"2. 【增收潜力】若全部大区达最佳转化率({best_rate:.1f}%)，月度提审可+{potential}个",
            f"3. 【经销商分层】高转化{len(df[df['转化分层'] == '高转化'])}家 | 中转化{len(df[df['转化分层'] == '中转化'])}家 | 低转化{len(df[df['转化分层'] == '低转化'])}家 | 有渲染无提审{len(df[df['转化分层'] == '有渲染无提审'])}家 | 无深化{len(df[df['转化分层'] == '无深化'])}家 | 零活跃{len(df[df['转化分层'] == '零活跃'])}家",
            f"4. 【最佳大区】{region_sorted.index[0]}（效率指数{region_sorted.iloc[0]['效率指数']}），最需关注：{region_sorted.index[-1]}",
            f"5. 【标杆可学】TOP3经销商：" + "、".join([f"{r['大区']} {r['经销商']}（{r['提审方案数']}提审/{r['提审率_创建']:.0f}%）" for _, r in df.nlargest(3, '提审方案数').iterrows()]),
        ]

    r += 1
    for insight in insights:
        ws.merge_cells(f'B{r}:H{r}')
        c = ws.cell(r, 2, insight)
        c.font = ft_data
        c.alignment = al_left
        c.border = bd_thin
        ws.row_dimensions[r].height = 28
        r += 1

    # ========== Sheet 2: Region Diagnosis ==========
    ws2 = wb.create_sheet("大区深度诊断")
    ws2.column_dimensions['A'].width = 2
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws2.column_dimensions[col].width = 12
    ws2.column_dimensions['M'].width = 40

    ws2.merge_cells('B2:M2')
    c = ws2.cell(2, 2, "大区深度诊断")
    c.font = ft_title
    c.alignment = al_center
    ws2.row_dimensions[2].height = 40

    ws2.merge_cells('B3:M3')
    c = ws2.cell(3, 2, "基于活跃度、转化率、人均产能的多维诊断")
    c.font = ft_subtitle
    c.alignment = al_center

    headers2 = ['大区', '经销商数', '活跃数', '活跃度', '创建总数', '渲染总数', '提审总数',
                '渲染率', '提审率', '人均创建', '人均提审', '效率指数', 'AI诊断与建议']
    for i, h in enumerate(headers2):
        c = ws2.cell(5, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws2.row_dimensions[5].height = 30

    r = 6
    ai_region_diag = ai_insights.get('大区诊断', {}) if ai_insights else {}
    for region, data in region_sorted.iterrows():
        diagnosis = ai_region_diag.get(region, data['AI诊断'])
        vals = [region, int(data['经销商数']), int(data['活跃经销商数']), f"{data['活跃度']:.1f}%",
                int(data['创建总和']), int(data['渲染总和']), int(data['提审总和']),
                f"{data['渲染率']:.1f}%", f"{data['提审率(创建)']:.1f}%",
                data['人均创建'], data['人均提审'], data['效率指数'], diagnosis]
        for i, v in enumerate(vals):
            c = ws2.cell(r, 2 + i, v)
            c.font = ft_bold if i == 0 else ft_data
            c.alignment = al_center if i < 12 else al_left
            c.border = bd_thin
            if i == 3 and data['活跃度'] < 60:
                c.fill = fill_red
            elif i == 3 and data['活跃度'] >= 90:
                c.fill = fill_green
            elif i == 8 and data['提审率(创建)'] >= 20:
                c.fill = fill_green
            elif i == 8 and data['提审率(创建)'] < 10:
                c.fill = fill_red
            elif i == 9 and data['人均创建'] >= 20:
                c.fill = fill_yellow
        ws2.row_dimensions[r].height = 32
        r += 1

    # ========== Sheet 3: Dealer Detail ==========
    ws3 = wb.create_sheet("经销商明细")
    ws3.column_dimensions['A'].width = 2
    ws3.column_dimensions['B'].width = 8
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 28
    for col in ['E', 'F', 'G']:
        ws3.column_dimensions[col].width = 12
    ws3.column_dimensions['H'].width = 10
    ws3.column_dimensions['I'].width = 12
    ws3.column_dimensions['J'].width = 14
    ws3.column_dimensions['K'].width = 32

    ws3.merge_cells('B2:K2')
    c = ws3.cell(2, 2, f"经销商全量明细（{total_count}家）")
    c.font = ft_title
    c.alignment = al_center
    ws3.row_dimensions[2].height = 40

    ws3.merge_cells('B3:K3')
    c = ws3.cell(3, 2, "包含转化分层标签和个性化AI诊断建议，可按大区或分层筛选")
    c.font = ft_subtitle
    c.alignment = al_center

    headers3 = ['序号', '大区', '经销商', '创建方案数', '渲染方案数', '提审方案数',
                '渲染率', '提审率(创建)', '转化分层', 'AI诊断建议']
    for i, h in enumerate(headers3):
        c = ws3.cell(5, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws3.row_dimensions[5].height = 30

    layer_colors = {
        '高转化': fill_green,
        '中转化': fill_vlight,
        '低转化': fill_orange,
        '有渲染无提审': fill_red,
        '无深化': fill_gray,
        '零活跃': PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    }

    def get_diagnosis(row):
        if row['转化分层'] == '高转化':
            return "标杆经销商，邀请分享经验"
        elif row['转化分层'] == '中转化':
            return "重点跟进，推动向高转化跃迁"
        elif row['转化分层'] == '低转化':
            return "排查报价/设计/客户决策卡点"
        elif row['转化分层'] == '有渲染无提审':
            if row['创建方案数'] >= 20:
                return "沉睡经销商！创建多但无提审，紧急复盘"
            elif row['渲染方案数'] >= 5:
                return "有渲染无提审，卡在报价/客户确认环节"
            else:
                return "低渲染无提审，推动深化出图"
        elif row['转化分层'] == '无深化':
            return "推动使用渲染工具，培训效果图技能"
        else:
            return "零活跃！确认是否停止合作"

    df_sorted = df.sort_values(['大区', '创建方案数'], ascending=[True, False])
    r = 6
    for idx, (_, d) in enumerate(df_sorted.iterrows(), 1):
        diagnosis = get_diagnosis(d)
        vals = [idx, d['大区'], d['经销商'], d['创建方案数'], d['渲染方案数'], d['提审方案数'],
                f"{d['渲染率']:.0f}%", f"{d['提审率_创建']:.0f}%", d['转化分层'], diagnosis]
        fl = layer_colors.get(d['转化分层'], fill_white)
        for i, v in enumerate(vals):
            c = ws3.cell(r, 2 + i, v)
            c.font = ft_bold if i in [1, 8] else ft_data
            c.alignment = al_center if i < 8 else al_left
            c.border = bd_thin
            c.fill = fl
        if d['提审率_创建'] >= 30:
            ws3.cell(r, 9).font = ft_positive
        elif d['提审率_创建'] == 0 and d['创建方案数'] > 0:
            ws3.cell(r, 9).font = ft_insight
        ws3.row_dimensions[r].height = 24
        r += 1

    # 为经销商明细表头添加筛选
    last_data_row = r - 1
    ws3.auto_filter.ref = f'B5:K{last_data_row}'

    # ========== Sheet 4: Action Recommendations ==========
    ws4 = wb.create_sheet("行动建议")
    ws4.column_dimensions['A'].width = 2
    ws4.column_dimensions['B'].width = 8
    ws4.column_dimensions['C'].width = 30
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 18
    ws4.column_dimensions['F'].width = 16
    ws4.column_dimensions['G'].width = 14
    ws4.column_dimensions['H'].width = 22
    ws4.column_dimensions['I'].width = 14
    ws4.column_dimensions['J'].width = 26

    ws4.merge_cells('B2:J2')
    c = ws4.cell(2, 2, "行动清单与影响预估")
    c.font = ft_title
    c.alignment = al_center
    ws4.row_dimensions[2].height = 40

    ws4.merge_cells('B3:J3')
    c = ws4.cell(3, 2, "每项行动都有明确的预期影响、投入估算和验收标准")
    c.font = ft_subtitle
    c.alignment = al_center

    # Identify sleeping dealers and zero-active regions
    sleeping = df[(df['创建方案数'] >= 20) & (df['提审方案数'] == 0)].sort_values('创建方案数', ascending=False)
    worst_region = region_sorted.index[-1]
    best_dealers = df.nlargest(3, '提审方案数')

    # P0 Actions
    r = 5
    ws4.merge_cells(f'B{r}:J{r}')
    c = ws4.cell(r, 2, "P0 紧急行动（本周启动）")
    c.font = ft_section
    c.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    c.alignment = al_left
    for col in range(2, 11):
        ws4.cell(r, col).fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
        ws4.cell(r, col).border = bd_thin
    ws4.row_dimensions[r].height = 32

    action_headers = ['优先级', '行动项', '目标对象', '预期指标变化', '预估影响', '投入估算', '预期ROI', '负责人', '验收标准']
    for i, h in enumerate(action_headers):
        c = ws4.cell(6, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws4.row_dimensions[6].height = 28

    p0_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    target_submit_rate = min(submit_rate_of_render * 100 + 15, 45)
    worst_region_zero = len(df[(df['大区'] == worst_region) & (df['创建方案数'] == 0)])
    sleeping_target = min(20, max(10, submit_rate_of_render * 100))
    sleeping_impact = max(1, int(len(sleeping) * sleeping_target / 100)) if len(sleeping) > 0 else 0
    zero_active_count = len(df[df['创建方案数'] == 0])

    # 本地默认行动建议
    local_p0_actions = [
        ('P0', '建立"渲染→提审"周跟进机制', f"全体{active_count}家活跃经销商",
         f"提审率{submit_rate_of_render*100:.1f}%→{target_submit_rate:.0f}%", f"提审+{int((total_render - total_submit) * (target_submit_rate/100 - submit_rate_of_render))}个/月", '运营专员1名', '见效后估算', '运营总监', f'周跟进表+提审率提升{target_submit_rate-submit_rate_of_render*100:.0f}%'),
        ('P0', f"{worst_region}零活跃经销商逐一拜访", f"{worst_region}零活跃{worst_region_zero}家",
         f"激活率{(active_count/total_count*100):.0f}%→{min(90, active_count/total_count*100+10):.0f}%", '创建量增长', '大区经理差旅', '见效后估算', f'{worst_region}区总', f'拜访记录+{max(1, worst_region_zero//3)}家恢复活跃'),
    ]
    if len(sleeping) > 0:
        local_p0_actions.append(
            ('P0', f"{len(sleeping)}家沉睡经销商专项复盘", f"沉睡{len(sleeping)}家",
             f'提审率0%→{sleeping_target:.0f}%', f"提审+{sleeping_impact}个/月", '销售支持', '见效后估算', '销售支持', '复盘报告+产生提审')
        )

    local_p1_actions = [
        ('P1', '推广标杆经销商提审经验', '全体大区',
         f'提审率+{min(5, max(2, (best_rate - submit_rate_of_create*100) / 2)):.0f}%', f"提审+{int(total_render * min(5, max(2, (best_rate - submit_rate_of_create*100) / 2)) / 100)}个/月", '培训部', '见效后估算', '培训经理', '最佳实践手册+分享会'),
        ('P1', '零活跃经销商批量激活', f"{zero_active_count}家零活跃",
         f'激活率{(active_count+zero_active_count*0.3)/total_count*100:.0f}%', '创建量增长', '运营部+大区', '见效后估算', '运营总监', f'激活方案+{max(1, zero_active_count//5)}家恢复活跃'),
    ]

    # 如果 AI 返回了行动建议，优先使用
    ai_actions = ai_insights.get('行动建议', {}) if ai_insights else {}
    p0_actions = []
    for item in ai_actions.get('P0紧急行动', []):
        p0_actions.append((
            item.get('优先级', 'P0'),
            item.get('行动项', ''),
            item.get('目标对象', ''),
            item.get('预期指标变化', item.get('预期效果', '')),
            item.get('预估影响', ''),
            item.get('投入估算', ''),
            item.get('预期ROI', ''),
            item.get('负责人', ''),
            item.get('验收标准', ''),
        ))
    if not p0_actions:
        p0_actions = local_p0_actions

    r = 7
    for action in p0_actions:
        for i, v in enumerate(action):
            c = ws4.cell(r, 2 + i, v)
            c.font = ft_data
            c.alignment = al_center if i in [0, 4, 5, 6, 7] else al_left
            c.border = bd_thin
            c.fill = p0_fill
        ws4.cell(r, 2).font = ft_insight
        ws4.row_dimensions[r].height = 32
        r += 1

    # P1 Actions
    r += 1
    ws4.merge_cells(f'B{r}:J{r}')
    c = ws4.cell(r, 2, "P1 重要行动（本月启动）")
    c.font = ft_section
    c.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    c.alignment = al_left
    for col in range(2, 11):
        ws4.cell(r, col).fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        ws4.cell(r, col).border = bd_thin
    ws4.row_dimensions[r].height = 32

    r += 1
    for i, h in enumerate(action_headers):
        c = ws4.cell(r, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws4.row_dimensions[r].height = 28
    r += 1

    p1_actions = []
    for item in ai_actions.get('P1重要行动', []):
        p1_actions.append((
            item.get('优先级', 'P1'),
            item.get('行动项', ''),
            item.get('目标对象', ''),
            item.get('预期指标变化', item.get('预期效果', '')),
            item.get('预估影响', ''),
            item.get('投入估算', ''),
            item.get('预期ROI', ''),
            item.get('负责人', ''),
            item.get('验收标准', ''),
        ))
    if not p1_actions:
        p1_actions = local_p1_actions

    for action in p1_actions:
        for i, v in enumerate(action):
            c = ws4.cell(r, 2 + i, v)
            c.font = ft_data
            c.alignment = al_center if i in [0, 4, 5, 6, 7] else al_left
            c.border = bd_thin
            c.fill = fill_orange
        ws4.cell(r, 2).font = Font(name='微软雅黑', size=11, bold=True, color=ORANGE)
        ws4.row_dimensions[r].height = 32
        r += 1

    # Star dealers section
    r += 1
    ws4.merge_cells(f'B{r}:J{r}')
    c = ws4.cell(r, 2, "标杆经销商案例（可复制的成功经验）")
    c.font = ft_section
    c.fill = fill_dark
    c.alignment = al_left
    for col in range(2, 11):
        ws4.cell(r, col).fill = fill_dark
        ws4.cell(r, col).border = bd_thin
    ws4.row_dimensions[r].height = 32

    star_headers = ['排名', '大区', '经销商', '创建', '渲染', '提审', '提审率', '成功模式', '可复制要点']
    for i, h in enumerate(star_headers):
        c = ws4.cell(r + 1, 2 + i, h)
        c.font = ft_header
        c.fill = fill_dark
        c.alignment = al_center
        c.border = bd_thin
    ws4.row_dimensions[r + 1].height = 28

    # AI 标杆案例或本地规则
    ai_star_cases = ai_actions.get('标杆案例', [])
    r += 2
    if ai_star_cases:
        for rank, item in enumerate(ai_star_cases[:5], 1):
            vals = [
                rank,
                item.get('大区', ''),
                item.get('经销商', ''),
                item.get('创建', 0),
                item.get('渲染', 0),
                item.get('提审', 0),
                item.get('提审率', ''),
                item.get('成功模式', ''),
                item.get('可复制要点', ''),
            ]
            for i, v in enumerate(vals):
                c = ws4.cell(r, 2 + i, v)
                c.font = ft_bold if i in [1, 2] else ft_data
                c.alignment = al_center if i < 7 else al_left
                c.border = bd_thin
                c.fill = fill_green if rank <= 3 else fill_white
            ws4.row_dimensions[r].height = 28
            r += 1
    else:
        star_cases = []
        for _, d in best_dealers.iterrows():
            if d['创建方案数'] >= 30 and d['提审率_创建'] >= 20:
                mode = "规模+转化双优"
                key = "高频创建+每张渲染都跟报价"
            elif d['提审率_创建'] >= 50:
                mode = "精准提审型"
                key = "精准方案+深度服务+高复购"
            elif d['渲染率'] >= 60:
                mode = "渲染驱动型"
                key = "重视效果图呈现+跟进闭环"
            else:
                mode = "高产出型"
                key = "大量方案覆盖+标准化流程"
            star_cases.append((d, mode, key))

        for rank, (dealer, mode, key) in enumerate(star_cases, 1):
            vals = [rank, dealer['大区'], dealer['经销商'], dealer['创建方案数'], dealer['渲染方案数'], dealer['提审方案数'],
                    f"{dealer['提审率_创建']:.0f}%", mode, key]
            for i, v in enumerate(vals):
                c = ws4.cell(r, 2 + i, v)
                c.font = ft_bold if i in [1, 2] else ft_data
                c.alignment = al_center if i < 7 else al_left
                c.border = bd_thin
                c.fill = fill_green if rank <= 3 else fill_white
            ws4.row_dimensions[r].height = 28
            r += 1

    # Problem dealers
    if len(sleeping) > 0:
        r += 1
        ws4.merge_cells(f'B{r}:J{r}')
        c = ws4.cell(r, 2, "问题清单（需紧急关注的经销商）")
        c.font = ft_section
        c.fill = fill_dark
        c.alignment = al_left
        for col in range(2, 11):
            ws4.cell(r, col).fill = fill_dark
            ws4.cell(r, col).border = bd_thin
        ws4.row_dimensions[r].height = 32

        prob_headers = ['类型', '大区', '经销商', '创建', '渲染', '提审', '转化率', '问题诊断', '跟进建议']
        for i, h in enumerate(prob_headers):
            c = ws4.cell(r + 1, 2 + i, h)
            c.font = ft_header
            c.fill = fill_dark
            c.alignment = al_center
            c.border = bd_thin
        ws4.row_dimensions[r + 1].height = 28

        r += 2
        ai_problem_list = ai_actions.get('问题清单', [])
        if ai_problem_list:
            for item in ai_problem_list[:5]:
                vals = [
                    item.get('类型', '沉睡'),
                    item.get('大区', ''),
                    item.get('经销商', ''),
                    item.get('创建', 0),
                    item.get('渲染', 0),
                    item.get('提审', 0),
                    item.get('转化率', '0%'),
                    item.get('问题诊断', ''),
                    item.get('跟进建议', ''),
                ]
                for i, v in enumerate(vals):
                    c = ws4.cell(r, 2 + i, v)
                    c.font = ft_data
                    c.alignment = al_center if i < 7 else al_left
                    c.border = bd_thin
                    c.fill = fill_orange
                ws4.row_dimensions[r].height = 28
                r += 1
        else:
            for _, d in sleeping.head(5).iterrows():
                if d['渲染方案数'] == 0:
                    diag = "无渲染：可能客户尚未确认方案"
                    advice = "推动渲染技能培训，检查工具使用"
                elif d['渲染方案数'] < 5:
                    diag = "低渲染：效果图产出不足"
                    advice = "推动效果图产出，培训渲染技能"
                else:
                    diag = "有渲染：卡在报价/客户决策环节"
                    advice = "排查报价/客户确认流程卡点"

                vals = ['沉睡', d['大区'], d['经销商'], d['创建方案数'], d['渲染方案数'], d['提审方案数'],
                        '0%', diag, advice]
                for i, v in enumerate(vals):
                    c = ws4.cell(r, 2 + i, v)
                    c.font = ft_data
                    c.alignment = al_center if i < 7 else al_left
                    c.border = bd_thin
                    c.fill = fill_orange
                ws4.row_dimensions[r].height = 28
                r += 1

    # Summary
    r += 1
    ws4.merge_cells(f'B{r}:J{r}')
    potential_from_followup = int((total_render - total_submit) * (target_submit_rate/100 - submit_rate_of_render))
    potential_from_sleeping = sleeping_impact
    potential_from_zero = max(1, zero_active_count // 5)
    total_potential = potential_from_followup + potential_from_sleeping + potential_from_zero
    c = ws4.cell(r, 2, f"汇总估计：若P0+P1行动全部落地，预期月度提审增长 +{total_potential}个")
    c.font = ft_positive
    c.fill = fill_green
    c.alignment = al_center
    for col in range(2, 11):
        ws4.cell(r, col).fill = fill_green
        ws4.cell(r, col).border = bd_thin
    ws4.row_dimensions[r].height = 32

    print("数据分析 sheets 已生成：数据总览、大区深度诊断、经销商明细、行动建议")
