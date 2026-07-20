import sys
from collections import defaultdict
from datetime import time
from openpyxl import load_workbook

sys.path.insert(0, '/app/src')
from generate_attendance import parse_input, _format_dist_time


def expected_main_value(day_data):
    times = sorted(set(day_data['all_times']))
    if not times:
        return None
    if len(times) == 1:
        return f"{_format_dist_time(times[0])}\n     "
    return f"{_format_dist_time(times[0])}\n{_format_dist_time(times[1])}"


def expected_extra_value(day_data):
    times = sorted(set(day_data['all_times']))
    if len(times) == 3:
        return f"{_format_dist_time(times[2])}\n     "
    if len(times) >= 4:
        return f"{_format_dist_time(times[2])}\n{_format_dist_time(times[3])}"
    return None


def load_generated_dist(path, sheet_name=None, max_days=31):
    wb = load_workbook(path, data_only=False)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    main_rows = defaultdict(dict)
    extra_rows = defaultdict(dict)
    # 3 月模板数据从第 2 行开始；基模板从第 3 行开始
    start_row = 2 if '3月' in sheet_name else 3
    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or name == '姓名':
            continue
        if name in main_rows:
            target = extra_rows[name]
        else:
            target = main_rows[name]
        for day in range(1, max_days + 1):
            col = 6 + (day - 1)
            val = ws.cell(row=r, column=col).value
            if val:
                target[day] = str(val)
    wb.close()
    return dict(main_rows), dict(extra_rows)


def main():
    input_path = sys.argv[1]
    gen_path = sys.argv[2]
    max_days = int(sys.argv[3]) if len(sys.argv) > 3 else 31
    employees = parse_input(input_path)
    main_rows, extra_rows = load_generated_dist(gen_path, max_days=max_days)

    diffs = []
    for name, emp in employees.items():
        if name not in main_rows:
            diffs.append((name, 'missing in generated main row'))
            continue
        for day, day_data in emp['days'].items():
            if day > max_days:
                continue
            expected_main = expected_main_value(day_data)
            actual_main = main_rows[name].get(day)
            if expected_main != actual_main:
                diffs.append((name, day, 'main', expected_main, actual_main))
            expected_extra = expected_extra_value(day_data)
            actual_extra = extra_rows.get(name, {}).get(day)
            if expected_extra != actual_extra:
                diffs.append((name, day, 'extra', expected_extra, actual_extra))

    for name in main_rows:
        if name not in employees:
            diffs.append((name, 'extra in generated'))

    print(f'Total diffs vs input: {len(diffs)}')
    for d in diffs[:30]:
        print(d)
    if not diffs:
        print('OK: generated distribution matches input')


if __name__ == '__main__':
    main()
