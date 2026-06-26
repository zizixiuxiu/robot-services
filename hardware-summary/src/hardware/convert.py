import re
import os
from collections import OrderedDict
from copy import copy
import pandas as pd
import openpyxl
import xlrd


def _get_merged_value(sheet, row_idx, col_idx):
    """获取单元格值，自动处理合并单元格。
    
    如果单元格本身有值，直接返回；
    如果单元格在合并区域内且为空，返回合并区域左上角的值。
    """
    val = sheet.cell(row_idx, col_idx).value
    if val != '' and val is not None:
        return val
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if clo <= col_idx < chi and rlo <= row_idx < rhi:
            top = sheet.cell(rlo, clo).value
            return top if top != '' and top is not None else None
    return None


def _find_value_by_label(sheet, label, max_offset=3, exclude_labels=None):
    """Find a value in sheet by searching for a label cell.
    
    Scans the sheet for a cell containing the exact label text,
    then searches to the right up to max_offset columns for a non-empty value.
    """
    exclude_set = set(exclude_labels or [])
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_val = _get_merged_value(sheet, row_idx, col_idx)
            if cell_val is not None and str(cell_val).strip() == label:
                for offset in range(1, max_offset + 1):
                    target_col = col_idx + offset
                    if target_col < sheet.ncols:
                        target_val = _get_merged_value(sheet, row_idx, target_col)
                        if target_val is not None:
                            val_str = str(target_val)
                            if val_str.strip() and val_str.strip() != label and val_str.strip() not in exclude_set:
                                return val_str
    return None


def extract_header_info(source_file):
    """Extract header information from source file.
    
    Returns dict with:
    - customer_name: 经销商姓名
    - customer_address: 经销商地址  
    - order_number: 订单编号 (base part)
    - summary_date: 汇总日期
    - delivery_date: 预计交货日期
    - brand: 品牌
    """
    book = xlrd.open_workbook(source_file, formatting_info=True)
    
    # Try to extract info from all sheets
    customer_name = None
    customer_address = None
    order_number_full = None
    brand = None
    summary_date = None
    delivery_date = None
    
    # First pass: search all sheets for basic info
    for sheet_name in book.sheet_names():
        sheet = book.sheet_by_name(sheet_name)
        
        if not customer_name:
            customer_name = _find_value_by_label(sheet, '经销商姓名')
        if not customer_address:
            customer_address = _find_value_by_label(sheet, '经销商地址')
        if not order_number_full:
            order_number_full = _find_value_by_label(sheet, '订单编号')
        if not brand:
            brand = _find_value_by_label(sheet, '品牌：', exclude_labels=['经销商姓名', '经销商地址', '订单编号', '接单日期'])
    
    # Order number: strip suffix like -22 from B2604-4125-22
    if order_number_full:
        order_number = re.sub(r'-\d+$', '', order_number_full)
    else:
        order_number = ''
    
    # Second pass: search for dates
    # Prioritize 五金 sheets for dates (they often have the correct summary dates)
    wu_sheets = [s for s in book.sheet_names() if '五' in s]
    other_sheets = [s for s in book.sheet_names() if '五' not in s]
    
    for sheet_list in [wu_sheets, other_sheets]:
        for sheet_name in sheet_list:
            sheet = book.sheet_by_name(sheet_name)
            
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell_val = _get_merged_value(sheet, row_idx, col_idx)
                    if cell_val is not None:
                        cell_str = str(cell_val).strip()
                        
                        # Look for order date (下单日期)
                        if '下单日期' in cell_str and not summary_date:
                            for c in range(col_idx + 1, min(col_idx + 8, sheet.ncols)):
                                val = _get_merged_value(sheet, row_idx, c)
                                if val is not None:
                                    val_str = str(val).strip()
                                    match = re.match(r'(\d{2,4})\.(\d{1,2})\.(\d{1,2})$', val_str)
                                    if match:
                                        year, month, day = match.groups()
                                        if len(year) == 2:
                                            year = '20' + year
                                        summary_date = f'{year}.{int(month)}.{int(day)}'
                                        break
                        
                        # Look for delivery date (包装预计交货日期 or 预计交货期)
                        if ('预计交货日期' in cell_str or '预计交货期' in cell_str) and not delivery_date:
                            for c in range(col_idx + 1, min(col_idx + 8, sheet.ncols)):
                                val = _get_merged_value(sheet, row_idx, c)
                                if val is not None:
                                    val_str = str(val).strip()
                                    match = re.match(r'(\d{2,4})\.(\d{1,2})\.(\d{1,2})$', val_str)
                                    if match:
                                        year, month, day = match.groups()
                                        if len(year) == 2:
                                            year = '20' + year
                                        delivery_date = f'{year}.{int(month)}.{int(day)}'
                                        break
            
            if summary_date and delivery_date:
                break
        if summary_date and delivery_date:
            break
    
    # Clean brand: take first word, preserve leading space (e.g., " 逸品 不贴商标..." -> " 逸品")
    if brand:
        first_word = brand.split()[0]
        if brand.startswith(' '):
            brand = ' ' + first_word
        else:
            brand = first_word
    else:
        brand = '奢匠'
    
    return {
        'customer_name': customer_name or '',
        'customer_address': customer_address or '',
        'order_number': order_number,
        'summary_date': summary_date or '',
        'delivery_date': delivery_date or '',
        'brand': brand,
    }


def _is_header_row(name):
    """Check if a row is a header row."""
    return str(name).strip() in ['五金名称', '序号']


def _is_summary_or_meta_row(name):
    """Check if a row is a summary, total, or metadata row."""
    skip_keywords = ['汇总', '总计', '数量总计', '此单共', '小计', '总合计', '数量']
    name_str = str(name).strip()
    return any(kw in name_str for kw in skip_keywords)


def _extract_data_from_hardware_sheet(sheet):
    """Extract hardware data rows from a single 五金 sheet."""
    items = []
    
    for row_idx in range(sheet.nrows):
        name = _get_merged_value(sheet, row_idx, 1)
        if name is None:
            continue
        
        name_str = str(name).strip()
        if _is_header_row(name_str) or _is_summary_or_meta_row(name_str):
            continue
        
        quantity = _get_merged_value(sheet, row_idx, 7)
        if quantity is None:
            continue
        
        try:
            qty_val = float(quantity)
            if qty_val <= 0:
                continue
        except (ValueError, TypeError):
            continue
        
        unit = _get_merged_value(sheet, row_idx, 9)
        length = _get_merged_value(sheet, row_idx, 11)
        width = _get_merged_value(sheet, row_idx, 14)
        remark = _get_merged_value(sheet, row_idx, 18)
        remark2 = _get_merged_value(sheet, row_idx, 19)
        
        items.append({
            'source_sheet': sheet.name,
            'hardware_name': name_str,
            'quantity': qty_val,
            'unit': str(unit).strip() if unit is not None else '',
            'length': length if length is not None else '',
            'width': width if width is not None else '',
            'remark': str(remark).strip() if remark is not None else '',
            'remark2': str(remark2).strip() if remark2 is not None else '',
        })
    
    return items


def extract_hardware_data(source_file):
    """Extract hardware data from all 五金 sheets."""
    book = xlrd.open_workbook(source_file, formatting_info=True)
    all_items = []
    
    for sheet_name in book.sheet_names():
        if '五' not in sheet_name:
            continue
        
        sheet = book.sheet_by_name(sheet_name)
        items = _extract_data_from_hardware_sheet(sheet)
        all_items.extend(items)
    
    return all_items


def _extract_special_products(source_file):
    """Extract data from 特殊产品表 and 外购表 sheets."""
    book = xlrd.open_workbook(source_file, formatting_info=True)
    
    # Find all matching sheets (特殊产品表 or 外购表*), skip hidden sheets
    special_sheet_names = []
    for i, name in enumerate(book.sheet_names()):
        sheet = book.sheet_by_index(i)
        if hasattr(sheet, 'visibility') and sheet.visibility != 0:
            continue  # Skip hidden sheets
        if name == '特殊产品表' or name.startswith('外购表'):
            special_sheet_names.append(name)
    
    if not special_sheet_names:
        return []
    
    all_items = []
    
    for sheet_name in special_sheet_names:
        sheet = book.sheet_by_name(sheet_name)
        
        for row_idx in range(sheet.nrows):
            if row_idx < 5:
                continue
            
            name = _get_merged_value(sheet, row_idx, 1)
            if name is None or _is_header_row(name) or _is_summary_or_meta_row(name):
                continue
            
            quantity = _get_merged_value(sheet, row_idx, 6)
            unit = _get_merged_value(sheet, row_idx, 7)
            length = _get_merged_value(sheet, row_idx, 8)
            width = _get_merged_value(sheet, row_idx, 9)
            thickness = _get_merged_value(sheet, row_idx, 10)
            remark1 = _get_merged_value(sheet, row_idx, 13)
            remark2 = _get_merged_value(sheet, row_idx, 14)
            
            # 新规则：A列放订单号(remark2)，G列放厚度|remark1
            # remark = 厚度 | remark1(如"窄边框 哑黑+银镜")
            remark_parts = []
            if thickness is not None and str(thickness).strip() and str(thickness).strip() != '0':
                remark_parts.append(f"厚度{thickness}mm")
            if remark1 is not None and str(remark1).strip():
                remark_parts.append(str(remark1).strip())
            remark = ' | '.join(remark_parts) if remark_parts else ''
            
            # A列用remark2（订单号）
            workbook_remark = str(remark2).strip() if remark2 is not None else ''
            
            if quantity is None:
                continue
            try:
                qty_val = float(quantity)
                if qty_val <= 0:
                    continue
            except (ValueError, TypeError):
                continue
            
            all_items.append({
                'source_sheet': sheet_name if sheet_name.startswith('外购表') else '特殊产品',
                'hardware_name': str(name).strip(),
                'quantity': qty_val,
                'unit': str(unit).strip() if unit is not None else '',
                'length': length if length is not None else '',
                'width': width if width is not None else '',
                'remark': remark,
                'remark2': str(remark2).strip() if remark2 is not None else '',
                'workbook_remark': workbook_remark,  # A列用订单号
            })
    
    return all_items


def _get_workbook_name(source_file, sheet_name, is_special=False):
    """Generate workbook name for output.
    
    For special products:
    - If sheet_name is exactly '特殊产品', return empty string (A column is blank)
    - If sheet_name contains '_' (custom workbook name), return it directly
    """
    if is_special:
        if str(sheet_name) == '特殊产品':
            return ''
        if '_' in str(sheet_name):
            return str(sheet_name)
    base_name = os.path.splitext(os.path.basename(source_file))[0]
    # Clean doc_ prefix from base_name
    base_name = re.sub(r'^doc_[a-f0-9]{12}_', '', base_name)
    return f'[{base_name}]{sheet_name}'


def _extract_sheet_number(sheet_name):
    """Extract numeric part from sheet name like '11五金' -> 11, '1五' -> 1."""
    match = re.search(r'(\d+)', str(sheet_name))
    return int(match.group(1)) if match else 0


def generate_summary(hardware_data, source_file):
    """Generate summary rows grouped by hardware name.
    
    Regular hardware items come first (sorted by name), followed by special products (sorted by name).
    Each group is followed by a summary row.
    """
    # Separate regular hardware and special products
    regular_items = [item for item in hardware_data if '五' in str(item['source_sheet'])]
    special_items = [item for item in hardware_data if '五' not in str(item['source_sheet'])]
    
    def _make_groups(items):
        groups = OrderedDict()
        for item in items:
            name = item['hardware_name']
            if name not in groups:
                groups[name] = []
            groups[name].append(item)
        return groups
    
    def _render_groups(groups, is_special=False):
        rows = []
        if is_special:
            # For special products: groups whose items are from "特殊产品" or "外购表" sheets come last
            sorted_groups = sorted(groups.items(), key=lambda x: (
                1 if any(item['source_sheet'] in ('特殊产品', '外购表') or item['source_sheet'].startswith('外购表') for item in x[1]) else 0,
                x[0].replace('_', '\x00'),
            ))
        else:
            sorted_groups = sorted(groups.items(), key=lambda x: x[0].replace('_', '\x00'))
        for name, items in sorted_groups:
            if is_special:
                # For special products: items from "特殊产品" or "外购表" sheets come last
                items_sorted = sorted(items, key=lambda x: (
                    1 if x['source_sheet'] in ('特殊产品', '外购表') or x['source_sheet'].startswith('外购表') else 0,
                    _extract_sheet_number(x['source_sheet']),
                ))
            else:
                items_sorted = sorted(items, key=lambda x: _extract_sheet_number(x['source_sheet']))
            
            for item in items_sorted:
                # 新规则：A列用workbook_remark(订单号，如B2604_4098_05打木架)
                workbook_name = item.get('workbook_remark', '')
                if not workbook_name:
                    workbook_name = _get_workbook_name(source_file, item['source_sheet'], is_special=is_special)
                # 统一8列输出：A工作簿名 B名称 C数量 D单位 E长度 F宽度 G备注 H摘要(空)
                rows.append([
                    workbook_name,
                    item['hardware_name'],
                    item['quantity'],
                    item['unit'],
                    item['length'],
                    item['width'],
                    item.get('remark', ''),
                    '',  # H列摘要留空
                ])
            
            total_qty = sum(item['quantity'] for item in items)
            rows.append([
                '',
                f'{name} 汇总',
                total_qty,
                '',
                '',
                '',
                '',
                '',
            ])
        return rows
    
    regular_groups = _make_groups(regular_items)
    special_groups = _make_groups(special_items)
    
    rows = _render_groups(regular_groups, is_special=False)
    rows.extend(_render_groups(special_groups, is_special=True))
    
    return rows


def _count_cabinets(source_file):
    """Count number of cabinets from 汇总 sheet."""
    book = xlrd.open_workbook(source_file, formatting_info=True)
    if '汇总' not in book.sheet_names():
        return 0
    
    sheet = book.sheet_by_name('汇总')
    count = 0
    for row_idx in range(2, sheet.nrows):
        cabinet_num = _get_merged_value(sheet, row_idx, 0)
        cabinet_body = _get_merged_value(sheet, row_idx, 1)
        
        if cabinet_num is not None and isinstance(cabinet_num, (int, float)):
            if cabinet_num > 0:
                if cabinet_body is not None:
                    try:
                        val = float(cabinet_body)
                        if val > 0:
                            count += 1
                    except (ValueError, TypeError):
                        pass
    
    return count


def _copy_cell_format(src_cell, dst_cell):
    """Copy format from source cell to destination cell."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format


def convert_hardware_summary(source_file, output_file, template_file=None):
    """Convert source XLS to hardware summary XLSX.
    
    If template_file is provided, uses it as a base for formatting.
    """
    # Extract data
    header = extract_header_info(source_file)
    hardware_data = extract_hardware_data(source_file)
    special_products = _extract_special_products(source_file)
    all_data = hardware_data + special_products
    data_rows = generate_summary(all_data, source_file)
    
    total_quantity = sum(
        row[2] for row in data_rows 
        if isinstance(row[2], (int, float)) and '汇总' not in str(row[1])
    )
    
    cabinet_count = _count_cabinets(source_file)
    if cabinet_count > 0 and header['brand'] == '奢匠':
        data_rows.append(['', '奢匠商标', cabinet_count, '', '', '', '', ''])
        total_quantity += cabinet_count
    
    data_rows.append(['', '总计', total_quantity, '', '', '', '', ''])
    
    # Default to format template if available
    default_template = os.path.join(os.path.dirname(__file__), '..', 'template_format.xlsx')
    if template_file and os.path.exists(template_file):
        _write_with_template(source_file, output_file, template_file, header, data_rows)
    elif os.path.exists(default_template):
        _write_with_template(source_file, output_file, default_template, header, data_rows)
    else:
        _write_plain(source_file, output_file, header, data_rows)


def _write_plain(source_file, output_file, header, data_rows):
    """Write output without template (plain data)."""
    output_data = [
        ['五金清单汇总', '', '', '', '', '', '', ''],
        ['客户名称', header['customer_name'], '订单编号', header['order_number'], '', '品牌', header['brand'], ''],
        ['客户地址', header['customer_address'], '汇总日期', header['summary_date'], '', '预计交货日期', header['delivery_date'], ''],
        ['工作簿名', '五金名称', '数量', '单位', '长度（mm）', '宽度（mm）', '备注', '摘要'],
    ]
    output_data.extend(data_rows)
    
    df_output = pd.DataFrame(output_data)
    df_output.to_excel(output_file, index=False, header=False)


# ========== 用户自定义配置区域 ==========
# 修改以下数值来调整行高和列宽
ROW_HEIGHT_HEADER = 35         # 表头往上的行高（第1-4行）
ROW_HEIGHT_DATA = 15           # 普通数据行的行高
ROW_HEIGHT_SUMMARY = 26        # 汇总行的行高
ENABLE_SHRINK_TO_FIT = True    # 是否启用缩小字体填充

# 各列固定列宽（字符宽度）
COL_WIDTHS = {
    'A': 22,   # 工作簿名
    'B': 58,   # 五金名称
    'C': 13,   # 数量
    'D': 9,    # 单位
    'E': 12,   # 长度
    'F': 16,   # 宽度
    'G': 19,   # 备注
    'H': 19,   # 摘要
}
# ======================================


def _write_with_template(source_file, output_file, template_file, header, data_rows):
    """Write output by filling data into a template workbook.
    
    This preserves all formatting from the template including fonts, fills,
    borders, alignments, and merge cells.
    """
    import os
    
    # Load template workbook directly (we'll modify it in place)
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    
    # Find the format sample rows in template
    template_total_row = ws.max_row
    
    # Collect format samples from template
    fmt_samples = {
        'data': {},
        'summary': {},
        'total': {},
    }
    
    for col in range(1, 9):
        fmt_samples['data'][col] = {
            'font': copy(ws.cell(row=5, column=col).font),
            'fill': copy(ws.cell(row=5, column=col).fill),
            'alignment': copy(ws.cell(row=5, column=col).alignment),
            'border': copy(ws.cell(row=5, column=col).border),
            'number_format': ws.cell(row=5, column=col).number_format,
        }
        fmt_samples['summary'][col] = {
            'font': copy(ws.cell(row=6, column=col).font),
            'fill': copy(ws.cell(row=6, column=col).fill),
            'alignment': copy(ws.cell(row=6, column=col).alignment),
            'border': copy(ws.cell(row=6, column=col).border),
            'number_format': ws.cell(row=6, column=col).number_format,
        }
        fmt_samples['total'][col] = {
            'font': copy(ws.cell(row=template_total_row, column=col).font),
            'fill': copy(ws.cell(row=template_total_row, column=col).fill),
            'alignment': copy(ws.cell(row=template_total_row, column=col).alignment),
            'border': copy(ws.cell(row=template_total_row, column=col).border),
            'number_format': ws.cell(row=template_total_row, column=col).number_format,
        }
    
    # Calculate required rows: 4 header + data_rows
    required_rows = 4 + len(data_rows)
    
    # If template doesn't have enough rows, insert new rows before the total row
    if required_rows > template_total_row:
        rows_to_insert = required_rows - template_total_row
        ws.insert_rows(template_total_row, rows_to_insert)
        for new_row in range(template_total_row, template_total_row + rows_to_insert):
            for col in range(1, 9):
                _copy_cell_format(ws.cell(row=5, column=col), ws.cell(row=new_row, column=col))
            if 5 in ws.row_dimensions:
                ws.row_dimensions[new_row].height = ROW_HEIGHT_DATA
        template_total_row = required_rows
    
    # Clear all data rows (rows 5 to template_total_row)
    for row in range(5, template_total_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).value = None
    
    # Remove old merge cells in data area
    merges_to_remove = []
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row >= 5:
            merges_to_remove.append(str(merge))
    for merge_str in merges_to_remove:
        ws.unmerge_cells(merge_str)
    
    # Write header info (rows 2-3)
    ws.cell(row=2, column=2).value = header['customer_name']
    ws.cell(row=2, column=4).value = header['order_number']
    ws.cell(row=2, column=7).value = header['brand']
    ws.cell(row=3, column=2).value = header['customer_address']
    ws.cell(row=3, column=4).value = header['summary_date']
    ws.cell(row=3, column=7).value = header['delivery_date']
    
    # 设置表头区域行高（第1-4行）
    for header_row in range(1, 5):
        ws.row_dimensions[header_row].height = ROW_HEIGHT_HEADER
    
    # 设置各列固定列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    
    # Write data rows starting from row 5
    current_row = 5
    
    for i, row_data in enumerate(data_rows):
        row_type = _get_row_type(row_data)
        
        # Determine if this is a special product row
        is_special = False
        if row_data[0] and not str(row_data[0]).startswith('[') and str(row_data[0]).strip():
            is_special = True
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            
            # Apply format from template
            if row_type == 'data' and is_special:
                src = fmt_samples['data'][col_idx]
            elif row_type == 'data':
                src = fmt_samples['data'][col_idx]
            elif row_type == 'summary':
                src = fmt_samples['summary'][col_idx]
            elif row_type == 'footer':
                src = fmt_samples['data'][col_idx]
            elif row_type == 'total':
                src = fmt_samples['total'][col_idx]
            else:
                src = fmt_samples['data'][col_idx]
            
            cell.font = copy(src['font'])
            cell.fill = copy(src['fill'])
            
            # 设置对齐：自动换行 + 垂直居中
            new_alignment = copy(src['alignment'])
            new_alignment.wrap_text = True           # 开启自动换行
            new_alignment.vertical = 'center'        # 垂直居中
            new_alignment.shrink_to_fit = False      # 关闭缩小字体填充
            cell.alignment = new_alignment
            
            cell.border = copy(src['border'])
            cell.number_format = src['number_format']
        
        # 根据行类型设置不同行高
        if row_type == 'summary':
            ws.row_dimensions[current_row].height = ROW_HEIGHT_SUMMARY
        else:
            ws.row_dimensions[current_row].height = ROW_HEIGHT_DATA
        
        current_row += 1
    
    # Save
    wb.save(output_file)


def _get_row_type(row_data):
    """Determine row type based on content."""
    name = str(row_data[1]) if len(row_data) > 1 and row_data[1] else ''
    
    if name == '总计':
        return 'total'
    elif name == '奢匠商标':
        return 'footer'
    elif '汇总' in name:
        return 'summary'
    else:
        return 'data'
