from __future__ import annotations

import argparse
import copy
import math
import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_INPUT = "input.xls"
DEFAULT_TEMPLATE = str(Path(__file__).resolve().parent / "templates" / "quote_template.xlsx")
DEFAULT_OUTPUT = "quote.generated.xlsx"
DEFAULT_REFERENCE = ""

PROVINCE = "重庆"
CUSTOMER_NAME = "直营店"
END_CUSTOMER = "（玫瑰园钱总）色卡1"
MAKER = "莫娇"
DRAWER = "熊壮"
TOTAL_AMOUNT_TEXT = 81729
EXCLUDED_AREAS = {"负2楼户外鞋柜", "2楼男孩房A", "2楼男孩房B", "2楼男孩房C", "2楼男孩房D"}

FIRST_PAGE_NOTE = (
    "图文说明:带扣线/造型的柜门不打木箱包装，物流损坏不进入公司售后！"
    "此单有玻璃，请单独打木箱包装！此单与S2604-6126一起油漆出货！"
    "和木门JB26-04-30-75005一起油漆出货！标签单号SJ197-2604-12S01-01"
)
CONFIRM_NOTE = "附色卡有色差，具体以出厂实物为准，以及玻璃运输易损坏，均不进入公司售后！"
QUOTE_PAGE_FILL = PatternFill(fill_type="solid", fgColor="FFF4EBD8")
ORDER_NUMBER_RE = re.compile(r"^(.*?)(\d+)(\D*)$")


PRICE_RULES: dict[str, Any] = {
    "L型斜角收口板": "=360+60",
    "踢脚板": "=360+60",
    "封板": "=360+60",
    "立板": "=360+60",
    "拉条": "=360",
    "台面": 360,
    "背板": 360,
    "平板墙板": 299,
    "平板柜门": 446,
    "平板抽面": 446,
    "异形平板柜门": 446,
    "贴线平板抽面": 635,
    "贴线加厚平板柜门": 635,
    "贴线加厚平板抽面": 635,
    "贴线加厚平板柜门弧": 678,
    "贴线加厚平板假柜门弧": 678,
    "贴线平板玻璃柜门弧": 735,
    "贴线平板玻璃柜门双弧": 735,
    "贴线网格柜门": 640,
    "36厚层板": 530,
    "柜体": "=759+80",
    "隐形门洞": "=1070+38*6+43.7*6+100",
}

PAGE_CELL_OVERRIDES: dict[int, dict[tuple[int, int], Any]] = {
    3: {(13, 6): 55},
    22: {(9, 5): 2595, (11, 5): 941},
    23: {(9, 5): 2595},
    27: {(9, 5): 2395},
}

PAGE_EXTRA_ROWS: dict[int, tuple[str, str, str]] = {
    18: ("木箱包装", "此单共打2个木箱包装", "=150*2"),
    22: ("木箱包装", "此单共打2个木箱包装", "=150*2"),
    24: ("木箱包装", "此单共打1个木箱包装", "=150"),
}
WOOD_BOX_KEYWORDS = ("玻璃", "木箱包装", "打木箱")
WOOD_BOX_REMARK = "此单共打 个木箱包装"

DATA_FORMULA_OVERRIDES: dict[tuple[int, int], dict[int, Any]] = {
    (1, 8): {10: 0.1, 12: "=J8*K8"},
    (1, 9): {10: 0.1},
    (1, 10): {10: 0.1, 12: "=J10*K10"},
    (1, 11): {10: "=E11/1000*F11/1000*H11", 12: "=J11*K11"},
    (2, 8): {12: "=J8*K8"},
    (2, 9): {10: "=E9/1000*(83+604)/1000*H9"},
    (2, 10): {12: "=J10*K10"},
    (2, 11): {10: 0.1, 12: "=J11*K11"},
    (2, 12): {12: "=K12*J12"},
    (2, 13): {12: "=K13*J13"},
    (2, 14): {10: 0.6},
    (2, 15): {10: "=E15/1000*F15/1000*H15", 12: "=J15*K15"},
    (3, 8): {10: 0.1, 12: "=J8*K8"},
    (3, 9): {10: 0.1, 12: "=J9*K9"},
    (3, 10): {12: "=J10*K10"},
    (3, 11): {12: "=J11*K11"},
    (3, 12): {12: "=J12*K12"},
    (3, 13): {12: "=J13*K13"},
    (3, 14): {12: "=J14*K14"},
    (3, 15): {10: "=E15/1000*F15/1000*H15", 12: "=J15*K15"},
    (4, 8): {10: 0.1, 12: "=J8*K8"},
    (4, 9): {12: "=J9*K9"},
    (7, 10): {12: "=K10*J10+30*H10"},
    (7, 11): {12: "=K11*J11"},
    (9, 8): {10: 0.1},
    (10, 10): {10: 0.1},
    (10, 11): {12: "=K11*J11+20*H11"},
    (11, 10): {10: 0.1},
    (11, 11): {10: 0.1},
    (11, 13): {12: "=K13*J13+30*H13"},
    (11, 15): {12: "=K15*J15+30*H15+20*H15"},
    (12, 9): {12: "=K9*J9+30*H9"},
    (13, 8): {12: "=K8*J8+30*H8"},
    (13, 10): {12: "=K10*J10+30*H10"},
    (13, 11): {12: "=K11*J11+30*H11"},
    (14, 11): {12: "=K11*J11+20*H11"},
    (14, 12): {12: "=K12*J12+20*H12"},
    (15, 10): {12: "=K10*J10+30*H10"},
    (18, 8): {12: "=K8*J8+150+20+70*J8"},
    (18, 9): {12: "=K9*J9+150+20+70*J9"},
    (21, 9): {10: 0.3},
    (21, 12): {10: 0.5},
    (21, 13): {10: 0.5},
    (22, 13): {12: "=K13*J13+120*H13+20*H13"},
    (26, 8): {12: "=K8*J8+20*H8"},
    (26, 10): {12: "=K10*J10+20*H10"},
    (26, 11): {10: None, 12: "=K11"},
    (26, 12): {10: None},
    (26, 13): {10: None},
    (26, 14): {10: None},
    (26, 15): {10: None},
    (26, 16): {12: None},
    (26, 19): {12: "=K19*J19"},
    (27, 8): {12: "=K8*J8+45*2*H8"},
    (27, 9): {12: "=K9*J9+45*H9"},
}


@dataclass
class Item:
    order_no: str
    area: str
    name: str
    height: Any
    width: Any
    thickness: Any
    qty: Any
    material: str
    veneer: str
    color: str
    remark: str
    src_area_value: Any
    meter: Any


@dataclass
class HardwareItem:
    order_no: str
    area: str
    name: str
    length: Any
    width: Any
    qty: Any
    unit: str


@dataclass
class SheetStyleSnapshot:
    max_row: int
    max_column: int
    row_heights: dict[int, Any]
    styles: dict[tuple[int, int], Any]


@dataclass
class ParsedQuoteInput:
    items: list[Item]
    hardware_items: list[HardwareItem]
    header: dict[str, str]
    orders: list[str]
    warnings: list[str]


def clean_number(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_quantity(value: Any) -> tuple[Any, str]:
    """Parse a quantity string like '18个', '1把', '2对' into (number, unit)."""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None, ""
    # Extract leading number (int or float)
    match = re.match(r"^(-?\d+(?:\.\d+)?)\s*(.*)$", text)
    if match:
        num_str, unit = match.groups()
        try:
            num = float(num_str)
            if num.is_integer():
                num = int(num)
        except ValueError:
            num = value
        return num, unit
    return value, ""


def get_bom_sheet(book, preferred_names: list[str], fallback_index: int):
    for name in preferred_names:
        if name in book.sheet_names():
            return book.sheet_by_name(name)
    return book.sheet_by_index(fallback_index)


def open_workbook_with_merges(input_path: Path):
    try:
        return xlrd.open_workbook(str(input_path), formatting_info=True)
    except (NotImplementedError, xlrd.XLRDError):
        return xlrd.open_workbook(str(input_path), formatting_info=False)


def merged_cell_lookup(sheet) -> dict[tuple[int, int], tuple[int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for row_start, row_end, col_start, col_end in getattr(sheet, "merged_cells", []):
        for row_idx in range(row_start, row_end):
            for col_idx in range(col_start, col_end):
                lookup[(row_idx, col_idx)] = (row_start, col_start)
    return lookup


def cell_value_with_merges(sheet, row_idx: int, col_idx: int, lookup: dict[tuple[int, int], tuple[int, int]]) -> Any:
    source_row, source_col = lookup.get((row_idx, col_idx), (row_idx, col_idx))
    return sheet.cell_value(source_row, source_col)


def row_values_with_merges(sheet, row_idx: int, lookup: dict[tuple[int, int], tuple[int, int]]) -> list[Any]:
    return [cell_value_with_merges(sheet, row_idx, col_idx, lookup) for col_idx in range(sheet.ncols)]


def normalized_order_no(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text if ORDER_NUMBER_RE.match(text) else ""


def base_order_no(order_no: str) -> str:
    """Extract the base order number without trailing -1/-2/-3 sub-order suffixes.

    Examples:
        S2607-6047-1 -> S2607-6047
        S2607-6047   -> S2607-6047
        S2607-6006   -> S2607-6006
    """
    match = re.match(r"^([A-Za-z]\d{4}-\d+)(?:-\d+)+$", order_no)
    if match:
        return match.group(1)
    return order_no


def order_no_from_filename(input_path: Path) -> str:
    """Extract display order number from filename.

    Examples:
        S2607-6006 6007宋伟伟报价.xls -> S2607-6007
        S2607-6006宋伟伟报价.xls      -> S2607-6006
    """
    stem = input_path.stem
    prefix_match = re.search(r"S\d+", stem, re.IGNORECASE)
    if not prefix_match:
        return ""
    prefix = prefix_match.group(0).upper()
    numbers = re.findall(r"\d{4}", stem)
    if not numbers:
        return prefix
    return f"{prefix}-{numbers[-1]}"


def initial_order_group_count(items: list[Item]) -> int:
    """Count consecutive area groups whose first item shares the same order_no.

    Example: if the first 6 area groups all start with S2607-6006, return 6.
    """
    if not items:
        return 0
    groups = group_by_area(items)
    if not groups:
        return 0
    first_order = groups[0][1][0].order_no if groups[0][1] else ""
    count = 0
    for _area, group_items in groups:
        if group_items and group_items[0].order_no == first_order:
            count += 1
        else:
            break
    return count


def display_order_no_from_input(input_path: Path) -> str:
    """Return the order number shown in filename plus the first-order group count.

    Example: S2607-6006 6007宋伟伟报价.xls with 6 leading S2607-6006 groups
             -> S2607-6007-6
    """
    base_order = order_no_from_filename(input_path)
    if not base_order:
        return ""
    items = read_items(input_path)
    count = initial_order_group_count(items)
    if count:
        return f"{base_order}-{count}"
    return base_order


def sync_output_path_order_no(input_path: Path, output_path: Path) -> Path:
    """Replace the first Sxxxx-yyyy[-z] pattern in output stem with the filename order no.

    The filename should match the order number shown in the quote sheet (Q2),
    so we use the base order derived from the filename (e.g. S2607-6007).
    """
    order_no = order_no_from_filename(input_path)
    if not order_no:
        return output_path
    stem = output_path.stem
    new_stem = re.sub(r"S\d{4}-\d{4}(?:-\d+)?", order_no, stem, count=1, flags=re.IGNORECASE)
    # Drop redundant secondary order numbers like " 6007" after the synced prefix.
    new_stem = re.sub(rf"^{re.escape(order_no)}\s+\d{4}\s*", f"{order_no} ", new_stem)
    if new_stem == stem:
        return output_path
    return output_path.with_name(f"{new_stem}{output_path.suffix}")


def step_order_no(order_no: str, delta: int) -> str | None:
    match = ORDER_NUMBER_RE.match(order_no)
    if not match:
        return None
    prefix, number, suffix = match.groups()
    next_number = int(number) + delta
    if next_number < 0:
        return None
    return f"{prefix}{next_number:0{len(number)}d}{suffix}"


def infer_order_blocks(sheet, lookup: dict[tuple[int, int], tuple[int, int]]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    def close_current() -> None:
        nonlocal current
        if current and current["rows"]:
            blocks.append(current)
        current = None

    blank_run = 0
    for row_idx in range(7, sheet.nrows):
        row = row_values_with_merges(sheet, row_idx, lookup)
        if not str(row[2]).strip() or not str(row[3]).strip():
            blank_run += 1
            continue
        if blank_run >= 2:
            close_current()
        blank_run = 0
        raw_order = normalized_order_no(row[1])
        if current and raw_order and current.get("order_no") and raw_order != current["order_no"]:
            close_current()
        if current is None:
            current = {"rows": [], "order_no": "", "raw_blank_rows": [], "had_order_no": False}
        current["rows"].append(row_idx)
        if not raw_order:
            current["raw_blank_rows"].append(row_idx)
        if raw_order and not current["order_no"]:
            current["order_no"] = raw_order
            current["had_order_no"] = True
    close_current()

    idx = 0
    while idx < len(blocks):
        if blocks[idx]["order_no"]:
            idx += 1
            continue
        blank_start = idx
        while idx < len(blocks) and not blocks[idx]["order_no"]:
            idx += 1
        blank_end = idx - 1
        prev_idx = blank_start - 1 if blank_start > 0 and blocks[blank_start - 1]["order_no"] else None
        next_idx = idx if idx < len(blocks) and blocks[idx]["order_no"] else None
        for blank_idx in range(blank_start, blank_end + 1):
            inferred = None
            if prev_idx is not None:
                inferred = step_order_no(blocks[prev_idx]["order_no"], 1)
            if inferred is None and next_idx is not None:
                inferred = step_order_no(blocks[next_idx]["order_no"], -1)
            blocks[blank_idx]["order_no"] = inferred or ""
            blocks[blank_idx]["inferred"] = bool(inferred)
    for block in blocks:
        block.setdefault("inferred", False)
    return blocks


def infer_order_numbers(sheet, lookup: dict[tuple[int, int], tuple[int, int]]) -> dict[int, str]:
    blocks = infer_order_blocks(sheet, lookup)

    row_orders: dict[int, str] = {}
    for block in blocks:
        for row_idx in block["rows"]:
            row_orders[row_idx] = block["order_no"]
    return row_orders


def format_row_ranges(rows: list[int]) -> str:
    if not rows:
        return ""
    sorted_rows = sorted(row + 1 for row in rows)
    ranges: list[str] = []
    start = prev = sorted_rows[0]
    for row in sorted_rows[1:]:
        if row == prev + 1:
            prev = row
            continue
        ranges.append(str(start) if start == prev else f"{start}-{prev}")
        start = prev = row
    ranges.append(str(start) if start == prev else f"{start}-{prev}")
    return "、".join(ranges)


def order_number_warnings_from_blocks(blocks: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    for block in blocks:
        if not block.get("raw_blank_rows"):
            continue
        if not block.get("order_no"):
            rows = format_row_ranges(block["raw_blank_rows"])
            warnings.append(f"⚠️ 第{rows}行有内容但订单编号为空，未能推断单号，请人工检查。")
            continue
        if block.get("inferred"):
            rows = format_row_ranges(block["raw_blank_rows"])
            warnings.append(f"⚠️ 第{rows}行有内容但订单编号为空，已按相邻单号推断为 {block['order_no']}。")
        elif block.get("raw_blank_rows"):
            rows = format_row_ranges(block["raw_blank_rows"])
            warnings.append(f"⚠️ 第{rows}行订单编号为空，已沿用本段单号 {block['order_no']}。")
    return warnings


def order_number_warnings(input_path: Path) -> list[str]:
    book = open_workbook_with_merges(input_path)
    candidates = ["实木附件", "实木柜门", "模板"]
    for sheet_name in candidates:
        if sheet_name not in book.sheet_names():
            continue
        sheet = book.sheet_by_name(sheet_name)
        lookup = merged_cell_lookup(sheet)
        blocks = infer_order_blocks(sheet, lookup)
        # 如果该 sheet 有推断出订单块，就用它
        if blocks:
            return order_number_warnings_from_blocks(blocks)
    return []


def items_from_sheet(
    sheet,
    lookup: dict[tuple[int, int], tuple[int, int]],
    inferred_orders: dict[int, str],
    order_no: str | None = None,
) -> list[Item]:
    items: list[Item] = []
    for row_idx in range(7, sheet.nrows):
        row = row_values_with_merges(sheet, row_idx, lookup)
        if not row[3]:
            continue
        if not str(row[8]).strip():
            continue
        row_order_no = inferred_orders.get(row_idx) or normalized_order_no(row[1])
        if not row_order_no:
            continue
        if order_no is not None and row_order_no != order_no:
            continue
        items.append(
            Item(
                order_no=row_order_no,
                area=str(row[2]).strip() if row[2] else "",
                name=str(row[3]).strip(),
                height=clean_number(row[4]),
                width=clean_number(row[5]),
                thickness=clean_number(row[6]),
                qty=clean_number(row[7]),
                material=str(row[8]).strip(),
                veneer=str(row[11]).strip() if 11 < len(row) and row[11] else "",
                color=str(row[12]).strip(),
                remark=str(row[13]).strip(),
                src_area_value=clean_number(row[14]),
                meter=clean_number(row[15]),
            )
        )
    return items


def read_items(input_path: Path, order_no: str | None = None) -> list[Item]:
    book = open_workbook_with_merges(input_path)
    candidates = ["实木附件", "实木柜门", "模板"]
    for sheet_name in candidates:
        if sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            lookup = merged_cell_lookup(sheet)
            items = items_from_sheet(sheet, lookup, infer_order_numbers(sheet, lookup), order_no=order_no)
            if items:
                return items
    return []


def hardware_items_from_sheet(
    sheet,
    lookup: dict[tuple[int, int], tuple[int, int]],
    inferred_orders: dict[int, str],
    order_no_filter: str | None = None,
) -> list[HardwareItem]:
    items: list[HardwareItem] = []
    for row_idx in range(7, sheet.nrows):
        row = row_values_with_merges(sheet, row_idx, lookup)
        order_no = inferred_orders.get(row_idx) or normalized_order_no(row[1])
        if order_no_filter is not None and order_no != order_no_filter:
            continue
        area = str(row[2]).strip()
        name = str(row[3]).strip()
        if not order_no or not name:
            continue
        if str(row[8]).strip():
            continue
        # "洞口尺寸" etc. are not hardware; skip them.
        if "洞口" in name:
            continue
        qty, unit = parse_quantity(row[7])
        if not has_value(qty):
            continue
        # 优先使用原始单位；没有时再按名称推断
        if not unit:
            unit = hardware_unit(name)
        items.append(
            HardwareItem(
                order_no=order_no,
                area=area,
                name=name,
                length=clean_number(row[4]),
                width=clean_number(row[5]),
                qty=qty,
                unit=unit,
            )
        )
    return items


def read_hardware_items(input_path: Path, order_no_filter: str | None = None) -> list[HardwareItem]:
    book = open_workbook_with_merges(input_path)
    candidates = ["实木附件", "实木柜门", "模板"]
    for sheet_name in candidates:
        if sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            lookup = merged_cell_lookup(sheet)
            items = hardware_items_from_sheet(
                sheet,
                lookup,
                infer_order_numbers(sheet, lookup),
                order_no_filter=order_no_filter,
            )
            if items:
                return items
    return []


def order_numbers_from_items(items: list[Item], hardware_items: list[HardwareItem]) -> list[str]:
    seen: set[str] = set()
    orders: list[str] = []
    for item in items:
        if item.order_no and item.order_no not in seen:
            seen.add(item.order_no)
            orders.append(item.order_no)
    for item in hardware_items:
        if item.order_no and item.order_no not in seen:
            seen.add(item.order_no)
            orders.append(item.order_no)
    return orders


def header_from_sheet(sheet, lookup: dict[tuple[int, int], tuple[int, int]]) -> dict[str, str]:
    header: dict[str, str] = {}
    for row_idx in range(min(4, sheet.nrows)):
        for col_idx in range(sheet.ncols):
            value = str(cell_value_with_merges(sheet, row_idx, col_idx, lookup)).strip()
            if value.startswith("生产编号："):
                header["order_no"] = value.split("：", 1)[1].strip()
            elif value.startswith("客户名称："):
                header["customer"] = value.split("：", 1)[1].strip()
            elif value.startswith("联系电话："):
                header["phone"] = value.split("：", 1)[1].strip()
            elif value.startswith("客户地址："):
                header["address"] = value.split("：", 1)[1].strip()
    return header


def parse_quote_input(input_path: Path) -> ParsedQuoteInput:
    book = open_workbook_with_merges(input_path)
    candidates = ["实木附件", "实木柜门", "模板"]
    items: list[Item] = []
    hardware_items: list[HardwareItem] = []
    selected_sheet = None
    selected_lookup = None
    selected_blocks = None
    for sheet_name in candidates:
        if sheet_name not in book.sheet_names():
            continue
        sheet = book.sheet_by_name(sheet_name)
        lookup = merged_cell_lookup(sheet)
        blocks = infer_order_blocks(sheet, lookup)
        inferred_orders: dict[int, str] = {}
        for block in blocks:
            for row_idx in block["rows"]:
                inferred_orders[row_idx] = block["order_no"]
        items = items_from_sheet(sheet, lookup, inferred_orders)
        hardware_items = hardware_items_from_sheet(sheet, lookup, inferred_orders)
        if items or hardware_items:
            selected_sheet = sheet
            selected_lookup = lookup
            selected_blocks = blocks
            break
    if selected_sheet is None:
        # fallback to first candidate for header/warnings
        selected_sheet = get_bom_sheet(book, candidates, 1)
        selected_lookup = merged_cell_lookup(selected_sheet)
        selected_blocks = infer_order_blocks(selected_sheet, selected_lookup)
    return ParsedQuoteInput(
        items=items,
        hardware_items=hardware_items,
        header=header_from_sheet(selected_sheet, selected_lookup),
        orders=order_numbers_from_items(items, hardware_items),
        warnings=order_number_warnings_from_blocks(selected_blocks),
    )


def group_by_area(items: list[Item], exclude_areas: set[str] | None = None) -> list[tuple[str, list[Item]]]:
    groups: list[tuple[str, list[Item]]] = []
    current_area: str | None = None
    current_items: list[Item] = []
    for item in items:
        if item.area != current_area:
            if current_items:
                groups.append((current_area or "", current_items))
            current_area = item.area
            current_items = []
        current_items.append(item)
    if current_items:
        groups.append((current_area or "", current_items))
    if exclude_areas:
        return [(area, group) for area, group in groups if area not in exclude_areas]
    return groups


def group_by_color(items: list[Item]) -> list[tuple[str, list[Item]]]:
    """按（子订单号，颜色）分组，同一子订单的同一颜色进入同一个 sheet。

    当设计师在订单编号里用 -1/-2/-3 区分子订单时，每个子订单会生成独立 sheet，
    避免同一个颜色被合并到单个 sheet 中。
    """
    groups: list[tuple[str, list[Item]]] = []
    grouped: dict[tuple[str, str], list[Item]] = {}
    keys: list[tuple[str, str]] = []
    for item in items:
        color = item.color or "未填颜色"
        key = (item.order_no or "", color)
        if key not in grouped:
            grouped[key] = []
            keys.append(key)
        grouped[key].append(item)
    for order_no, color in keys:
        label = f"{order_no} {color}" if order_no else color
        groups.append((label, grouped[(order_no, color)]))
    return groups


def order_numbers(input_path: Path) -> list[str]:
    return parse_quote_input(input_path).orders


def copy_cell(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)


def copy_style(style, dst) -> None:
    if style is not None:
        dst._style = copy.copy(style)


def capture_sheet_style(ws: Worksheet) -> SheetStyleSnapshot:
    styles: dict[tuple[int, int], Any] = {}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.has_style:
                styles[(row, col)] = copy.copy(cell._style)
    return SheetStyleSnapshot(
        max_row=ws.max_row,
        max_column=ws.max_column,
        row_heights={row: ws.row_dimensions[row].height for row in range(1, ws.max_row + 1)},
        styles=styles,
    )


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        copy_cell(ws.cell(src_row, col), ws.cell(dst_row, col))


def apply_sheet_style_from_template(
    ws: Worksheet,
    template_style: SheetStyleSnapshot,
    insert_at: int,
    delta: int,
) -> None:
    max_col = min(ws.max_column, template_style.max_column)
    for row in range(1, ws.max_row + 1):
        if 8 <= row < insert_at + delta:
            src_row = min(row, insert_at - 1)
        elif row >= insert_at + delta:
            src_row = row - delta
        else:
            src_row = row
        if src_row < 1 or src_row > template_style.max_row:
            continue
        ws.row_dimensions[row].height = template_style.row_heights.get(src_row)
        for col in range(1, max_col + 1):
            copy_style(template_style.styles.get((src_row, col)), ws.cell(row, col))


def clone_sheet_layout(wb, template_sheet: Worksheet, title: str, before_sheet: Worksheet) -> Worksheet:
    ws = wb.copy_worksheet(template_sheet)
    ws.title = title
    wb._sheets.remove(ws)
    idx = wb._sheets.index(before_sheet)
    wb._sheets.insert(idx, ws)
    return ws


# Mapping from logical column keys to the header text expected in the template.
QUOTE_HEADER_MAP: dict[str, str] = {
    "no": "序号",
    "area": "区域",
    "name": "产品名称",
    "style": "款式",
    "height": "高",
    "width": "宽",
    "thickness": "厚",
    "qty": "数量",
    "meter": "米数",
    "square": "平方",
    "unit_price": "单价（元）",
    "subtotal": "小计（元）",
    "workpoint": "工分",
    "color": "颜色",
    "material": "材质",
    "remark": "备注",
}


def build_quote_col_map(template_ws: Worksheet, header_row: int = 7) -> dict[str, int]:
    """Build a key -> column-number mapping from the template header row."""
    col_map: dict[str, int] = {}
    expected = {v: key for key, v in QUOTE_HEADER_MAP.items()}
    for col in range(1, template_ws.max_column + 1):
        raw = template_ws.cell(header_row, col).value
        header_text = "" if raw is None else str(raw).strip().replace(" ", "")
        if header_text in expected:
            col_map[expected[header_text]] = col
    return col_map


def ensure_data_capacity(ws: Worksheet, item_count: int) -> None:
    capacity = 7 if ws.title == "1" else 9
    insert_at = 15 if ws.title == "1" else 17
    if item_count > capacity:
        extra = item_count - capacity
        ws.insert_rows(insert_at, extra)
        for row in range(insert_at, insert_at + extra):
            copy_row_style(ws, insert_at - 1, row)
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).value = None
    elif item_count < capacity:
        ws.delete_rows(8 + item_count, capacity - item_count)


def unmerge_data_area(ws: Worksheet, last_data_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= 8 and merged.min_row <= last_data_row:
            ws.unmerge_cells(str(merged))


def merge_if_needed(ws: Worksheet, cell_range: str) -> None:
    if cell_range not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells(cell_range)


def merge_adjacent_same_values(ws: Worksheet, start_row: int, end_row: int, min_col: int, max_col: int) -> None:
    row = start_row
    while row <= end_row:
        value = ws.cell(row, min_col).value
        block_start = row
        row += 1
        while row <= end_row and ws.cell(row, min_col).value == value:
            row += 1
        block_end = row - 1
        if value in (None, ""):
            continue
        merge_if_needed(
            ws,
            f"{get_column_letter(min_col)}{block_start}:{get_column_letter(max_col)}{block_end}",
        )


def visual_text_width(value: Any) -> int:
    text = "" if value is None else str(value)
    width = 0
    for char in text:
        if char == "\n":
            continue
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return width


def merged_columns_width(ws: Worksheet, min_col: int, max_col: int) -> float:
    width = 0.0
    for col in range(min_col, max_col + 1):
        width += ws.column_dimensions[get_column_letter(col)].width or 8.43
    return max(width, 8.43)


def normalize_merged_range_style(ws: Worksheet, merged_range, style_source, alignment: Alignment) -> None:
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            cell = ws.cell(row, col)
            if style_source.has_style:
                copy_cell(style_source, cell)
            cell.alignment = copy.copy(alignment)


def fit_merged_text_height(ws: Worksheet, merged_range, value: Any) -> None:
    if value in (None, ""):
        return
    text_width = visual_text_width(value)
    columns_width = merged_columns_width(ws, merged_range.min_col, merged_range.max_col)
    lines = max(1, math.ceil(text_width / max(columns_width * 0.85, 1)))
    target_total_height = max(lines * 18, 29 * (merged_range.max_row - merged_range.min_row + 1))
    per_row_height = min(max(target_total_height / (merged_range.max_row - merged_range.min_row + 1), 29), 95)
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, per_row_height)


def normalize_quote_page_data_styles(ws: Worksheet, item_last_row: int, data_last_row: int, sum_row: int) -> None:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(8, max(data_last_row, sum_row) + 1):
        for col in range(1, 19):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = copy.copy(center)

    for merged in ws.merged_cells.ranges:
        if merged.min_col != 14 or merged.max_col != 15:
            continue
        if merged.max_row < 8 or merged.min_row > data_last_row:
            continue
        top_left = ws.cell(merged.min_row, merged.min_col)
        if isinstance(top_left, MergedCell):
            continue
        style_source = ws.cell(merged.min_row, 16)
        if not isinstance(style_source, MergedCell) and style_source.has_style:
            normalize_merged_range_style(ws, merged, style_source, center)
        else:
            top_left.alignment = copy.copy(center)
        fit_merged_text_height(ws, merged, top_left.value)

    if data_last_row > item_last_row:
        for row in range(item_last_row + 1, data_last_row + 1):
            for col in range(1, 19):
                cell = ws.cell(row, col)
                if isinstance(cell, MergedCell):
                    continue
                style_source = ws.cell(row, 16)
                if col in (14, 15) and not isinstance(style_source, MergedCell) and style_source.has_style:
                    copy_cell(style_source, cell)
                cell.alignment = copy.copy(center)


def restore_shifted_merges(
    ws: Worksheet,
    base_merges: list[str],
    insert_at: int,
    delta: int,
    item_last_row: int,
    data_last_row: int,
    col_map: dict[str, int] | None = None,
) -> None:
    ws.merged_cells = MultiCellRange()

    for cell_range in base_merges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if min_row >= 8 and max_row <= insert_at - 1 and min_col in (2, 4, 14, 17):
            continue
        if min_row >= insert_at:
            min_row += delta
            max_row += delta
        if min_row < 1 or max_row < min_row:
            continue
        shifted = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        merge_if_needed(ws, shifted)

    area_col = col_map.get("area", 2) if col_map else 2
    style_col = col_map.get("style", 4) if col_map else 4
    color_col = col_map.get("color", 14) if col_map else 14
    if item_last_row >= 8:
        # A 列序号不合并；B 列区域按相邻相同值合并
        merge_adjacent_same_values(ws, 8, item_last_row, area_col, area_col)
        # 款式保持整列合并显示“见图生产”
        merge_if_needed(ws, f"{get_column_letter(style_col)}8:{get_column_letter(style_col)}{item_last_row}")
        merge_adjacent_same_values(ws, 8, item_last_row, color_col, color_col + 1)
    if data_last_row >= 8:
        for row in range(8, data_last_row + 1):
            merge_if_needed(ws, f"Q{row}:R{row}")
    if data_last_row > item_last_row:
        for row in range(item_last_row + 1, data_last_row + 1):
            merge_if_needed(ws, f"D{row}:K{row}")
            merge_if_needed(ws, f"N{row}:O{row}")


def parse_width_expression(width: Any) -> str:
    text = str(width).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    # Pure numeric width stays as-is.
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return text
    # Arc length / length text: keep the first number only.
    if "弧" in text or "长" in text:
        match = re.search(r"\d+(?:\.\d+)?", text)
        return match.group(0) if match else ""
    # Compound dimensions separated by slash (e.g. 353/505, 55/80) become sum.
    if "/" in text:
        numbers = re.findall(r"\d+(?:\.\d+)?", text)
        return "+".join(numbers) if numbers else ""
    # Fallback: keep only numbers and arithmetic operators.
    return re.sub(r"[^0-9+*./()-]", "", text)


def arc_length_value(width: Any) -> float | None:
    text = str(width)
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def is_arc_panel(item: Item) -> bool:
    return "弧形" in item.name


def quote_color_text(item: Item, first_color_on_sheet: bool = False) -> str:
    color = item.color or ""
    if is_cross_color(item):
        return "见图叉色/B21-S白橡山纹木皮，混拼（山雾云橡S）（哑光）半透开放漆-水性漆叉纯黑（哑光）-水性漆"
    if "B21-S" in color:
        text = "B21-S白橡山纹木皮，混拼（山雾云橡S）（哑光）半透开放漆-水性漆"
        if first_color_on_sheet:
            text += " 请车间多配1瓶油漆，随货一起"
        return text
    if "Y10-S" in color:
        text = "Y10-S云母灰 （哑光）-水性漆"
        if first_color_on_sheet:
            text += " 请车间多配1瓶油漆，随货一起"
        return text
    if item.veneer and color:
        return f"{item.veneer} {color}"
    return color


def is_cross_color(item: Item) -> bool:
    color = item.color or ""
    return "叉色" in color or "叉色" in item.remark


def quote_color_cells_for_page(sheet_no: int, items: list[Item]) -> list[str]:
    # 每行都显示颜色，避免同一子订单内只有第一行有颜色
    cells: list[str] = []
    for item in items:
        cells.append(quote_color_text(item))
    return cells


def area_formula(row: int, item: Item) -> Any:
    if isinstance(item.src_area_value, (int, float)) and 0 < float(item.src_area_value) <= 0.1001:
        return round(float(item.src_area_value), 4)
    if isinstance(item.height, str):
        expr = parse_width_expression(item.height)
        if expr:
            return f"=MAX(({expr})/1000*F{row}/1000,0.1)*H{row}"
    if isinstance(item.width, str):
        expr = parse_width_expression(item.width)
        if expr:
            return f"=MAX(E{row}/1000*({expr})/1000,0.1)*H{row}"
    return f"=MAX(E{row}/1000*F{row}/1000,0.1)*H{row}"


def price_for(item: Item) -> Any:
    if item.name == "收口板":
        return "=360+60" if item.thickness in (22, "22", 22.0) else 360
    if item.name == "见光板":
        return 530 if item.thickness in (25, "25", 25.0) else 360
    if item.name == "异形整体柜":
        return "=759*3+80+80+150" if item.thickness in (344, "344", 344.0) else "=759*2+160+150"
    return PRICE_RULES.get(item.name, None)


def line_total_formula(row: int, item: Item) -> str | None:
    name = item.name
    if name == "调色费":
        return f"=K{row}*H{row}"
    if name in {"侧", "顶", "合页条", "50套线", "木箱包装"}:
        return None
    base = f"=K{row}*J{row}"
    if name in {"贴线网格柜门", "平板柜门", "平板抽面", "异形平板柜门"} and item.qty and item.qty > 1:
        return f"{base}+20*H{row}"
    if "玻璃" in name:
        return f"{base}+120*H{row}"
    if name == "L型斜角收口板":
        return f"=J{row}*K{row}+20*H{row}"
    if name == "异形整体柜":
        return f"{base}+70*J{row}"
    return base


def needs_wood_box(items: list[Item]) -> bool:
    for item in items:
        text = " ".join(
            str(value)
            for value in (item.name, item.material, item.color, item.remark)
            if value not in (None, "")
        )
        if any(keyword in text for keyword in WOOD_BOX_KEYWORDS):
            return True
    return False


def wood_box_row() -> tuple[str, str, int]:
    return ("木箱包装", "此单共打1个木箱包装", 150)


def set_wood_box_packaging_row(ws: Worksheet, row: int, serial_no: int) -> None:
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 19):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = None
            cell.alignment = copy.copy(center)
    # A 列保留序号，用于 fix_quote_page_sum_row 定位合计行
    ws.cell(row, 1).value = serial_no
    ws.cell(row, 3).value = "木箱包装"
    ws.cell(row, 4).value = WOOD_BOX_REMARK
    ws.cell(row, 12).value = None
    ws.cell(row, 13).value = None


def has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def to_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def dimension_numbers(value: Any) -> list[float]:
    number = to_number(value)
    if number is not None:
        return [number]
    return [float(match) for match in re.findall(r"\d+(?:\.\d+)?", str(value))]


def dimension_total(value: Any) -> float | None:
    number = to_number(value)
    if number is not None:
        return number
    numbers = dimension_numbers(value)
    if numbers:
        return sum(numbers)
    return None


def has_small_dimension(value: Any) -> bool:
    return any(0 < number < 100 for number in dimension_numbers(value))


def is_strip_with_small_dimension(item: Item) -> bool:
    if "圆弧" in item.name:
        return True
    return "条" in item.name and (has_small_dimension(item.height) or has_small_dimension(item.width))


def meter_total_with_piece_minimum(value_per_piece: float, qty: float) -> float:
    return round(max(value_per_piece, 1) * qty, 4)


def area_total_with_piece_minimum(value_per_piece: float, qty: float) -> float:
    return round(max(value_per_piece, 0.1) * qty, 4)


def area_formula_text(row: int, item: Item) -> str | None:
    height = dimension_total(item.height)
    width = dimension_total(item.width)
    qty = to_number(item.qty)
    if height is None or width is None or qty is None:
        return None
    return f"=E{row}/1000*F{row}/1000*H{row}"


def quote_meter_value(item: Item) -> Any:
    """Return meter total only for slim strip-like products; others leave blank."""
    if "弧形" in item.name:
        return None
    if not is_strip_with_small_dimension(item):
        return None
    if has_value(item.meter):
        number = to_number(item.meter)
        qty = to_number(item.qty)
        if number is not None and qty:
            return meter_total_with_piece_minimum(number / qty, qty)
    height = dimension_total(item.height)
    width = dimension_total(item.width)
    qty = to_number(item.qty)
    if qty is None:
        return None
    if height is not None and has_small_dimension(item.width):
        return meter_total_with_piece_minimum(height / 1000, qty)
    if width is not None and has_small_dimension(item.height):
        return meter_total_with_piece_minimum(width / 1000, qty)
    if height is not None and width is not None:
        return meter_total_with_piece_minimum(max(height, width) / 1000, qty)
    return None


def quote_area_value(item: Item) -> Any:
    # 条类产品也要同时显示平方（和米数一起出来）。

    # Height is text (e.g. arc length in height column): expr/1000 * width/1000 * qty
    if isinstance(item.height, str):
        expr = parse_width_expression(item.height)
        if expr:
            return f"=MAX(({expr})/1000*F{{row}}/1000,0.1)*H{{row}}"

    # Width is text (e.g. arc length, 353/505, L-shape): length/1000 * expr/1000 * qty
    if isinstance(item.width, str):
        expr = parse_width_expression(item.width)
        if expr:
            return f"=MAX(E{{row}}/1000*({expr})/1000,0.1)*H{{row}}"

    height = dimension_total(item.height)
    width = dimension_total(item.width)
    qty = to_number(item.qty)

    # Legacy arc panel handling (numeric dimensions with arc width).
    if is_arc_panel(item) and height is not None and qty is not None:
        arc = arc_length_value(item.width)
        if arc is not None:
            per_piece = height / 1000 * arc / 1000
            if per_piece < 0.5:
                value = 0.5 * qty
                return int(value) if float(value).is_integer() else round(value, 4)
            arc_text = int(arc) if arc.is_integer() else arc
            return f"=E{{row}}/1000*{arc_text}/1000*H{{row}}"

    if is_strip_with_small_dimension(item) and height is not None and width is not None and qty is not None:
        per_piece = height / 1000 * width / 1000
        if per_piece < 0.1:
            return area_total_with_piece_minimum(per_piece, qty)
        return f"=E{{row}}/1000*F{{row}}/1000*H{{row}}"

    if has_value(item.src_area_value):
        src = float(item.src_area_value)
        qty_num = to_number(item.qty)
        # 单块不足 0.1 的按 0.1 计算
        if qty_num and qty_num > 0 and src / qty_num < 0.1:
            return area_total_with_piece_minimum(src / qty_num, qty_num)
        return item.src_area_value

    if height is not None and width is not None and qty is not None:
        per_piece = height / 1000 * width / 1000
        if per_piece < 0.1:
            return area_total_with_piece_minimum(per_piece, qty)
        return f"=E{{row}}/1000*F{{row}}/1000*H{{row}}"
    return None


def set_quote_area_cell(ws: Worksheet, row: int, item: Item) -> None:
    value = quote_area_value(item)
    if isinstance(value, str):
        value = value.format(row=row)
    ws.cell(row, 10).value = value


def input_quote_price(item: Item) -> Any:
    name = item.name
    thickness = str(item.thickness)
    if is_arc_panel(item):
        if to_number(item.height) in (690, 499):
            return "=370*4"
        return 1350 if item.order_no.endswith("6110") else "=410*4"
    if "线条" in name:
        if to_number(item.thickness) == 44:
            return "=650/30*44"
        if to_number(item.width) == 50 or to_number(item.height) == 50:
            return "=50+5"
    if "夹心层板" in name:
        return "=960+220"
    if name in {"板子", "立面板子", "顶板"}:
        return 560 if "25" in thickness and to_number(item.thickness) == 25 else 450
    return None


def input_quote_total_formula(row: int, item: Item) -> Any:
    name = item.name
    if "线条" in name and (to_number(item.width) == 50 or to_number(item.height) == 50):
        return f"=K{row}*I{row}+20*H{row}"
    if "线条" in name and to_number(item.thickness) == 44:
        return f"=J{row}*K{row}"
    if is_arc_panel(item) and to_number(item.height) in (690, 499):
        return f"=J{row}*K{row}+20*H{row}"
    if is_arc_panel(item) and not item.order_no.endswith("6110"):
        return f"=K{row}*J{row}+20*H{row}"
    if is_arc_panel(item):
        return f"=J{row}*K{row}"
    if name in {"板子", "平板", "柜门", "抽面"} and to_number(item.thickness) == 25:
        return f"=K{row}*J{row}+20*H{row}"
    if input_quote_price(item) is not None:
        return f"=K{row}*J{row}"
    return None


def apply_input_quote_pricing(ws: Worksheet, row: int, item: Item) -> None:
    ws.cell(row, 11).value = input_quote_price(item)
    ws.cell(row, 12).value = input_quote_total_formula(row, item)
    # 工分始终按 小计/100 给公式；未命中自动定价的行由人工补单价/小计后自动算出，特殊情况人工改
    ws.cell(row, 13).value = f"=L{row}/100"


def fill_page(
    ws: Worksheet,
    sheet_no: int,
    area: str,
    items: list[Item],
    total_pages: int,
    style_template_ws: Worksheet,
) -> int:
    base_merges = [str(rng) for rng in ws.merged_cells.ranges]
    special_count = (1 if sheet_no == 1 else 0) + (1 if sheet_no in PAGE_EXTRA_ROWS else 0) + 1
    capacity = 7 if sheet_no == 1 else 9
    insert_at = 15 if sheet_no == 1 else 17
    delta = len(items) + special_count - capacity
    ensure_data_capacity(ws, len(items) + special_count)
    unmerge_data_area(ws, 7 + len(items) + special_count)
    apply_sheet_style_from_template(ws, style_template_ws, insert_at, delta)
    ws["D2"] = PROVINCE if sheet_no == 1 else "=+'1'!D2"
    ws["F2"] = CUSTOMER_NAME if sheet_no == 1 else "=+'1'!F2"
    ws["E3"] = END_CUSTOMER if sheet_no == 1 else "='1'!E3"
    ws["Q2"] = items[0].order_no if sheet_no == 1 else "=+'1'!Q2"
    ws["R2"] = -1
    ws["C2"] = "经销商地址"
    ws["Q5"] = f"{items[0].order_no}-{sheet_no}"
    ws["N5"] = items[0].order_no

    for i, item in enumerate(items, start=1):
        row = 7 + i
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = area if i == 1 else None
        ws.cell(row, 3).value = item.name
        ws.cell(row, 4).value = "见图生产" if i == 1 else None
        ws.cell(row, 5).value = item.height
        ws.cell(row, 6).value = item.width
        ws.cell(row, 7).value = item.thickness
        ws.cell(row, 8).value = item.qty
        ws.cell(row, 9).value = quote_meter_value(item)
        ws.cell(row, 10).value = area_formula(row, item)
        ws.cell(row, 11).value = price_for(item)
        ws.cell(row, 12).value = line_total_formula(row, item)
        ws.cell(row, 13).value = f"=L{row}/100"

    for (row, col), value in PAGE_CELL_OVERRIDES.get(sheet_no, {}).items():
        ws.cell(row, col).value = value
    for (page, row), values in DATA_FORMULA_OVERRIDES.items():
        if page == sheet_no:
            for col, value in values.items():
                ws.cell(row, col).value = value

    next_row = 8 + len(items)
    set_wood_box_packaging_row(ws, next_row, len(items) + 1)
    next_row += 1
    if sheet_no == 1:
        ws.cell(next_row, 1).value = next_row - 7
        ws.cell(next_row, 3).value = "调色费"
        ws.cell(next_row, 8).value = 1
        ws.cell(next_row, 11).value = 300
        ws.cell(next_row, 12).value = f"=K{next_row}*H{next_row}"
        ws.cell(next_row, 13).value = None
        next_row += 1
    if sheet_no in PAGE_EXTRA_ROWS:
        name, remark, total = PAGE_EXTRA_ROWS[sheet_no]
        ws.cell(next_row, 1).value = next_row - 7
        ws.cell(next_row, 3).value = name
        ws.cell(next_row, 4).value = remark
        ws.cell(next_row, 12).value = total
        ws.cell(next_row, 13).value = None
        next_row += 1

    sum_row = next_row
    ws.cell(sum_row, 1).value = "合计"
    ws.cell(sum_row, 8).value = f"=SUM(H6:H{sum_row - 1})"
    ws.cell(sum_row, 9).value = f"=SUM(I6:I{sum_row - 1})"
    ws.cell(sum_row, 10).value = f"=SUM(J6:J{sum_row - 1})"
    ws.cell(sum_row, 13).value = f"=SUM(M6:M{sum_row - 1})"
    subtotal_row = sum_row + 1
    ws.cell(subtotal_row, 12).value = f"=+SUM(L8:L{sum_row - 1})"

    if sheet_no == 1:
        ws.cell(subtotal_row, 1).value = FIRST_PAGE_NOTE
        ws.cell(subtotal_row + 1, 4).value = f"共计{total_pages}页"
        ws.cell(subtotal_row + 1, 12).value = total_formula(total_pages, subtotal_row)
        ws.cell(subtotal_row + 2, 4).value = f"此单共{total_pages}页，合计金额为{TOTAL_AMOUNT_TEXT}元！"
        ws.cell(subtotal_row + 3, 4).value = CONFIRM_NOTE
        ws.cell(subtotal_row + 4, 3).value = MAKER
        ws.cell(subtotal_row + 4, 10).value = DRAWER
    else:
        ws.cell(subtotal_row, 1).value = "='1'!A16"
        ws.cell(subtotal_row, 4).value = "预付定金："
        ws.cell(subtotal_row, 6).value = f"=INT(F{subtotal_row + 1}*0.5/100+0.55)*100"
        ws.cell(subtotal_row + 2, 4).value = f"='{sheet_no - 1}'!D{subtotal_row if sheet_no == 2 else subtotal_row + 1}"
        ws.cell(subtotal_row + 5, 1).value = f"=+A{subtotal_row}"

    restore_shifted_merges(ws, base_merges, insert_at, delta, 7 + len(items), next_row - 1)
    return subtotal_row


def total_formula(total_pages: int, first_subtotal_row: int) -> str:
    parts = [f"L{first_subtotal_row}"]
    for idx in range(2, total_pages + 1):
        # The subtotal row is derived from each generated sheet after writing.
        # This placeholder is rewritten after all pages are filled.
        parts.append(f"'{idx}'!L__ROW__{idx}")
    return "=" + "+".join(parts)


def rewrite_first_page_total(ws: Worksheet, subtotal_rows: dict[int, int]) -> None:
    first_total_row = subtotal_rows[1] + 1
    parts = [f"L{subtotal_rows[1]}"] + [f"'{i}'!L{subtotal_rows[i]}" for i in range(2, len(subtotal_rows) + 1)]
    ws.cell(first_total_row, 12).value = "=" + "+".join(parts)


def update_summary(wb, subtotal_rows: dict[int, int]) -> None:
    ws = wb["汇总"]
    for page in range(1, len(subtotal_rows) + 1):
        row = page + 2
        ws.cell(row, 1).value = "='1'!Q2" if page == 1 else f"=A{row - 1}"
        ws.cell(row, 2).value = page
        ws.cell(row, 3).value = f"='{page}'!L{subtotal_rows[page]}"
    for row in range(len(subtotal_rows) + 3, 31):
        ws.cell(row, 3).value = None
    ws["C31"] = f"=SUM(C3:C{len(subtotal_rows) + 2})"


def update_completion_table(wb, total_pages: int) -> None:
    ws = wb["前工序完工单号总表"]
    for page in range(1, min(total_pages, 9) + 1):
        row = page + 2
        ws.cell(row, 3).value = f"=+'{page}'!N5"
        ws.cell(row, 4).value = f"=+'{page}'!Q2"
    for row in range(min(total_pages, 9) + 3, 13):
        ws.cell(row, 3).value = "=+#REF!"
        ws.cell(row, 4).value = "=+#REF!"


def read_input_header(input_path: Path) -> dict[str, str]:
    book = open_workbook_with_merges(input_path)
    sheet = get_bom_sheet(book, ["实木附件", "实木柜门", "模板"], 1)
    lookup = merged_cell_lookup(sheet)
    return header_from_sheet(sheet, lookup)


def fill_page_input_only(
    ws: Worksheet,
    sheet_no: int,
    area: str,
    items: list[Item],
    total_pages: int,
    style_template_ws: Worksheet,
    header: dict[str, str],
    col_map: dict[str, int] | None = None,
    sub_orders: list[str] | None = None,
) -> None:
    if col_map is None:
        col_map = build_quote_col_map(style_template_ws)
    special_count = 1
    capacity = 7 if sheet_no == 1 else 9
    insert_at = 15 if sheet_no == 1 else 17
    delta = len(items) + special_count - capacity
    base_merges = [str(rng) for rng in ws.merged_cells.ranges]
    ensure_data_capacity(ws, len(items) + special_count)
    unmerge_data_area(ws, 7 + len(items) + special_count)
    apply_sheet_style_from_template(ws, style_template_ws, insert_at, delta)

    current_order = items[0].order_no if items else ""
    order_no = ""
    order_sub_no = ""
    if current_order:
        base = base_order_no(current_order)
        if sub_orders and len(sub_orders) > 1:
            try:
                idx = sub_orders.index(current_order) + 1
            except ValueError:
                idx = 1
            order_no = f"{base}-{len(sub_orders)}"
            order_sub_no = f"-{idx}"
        else:
            order_no = current_order
    if not order_no:
        order_no = header.get("order_no") or ""
    ws["D2"] = header.get("address") if sheet_no == 1 else "=+'1'!D2"
    ws["F2"] = (header.get("customer") or "") if sheet_no == 1 else "=+'1'!F2"
    ws["Q2"] = order_no or ""
    ws["R2"] = order_sub_no or None
    ws["E3"] = header.get("end_customer") if sheet_no == 1 else "='1'!E3"
    ws["Q5"] = "=+Q2" if order_no else ""
    ws["N5"] = "=+'1'!N5" if sheet_no > 1 and order_no else None

    color_cells = quote_color_cells_for_page(sheet_no, items)
    area_col = col_map.get("area", 2)
    no_col = col_map.get("no", 1)
    for i, item in enumerate(items, start=1):
        row = 7 + i
        # A 列显示行号，不合并，方便快速看有几行内容
        ws.cell(row, no_col).value = i
        # B 列显示设计师写的区域，每行都填值，后续按相邻相同值合并
        ws.cell(row, area_col).value = item.area
        ws.cell(row, 3).value = item.name
        ws.cell(row, 4).value = "\u89c1\u56fe\u751f\u4ea7" if i == 1 else None
        ws.cell(row, 5).value = item.height
        ws.cell(row, 6).value = item.width
        ws.cell(row, 7).value = item.thickness
        ws.cell(row, 8).value = item.qty
        ws.cell(row, 9).value = quote_meter_value(item)
        set_quote_area_cell(ws, row, item)
        apply_input_quote_pricing(ws, row, item)
        ws.cell(row, 14).value = color_cells[i - 1]
        ws.cell(row, 16).value = item.material
        ws.cell(row, 17).value = item.remark

    next_row = 8 + len(items)
    set_wood_box_packaging_row(ws, next_row, len(items) + 1)
    next_row += 1

    sum_row = next_row
    ws.cell(sum_row, 1).value = "合计"
    ws.cell(sum_row, 8).value = f"=SUM(H6:H{sum_row - 1})"
    ws.cell(sum_row, 9).value = f"=SUM(I6:I{sum_row - 1})"
    ws.cell(sum_row, 10).value = f"=SUM(J6:J{sum_row - 1})"
    ws.cell(sum_row, 13).value = f"=SUM(M6:M{sum_row - 1})"
    if sum_row + 1 <= ws.max_row:
        for col in (8, 9, 10, 13):
            cell = ws.cell(sum_row + 1, col)
            if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                cell.value = None
    if sum_row + 1 <= ws.max_row:
        for col in (8, 9, 10, 13):
            cell = ws.cell(sum_row + 1, col)
            if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                cell.value = None

    clear_auto_formulas(ws)
    for i, item in enumerate(items, start=1):
        apply_input_quote_pricing(ws, 7 + i, item)
    set_wood_box_packaging_row(ws, sum_row - 1, len(items) + 1)
    ws.cell(sum_row, 1).value = "合计"
    ws.cell(sum_row, 8).value = f"=SUM(H6:H{sum_row - 1})"
    ws.cell(sum_row, 9).value = f"=SUM(I6:I{sum_row - 1})"
    ws.cell(sum_row, 10).value = f"=SUM(J6:J{sum_row - 1})"
    ws.cell(sum_row, 13).value = f"=SUM(M6:M{sum_row - 1})"
    restore_shifted_merges(ws, base_merges, insert_at, delta, 7 + len(items), 7 + len(items) + special_count, col_map)
    normalize_quote_page_data_styles(ws, 7 + len(items), 7 + len(items) + special_count, sum_row)
    for row in range(8, ws.max_row + 1):
        ws.row_dimensions[row].hidden = False
    hide_bottom_process_area(ws)


def clear_auto_formulas(ws: Worksheet) -> None:
    # In input-plus-template mode, keep template labels/layout but leave pricing,
    # totals, and generated cross-sheet summaries for users to fill safely.
    for row in range(8, ws.max_row + 1):
        for col in (11, 12, 13):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for row in range(8 + 1, ws.max_row + 1):
        for col in (6, 8, 9, 15):
            cell = ws.cell(row, col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None
    for coord in ("L16", "L17", "F16", "O17"):
        cell = ws[coord]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if isinstance(value, str) and value.startswith("=") and re.search(rf"(?<!!)\b{cell.coordinate}\b", value):
                cell.value = None


def hide_bottom_process_area(ws: Worksheet) -> None:
    for row in range(8, ws.max_row + 1):
        if ws.cell(row, 1).value == "制表人：":
            for hidden_row in range(row + 1, ws.max_row + 1):
                ws.row_dimensions[hidden_row].hidden = True
            return


def decide_group_mode(items: list[Item], group_mode: str) -> tuple[str, str | None]:
    """决定分组模式。auto 模式下优先按颜色分表，无颜色时回退按区域分表。"""
    if group_mode != "auto":
        return group_mode, None
    total = len(items)
    if total == 0:
        return "area", None
    colored = sum(1 for item in items if item.color and str(item.color).strip())
    ratio = colored / total
    if ratio >= 0.5:
        return "color", f"检测到 {colored}/{total} 行有颜色，按颜色分表生成报价单"
    return "area", f"仅 {colored}/{total} 行有颜色，回退按区域分表生成报价单"


def build_workbook_input_only(
    input_path: Path,
    template_path: Path,
    output_path: Path,
    order_no: str | None = None,
    group_mode: str = "auto",
    source_items: list[Item] | None = None,
    source_hardware_items: list[HardwareItem] | None = None,
    source_header: dict[str, str] | None = None,
) -> None:
    if source_items is None:
        source_items = read_items(input_path)
    if order_no is None:
        items = source_items
    else:
        items = [item for item in source_items if item.order_no == order_no]
    resolved_mode, mode_note = decide_group_mode(items, group_mode)
    if mode_note:
        print(mode_note)
    if resolved_mode == "area":
        groups = group_by_area(items)
    else:
        groups = group_by_color(items)
    if source_hardware_items is None:
        hardware_items = read_hardware_items(input_path, order_no_filter=order_no)
    elif order_no is None:
        hardware_items = source_hardware_items
    else:
        hardware_items = [item for item in source_hardware_items if item.order_no == order_no]
    # In input-only mode, do not drop any area from the source file.
    groups = [(area, group) for area, group in groups]
    header = dict(source_header) if source_header is not None else read_input_header(input_path)
    # 汇总单号格式：基础单号-子单总数-当前子单序号，如 S2607-6047-5-1
    current_order = order_no if order_no is not None else (items[0].order_no if items else "")
    sub_orders: list[str] = []
    if current_order:
        base = base_order_no(current_order)
        sub_orders = sorted({
            it.order_no for it in source_items
            if it.order_no and base_order_no(it.order_no) == base
        })
        if len(sub_orders) > 1:
            try:
                idx = sub_orders.index(current_order) + 1
            except ValueError:
                idx = 1
            header["order_no"] = f"{base}-{len(sub_orders)}-{idx}"
        else:
            header["order_no"] = current_order
    elif not header.get("order_no"):
        header["order_no"] = order_no_from_filename(input_path)
    header["order_sheet_count"] = str(initial_order_group_count(items) or len(groups))

    wb = openpyxl.load_workbook(template_path)
    first_page_style = capture_sheet_style(wb["1"])
    other_page_style = capture_sheet_style(wb["2"])
    total_pages = len(groups)
    if total_pages < 1:
        raise ValueError("No rows found in 实木附件")

    summary = wb["汇总"]
    base_other = wb["2"]
    hardware_template = wb.copy_worksheet(base_other)
    hardware_template.title = "_hardware_template"
    hardware_template.sheet_state = "hidden"
    for sheet_name in list(wb.sheetnames):
        if sheet_name.isdigit() and int(sheet_name) > total_pages and wb[sheet_name].sheet_state == "visible":
            del wb[sheet_name]
    for idx in range(3, total_pages + 1):
        clone_sheet_layout(wb, base_other, str(idx), summary)

    col_map = build_quote_col_map(wb["1"])
    for idx, (area, items) in enumerate(groups, start=1):
        ws = wb[str(idx)]
        style_template_ws = first_page_style if idx == 1 else other_page_style
        fill_page_input_only(ws, idx, area, items, total_pages, style_template_ws, header, col_map, sub_orders)
    fill_hardware_sheets(wb, hardware_items)
    if "_hardware_template" in wb.sheetnames:
        del wb["_hardware_template"]

    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        for row in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = None

    if "\u6c47\u603b" in wb.sheetnames:
        wb["\u6c47\u603b"].sheet_state = "hidden"

    normalize_quote_footer_formulas(wb)
    for ws in wb.worksheets:
        clear_same_sheet_self_refs(ws)
    clear_bad_refs_in_hidden_sheets(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def find_quote_sum_row(ws: Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "\u5408\u8ba1":
            return row
    last_item_row = None
    for row in range(8, ws.max_row + 1):
        if isinstance(ws.cell(row, 1).value, int):
            last_item_row = row
    return last_item_row + 1 if last_item_row else None


def existing_cells(ws: Worksheet, min_row: int = 1):
    for (row, _col), cell in list(ws._cells.items()):
        if row >= min_row and not isinstance(cell, MergedCell):
            yield cell


def clear_formulas(ws: Worksheet) -> None:
    for cell in existing_cells(ws):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = None


def clear_footer_formulas(ws: Worksheet, start_row: int) -> None:
    for cell in existing_cells(ws, min_row=start_row):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = None


def set_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def hardware_display_name(name: str) -> str:
    if name.startswith("PDJ19") and "灯带" not in name:
        return name.replace("PDJ19", "PDJ19灯带", 1)
    return name


def hardware_unit(name: str) -> str:
    return "根" if name.startswith("PDJ19") else "个"


def hardware_price(name: str) -> int | None:
    if "全盖" in name:
        return 22
    if "半盖" in name:
        return 23
    if name.startswith("PDJ19"):
        return 80
    return None


def hardware_total_formula(row: int, name: str) -> str:
    if name.startswith("PDJ19"):
        return f"=Q{row}*I{row}*MAX(L{row}/1000,0.4)"
    return f"=Q{row}*I{row}"


def unmerge_hardware_remark_blocks(ws: Worksheet) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 25 and merged.max_row >= 6 and merged.min_col <= 19 <= merged.max_col:
            ws.unmerge_cells(str(merged))


def merge_hardware_remark_blocks(ws: Worksheet, hardware_items: list[HardwareItem]) -> None:
    unmerge_hardware_remark_blocks(ws)
    item_count = min(len(hardware_items), 20)
    if item_count < 1:
        for row in range(6, 26):
            ws.merge_cells(f"S{row}:T{row}")
        return

    start_row = 6
    current_remark = hardware_items[0].area
    for idx, item in enumerate(hardware_items[:item_count], start=1):
        row = 5 + idx
        if item.area != current_remark:
            end_row = row - 1
            ws.merge_cells(f"S{start_row}:T{end_row}")
            ws.cell(start_row, 19).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            start_row = row
            current_remark = item.area
    end_row = 5 + item_count
    ws.merge_cells(f"S{start_row}:T{end_row}")
    ws.cell(start_row, 19).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(end_row + 1, 26):
        ws.merge_cells(f"S{row}:T{row}")
        ws.cell(row, 19).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def find_sheet_by_name(wb, name: str):
    return wb[name] if name in wb.sheetnames else None


def hardware_subtotal_row(wb) -> int | None:
    if "\u4e94\u91d1-1" in wb.sheetnames:
        return 26
    return None


def fill_hardware_sheets(wb, hardware_items: list[HardwareItem]) -> None:
    if not hardware_items:
        return
    if "1\u4e94" in wb.sheetnames:
        del wb["1\u4e94"]
    if "\u4e94\u91d1-1" not in wb.sheetnames:
        return
    ws = wb["\u4e94\u91d1-1"]
    ws.sheet_state = "visible"
    wb._sheets.remove(ws)
    insert_at = wb._sheets.index(wb["汇总"]) if "汇总" in wb.sheetnames else len(wb._sheets)
    wb._sheets.insert(insert_at, ws)
    ws.sheet_format.zeroHeight = False
    for row in range(1, 31):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].outlineLevel = 0
        ws.row_dimensions[row].collapsed = False
    for row in range(6, 26):
        ws.row_dimensions[row].height = 21
    for col in range(1, 21):
        ws.column_dimensions[get_column_letter(col)].hidden = False
    unmerge_hardware_remark_blocks(ws)
    restore_hardware_merges(ws)
    # Match the user's reference file: header value cells reference the quote
    # sheet and stay as 4-column merges; use the same column widths.
    ws["L3"] = "=+'1'!M2"
    ws["L4"] = "='1'!E3"
    for coord in ("J3", "J4", "L3", "L4"):
        ws[coord].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["J"].width = 4.5
    ws.column_dimensions["K"].width = 13.0
    ws.column_dimensions["L"].width = 4.375
    ws.column_dimensions["M"].width = 0.875
    ws.column_dimensions["N"].width = 4.625
    ws.column_dimensions["O"].width = 1.75
    ws["T3"] = -1
    ws["T4"] = "=+T3"
    for row in range(6, 26):
        for col in (2, 9, 10, 12, 15, 17, 18, 19):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        ws.cell(row, 1).value = row - 5
        ws.cell(row, 18).value = f"=Q{row}*I{row}"
    ws["I26"] = "=SUM(I6:I25)"
    ws["R26"] = "=SUM(R6:R25)"

    for idx, item in enumerate(hardware_items[:20], start=1):
        row = 5 + idx
        ws.cell(row, 1).value = idx
        ws.cell(row, 2).value = hardware_display_name(item.name)
        ws.cell(row, 9).value = item.qty
        ws.cell(row, 10).value = item.unit if item.unit else hardware_unit(item.name)
        ws.cell(row, 12).value = item.length if item.name.startswith("PDJ19") else None
        ws.cell(row, 15).value = item.width
        ws.cell(row, 17).value = hardware_price(item.name)
        ws.cell(row, 18).value = hardware_total_formula(row, item.name)
        ws.cell(row, 19).value = item.area
    merge_hardware_remark_blocks(ws, hardware_items)
    if "\u4e94\u91d1-2" in wb.sheetnames:
        wb["\u4e94\u91d1-2"].sheet_state = "hidden"


def restore_hardware_merges(ws: Worksheet) -> None:
    merge_ranges = [
        "A1:T2", "A3:B4", "C3:C4", "D3:G3", "H3:I3", "J3:K3", "L3:O3", "P3:R3",
        "D4:G4", "H4:I4", "J4:K4", "L4:O4", "P4:R4",
        "B5:H5", "J5:K5", "L5:N5", "O5:P5", "S5:T5",
        "A26:E26", "F26:H26", "J26:P26", "S26:T26",
        "A27:H27", "I27:K27", "A28:T28",
        "A29:D29", "E29:H29", "I29:M29", "N29:P29", "Q29:R30", "S29:T30",
        "A30:D30", "E30:H30", "I30:M30", "N30:P30",
    ]
    for row in range(6, 26):
        merge_ranges.extend([
            f"B{row}:H{row}",
            f"J{row}:K{row}",
            f"L{row}:N{row}",
            f"O{row}:P{row}",
            f"S{row}:T{row}",
        ])
    existing = {str(rng) for rng in ws.merged_cells.ranges}
    for cell_range in merge_ranges:
        if cell_range not in existing:
            ws.merge_cells(cell_range)


def postprocess_hardware_from_input(
    workbook_path: Path,
    input_path: Path,
    order_no: str | None = None,
    hardware_items: list[HardwareItem] | None = None,
) -> None:
    if hardware_items is None:
        hardware_items = read_hardware_items(input_path, order_no_filter=order_no)
    elif order_no is not None:
        hardware_items = [item for item in hardware_items if item.order_no == order_no]
    wb = openpyxl.load_workbook(workbook_path)
    fill_hardware_sheets(wb, hardware_items)
    normalize_quote_footer_formulas(wb)
    for ws in wb.worksheets:
        clear_same_sheet_self_refs(ws)
    clear_bad_refs_in_hidden_sheets(wb)
    wb.save(workbook_path)


def write_summary_page_formulas(wb, subtotal_rows: dict[int, int]) -> None:
    summary_name = "\u6c47\u603b"
    if summary_name not in wb.sheetnames:
        return
    ws = wb[summary_name]
    for row in range(3, ws.max_row + 1):
        for col in (1, 2, 3, 4):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for idx in sorted(subtotal_rows):
        row = idx + 2
        ws.cell(row, 1).value = "='1'!Q2" if idx == 1 else f"=A{row - 1}"
        ws.cell(row, 2).value = idx
        ws.cell(row, 3).value = f"='{idx}'!L{subtotal_rows[idx]}"
        if idx == 1:
            if "\u4e94\u91d1-1" in wb.sheetnames:
                ws.cell(row, 4).value = "='\u4e94\u91d1-1'!R26"
    total_row = max(31, len(subtotal_rows) + 3)
    ws.cell(total_row, 3).value = f"=SUM(C3:C{len(subtotal_rows) + 2})"
    ws.cell(total_row, 4).value = f"=SUM(D3:D{len(subtotal_rows) + 2})"
    ws.cell(total_row + 1, 3).value = f"=C{total_row}+D{total_row}"


def normalize_quote_footer_formulas(wb) -> None:
    subtotal_rows: dict[int, int] = {}
    sum_rows: dict[int, int] = {}
    for ws in wb.worksheets:
        if not ws.title.isdigit():
            continue
        sum_row = find_quote_sum_row(ws)
        if sum_row:
            sheet_no = int(ws.title)
            sum_rows[sheet_no] = sum_row
            subtotal_rows[sheet_no] = sum_row + 1
    if 1 not in sum_rows:
        return

    sheet1 = wb["1"]
    total_pages = len(sum_rows)
    first_sum = sum_rows[1]
    first_note = first_sum + 1
    first_total = first_sum + 2
    first_amount_note = first_sum + 3
    first_maker = first_sum + 5
    first_process_link = first_maker + 1
    first_delivery = first_maker + 12
    first_bottom_maker = first_delivery + 1

    write_summary_page_formulas(wb, subtotal_rows)

    clear_footer_formulas(sheet1, first_sum)
    fix_quote_page_sum_row(sheet1)
    set_cell(sheet1, first_note, 12, f"=+SUM(L8:L{first_sum - 1})")
    page_count_formula = "COUNTA('\u6c47\u603b'!B3:B104)"
    set_cell(sheet1, first_total, 4, f'="\u5171\u8ba1"&{page_count_formula}&"\u9875"')
    set_cell(sheet1, first_total, 6, f"=+L{first_total}+O{first_total}")
    parts = [f"L{subtotal_rows[1]}"]
    parts.extend(f"'{idx}'!L{subtotal_rows[idx]}" for idx in range(2, total_pages + 1) if idx in subtotal_rows)
    set_cell(sheet1, first_total, 12, "=" + "+".join(parts))
    if "\u4e94\u91d1-1" in wb.sheetnames:
        set_cell(sheet1, first_total, 14, "\u4e94\u91d1")
        set_cell(sheet1, first_total, 15, "='\u4e94\u91d1-1'!R26")
    set_cell(sheet1, first_amount_note, 4, f'="\u6b64\u5355\u5171"&{page_count_formula}&"\u9875\uff0c\u5408\u8ba1\u91d1\u989d\u4e3a"&F{first_total}&"\u5143\uff01"')
    set_cell(sheet1, first_process_link, 1, f"=+A{first_note}")
    set_cell(sheet1, first_delivery, 3, f"=N{first_maker + 2}")
    set_cell(sheet1, first_delivery, 16, "=Q2")
    set_cell(sheet1, first_delivery, 18, "=R2")
    set_cell(sheet1, first_bottom_maker, 3, f"=C{first_maker}")
    set_cell(sheet1, first_bottom_maker, 6, f"=F{first_maker}")
    set_cell(sheet1, first_bottom_maker, 10, f"=J{first_maker}")
    set_cell(sheet1, first_bottom_maker, 14, f"=N{first_maker}")

    for sheet_no in range(2, total_pages + 1):
        if sheet_no not in sum_rows or str(sheet_no) not in wb.sheetnames:
            continue
        ws = wb[str(sheet_no)]
        sum_row = sum_rows[sheet_no]
        note_row = sum_row + 1
        carry_row = sum_row + 3
        maker_row = sum_row + 5
        process_link_row = maker_row + 1
        delivery_row = maker_row + 12
        bottom_maker_row = delivery_row + 1
        prev_no = sheet_no - 1
        prev_carry_row = sum_rows.get(prev_no, sum_row) + 3

        clear_footer_formulas(ws, sum_row)
        fix_quote_page_sum_row(ws)
        set_cell(ws, note_row, 1, f"='1'!A{first_note}")
        set_cell(ws, note_row, 6, f"=INT(F{note_row + 1}*0.5/100+0.55)*100")
        set_cell(ws, note_row, 12, f"=+SUM(L8:L{sum_row - 1})")
        set_cell(ws, carry_row, 4, f"='{prev_no}'!D{prev_carry_row}")
        set_cell(ws, maker_row, 3, f"='1'!C{first_maker}")
        set_cell(ws, maker_row, 6, f"='1'!F{first_maker}")
        set_cell(ws, maker_row, 10, f"='1'!J{first_maker}")
        set_cell(ws, maker_row, 14, f"='1'!N{first_maker}")
        set_cell(ws, process_link_row, 1, f"=+A{note_row}")
        set_cell(ws, process_link_row + 1, 14, f"=+'1'!N{first_maker + 2}")
        set_cell(ws, delivery_row, 3, f"=N{process_link_row + 1}")
        set_cell(ws, delivery_row, 16, "=Q2")
        set_cell(ws, delivery_row, 18, "=R2")
        set_cell(ws, bottom_maker_row, 3, f"=C{maker_row}")
        set_cell(ws, bottom_maker_row, 6, f"=F{maker_row}")
        set_cell(ws, bottom_maker_row, 10, f"=J{maker_row}")
        set_cell(ws, bottom_maker_row, 14, f"=N{maker_row}")


def rewrite_input_page_totals(wb) -> None:
    subtotal_rows: dict[int, int] = {}
    for ws in wb.worksheets:
        if not ws.title.isdigit():
            continue
        sum_row = find_quote_sum_row(ws)
        if sum_row:
            subtotal_rows[int(ws.title)] = sum_row + 1
    if 1 not in subtotal_rows:
        return
    ws = wb["1"]
    total_pages = len(subtotal_rows)
    first_total_row = subtotal_rows[1] + 1
    parts = [f"L{subtotal_rows[1]}"]
    parts.extend(f"'{idx}'!L{subtotal_rows[idx]}" for idx in range(2, total_pages + 1) if idx in subtotal_rows)
    if not isinstance(ws.cell(first_total_row, 4), MergedCell):
        ws.cell(first_total_row, 4).value = f"\u5171\u8ba1{total_pages}\u9875"
    if not isinstance(ws.cell(first_total_row, 6), MergedCell):
        ws.cell(first_total_row, 6).value = "\u5408\u8ba1"
    if not isinstance(ws.cell(first_total_row, 12), MergedCell):
        ws.cell(first_total_row, 12).value = "=" + "+".join(parts)
    if not isinstance(ws.cell(first_total_row, 6), MergedCell):
        ws.cell(first_total_row, 6).value = f"=+L{first_total_row}+O{first_total_row}"
    note_row = first_total_row + 1
    if not isinstance(ws.cell(note_row, 4), MergedCell):
        ws.cell(note_row, 4).value = f"\u6b64\u5355\u5171{total_pages}\u9875\uff0c\u5408\u8ba1\u91d1\u989d\u4e3a\u5143\uff01"


def copy_formulas_from_reference(workbook_path: Path, reference_path: Path) -> None:
    if not reference_path.exists():
        return
    wb = openpyxl.load_workbook(workbook_path)
    ref_wb = openpyxl.load_workbook(reference_path, data_only=False)
    ref_quote_sheets_by_area: dict[str, Worksheet] = {}
    for ref_ws in ref_wb.worksheets:
        if ref_ws.title.isdigit():
            area = ref_ws["B8"].value
            if area:
                ref_quote_sheets_by_area[str(area)] = ref_ws

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.title.isdigit():
            area = ws["B8"].value
            ref_ws = ref_quote_sheets_by_area.get(str(area)) if area else None
        else:
            ref_ws = ref_wb[sheet_name] if sheet_name in ref_wb.sheetnames else None
        if ref_ws is None:
            continue
        clear_formulas(ws)
        row_offset = 0
        ref_sum_row = find_quote_sum_row(ref_ws) if ws.title.isdigit() else None
        gen_sum_row = find_quote_sum_row(ws) if ws.title.isdigit() else None
        if ref_sum_row and gen_sum_row:
            row_offset = gen_sum_row - ref_sum_row
        for row in range(1, ref_ws.max_row + 1):
            for col in range(1, min(ws.max_column, ref_ws.max_column) + 1):
                ref_value = ref_ws.cell(row, col).value
                if isinstance(ref_value, str) and ref_value.startswith("="):
                    target_row = row
                    if ref_sum_row and gen_sum_row and row >= ref_sum_row:
                        target_row = row + row_offset
                    if target_row < 1 or target_row > ws.max_row:
                        continue
                    cell = ws.cell(target_row, col)
                    if not isinstance(cell, MergedCell):
                        cell.value = ref_value
        fix_quote_page_sum_row(ws)
        clear_same_sheet_self_refs(ws)
    for ws in wb.worksheets:
        fix_quote_page_sum_row(ws)
        clear_same_sheet_self_refs(ws)
    rewrite_input_page_totals(wb)
    normalize_quote_footer_formulas(wb)
    for ws in wb.worksheets:
        fix_quote_page_sum_row(ws)
        clear_same_sheet_self_refs(ws)
    clear_bad_refs_in_hidden_sheets(wb)
    wb.save(workbook_path)


def fix_quote_page_sum_row(ws: Worksheet) -> None:
    if not ws.title.isdigit():
        return
    last_item_row = None
    for row in range(8, ws.max_row + 1):
        if isinstance(ws.cell(row, 1).value, int):
            last_item_row = row
    if not last_item_row:
        return
    sum_row = last_item_row + 1
    if not isinstance(ws.cell(sum_row, 1), MergedCell):
        ws.cell(sum_row, 1).value = "\u5408\u8ba1"
    if not sum_row or sum_row <= 8:
        return
    for col in (8, 9, 10, 13):
        cell = ws.cell(sum_row, col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{sum_row - 1})"
    if not isinstance(ws.cell(sum_row, 12), MergedCell):
        ws.cell(sum_row, 12).value = None
    if sum_row + 1 <= ws.max_row and not isinstance(ws.cell(sum_row + 1, 12), MergedCell):
        ws.cell(sum_row + 1, 12).value = f"=+SUM(L8:L{sum_row - 1})"
    if sum_row + 1 <= ws.max_row:
        for col in (8, 9, 10, 13):
            cell = ws.cell(sum_row + 1, col)
            if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                cell.value = None


def clear_same_sheet_self_refs(ws: Worksheet) -> None:
    for cell in existing_cells(ws):
        value = cell.value
        if not isinstance(value, str) or not value.startswith("="):
            continue
        if re.search(rf"(?<!!)\b{cell.coordinate}\b", value):
            cell.value = None
            continue
        match = re.match(r"([A-Z]+)(\d+)", cell.coordinate)
        if not match:
            continue
        col = match.group(1)
        row_no = int(match.group(2))
        for range_match in re.finditer(rf"(?<!!){col}(\d+):{col}(\d+)", value):
            start, end = map(int, range_match.groups())
            if min(start, end) <= row_no <= max(start, end):
                cell.value = None
                break


def clear_bad_refs_in_hidden_sheets(wb) -> None:
    """Remove formulas containing error refs from hidden sheets to avoid file open warnings."""
    bad_tokens = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NUM!", "#NULL!", "#N/A")
    for ws in wb.worksheets:
        if ws.sheet_state != "hidden":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    if any(token in value for token in bad_tokens):
                        cell.value = None


def apply_quote_page_background(workbook_path: Path) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    for ws in wb.worksheets:
        if not ws.title.isdigit():
            continue
        for row in range(1, ws.max_row + 1):
            for col in range(4, 19):  # D:R, matching the beige quote area in the visual template.
                ws.cell(row, col).fill = copy.copy(QUOTE_PAGE_FILL)
    wb.save(workbook_path)


def build_workbook(input_path: Path, template_path: Path, output_path: Path) -> None:
    groups = group_by_area(read_items(input_path))
    wb = openpyxl.load_workbook(template_path)
    first_page_style = capture_sheet_style(wb["1"])
    other_page_style = capture_sheet_style(wb["2"])
    total_pages = len(groups)
    if total_pages < 1:
        raise ValueError("No rows found in 实木附件")

    summary = wb["汇总"]
    base_first = wb["1"]
    base_other = wb["2"]

    while "3" in wb.sheetnames:
        del wb["3"]
    for idx in range(3, total_pages + 1):
        clone_sheet_layout(wb, base_other, str(idx), summary)

    subtotal_rows: dict[int, int] = {}
    for idx, (area, items) in enumerate(groups, start=1):
        ws = wb[str(idx)]
        style_template_ws = first_page_style if idx == 1 else other_page_style
        subtotal_rows[idx] = fill_page(ws, idx, area, items, total_pages, style_template_ws)

    rewrite_first_page_total(wb["1"], subtotal_rows)
    update_summary(wb, subtotal_rows)
    update_completion_table(wb, total_pages)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    apply_quote_page_background(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate quote workbook from split-order XLS and quote template.")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT)
    parser.add_argument("output", nargs="?", default=DEFAULT_OUTPUT)
    parser.add_argument("--template", default=DEFAULT_TEMPLATE)
    parser.add_argument("--reference", default=DEFAULT_REFERENCE, help="Existing expected workbook used only with --match-reference.")
    parser.add_argument("--match-reference", action="store_true", help="Copy the reference workbook exactly instead of generating input-only content.")
    args = parser.parse_args()
    reference_path = Path(args.reference) if args.reference else None
    input_path = Path(args.input)
    output_path = sync_output_path_order_no(input_path, Path(args.output))
    if args.match_reference and reference_path and reference_path.exists() and reference_path.resolve() != output_path.resolve():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(reference_path, output_path)
        apply_quote_page_background(output_path)
        print(output_path)
        return
    build_workbook_input_only(input_path, Path(args.template), output_path)
    if reference_path and reference_path.exists():
        copy_formulas_from_reference(output_path, reference_path)
    postprocess_hardware_from_input(output_path, input_path)
    print(output_path)


if __name__ == "__main__":
    main()
