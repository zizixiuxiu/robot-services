import sys
from collections import defaultdict
from datetime import time
from openpyxl import load_workbook

sys.path.insert(0, '/app/src')
from generate_attendance import parse_input, _format_dist_time, SUMMARY_DAY_START_COL, SUMMARY_DAY_COLS_PER_DAY


def parse_summary_time(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Normalize time strings to HH:MM
    import re
    m = re.match(r'(\d{1,2}):(\d{2})(?::\d{2})?', s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def load_generated_summary(path, sheet_name=None, max_days=31):
    wb = load_workbook(path, data_only=False)
    if sheet_name is None:
        for sn in wb.sheetnames:
            if '部门分发' in sn:
                sheet_name = sn
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    main_rows = defaultdict(dict)
    extra_rows = defaultdict(dict)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or name == '姓名':
            continue
        target = extra_rows[name] if name in main_rows else main_rows[name]
        for day in range(1, max_days + 1):
            cin_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (day - 1)
            cout_col = cin_col + 1
            cin = parse_summary_time(ws.cell(row=r, column=cin_col).value)
            cout = parse_summary_time(ws.cell(row=r, column=cout_col).value)
            if cin or cout:
                target[day] = (cin, cout)
    wb.close()
    return dict(main_rows), dict(extra_rows)


def expected_main_value(day_data):
    """汇总表 部门分发 主行：取最早和最晚两个时间"""
    times = sorted(set(day_data['all_times']))
    if not times:
        return (None, None)
    if len(times) == 1:
        return (_format_dist_time(times[0]), None)
    return (_format_dist_time(times[0]), _format_dist_time(times[-1]))


def expected_extra_value(day_data):
    """汇总表 部门分发 续行：中间时间（3 个时间取第 2 个，4 个取第 2、3 个）"""
    times = sorted(set(day_data['all_times']))
    if len(times) == 3:
        return (_format_dist_time(times[1]), None)
    if len(times) >= 4:
        return (_format_dist_time(times[1]), _format_dist_time(times[2]))
    return (None, None)


def main():
    input_path = sys.argv[1]
    gen_path = sys.argv[2]
    max_days = int(sys.argv[3]) if len(sys.argv) > 3 else 31
    employees = parse_input(input_path)
    main_rows, extra_rows = load_generated_summary(gen_path, max_days=max_days)

    diffs = []
    for name, emp in employees.items():
        if name not in main_rows:
            diffs.append((name, 'missing in generated'))
            continue
        for day, day_data in emp['days'].items():
            if day > max_days:
                continue
            expected = expected_main_value(day_data)
            actual = main_rows[name].get(day, (None, None))
            if expected != actual:
                diffs.append((name, day, 'main', expected, actual))
            expected_e = expected_extra_value(day_data)
            actual_e = extra_rows.get(name, {}).get(day, (None, None))
            if expected_e != actual_e:
                diffs.append((name, day, 'extra', expected_e, actual_e))

    for name in main_rows:
        if name not in employees:
            diffs.append((name, 'extra in generated'))

    print(f'Total diffs vs input: {len(diffs)}')
    for d in diffs[:30]:
        print(d)
    if not diffs:
        print('OK: generated summary distribution matches input')


if __name__ == '__main__':
    main()
