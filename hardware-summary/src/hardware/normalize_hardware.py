"""
五金表标准化转换脚本

功能：将各种格式的五金表转换为标准格式
标准格式参考：1五、2五、3五（作为期望结果）

TDD开发方式：
1. 分析1-3五的标准格式 → 定义期望结果
2. 分析4-27五金的实际格式 → 定义输入
3. 编写转换逻辑 → 使输入匹配期望结果
"""

import xlrd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from copy import copy
from pathlib import Path
import re


def analyze_standard_format(book, sheet_name='1五'):
    """
    分析标准五金表的格式结构（TDD: 期望结果）
    返回标准格式的元数据
    """
    sheet = book.sheet_by_name(sheet_name)
    
    standard = {
        'title': sheet.cell_value(0, 0),  # 标题
        'header_rows': [],  # 表头信息行（品牌、经销商等）
        'column_headers': {},  # 列标题映射
        'data_start_row': 5,  # 数据开始行
        'footer_rows': []  # 底部信息行
    }
    
    # 分析表头信息行（Row 2-3）
    for row in range(2, 4):
        row_info = {}
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val:
                row_info[col] = val
        standard['header_rows'].append(row_info)
    
    # 分析列标题（Row 4）
    for col in range(sheet.ncols):
        val = sheet.cell_value(4, col)
        if val:
            standard['column_headers'][col] = val
    
    # 分析数据行结构
    data_rows = []
    for row in range(5, sheet.nrows):
        row_data = {}
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val:
                row_data[col] = val
        if row_data:
            data_rows.append(row_data)
    standard['data_rows'] = data_rows
    
    return standard


def extract_hardware_data(book, product_sheet_name):
    """
    从产品表中提取五金相关信息
    
    由于8-27的产品表是空的，这个函数主要用于：
    1. 读取现有五金表的数据
    2. 标准化格式
    """
    # 获取对应的五金表名
    if product_sheet_name.isdigit():
        hardware_sheet_name = product_sheet_name + '五'
    else:
        return None
    
    if hardware_sheet_name not in book.sheet_names():
        return None
    
    sheet = book.sheet_by_name(hardware_sheet_name)
    
    # 提取数据行
    data = []
    for row in range(5, sheet.nrows):
        # 检查是否有实际数据（不只是序号）
        name = sheet.cell_value(row, 1)
        if name and str(name).strip():
            data.append({
                'row': row,
                'seq': sheet.cell_value(row, 0),
                'name': name,
                'quantity': sheet.cell_value(row, 7),
                'unit': sheet.cell_value(row, 9),
                'length': sheet.cell_value(row, 11),
                'width': sheet.cell_value(row, 14),
                'price': sheet.cell_value(row, 16),
                'total': sheet.cell_value(row, 17),
                'remark': sheet.cell_value(row, 18)
            })
    
    return data


def create_standard_hardware_sheet(wb, sheet_name, source_data=None, 
                                   dealer_info=None, order_info=None):
    """
    创建标准格式的五金表
    
    Args:
        wb: openpyxl Workbook
        sheet_name: 工作表名称
        source_data: 源数据（从现有表提取的数据）
        dealer_info: 经销商信息
        order_info: 订单信息
    """
    ws = wb.create_sheet(title=sheet_name)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 3
    ws.column_dimensions['J'].width = 6
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 3
    ws.column_dimensions['M'].width = 3
    ws.column_dimensions['N'].width = 3
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 3
    ws.column_dimensions['Q'].width = 8
    ws.column_dimensions['R'].width = 8
    ws.column_dimensions['S'].width = 15
    ws.column_dimensions['T'].width = 3
    
    # Row 0: 标题
    ws.merge_cells('A1:T1')
    ws['A1'] = '定制家居五金清单'
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Row 1: 空行
    ws.row_dimensions[2].height = 15
    
    # Row 2: 品牌、经销商、日期、订单号
    ws['A3'] = '品牌：'
    ws['D3'] = '经销商姓名'
    ws['G3'] = dealer_info.get('name', '') if dealer_info else ''
    ws['J3'] = '接单日期'
    ws['L3'] = order_info.get('date', '') if order_info else ''
    ws['P3'] = '订单编号'
    ws['S3'] = order_info.get('order_no', '') if order_info else ''
    
    # Row 3: 地址、客户、区域
    ws['D4'] = '经销商地址'
    ws['G4'] = dealer_info.get('address', '') if dealer_info else ''
    ws['J4'] = '终端客户'
    ws['L4'] = order_info.get('customer', '') if order_info else ''
    ws['P4'] = '区域'
    ws['S4'] = order_info.get('area', '') if order_info else ''
    
    # Row 4: 表头
    headers = ['序号', '五金名称', '', '', '', '', '', '数量', '', '单位', 
               '长度(mm)', '', '', '宽度（mm）', '', '单价', '总价', '备注', '', '']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(name='微软雅黑', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 20
    
    # Row 5+: 数据行
    if source_data:
        for idx, item in enumerate(source_data, 1):
            row = 5 + idx
            ws.cell(row=row, column=1, value=idx)  # 序号
            ws.cell(row=row, column=2, value=item.get('name', ''))  # 五金名称
            ws.cell(row=row, column=8, value=item.get('quantity', ''))  # 数量
            ws.cell(row=row, column=10, value=item.get('unit', ''))  # 单位
            ws.cell(row=row, column=11, value=item.get('length', ''))  # 长度
            ws.cell(row=row, column=15, value=item.get('width', ''))  # 宽度
            ws.cell(row=row, column=17, value=item.get('price', ''))  # 单价
            ws.cell(row=row, column=18, value=item.get('total', ''))  # 总价
            ws.cell(row=row, column=19, value=item.get('remark', ''))  # 备注
            ws.row_dimensions[row].height = 18
    
    # 底部: 数量总计
    footer_row = 5 + len(source_data) + 1 if source_data else 25
    ws.cell(row=footer_row, column=5, value='数量总计：')
    ws.cell(row=footer_row, column=17, value='总价格')
    
    # 制表人
    ws.cell(row=footer_row + 2, column=1, value='制表人')
    ws.cell(row=footer_row + 2, column=4, value='陈昌容')
    ws.cell(row=footer_row + 2, column=9, value='设计师')
    ws.cell(row=footer_row + 2, column=17, value='经销商确认')
    
    return ws


def normalize_hardware_sheets(input_path, output_path):
    """
    主函数：标准化所有五金表
    
    1. 分析1-3五的标准格式（期望结果）
    2. 读取所有五金表
    3. 转换为标准格式
    4. 输出到新的Excel文件
    """
    book = xlrd.open_workbook(input_path)
    
    # TDD Step 1: 分析标准格式
    print("TDD Step 1: 分析标准格式...")
    standard = analyze_standard_format(book, '1五')
    print(f"  标准格式: {standard['title']}")
    print(f"  列标题: {standard['column_headers']}")
    print(f"  数据行数: {len(standard['data_rows'])}")
    
    # 创建新的Workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 提取经销商和订单信息（从1五）
    dealer_info = {
        'name': book.sheet_by_name('1五').cell_value(2, 6),
        'address': book.sheet_by_name('1五').cell_value(3, 6)
    }
    order_info = {
        'date': book.sheet_by_name('1五').cell_value(2, 11),
        'order_no': book.sheet_by_name('1五').cell_value(2, 18),
        'customer': book.sheet_by_name('1五').cell_value(3, 11),
        'area': book.sheet_by_name('1五').cell_value(3, 18)
    }
    
    # 处理所有五金表
    print("\nTDD Step 2: 处理所有五金表...")
    for sheet_name in book.sheet_names():
        if '五' in sheet_name or '五金' in sheet_name:
            print(f"  处理: {sheet_name}")
            
            # 提取现有数据
            product_num = sheet_name.replace('五', '').replace('五金', '')
            data = extract_hardware_data(book, product_num)
            
            # 创建标准格式sheet
            create_standard_hardware_sheet(
                wb, 
                sheet_name,
                source_data=data,
                dealer_info=dealer_info,
                order_info=order_info
            )
    
    # 保存
    wb.save(output_path)
    print(f"\n完成！输出文件: {output_path}")
    print(f"共处理 {len(wb.sheetnames)} 个五金表")


if __name__ == '__main__':
    input_path = "/mnt/c/Users/Administrator/Downloads/B2605-4019陈时伟（滨河8楼）.xls"
    output_path = "/mnt/c/Users/Administrator/Downloads/B2605-4019陈时伟（滨河8楼）_五金标准化.xlsx"
    
    normalize_hardware_sheets(input_path, output_path)
