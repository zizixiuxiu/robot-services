#!/usr/bin/env python3
"""
以 6 月模板为基础，生成 31 天的通用 base 模板：
- base/汇总.xlsx
- base/分发.xlsx

base 模板用于没有专用模板的月份，保证格式、公式与 6 月一致。
"""
import os
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
JUNE_DIR = BASE_DIR / "templates" / "6月"
BASE_OUT = BASE_DIR / "templates" / "base"
BASE_OUT.mkdir(parents=True, exist_ok=True)


def _copy_col_style(ws, src_col: int, dst_col: int):
    """复制整列样式（列宽 + 每个单元格的样式）"""
    src_letter = get_column_letter(src_col)
    dst_letter = get_column_letter(dst_col)
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    for r in range(1, ws.max_row + 1):
        src = ws.cell(row=r, column=src_col)
        dst = ws.cell(row=r, column=dst_col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def build_base_summary():
    src = JUNE_DIR / "汇总.xlsx"
    wb = load_workbook(src)
    ws = wb['6月部门分发']

    # 6 月模板：day 30 占用 4 列（122-125），之后是总小时/总天数列（126-127）
    # 在 day30 后面插入 4 列作为 day31
    insert_after = 125  # day30 的最后一列
    ws.insert_cols(insert_after + 1, 4)

    # 把 day30 的 4 列样式复制给 day31
    for offset in range(4):
        _copy_col_style(ws, 122 + offset, 126 + offset)

    # 更新 day31 表头：第 1 行日期、第 2 行标签
    ws.cell(row=1, column=126).value = None  # 日期由生成时根据月份填充，这里清空
    ws.cell(row=1, column=127).value = None
    ws.cell(row=1, column=128).value = "小时数"
    ws.cell(row=1, column=129).value = "天数"

    # 更新汇总表公式到 day31
    ws_sum = wb['汇总']
    for r in range(2, ws_sum.max_row + 1):
        ws_sum.cell(row=r, column=8).value = "=SUMIF('6月部门分发'!B:B,A:A,'6月部门分发'!DX:DX)"
        ws_sum.cell(row=r, column=9).value = f"=H{r}*60"
        ws_sum.cell(row=r, column=10).value = "=SUMIF('6月部门分发'!B:B,A:A,'6月部门分发'!DY:DY)"
        ws_sum.cell(row=r, column=18).value = f"=J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}+Q{r}"

    out = BASE_OUT / "汇总.xlsx"
    if out.exists():
        out.unlink()
    wb.save(out)
    wb.close()
    print("Created", out)


def build_base_distribution():
    src = JUNE_DIR / "分发.xlsx"
    wb = load_workbook(src)
    ws = wb['6月部门分发']

    # 6 月模板：day 30 在 col 35 (AI)，在 col 35 后插入 1 列作为 day31
    ws.insert_cols(36, 1)
    _copy_col_style(ws, 35, 36)

    # 表头：第 2 行是数字表头，第 1 行是标题
    ws.cell(row=2, column=36).value = "31"

    out = BASE_OUT / "分发.xlsx"
    if out.exists():
        out.unlink()
    wb.save(out)
    wb.close()
    print("Created", out)


if __name__ == "__main__":
    build_base_summary()
    build_base_distribution()
