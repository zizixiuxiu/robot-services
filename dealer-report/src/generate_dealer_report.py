#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
经销商数据报表生成脚本
从"账号指标"和"设计师数据统计"两个数据源生成格式统一的经销商报表。
支持两种格式：
  - old / 旧版：与原 3月/4月/6.2 经销商数据表格式一致
  - new / 新版：与 6月经销商酷家乐数据(新) 格式一致，含大区合计、总合计、美化样式
"""

import argparse
import re
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import urllib.request
import os
import json

from dealer_analysis import prepare_dealer_df, add_analysis_sheets


# ==================== 配置项 ====================

DEFAULT_REGION_ORDER = ['东北区', '华北区', '华东区', '华南区', '华中区', '西北区', '西南区']

# 新版格式颜色方案
COLOR_TITLE_BG = '1F4E78'      # 深蓝
COLOR_HEADER_BG = '4472C4'     # 蓝色
COLOR_REGION_TOTAL_BG = 'D9E1F2'  # 浅蓝
COLOR_GRAND_TOTAL_BG = '1F4E78'   # 深蓝（总合计，与标题呼应）
COLOR_ZEBRA = 'F2F2F2'         # 浅灰
COLOR_WHITE = 'FFFFFF'
COLOR_BORDER = 'BFBFBF'

# 字体样式
FONT_TITLE = Font(name='黑体', size=18, bold=True, color='FFFFFF')
FONT_HEADER = Font(name='黑体', size=11, bold=True, color='FFFFFF')
FONT_DATA = Font(name='微软雅黑', size=10)
FONT_TOTAL = Font(name='微软雅黑', size=11, bold=True)
FONT_TOTAL_BLUE = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
FONT_TOTAL_WHITE = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')

# 边框样式
THIN_BORDER = Border(
    left=Side(style='thin', color='FF' + COLOR_BORDER),
    right=Side(style='thin', color='FF' + COLOR_BORDER),
    top=Side(style='thin', color='FF' + COLOR_BORDER),
    bottom=Side(style='thin', color='FF' + COLOR_BORDER)
)

# 居中对齐
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)


# ==================== 填充样式对象 ====================

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


# ==================== 数据清洗函数 ====================

def clean_dept6_name(row):
    """
    清洗数据源中的经销商名称。
    优先使用"部门6"，为空时回退到"账号名称"。
    去掉末尾的数字标记，如：
      '冯海东3/4' -> '冯海东'
      '张守芹4'   -> '张守芹'
      '张晋芳 2'  -> '张晋芳'
    """
    dept6 = row['部门6']
    account = row['账号名称']
    if pd.notna(dept6):
        s = str(dept6).strip()
    else:
        s = str(account).strip() if pd.notna(account) else ''

    s = re.sub(r'\s+\d+/\d+$', '', s)
    s = re.sub(r'\s+\d+$', '', s)
    s = re.sub(r'\d+/\d+$', '', s)
    s = re.sub(r'\d+$', '', s)
    s = s.strip()
    return s


def clean_designer_dept6(s):
    """
    清洗设计师数据中的部门名称最后一级。
    过滤掉纯区域名称（如"华东区+上海..."）和空值。
    """
    if pd.isna(s):
        return ''
    s = str(s).strip()
    if re.match(r'^[\u4e00-\u9fa5]{2}区\+', s) or s == '':
        return ''

    s = re.sub(r'\s+\d+/\d+$', '', s)
    s = re.sub(r'\s+\d+$', '', s)
    s = re.sub(r'\d+/\d+$', '', s)
    s = re.sub(r'\d+$', '', s)
    s = s.strip()
    return s


# ==================== 核心处理逻辑 ====================

def process_account_data(account_file, sheet_name='账号信息明细'):
    """
    读取账号指标数据，清洗并聚合。
    返回 DataFrame：['区域', '经销商', '创建方案数', '新增渲染方案数']

    过滤规则：
    - 只保留"部门5"包含"+"号的行（即正常的大区+省份格式），
      这样可以排除内部部门（如直营店、研发部、设计支持部等）
      以及其他非经销商账号。
    """
    df = pd.read_excel(account_file, sheet_name=sheet_name)

    df = df[df['部门5'].astype(str).str.contains(r'\+', na=False)].copy()

    df['经销商'] = df.apply(clean_dept6_name, axis=1)
    df['区域'] = df['部门5'].str.extract(r'^([^+]+)')

    agg = df.groupby(['区域', '经销商']).agg({
        '创建方案数': 'sum',
        '新增渲染方案数': 'sum'
    }).reset_index()

    agg['区域_order'] = agg['区域'].apply(
        lambda x: DEFAULT_REGION_ORDER.index(x) if x in DEFAULT_REGION_ORDER else 999
    )
    agg = agg.sort_values(['区域_order', '经销商']).drop(
        columns='区域_order'
    ).reset_index(drop=True)

    return agg


def process_designer_data(designer_file, sheet_name='设计师数据统计'):
    """
    读取设计师数据统计，提取并聚合提审方案数。
    返回 DataFrame：['经销商', '提审方案数']
    """
    df = pd.read_excel(designer_file, sheet_name=sheet_name)

    df['部门6_raw'] = df['部门名称'].str.split('/', n=4).str[-1]
    df['经销商'] = df['部门6_raw'].apply(clean_designer_dept6)
    df = df[df['经销商'] != '']

    agg = df.groupby('经销商').agg({'提审方案数': 'sum'}).reset_index()
    return agg


def build_output_df(account_agg, designer_agg):
    """
    合并账号数据和设计师数据，构建最终输出 DataFrame。
    """
    out = pd.DataFrame()
    out['区域'] = account_agg['区域']
    out['经销商'] = account_agg['经销商']
    out['创建方案数'] = account_agg['创建方案数']
    out['新增渲染方案数'] = account_agg['新增渲染方案数']
    out['方案深化占比'] = ''
    out['提审方案数'] = ''

    out = pd.merge(out, designer_agg, on='经销商', how='left')
    out['提审方案数'] = out['提审方案数_y'].fillna(0).astype(int)
    out = out.drop(columns=['提审方案数_x', '提审方案数_y'])

    return out


# ==================== 通用写入辅助 ====================

def _generate_analysis_data(df):
    """生成本地分析数据摘要"""
    data_df = df.copy()
    for col in ['创建方案数', '新增渲染方案数', '提审方案数']:
        if col in data_df.columns:
            data_df[col] = pd.to_numeric(data_df[col], errors='coerce').fillna(0)

    total_create = int(data_df['创建方案数'].sum())
    total_render = int(data_df['新增渲染方案数'].sum())
    total_submit = int(data_df['提审方案数'].sum())
    ratio_create = total_submit / total_create if total_create > 0 else 0
    ratio_render = total_submit / total_render if total_render > 0 else 0

    region_stats = data_df.groupby('区域').agg({
        '创建方案数': 'sum',
        '新增渲染方案数': 'sum',
        '提审方案数': 'sum',
        '经销商': 'count'
    }).rename(columns={'经销商': '经销商数'})
    region_stats['提审/创建'] = region_stats['提审方案数'] / region_stats['创建方案数']
    region_stats['提审/渲染'] = region_stats['提审方案数'] / region_stats['新增渲染方案数']
    region_stats = region_stats.sort_values('提审方案数', ascending=False)

    top_dealers = data_df.nlargest(5, '提审方案数')[['区域', '经销商', '创建方案数', '新增渲染方案数', '提审方案数']]

    low_conv = data_df[data_df['创建方案数'] >= 20].copy()
    low_conv['转化率'] = low_conv['提审方案数'] / low_conv['创建方案数']
    low_conv = low_conv[low_conv['转化率'] <= 0.1].sort_values('创建方案数', ascending=False)

    # 漏斗数据
    funnel = {
        'total_create': total_create,
        'total_render': total_render,
        'total_submit': total_submit,
        'render_rate': total_render / total_create if total_create > 0 else 0,
        'submit_rate': total_submit / total_create if total_create > 0 else 0,
        'render_to_submit': total_submit / total_render if total_render > 0 else 0,
    }

    # 转化分层（仅统计有创建的经销商）
    dealer_conv = data_df[data_df['创建方案数'] > 0].copy()
    dealer_conv['转化率'] = dealer_conv['提审方案数'] / dealer_conv['创建方案数']
    tier_defs = [
        ('高转化（≥20%）', dealer_conv['转化率'] >= 0.2),
        ('中转化（10%-20%）', (dealer_conv['转化率'] >= 0.1) & (dealer_conv['转化率'] < 0.2)),
        ('低转化（<10% 且 >0%）', (dealer_conv['转化率'] > 0) & (dealer_conv['转化率'] < 0.1)),
        ('无转化（0%）', dealer_conv['转化率'] == 0),
    ]
    conversion_tiers = [(name, int(mask.sum())) for name, mask in tier_defs]

    # 沉睡经销商：创建>=20 且 提审=0
    sleeping = data_df[(data_df['创建方案数'] >= 20) & (data_df['提审方案数'] == 0)].sort_values(
        '创建方案数', ascending=False
    )[['区域', '经销商', '创建方案数', '新增渲染方案数', '提审方案数']]

    # 区域贡献
    region_contrib = region_stats.copy()
    region_contrib['创建占比'] = region_contrib['创建方案数'] / total_create if total_create > 0 else 0
    region_contrib['渲染占比'] = region_contrib['新增渲染方案数'] / total_render if total_render > 0 else 0
    region_contrib['提审占比'] = region_contrib['提审方案数'] / total_submit if total_submit > 0 else 0

    return {
        'total_create': total_create,
        'total_render': total_render,
        'total_submit': total_submit,
        'ratio_create': ratio_create,
        'ratio_render': ratio_render,
        'region_stats': region_stats,
        'top_dealers': top_dealers,
        'low_conv': low_conv,
        'funnel': funnel,
        'conversion_tiers': conversion_tiers,
        'sleeping': sleeping,
        'region_contrib': region_contrib,
    }


ANALYSIS_JSON_SCHEMA = """{
  "核心结论": "200字以内的总体判断，指出转化效率和主要问题",
  "关键指标": {
    "提审/创建": "如：17.00%",
    "提审/渲染": "如：28.17%",
    "最佳大区": "如：华东区（23.94%）",
    "最差大区": "如：西北区（11.93%）"
  },
  "漏斗分析": {
    "创建方案数": 1253,
    "渲染方案数": 756,
    "提审方案数": 213,
    "渲染率": "60.33%",
    "提审率": "17.00%",
    "渲染到提审": "28.17%",
    "ai建议": "针对漏斗流失环节的1-2句建议"
  },
  "大区排名": [
    {"大区": "西南区", "提审数": 57, "转化率": "18.45%", "评价": "..."}
  ],
  "区域贡献": [
    {"大区": "西南区", "创建占比": "20.1%", "渲染占比": "18.5%", "提审占比": "26.8%", "贡献评级": "A", "ai建议": "..."}
  ],
  "转化分层": [
    {"分层": "高转化（≥20%）", "经销商数": 20, "占比": "13.6%", "ai建议": "..."}
  ],
  "头部经销商": [
    {"排名": 1, "经销商": "...", "提审数": 26, "亮点": "..."}
  ],
  "沉睡经销商": [
    {"区域": "...", "经销商": "...", "创建方案数": 100, "渲染方案数": 50, "提审方案数": 0, "诊断": "..."}
  ],
  "问题诊断": [
    {"问题": "...", "涉及经销商": "...", "影响": "...", "建议": "..."}
  ],
  "行动建议": [
    "建议1",
    "建议2",
    "建议3"
  ]
}"""


def _build_analysis_prompt(analysis_data):
    """构建给 AI 的分析 prompt，要求返回固定 JSON"""
    lines = [
        '你是一位资深的数据分析师，请对以下经销商数据进行分析。',
        '必须严格按下方 JSON Schema 返回结果，不要输出 JSON 以外的任何内容（不要加 markdown 代码块标记）：',
        '',
        ANALYSIS_JSON_SCHEMA,
        '',
        '## 输入数据',
        f'- 经销商总数：{int(analysis_data["region_stats"]["经销商数"].sum())} 家',
        f'- 创建方案数：{analysis_data["total_create"]}',
        f'- 渲染方案数：{analysis_data["total_render"]}',
        f'- 提审方案数：{analysis_data["total_submit"]}',
        f'- 整体提审/创建：{analysis_data["ratio_create"]:.2%}',
        f'- 整体提审/渲染：{analysis_data["ratio_render"]:.2%}',
        '',
        '## 按大区统计（按提审数降序）',
    ]
    for region, row in analysis_data['region_stats'].iterrows():
        ratio_c = row['提审/创建'] if not pd.isna(row['提审/创建']) else 0
        ratio_r = row['提审/渲染'] if not pd.isna(row['提审/渲染']) else 0
        lines.append(f'- {region}: 创建{int(row["创建方案数"])}, 渲染{int(row["新增渲染方案数"])}, 提审{int(row["提审方案数"])}, 提审/创建{ratio_c:.2%}, 提审/渲染{ratio_r:.2%}')

    lines.extend(['', '## 提审数 TOP5 经销商'])
    for idx, row in analysis_data['top_dealers'].iterrows():
        lines.append(f'- {row["区域"]} {row["经销商"]}: 创建{int(row["创建方案数"])}, 渲染{int(row["新增渲染方案数"])}, 提审{int(row["提审方案数"])}')

    lines.extend(['', '## 高创建低转化（创建>=20 且 提审/创建<=10%，需重点关注）'])
    if len(analysis_data['low_conv']) == 0:
        lines.append('- 无')
    else:
        for idx, row in analysis_data['low_conv'].iterrows():
            lines.append(f'- {row["区域"]} {row["经销商"]}: 创建{int(row["创建方案数"])}, 提审{int(row["提审方案数"])}, 转化率{row["转化率"]:.2%}')

    lines.extend(['', '## 漏斗数据'])
    funnel = analysis_data['funnel']
    lines.append(f'- 创建{int(funnel["total_create"])} → 渲染{int(funnel["total_render"])}（渲染率{funnel["render_rate"]:.2%}）→ 提审{int(funnel["total_submit"])}（提审率{funnel["submit_rate"]:.2%}，渲染到提审{funnel["render_to_submit"]:.2%}）')

    lines.extend(['', '## 经销商转化分层（按创建>0的经销商转化率）'])
    for name, count in analysis_data['conversion_tiers']:
        lines.append(f'- {name}: {count} 家')

    lines.extend(['', '## 沉睡经销商（创建>=20 且 提审=0）'])
    if len(analysis_data['sleeping']) == 0:
        lines.append('- 无')
    else:
        for idx, row in analysis_data['sleeping'].head(10).iterrows():
            lines.append(f'- {row["区域"]} {row["经销商"]}: 创建{int(row["创建方案数"])}, 渲染{int(row["新增渲染方案数"])}, 提审{int(row["提审方案数"])}')

    lines.extend(['', '## 区域贡献（按提审占比降序）'])
    contrib = analysis_data['region_contrib'].sort_values('提审占比', ascending=False)
    for region, row in contrib.iterrows():
        lines.append(f'- {region}: 创建占比{row["创建占比"]:.2%}, 渲染占比{row["渲染占比"]:.2%}, 提审占比{row["提审占比"]:.2%}')

    lines.extend(['', '请确保 JSON 合法、字段完整，分析透彻、建议具体。'])
    return '\n'.join(lines)


def _local_analysis_json(analysis_data):
    """没有 AI 配置时，生成固定 JSON 结构的本地分析"""
    best_region = analysis_data['region_stats'].iloc[0]
    worst_region = analysis_data['region_stats'].iloc[-1]

    region_ranking = []
    for region, row in analysis_data['region_stats'].iterrows():
        ratio_c = row['提审/创建'] if not pd.isna(row['提审/创建']) else 0
        if ratio_c >= 0.2:
            evaluation = '转化优秀'
        elif ratio_c >= 0.15:
            evaluation = '转化良好'
        elif ratio_c > 0:
            evaluation = '转化偏弱'
        else:
            evaluation = '无提审'
        region_ranking.append({
            '大区': region,
            '提审数': int(row['提审方案数']),
            '转化率': f'{ratio_c:.2%}',
            '评价': evaluation,
        })

    top_dealers = []
    for i, (idx, row) in enumerate(analysis_data['top_dealers'].iterrows(), 1):
        top_dealers.append({
            '排名': i,
            '经销商': f'{row["区域"]} {row["经销商"]}',
            '提审数': int(row['提审方案数']),
            '亮点': '提审规模领先' if i <= 2 else '提审贡献突出',
        })

    # 漏斗分析
    funnel = analysis_data['funnel']
    if funnel['render_rate'] < 0.5:
        funnel_advice = '创建到渲染流失严重，建议优化渲染工具培训或渲染流程引导。'
    elif funnel['submit_rate'] < 0.15:
        funnel_advice = '渲染到提审转化偏低，建议加强方案报价、客户跟进和下单激励。'
    else:
        funnel_advice = '漏斗整体健康，建议保持并复制高效打法。'

    # 区域贡献
    region_contrib_rows = []
    contrib_sorted = analysis_data['region_contrib'].sort_values('提审占比', ascending=False)
    for i, (region, row) in enumerate(contrib_sorted.iterrows(), 1):
        if i <= 2:
            rating = 'A'
            advice = '核心贡献区，重点资源倾斜并复制经验。'
        elif row['提审占比'] >= 0.15:
            rating = 'B'
            advice = '贡献稳定，可进一步挖掘潜力。'
        elif row['提审占比'] > 0:
            rating = 'C'
            advice = '贡献偏弱，需诊断转化卡点。'
        else:
            rating = 'D'
            advice = '本月无提审，需紧急激活。'
        region_contrib_rows.append({
            '大区': region,
            '创建占比': f'{row["创建占比"]:.2%}',
            '渲染占比': f'{row["渲染占比"]:.2%}',
            '提审占比': f'{row["提审占比"]:.2%}',
            '贡献评级': rating,
            'ai建议': advice,
        })

    # 转化分层
    total_with_create = sum(c for _, c in analysis_data['conversion_tiers'])
    tier_rows = []
    for name, count in analysis_data['conversion_tiers']:
        pct = count / total_with_create if total_with_create > 0 else 0
        if '高转化' in name:
            advice = '树立标杆，组织经验分享。'
        elif '中转化' in name:
            advice = '重点跟进，推动向高转化跃迁。'
        elif '低转化' in name:
            advice = '排查报价/设计/客户决策卡点，一对一复盘。'
        else:
            advice = '启动预警，排查是否已停止合作或需专项扶持。'
        tier_rows.append({
            '分层': name,
            '经销商数': count,
            '占比': f'{pct:.2%}',
            'ai建议': advice,
        })

    # 沉睡经销商
    sleeping_rows = []
    for idx, row in analysis_data['sleeping'].head(10).iterrows():
        reason = '有渲染无提审，可能在报价/客户确认环节卡单' if row['新增渲染方案数'] > 0 else '创建后未渲染，可能在设计效率或客户意向不足'
        sleeping_rows.append({
            '区域': row['区域'],
            '经销商': row['经销商'],
            '创建方案数': int(row['创建方案数']),
            '渲染方案数': int(row['新增渲染方案数']),
            '提审方案数': int(row['提审方案数']),
            '诊断': reason,
        })

    problems = []
    if len(analysis_data['low_conv']) > 0:
        names = '、'.join([f'{row["区域"]} {row["经销商"]}' for idx, row in analysis_data['low_conv'].head(5).iterrows()])
        problems.append({
            '问题': '高创建低转化',
            '涉及经销商': names,
            '影响': f'共 {len(analysis_data["low_conv"])} 家经销商创建>=20 但转化率<=10%，大量方案未形成订单',
            '建议': '逐家复盘，排查报价、设计、客户跟进等环节卡点',
        })

    zero_regions = analysis_data['region_stats'][analysis_data['region_stats']['提审方案数'] == 0]
    if len(zero_regions) > 0:
        problems.append({
            '问题': '部分大区无提审',
            '涉及经销商': '、'.join(zero_regions.index.tolist()),
            '影响': '该大区本月未产生任何正式订单',
            '建议': '大区经理重点走访，激活经销商下单',
        })

    problems.append({
        '问题': '整体转化率偏低',
        '涉及经销商': '全部经销商',
        '影响': f'整体提审/创建仅 {analysis_data["ratio_create"]:.2%}，大量设计资源投入未转化为营收',
        '建议': '建立转化跟进机制，提升渲染到提审的闭环效率',
    })

    suggestions = [
        f'重点提升 {worst_region.name} 等转化率偏低大区的跟进效率',
        '对高创建低转化经销商进行一对一复盘，找出报价/设计/客户决策卡点',
        f'推广 {best_region.name} 及头部经销商的转化经验',
        '建立"创建→渲染→提审"周跟进机制，减少方案流失',
        '对 0 提审经销商启动预警和专项扶持',
    ]

    return {
        '核心结论': (
            f'本月共 {len(analysis_data["region_stats"])} 个大区、{int(analysis_data["region_stats"]["经销商数"].sum())} 家经销商参与统计，'
            f'创建方案 {analysis_data["total_create"]} 个、渲染 {analysis_data["total_render"]} 个、提审 {analysis_data["total_submit"]} 个。'
            f'整体提审/创建 {analysis_data["ratio_create"]:.2%}、提审/渲染 {analysis_data["ratio_render"]:.2%}，转化率整体偏低。'
            f'{best_region.name} 表现最佳（转化率 {best_region["提审/创建"]:.2%}），{worst_region.name} 转化率最低（{worst_region["提审/创建"]:.2%}）。'
            '建议重点跟进高创建低转化经销商，提升从设计到下单的闭环效率。'
        ),
        '关键指标': {
            '提审/创建': f'{analysis_data["ratio_create"]:.2%}',
            '提审/渲染': f'{analysis_data["ratio_render"]:.2%}',
            '最佳大区': f'{best_region.name}（{best_region["提审/创建"]:.2%}）',
            '最差大区': f'{worst_region.name}（{worst_region["提审/创建"]:.2%}）',
        },
        '漏斗分析': {
            '创建方案数': int(funnel['total_create']),
            '渲染方案数': int(funnel['total_render']),
            '提审方案数': int(funnel['total_submit']),
            '渲染率': f'{funnel["render_rate"]:.2%}',
            '提审率': f'{funnel["submit_rate"]:.2%}',
            '渲染到提审': f'{funnel["render_to_submit"]:.2%}',
            'ai建议': funnel_advice,
        },
        '大区排名': region_ranking,
        '区域贡献': region_contrib_rows,
        '转化分层': tier_rows,
        '头部经销商': top_dealers,
        '沉睡经销商': sleeping_rows,
        '问题诊断': problems,
        '行动建议': suggestions,
    }


def call_ai_analysis(prompt: str) -> str:
    """调用 AI 进行数据分析，未配置 API 时返回空字符串"""
    api_key = os.getenv('AI_API_KEY')
    base_url = os.getenv('AI_BASE_URL', 'https://api.deepseek.com/v1')
    model = os.getenv('AI_MODEL', 'deepseek-v4-flash')

    if not api_key:
        return ''

    try:
        body = json.dumps({
            'model': model,
            'messages': [{'role': 'user', 'content': prompt}],
            'temperature': 0.3,
            'max_tokens': 2000,
        }, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(
            f'{base_url}/chat/completions',
            data=body,
            headers={
                'Content-Type': 'application/json; charset=utf-8',
                'Authorization': f'Bearer {api_key}',
            },
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode('utf-8'))
        return result['choices'][0]['message']['content']
    except Exception as e:
        print(f'AI 分析调用失败: {e}')
        return ''


def _safe_json_loads(text):
    """安全解析 AI 返回的 JSON，尝试去除常见包装"""
    if not text:
        return None
    text = text.strip()
    if text.startswith('```json'):
        text = text[7:]
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        return None


def generate_analysis_json(df) -> dict:
    """生成数据分析 JSON，优先使用 AI，失败时 fallback 本地"""
    analysis_data = _generate_analysis_data(df)
    prompt = _build_analysis_prompt(analysis_data)
    ai_text = call_ai_analysis(prompt)
    ai_json = _safe_json_loads(ai_text)
    if ai_json and isinstance(ai_json, dict) and '核心结论' in ai_json:
        return ai_json
    return _local_analysis_json(analysis_data)


def _set_cell_style(cell, font, fill=None, align=None, border=True):
    cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = THIN_BORDER


def _write_section_title(ws, row, text, col_start=1, col_end=8):
    """写入蓝色章节标题行，A 列显示标题，同行其他列留空但统一样式"""
    section_fill = fill(COLOR_HEADER_BG)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c, value=text if c == col_start else '')
        cell.font = FONT_HEADER
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def _write_content_row(ws, row, label, value, label_col=1, label_end=2, value_col=3, value_end=8):
    """写入标签+内容行"""
    ws.merge_cells(start_row=row, start_column=label_col, end_row=row, end_column=label_end)
    label_cell = ws.cell(row=row, column=label_col, value=label)
    label_cell.font = Font(name='微软雅黑', size=11, bold=True)
    label_cell.alignment = Alignment(horizontal='left', vertical='center')
    label_cell.border = THIN_BORDER
    label_cell.fill = fill(COLOR_ZEBRA)

    ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=value_end)
    value_cell = ws.cell(row=row, column=value_col, value=value)
    value_cell.font = FONT_DATA
    value_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    value_cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _write_ai_tip(ws, row, text, max_col=8):
    """在分析表格下方写入 AI 建议提示行"""
    tip_fill = fill('E7F3FF')
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = tip_fill
        cell.border = THIN_BORDER
    cell = ws.cell(row=row, column=1, value=f'AI 建议：{text}')
    cell.font = Font(name='微软雅黑', size=10, italic=True, color='1F4E78')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.row_dimensions[row].height = 36


def write_analysis_sheet(wb, analysis_json: dict):
    """按固定表样式写入数据分析 sheet，所有区块均占满 A-H 列"""
    ws = wb.create_sheet(title='数据分析')
    max_col = 8

    # 列宽
    widths = {'A': 10, 'B': 16, 'C': 16, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    def style_range(row, col_start, col_end, font, fill_color, align, border=True):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = font
            cell.fill = fill_color
            cell.alignment = align
            if border:
                cell.border = THIN_BORDER

    # 第1行：大标题（不合并，避免部分列格式丢失；A1 写标题，其余列写空值但同样式）
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c, value='经销商数据分析报告' if c == 1 else '')
        cell.font = FONT_TITLE
        cell.fill = fill(COLOR_TITLE_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 45

    current_row = 3

    # 核心结论
    _write_section_title(ws, current_row, '一、核心结论')
    current_row += 1
    style_range(current_row, 1, max_col, FONT_DATA, fill(COLOR_WHITE),
                Alignment(horizontal='left', vertical='top', wrap_text=True))
    ws.cell(row=current_row, column=1, value=analysis_json.get('核心结论', ''))
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
    ws.row_dimensions[current_row].height = 80
    current_row += 2

    # 关键指标（2行2列卡片，占满 A-H）
    _write_section_title(ws, current_row, '二、关键指标')
    current_row += 1
    indicators = analysis_json.get('关键指标', {})
    items = [
        ('提审/创建', indicators.get('提审/创建', '')),
        ('提审/渲染', indicators.get('提审/渲染', '')),
        ('最佳大区', indicators.get('最佳大区', '')),
        ('最差大区', indicators.get('最差大区', '')),
    ]
    headers = [('指标', 2), ('数值', 2), ('指标', 2), ('数值', 2)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for i in range(0, len(items), 2):
        row_items = items[i:i + 2]
        while len(row_items) < 2:
            row_items.append(('', ''))
        col_idx = 1
        for label, value in row_items:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + 1)
            label_cell = ws.cell(row=current_row, column=col_idx, value=label)
            label_cell.font = Font(name='微软雅黑', size=10, bold=True)
            label_cell.fill = fill(COLOR_ZEBRA)
            label_cell.alignment = CENTER_ALIGN
            label_cell.border = THIN_BORDER
            ws.merge_cells(start_row=current_row, start_column=col_idx + 2, end_row=current_row, end_column=col_idx + 3)
            value_cell = ws.cell(row=current_row, column=col_idx + 2, value=value)
            value_cell.font = FONT_DATA
            value_cell.fill = fill(COLOR_WHITE)
            value_cell.alignment = CENTER_ALIGN
            value_cell.border = THIN_BORDER
            col_idx += 4
        ws.row_dimensions[current_row].height = 28
        current_row += 1
    current_row += 1

    # 漏斗分析
    _write_section_title(ws, current_row, '三、漏斗分析')
    current_row += 1
    funnel = analysis_json.get('漏斗分析', {})
    funnel_items = [
        ('创建方案数', funnel.get('创建方案数', '')),
        ('渲染方案数', funnel.get('渲染方案数', '')),
        ('提审方案数', funnel.get('提审方案数', '')),
        ('渲染率', funnel.get('渲染率', '')),
        ('提审率', funnel.get('提审率', '')),
        ('渲染到提审', funnel.get('渲染到提审', '')),
    ]
    headers = [('指标', 2), ('数值', 2), ('指标', 2), ('数值', 2)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for i in range(0, len(funnel_items), 2):
        row_items = funnel_items[i:i + 2]
        while len(row_items) < 2:
            row_items.append(('', ''))
        col_idx = 1
        for label, value in row_items:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + 1)
            label_cell = ws.cell(row=current_row, column=col_idx, value=label)
            label_cell.font = Font(name='微软雅黑', size=10, bold=True)
            label_cell.fill = fill(COLOR_ZEBRA)
            label_cell.alignment = CENTER_ALIGN
            label_cell.border = THIN_BORDER
            ws.merge_cells(start_row=current_row, start_column=col_idx + 2, end_row=current_row, end_column=col_idx + 3)
            value_cell = ws.cell(row=current_row, column=col_idx + 2, value=value)
            value_cell.font = FONT_DATA
            value_cell.fill = fill(COLOR_WHITE)
            value_cell.alignment = CENTER_ALIGN
            value_cell.border = THIN_BORDER
            col_idx += 4
        ws.row_dimensions[current_row].height = 28
        current_row += 1
    _write_ai_tip(ws, current_row, funnel.get('ai建议', ''))
    current_row += 2

    # 大区排名
    _write_section_title(ws, current_row, '四、大区排名')
    current_row += 1
    headers = [('大区', 2), ('提审数', 1), ('转化率', 1), ('评价', 4)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for item in analysis_json.get('大区排名', []):
        values = [
            (item.get('大区', ''), 2),
            (item.get('提审数', ''), 1),
            (item.get('转化率', ''), 1),
            (item.get('评价', ''), 4),
        ]
        col_idx = 1
        for value, span in values:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.fill = fill(COLOR_WHITE)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            col_idx += span
        ws.row_dimensions[current_row].height = 24
        current_row += 1
    current_row += 1

    # 区域贡献
    _write_section_title(ws, current_row, '五、区域贡献')
    current_row += 1
    headers = [('大区', 2), ('创建占比', 1), ('渲染占比', 1), ('提审占比', 1), ('评级', 1), ('AI 建议', 2)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for item in analysis_json.get('区域贡献', []):
        values = [
            (item.get('大区', ''), 2),
            (item.get('创建占比', ''), 1),
            (item.get('渲染占比', ''), 1),
            (item.get('提审占比', ''), 1),
            (item.get('贡献评级', ''), 1),
            (item.get('ai建议', ''), 2),
        ]
        col_idx = 1
        for value, span in values:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.fill = fill(COLOR_WHITE)
            cell.alignment = Alignment(horizontal='left' if span >= 2 else 'center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            col_idx += span
        ws.row_dimensions[current_row].height = 30
        current_row += 1
    current_row += 1

    # 转化分层
    _write_section_title(ws, current_row, '六、转化分层')
    current_row += 1
    headers = [('分层', 3), ('经销商数', 2), ('占比', 1), ('AI 建议', 2)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for item in analysis_json.get('转化分层', []):
        values = [
            (item.get('分层', ''), 3),
            (item.get('经销商数', ''), 2),
            (item.get('占比', ''), 1),
            (item.get('ai建议', ''), 2),
        ]
        col_idx = 1
        for value, span in values:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.fill = fill(COLOR_WHITE)
            cell.alignment = Alignment(horizontal='left' if span >= 2 else 'center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            col_idx += span
        ws.row_dimensions[current_row].height = 30
        current_row += 1
    current_row += 1

    # 头部经销商
    _write_section_title(ws, current_row, '七、头部经销商')
    current_row += 1
    headers = [('排名', 1), ('经销商', 2), ('提审数', 1), ('亮点', 4)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for item in analysis_json.get('头部经销商', []):
        values = [
            (item.get('排名', ''), 1),
            (item.get('经销商', ''), 2),
            (item.get('提审数', ''), 1),
            (item.get('亮点', ''), 4),
        ]
        col_idx = 1
        for value, span in values:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.fill = fill(COLOR_WHITE)
            cell.alignment = Alignment(horizontal='left' if span >= 2 else 'center',
                                       vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            col_idx += span
        ws.row_dimensions[current_row].height = 24
        current_row += 1
    current_row += 1

    # 沉睡经销商
    _write_section_title(ws, current_row, '八、沉睡经销商（创建≥20 且 提审=0）')
    current_row += 1
    headers = [('区域', 1), ('经销商', 2), ('创建', 1), ('渲染', 1), ('提审', 1), ('AI 诊断', 2)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    sleeping = analysis_json.get('沉睡经销商', [])
    if not sleeping:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
        cell = ws.cell(row=current_row, column=1, value='本月无沉睡经销商，整体转化健康。')
        cell.font = FONT_DATA
        cell.fill = fill(COLOR_WHITE)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.row_dimensions[current_row].height = 28
        current_row += 1
    else:
        for item in sleeping:
            values = [
                (item.get('区域', ''), 1),
                (item.get('经销商', ''), 2),
                (item.get('创建方案数', ''), 1),
                (item.get('渲染方案数', ''), 1),
                (item.get('提审方案数', ''), 1),
                (item.get('诊断', ''), 2),
            ]
            col_idx = 1
            for value, span in values:
                ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = FONT_DATA
                cell.fill = fill(COLOR_WHITE)
                cell.alignment = Alignment(horizontal='left' if span >= 2 else 'center', vertical='center', wrap_text=True)
                cell.border = THIN_BORDER
                col_idx += span
            ws.row_dimensions[current_row].height = 30
            current_row += 1
    current_row += 1

    # 问题诊断
    _write_section_title(ws, current_row, '九、问题诊断')
    current_row += 1
    headers = [('问题', 1), ('涉及经销商', 2), ('影响', 2), ('建议', 3)]
    col_idx = 1
    for header, span in headers:
        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col_idx += span
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    for item in analysis_json.get('问题诊断', []):
        values = [
            (item.get('问题', ''), 1),
            (item.get('涉及经销商', ''), 2),
            (item.get('影响', ''), 2),
            (item.get('建议', ''), 3),
        ]
        col_idx = 1
        for value, span in values:
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + span - 1)
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.fill = fill(COLOR_WHITE)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            col_idx += span
        ws.row_dimensions[current_row].height = 45
        current_row += 1
    current_row += 1

    # 行动建议
    _write_section_title(ws, current_row, '十、行动建议')
    current_row += 1
    for i, suggestion in enumerate(analysis_json.get('行动建议', []), 1):
        cell_no = ws.cell(row=current_row, column=1, value=i)
        cell_no.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        cell_no.fill = fill(COLOR_HEADER_BG)
        cell_no.alignment = CENTER_ALIGN
        cell_no.border = THIN_BORDER
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=max_col)
        cell_text = ws.cell(row=current_row, column=2, value=suggestion)
        cell_text.font = FONT_DATA
        cell_text.fill = fill(COLOR_WHITE) if i % 2 == 0 else fill(COLOR_ZEBRA)
        cell_text.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_text.border = THIN_BORDER
        ws.row_dimensions[current_row].height = 32
        current_row += 1


def _write_header(ws, title, headers, max_col):
    """写入标题行和表头行，标题跨列居中合并"""
    title_fill = fill(COLOR_TITLE_BG)
    # 先为整行设置统一样式，再合并单元格
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col, value=title if col == 1 else '')
        c.font = FONT_TITLE
        c.alignment = CENTER_ALIGN
        c.fill = title_fill
        c.border = THIN_BORDER
    # 合并第一行所有列并居中显示标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.row_dimensions[1].height = 55.0

    # 列标题
    header_fill = fill(COLOR_HEADER_BG)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = FONT_HEADER
        cell.alignment = CENTER_ALIGN
        cell.fill = header_fill
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 40.0


def _set_column_widths(ws, widths):
    """设置列宽"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ==================== 旧版格式写入 ====================

def write_formatted_excel_old(df, output_path, title='经销商数据'):
    """
    旧版格式：与原 3月/4月/6.2 经销商数据表格式一致
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '账号信息明细'

    _set_column_widths(ws, {
        'A': 17.875, 'B': 32.875, 'C': 21.75,
        'D': 11.0, 'E': 16.25, 'F': 16.75
    })

    headers = ['区域', '经销商', '创建方案数', '新增渲染方案数', '方案深化占比', '提审方案数']
    _write_header(ws, title, headers, 6)

    # 数据行
    for r_idx, row in enumerate(df.itertuples(index=False), 3):
        cell_a = ws.cell(row=r_idx, column=1, value=row.区域 if row.区域 else None)
        cell_a.font = FONT_DATA
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER

        ws.cell(row=r_idx, column=2, value=row.经销商).font = FONT_DATA
        ws.cell(row=r_idx, column=2).alignment = CENTER_ALIGN
        ws.cell(row=r_idx, column=2).border = THIN_BORDER

        ws.cell(row=r_idx, column=3, value=row.创建方案数).font = FONT_DATA
        ws.cell(row=r_idx, column=3).alignment = CENTER_ALIGN
        ws.cell(row=r_idx, column=3).border = THIN_BORDER

        ws.cell(row=r_idx, column=4, value=row.新增渲染方案数).font = FONT_DATA
        ws.cell(row=r_idx, column=4).alignment = CENTER_ALIGN
        ws.cell(row=r_idx, column=4).border = THIN_BORDER

        if row.创建方案数 > 0:
            cell_e = ws.cell(row=r_idx, column=5, value=f'=D{r_idx}/C{r_idx}')
        else:
            cell_e = ws.cell(row=r_idx, column=5, value=0)
        cell_e.font = FONT_DATA
        cell_e.alignment = CENTER_ALIGN
        cell_e.border = THIN_BORDER
        cell_e.number_format = '0%'

        ws.cell(row=r_idx, column=6, value=row.提审方案数).font = FONT_DATA
        ws.cell(row=r_idx, column=6).alignment = CENTER_ALIGN
        ws.cell(row=r_idx, column=6).border = THIN_BORDER

        ws.row_dimensions[r_idx].height = 17.0

    # 合并区域列
    current_region = None
    start_row = 3
    for r_idx, region in enumerate(df['区域'], 3):
        if region != '' and region != current_region:
            if current_region is not None:
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=r_idx - 1, end_column=1)
            current_region = region
            start_row = r_idx
    if current_region is not None:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=len(df) + 2, end_column=1)

    wb.save(output_path)
    print(f'旧版报表已生成: {output_path}')
    print(f'共 {len(df)} 行经销商数据')


# ==================== 新版格式写入（含美化） ====================

def write_formatted_excel_new(df, output_path, title='经销商数据'):
    """
    新版格式：与 6月经销商酷家乐数据(新) 格式一致
    包含：大区合计、总合计、提审/渲染、提审/创建，并带美化样式
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '账号信息明细'

    _set_column_widths(ws, {
        'A': 24, 'B': 30, 'C': 14, 'D': 16,
        'E': 16, 'F': 16, 'G': 12, 'H': 12
    })

    headers = [
        '区域', '经销商', '创建方案数',
        '新增渲染方案数\n(出效果图)',
        '方案深化占比\n(渲染/创建)',
        '提审方案数\n(下单到工厂)',
        '提审/渲染', '提审/创建'
    ]
    _write_header(ws, title, headers, 8)

    # 构建区域块
    region_blocks = []
    current_region = None
    current_rows = []

    for idx, region in enumerate(df['区域']):
        if region != '' and region != current_region:
            if current_region is not None:
                region_blocks.append((current_region, current_rows))
            current_region = region
            current_rows = []
        current_rows.append(idx)
    if current_region is not None:
        region_blocks.append((current_region, current_rows))

    data_rows_info = []
    current_excel_row = 3

    for region_name, row_indices in region_blocks:
        block_start = current_excel_row
        data_idx_in_block = 0

        for data_idx in row_indices:
            row = df.iloc[data_idx]
            data_idx_in_block += 1
            data_rows_info.append({
                'excel_row': current_excel_row,
                'data_idx': data_idx,
                'is_data': True
            })

            # 斑马纹：隔行底色
            row_fill = fill(COLOR_ZEBRA) if data_idx_in_block % 2 == 0 else fill(COLOR_WHITE)

            cell_a = ws.cell(row=current_excel_row, column=1,
                             value=region_name if data_idx == row_indices[0] else None)
            cell_a.fill = row_fill
            cell_a.font = FONT_DATA
            cell_a.alignment = CENTER_ALIGN
            cell_a.border = THIN_BORDER

            for col, val in [(2, row['经销商']), (3, row['创建方案数']),
                             (4, row['新增渲染方案数']), (6, row['提审方案数'])]:
                cell = ws.cell(row=current_excel_row, column=col, value=val)
                cell.fill = row_fill
                cell.font = FONT_DATA
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

            # 方案深化占比公式
            if row['创建方案数'] > 0:
                cell_e = ws.cell(row=current_excel_row, column=5,
                                 value=f'=D{current_excel_row}/C{current_excel_row}')
            else:
                cell_e = ws.cell(row=current_excel_row, column=5, value=0)
            cell_e.fill = row_fill
            cell_e.font = FONT_DATA
            cell_e.alignment = CENTER_ALIGN
            cell_e.border = THIN_BORDER
            cell_e.number_format = '0%'

            # 提审/渲染：提审和渲染都为0时留空
            if row['新增渲染方案数'] > 0 and row['提审方案数'] > 0:
                cell_g = ws.cell(row=current_excel_row, column=7,
                                 value=f'=F{current_excel_row}/D{current_excel_row}')
                cell_g.number_format = '0.00%'
            else:
                cell_g = ws.cell(row=current_excel_row, column=7, value='')
            cell_g.fill = row_fill
            cell_g.font = FONT_DATA
            cell_g.alignment = CENTER_ALIGN
            cell_g.border = THIN_BORDER

            # 提审/创建：提审和创建都为0时留空
            if row['创建方案数'] > 0 and row['提审方案数'] > 0:
                cell_h = ws.cell(row=current_excel_row, column=8,
                                 value=f'=F{current_excel_row}/C{current_excel_row}')
                cell_h.number_format = '0.00%'
            else:
                cell_h = ws.cell(row=current_excel_row, column=8, value='')
            cell_h.fill = row_fill
            cell_h.font = FONT_DATA
            cell_h.alignment = CENTER_ALIGN
            cell_h.border = THIN_BORDER

            ws.row_dimensions[current_excel_row].height = 22.0
            current_excel_row += 1

        # 大区合计行
        total_row = current_excel_row
        sum_range_c = f'C{block_start}:C{total_row - 1}'
        sum_range_d = f'D{block_start}:D{total_row - 1}'
        sum_range_f = f'F{block_start}:F{total_row - 1}'

        total_fill = fill(COLOR_REGION_TOTAL_BG)

        for col in range(1, 9):
            ws.cell(row=total_row, column=col).fill = total_fill
            ws.cell(row=total_row, column=col).font = FONT_TOTAL_BLUE
            ws.cell(row=total_row, column=col).alignment = CENTER_ALIGN
            ws.cell(row=total_row, column=col).border = THIN_BORDER

        ws.cell(row=total_row, column=2, value='大区合计')
        ws.cell(row=total_row, column=3, value=f'=SUM({sum_range_c})')
        ws.cell(row=total_row, column=4, value=f'=SUM({sum_range_d})')
        ws.cell(row=total_row, column=5, value=f'=D{total_row}/C{total_row}')
        ws.cell(row=total_row, column=5).number_format = '0.00%'
        ws.cell(row=total_row, column=6, value=f'=SUM({sum_range_f})')
        ws.cell(row=total_row, column=7, value=f'=F{total_row}/D{total_row}')
        ws.cell(row=total_row, column=7).number_format = '0.00%'
        ws.cell(row=total_row, column=8, value=f'=F{total_row}/C{total_row}')
        ws.cell(row=total_row, column=8).number_format = '0.00%'

        ws.row_dimensions[total_row].height = 28.0
        current_excel_row += 1

    # 空行
    empty_row = current_excel_row
    for col in range(1, 9):
        ws.cell(row=empty_row, column=col).border = THIN_BORDER
    ws.row_dimensions[empty_row].height = 12.0
    current_excel_row += 1

    # 总合计行
    grand_total_row = current_excel_row
    total_create = int(df['创建方案数'].sum())
    total_render = int(df['新增渲染方案数'].sum())
    total_submit = int(df['提审方案数'].sum())

    total_fill = fill(COLOR_GRAND_TOTAL_BG)

    for col in range(1, 9):
        ws.cell(row=grand_total_row, column=col).fill = total_fill
        ws.cell(row=grand_total_row, column=col).font = FONT_TOTAL_WHITE
        ws.cell(row=grand_total_row, column=col).alignment = CENTER_ALIGN
        ws.cell(row=grand_total_row, column=col).border = THIN_BORDER

    ws.cell(row=grand_total_row, column=2, value='总合计')
    ws.cell(row=grand_total_row, column=3, value=total_create)
    ws.cell(row=grand_total_row, column=4, value=total_render)
    cell_e = ws.cell(row=grand_total_row, column=5,
                     value=f'=D{grand_total_row}/C{grand_total_row}')
    cell_e.number_format = '0.00%'
    ws.cell(row=grand_total_row, column=6, value=total_submit)
    cell_g = ws.cell(row=grand_total_row, column=7,
                     value=f'=F{grand_total_row}/D{grand_total_row}')
    cell_g.number_format = '0.00%'
    cell_h = ws.cell(row=grand_total_row, column=8,
                     value=f'=F{grand_total_row}/C{grand_total_row}')
    cell_h.number_format = '0.00%'

    ws.row_dimensions[grand_total_row].height = 45.0

    # 合并每个区域的A列
    for region_name, row_indices in region_blocks:
        first_row = data_rows_info[row_indices[0]]['excel_row']
        last_row = data_rows_info[row_indices[-1]]['excel_row']
        ws.merge_cells(start_row=first_row, start_column=1,
                       end_row=last_row, end_column=1)

    # 冻结窗格：冻结前两行
    ws.freeze_panes = 'A3'

    # 为账号信息明细表头添加筛选
    ws.auto_filter.ref = f'A2:H{grand_total_row}'

    # 生成数据分析 sheets（使用 dealer-data-analyzer skill 逻辑）
    try:
        analysis_df = prepare_dealer_df(df)
        add_analysis_sheets(wb, analysis_df)
    except Exception as e:
        print(f'数据分析 sheets 生成失败: {e}')

    wb.save(output_path)
    print(f'新版美化报表已生成: {output_path}')
    print(f'共 {len(df)} 行经销商数据，{len(region_blocks)} 个大区')


# ==================== 主函数 ====================

def main():
    parser = argparse.ArgumentParser(
        description='从账号指标和设计师数据生成经销商报表'
    )
    parser.add_argument(
        '--account',
        default=(r'D:\wechat\xwechat_files\wxid_0fh4oxng8dq212_f810\msg\file\2026-06\账号指标20260602.xlsx'),
        help='账号指标Excel文件路径'
    )
    parser.add_argument(
        '--designer',
        default=(r'D:\wechat\xwechat_files\wxid_0fh4oxng8dq212_f810\msg\file\2026-06\设计师数据统计 .xlsx'),
        help='设计师数据统计Excel文件路径'
    )
    parser.add_argument(
        '--output',
        default=(r'D:\wechat\xwechat_files\wxid_0fh4oxng8dq212_f810\msg\file\2026-06\经销商数据.xlsx'),
        help='输出报表路径'
    )
    parser.add_argument(
        '--title',
        default='6月经销商数据',
        help='报表标题（第一行显示内容）'
    )
    parser.add_argument(
        '--format',
        choices=['old', 'new'],
        default='new',
        help='输出格式：old=旧版（无合计），new=新版（含大区合计、总合计、美化样式）'
    )
    args = parser.parse_args()

    for path, name in [(args.account, '账号指标'), (args.designer, '设计师数据')]:
        if not Path(path).exists():
            raise FileNotFoundError(f'{name}文件不存在: {path}')

    account_agg = process_account_data(args.account)
    designer_agg = process_designer_data(args.designer)
    out_df = build_output_df(account_agg, designer_agg)

    if args.format == 'old':
        write_formatted_excel_old(out_df, args.output, title=args.title)
    else:
        write_formatted_excel_new(out_df, args.output, title=args.title)


if __name__ == '__main__':
    main()
