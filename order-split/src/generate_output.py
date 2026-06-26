"""
根据输入文件生成输出料单文件
"""
import os
import re
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell



# 需要忽略的名称
# 精确匹配用于单字/特定结构件，避免误伤“中竖板/横隔板”等正常板件名称。
IGNORED_EXACT_NAMES = {'竖', '横', '标准雕花'}

# 含以下关键词即过滤。
IGNORED_KEYWORDS = ['竖枋', '横枋', '上横枋', '下横枋', '芯板', '中横', '格条', '隔条', '门洞', '格栅条', '左右板子', '前后板子', '顶板']

# Xmm加厚规格过滤（如3mm加厚板）
MM_THICK_PATTERN = re.compile(r'\d+mm.*加厚', re.IGNORECASE)


def should_ignore_name(name: str) -> bool:
    """判断名称是否应该被过滤"""
    if not name:
        return False
    name = str(name).strip()
    if name in IGNORED_EXACT_NAMES:
        return True
    if any(kw in name for kw in IGNORED_KEYWORDS):
        return True
    if MM_THICK_PATTERN.search(name):
        return True
    return False

# 80 材质代码映射 — 实际应从结构件汇总，但先做兜底
MATERIAL_CODE_MAP = {'80': '红胡桃原木'}


def read_input_data(input_path):
    """读取输入文件的实木柜门和实木附件 sheet

    合并单元格处理：逐行按需从 xls 合并单元格获取值，不提前污染 DataFrame
    """
    import pandas as pd
    import xlrd

    df_door = pd.read_excel(input_path, sheet_name='实木柜门', header=None)
    df_attach = pd.read_excel(input_path, sheet_name='实木附件', header=None)

    book = xlrd.open_workbook(input_path, formatting_info=True)
    sheet_door = book.sheet_by_name('实木柜门')
    sheet_attach = book.sheet_by_name('实木附件')

    def get_merged_value(sheet, row_idx, col_idx):
        """获取单元格值，如果在合并区域内且原值为空，返回左上角值。
        xlrd 空单元格返回 ''，这里统一转为 None 以兼容 pandas 行为。"""
        val = sheet.cell(row_idx, col_idx).value
        if val == '' or val is None:
            val = None
        else:
            return val
        for rlo, rhi, clo, chi in sheet.merged_cells:
            if clo <= col_idx < chi and rlo <= row_idx < rhi:
                top = sheet.cell(rlo, clo).value
                return top if top != '' and top is not None else None
        return None

    def extract_number(val):
        """提取数值：字符串中的数字，支持 + - * / 运算，否则取第一个数字"""
        if isinstance(val, (int, float)):
            return val
        if isinstance(val, str):
            import re
            nums = re.findall(r'\d+(?:\.\d+)?', val)
            ops = re.findall(r'[+\-*/]', val)
            if len(nums) >= 2 and len(ops) >= 1:
                result = float(nums[0])
                for i, n in enumerate(nums[1:]):
                    if i < len(ops):
                        op = ops[i]
                        if op == '+':
                            result += float(n)
                        elif op == '-':
                            result -= float(n)
                        elif op == '*':
                            result *= float(n)
                        elif op == '/':
                            result /= float(n)
                return result
            if nums:
                num_str = nums[0]
                return float(num_str) if '.' in num_str else int(num_str)
        return val

    def extract_data(df, sheet):
        """从 DataFrame 提取有效数据行，同时处理合并单元格"""
        data = []
        for idx in range(6, len(df)):
            order_no = get_merged_value(sheet, idx, 1)
            area = get_merged_value(sheet, idx, 2)
            name = get_merged_value(sheet, idx, 3)

            # 跳过空行
            if not order_no or pd.isna(order_no) or order_no == '':
                continue
            if not area or pd.isna(area) or area == '':
                continue
            if not name or pd.isna(name) or name == '':
                continue
            order_str = str(order_no).strip()
            # 过滤说明行：order_no 必须以订单号格式开头（字母数字+'-'+数字，如 S2604-4106-1）
            if not re.match(r'^[A-Za-z0-9]+-\d+', order_str):
                continue

            data.append({
                'idx': idx,
                'order_no': str(order_no).strip(),
                'area': str(area).strip(),
                'name': str(name).strip(),
                'length': extract_number(get_merged_value(sheet, idx, 4)),
                'width': extract_number(get_merged_value(sheet, idx, 5)),
                'thickness': extract_number(get_merged_value(sheet, idx, 6)),
                'qty': extract_number(get_merged_value(sheet, idx, 7)),
                'material': get_merged_value(sheet, idx, 8),
                'wood_skin': get_merged_value(sheet, idx, 9),
                'veneer': get_merged_value(sheet, idx, 10),
                'edge': get_merged_value(sheet, idx, 11),
                'category': get_merged_value(sheet, idx, 11),
                'color': get_merged_value(sheet, idx, 12),
                'remark': get_merged_value(sheet, idx, 13),
            })
        return data

    door_data = extract_data(df_door, sheet_door)
    attach_data = extract_data(df_attach, sheet_attach)
    return door_data, attach_data


# 80 材质代码映射 — 实际应从结构件汇总，但先做兜底
MATERIAL_CODE_MAP = {'80': '红胡桃原木'}


def is_material_placeholder(val):
    """判断材质值是否是占位符（纯数字或空），需要从子件收集真实材质"""
    if pd.isna(val):
        return True
    s = str(val).strip()
    if not s:
        return True
    # 纯数字（包括整数和浮点数）都是占位符
    if re.match(r'^\d+(?:\.\d+)?$', s):
        return True
    return False


def resolve_material(raw_material, struct_materials):
    """根据结构件汇总材质 - 当材质列为数字或空时，从子件收集真实材质"""
    if is_material_placeholder(raw_material):
        # 从子件收集非空材质，去重
        mats = [m for m in struct_materials if not is_material_placeholder(m)]
        mats = list(dict.fromkeys(mats))
        if mats:
            return '、'.join(str(m) for m in mats)
        return ''
    s = str(raw_material).strip()
    if s in MATERIAL_CODE_MAP:
        return MATERIAL_CODE_MAP[s]
    if s:
        # 柜门材质统一加"实木"前缀（系统识别需要）
        if '集成实木' in s:
            s = '实木' + s
        return s
    return ''


def resolve_color(raw_color, struct_colors):
    """根据结构件汇总颜色"""
    if not pd.isna(raw_color):
        s = str(raw_color).strip()
        if s:
            return s
    cols = [c for c in struct_colors if not pd.isna(c)]
    cols = list(dict.fromkeys(cols))
    if cols:
        return '、'.join(str(c) for c in cols)
    return ''


def build_sheet1_rows(order_no, door_data, attach_data):
    """构建 Sheet1 的数据行"""
    rows = []

    # 附件数据（过滤结构件）
    for item in attach_data:
        if item['order_no'] != order_no:
            continue
        if should_ignore_name(item['name']):
            continue
        raw_remark = item['remark']
        if pd.isna(raw_remark):
            sheet1_remark = None
            guiti_remark = 0
        else:
            s = str(raw_remark).strip()
            if s == '0':
                sheet1_remark = None
                guiti_remark = 0
            else:
                sheet1_remark = s
                guiti_remark = s

        material = resolve_material(item['material'], [])
        color = resolve_color(item['color'], [])

        rows.append({
            'order_no': item['order_no'],
            'area': item['area'],
            'name': item['name'],
            'length': item['length'],
            'width': item['width'],
            'thickness': item['thickness'],
            'qty': item['qty'],
            'material': material,
            'wood_skin': item['wood_skin'],
            'veneer': item['veneer'],
            'edge': item['edge'],
            'category': item['category'],
            'color': color,
            'sheet1_remark': sheet1_remark,
            'guiti_remark': guiti_remark,
        })

    # 柜门数据：过滤结构件，按组汇总材质颜色
    door_items = [d for d in door_data if d['order_no'] == order_no]
    i = 0
    while i < len(door_items):
        item = door_items[i]
        if should_ignore_name(item['name']):
            i += 1
            continue

        # 收集后续结构件的材质和颜色
        struct_materials = []
        struct_colors = []
        j = i + 1
        while j < len(door_items):
            nxt = door_items[j]
            if should_ignore_name(nxt['name']):
                struct_materials.append(nxt['material'])
                struct_colors.append(nxt['color'])
                j += 1
            else:
                break

        material = resolve_material(item['material'], struct_materials)
        color = resolve_color(item['color'], struct_colors)

        raw_remark = item['remark']
        if pd.isna(raw_remark):
            sheet1_remark = None
            guiti_remark = '0'
        else:
            s = str(raw_remark).strip()
            if s == '0':
                sheet1_remark = None
                guiti_remark = '0'
            else:
                sheet1_remark = s
                guiti_remark = s

        rows.append({
            'order_no': item['order_no'],
            'area': item['area'],
            'name': item['name'],
            'length': item['length'],
            'width': item['width'],
            'thickness': item['thickness'],
            'qty': item['qty'],
            'material': material,
            'wood_skin': item['wood_skin'],
            'veneer': item['veneer'],
            'edge': item['edge'],
            'category': item['category'],
            'color': color,
            'sheet1_remark': sheet1_remark,
            'guiti_remark': guiti_remark,
        })
        i = j if j > i + 1 else i + 1

    return rows


def build_guiti_rows(order_no, sheet1_rows):
    """根据 Sheet1 行构建柜体 sheet 的展开行，并分配纹路编号

    纹路规则：全部设置为1
    """
    rows = []

    for i, srow in enumerate(sheet1_rows):
        qty = int(srow['qty']) if not pd.isna(srow['qty']) else 1
        area_key = srow['area'] + '_0'
        raw_material = srow['material']
        # 柜门材质统一加"实木"前缀（系统识别需要）
        material = raw_material
        if material and not str(material).startswith('实木'):
            material = '实木' + str(material)
        for _ in range(qty):
            rows.append({
                'area': area_key,
                'name': srow['name'],
                'material': material,
                'color': srow['color'],
                'length': srow['length'],
                'width': srow['width'],
                'thickness': srow['thickness'],
                'qty': 1,
                'texture': 1,  # 全部设置为1
                'guiti_remark': srow['guiti_remark'],
            })

    return rows


# 需要填充基材的组件类型（无材质/颜色时取下一行）
FILL_PARENT_TYPES = ('整体柜', '盒子', '弧形盒子', '抽屉盒', '地柜', '吊柜')

# ===== 填充整体柜/盒子的缺失基材和颜色（取下一行） =====
def fill_parent_material(sheet1_rows):
    """对整体柜/盒子等父级组件，填充下一行的材质/油漆面/颜色"""
    for i, r in enumerate(sheet1_rows):
        name = r.get('name', '')
        if not any(pt in name for pt in FILL_PARENT_TYPES):
            continue
        if i + 1 >= len(sheet1_rows):
            continue
        next_row = sheet1_rows[i + 1]
        # 填充材质
        if not r['material'] or str(r['material']).strip() == '':
            r['material'] = next_row.get('material')
        # 填充油漆面
        if not r['wood_skin'] or str(r['wood_skin']).strip() == '':
            r['wood_skin'] = next_row.get('wood_skin')
        # 填充颜色
        if not r['color'] or str(r['color']).strip() == '':
            r['color'] = next_row.get('color')


def fill_guiti_material(guiti_rows):
    """对柜体sheet的父级组件，填充下一行的材质/颜色"""
    for i, r in enumerate(guiti_rows):
        name = r.get('name', '')
        if not any(pt in name for pt in FILL_PARENT_TYPES):
            continue
        if i + 1 >= len(guiti_rows):
            continue
        next_row = guiti_rows[i + 1]
        # 填充材质
        if not r['material'] or str(r['material']).strip() == '':
            r['material'] = next_row.get('material')
        # 填充颜色
        if not r['color'] or str(r['color']).strip() == '':
            r['color'] = next_row.get('color')



def extract_name_from_path(input_path):
    """从输入文件路径提取姓名
    支持格式: S2604-4090 张姝媛.xls / B2604-6125熊安斌（张润玉）1.xls / S2603-4086佟广会料单.xls
    自动去除浏览器重复下载后缀如 (1)
    """
    import re
    base = os.path.splitext(os.path.basename(input_path))[0]
    # 去掉飞书缓存前缀
    base = re.sub(r'^doc_[a-f0-9]+_', '', base)
    # 去掉浏览器重复下载后缀：(1), (2) 等，如 "S2604-4090 张姝媛(1)"
    base = re.sub(r'\(\d+\)$', '', base)
    # 尝试匹配带"料单"后缀的格式: S2603-4086佟广会料单
    m = re.match(r'[SB]\d{4}-\d{4}(.+?)料单$', base)
    if m:
        return m.group(1)
    # 尝试匹配空格分隔的格式: S2604-4090 张姝媛
    m = re.match(r'[SB]\d{4}-\d{4}\s+(.+)$', base)
    if m:
        return m.group(1)
    # 尝试匹配无空格带名字的格式: B2604-6125熊安斌（张润玉）1
    # 排除纯子订单号格式如 -1, -2, -3（以-数字结尾且前面没有中文字符）
    m = re.match(r'[SB]\d{4}-\d{4}(-\d+)$', base)
    if m:
        return ''
    m = re.match(r'[SB]\d{4}-\d{4}([^\d\-].+?)$', base)
    if m:
        return m.group(1)
    # 兜底：-数字后面还有内容的，如 B2605-1623-1张三
    m = re.match(r'[SB]\d{4}-\d{4}-\d+(.+)$', base)
    if m and m.group(1).strip():
        return m.group(1)
    return ''


def find_template(input_path):
    """查找合适的模板文件，在脚本所在目录查找"""
    import re
    # 使用脚本所在目录而不是当前工作目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dir_name = script_dir
    base = os.path.splitext(os.path.basename(input_path))[0]
    m = re.match(r'(S\d{4}-\d{4})', base)
    project_no = m.group(1) if m else ''

    # 优先找同项目编号的 -01 模板
    for fname in sorted(os.listdir(dir_name)):
        if fname.startswith(project_no) and fname.endswith('.xlsx') and '-01' in fname:
            return os.path.join(dir_name, fname)

    # 否则找任意 -01 料单作为默认模板
    for fname in sorted(os.listdir(dir_name)):
        if fname.endswith('料单.xlsx') and '-01' in fname:
            return os.path.join(dir_name, fname)

    return None


def generate_one(input_path, template_path, order_no, output_dir, name=None):
    """生成单个订单的输出文件"""
    door_data, attach_data = read_input_data(input_path)
    sheet1_rows = build_sheet1_rows(order_no, door_data, attach_data)
    guiti_rows = build_guiti_rows(order_no, sheet1_rows)

    os.makedirs(output_dir, exist_ok=True)
    formatted_order = order_no

    if not name:
        name = extract_name_from_path(input_path)
    out_name = f'{formatted_order}{name}料单.xlsx'
    out_path = os.path.join(output_dir, out_name)

    wb = load_workbook(template_path)

    def get_style_row(ws, start_row):
        """获取某行每个单元格的样式，用于复制到新行"""
        from copy import copy
        styles = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=col)
            if not isinstance(cell, MergedCell):
                styles[col] = {
                    'font': copy(cell.font),
                    'alignment': copy(cell.alignment),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                }
        return styles

    def apply_style(ws, row, col, style):
        """将样式应用到指定单元格"""
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            cell.font = style['font']
            cell.alignment = style['alignment']
            cell.border = style['border']
            cell.fill = style['fill']
            cell.number_format = style['number_format']

    # ===== 填充整体柜/盒子的缺失基材和颜色（取下一行） =====
    fill_parent_material(sheet1_rows)
    fill_guiti_material(guiti_rows)

    # ===== 写入 Sheet1 =====
    ws1 = wb['Sheet1']
    # 保存模板数据行（Row 3）的样式
    sheet1_style = get_style_row(ws1, 3)

    # 清空数据行（保留表头前两行）
    for row in range(3, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, r in enumerate(sheet1_rows):
        row = i + 3
        data = [
            r['order_no'], r['area'], r['name'], r['length'], r['width'],
            r['thickness'], r['qty'], r['material'],
            r['wood_skin'] if not pd.isna(r['wood_skin']) else None,
            r['veneer'] if not pd.isna(r['veneer']) else None,
            r['edge'] if not pd.isna(r['edge']) else None,
            r['color'], r['sheet1_remark']
        ]
        for col, val in enumerate(data, start=1):
            ws1.cell(row=row, column=col, value=val)
            if col in sheet1_style:
                apply_style(ws1, row, col, sheet1_style[col])

    # 删除尾部多余的空行（先解除尾部合并单元格）
    last_data_row = len(sheet1_rows) + 2
    if ws1.max_row > last_data_row:
        for rng in list(ws1.merged_cells.ranges):
            if rng.min_row > last_data_row:
                ws1.unmerge_cells(str(rng))
        ws1.delete_rows(last_data_row + 1, ws1.max_row - last_data_row)

    # ===== 写入 柜体 =====
    ws2 = wb['柜体']
    # 保存模板数据行（Row 2）的样式
    guiti_style = get_style_row(ws2, 2)

    # 清空数据行（保留表头第一行）
    for row in range(2, ws2.max_row + 1):
        for col in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, r in enumerate(guiti_rows):
        row = i + 2
        data = {
            1: r['area'],
            2: None,
            3: r['name'],
            4: r['material'],
            5: r['color'],
            6: r['length'],
            7: r['width'],
            8: r['thickness'],
            9: r['qty'],
            22: r['texture'],
            23: r['guiti_remark'],
            24: None,
        }
        for col, val in data.items():
            ws2.cell(row=row, column=col, value=val)
            if col in guiti_style:
                apply_style(ws2, row, col, guiti_style[col])

    # ===== 五金 sheet 保持空表头 =====

    wb.save(out_path)
    wb.close()
    return out_path


def generate_all(input_path, output_dir):
    """生成所有订单的输出文件，每个订单单独一个文件夹"""
    door_data, attach_data = read_input_data(input_path)

    # 提取所有订单号
    order_nos = set()
    for d in door_data:
        order_nos.add(d['order_no'])
    for a in attach_data:
        order_nos.add(a['order_no'])

    if not order_nos:
        print('警告: 未找到任何订单数据')
        return

    template_path = find_template(input_path)
    if not template_path:
        print('错误: 未找到模板文件')
        return

    name = extract_name_from_path(input_path)

    for order_no in sorted(order_nos):
        # 每个订单单独一个子文件夹
        order_folder = os.path.join(output_dir, order_no)
        generate_one(input_path, template_path, order_no, order_folder, name)
        print(f'Generated: {order_no}')


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
        output_dir = sys.argv[2] if len(sys.argv) > 2 else 'output'
    else:
        input_path = 'S2604-4090 张姝媛.xls'
        output_dir = 'output'
    generate_all(input_path, output_dir)
