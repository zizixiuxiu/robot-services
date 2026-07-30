#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
考勤汇总生成服务 - FastAPI 后端
端口：8009

端点：
  POST /generate        上传 .xls，返回包含两份输出 XLSX 的 zip
  POST /generate-local  本地文件路径生成，输出到指定目录
  GET  /health          健康检查
"""
import os
import sys
import io
import re
import time
import zipfile
import tempfile
import shutil
import logging
import argparse
import base64
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
import uvicorn
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_attendance import generate
from workshop_attendance import transform_attendance

WORK_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
WORKSHOP_TEMPLATE = WORK_DIR / "templates_workshop" / "workshop_template.xlsx"
WORKSHOP_CONFIG = WORK_DIR / "templates_workshop" / "sample_adjustments_2026_01_06.json"

LOG_DIR = Path(os.getenv("LOG_DIR", str(WORK_DIR.parent / "logs")))
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "attendance_summary_http.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("attendance-summary")

app = FastAPI(title="考勤汇总生成服务", version="1.2.0")


WORKSHOP_REQUIRED_HEADERS = {"人员ID", "姓名", "设备名称", "方向", "识别方式", "时间"}


def _detect_workshop_input(input_path: Path) -> bool:
    """Detect workshop raw gate-log files by workbook headers, not by extension."""
    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(rows, ())
            headers = {str(v).strip() for v in header_row if v is not None}
            if WORKSHOP_REQUIRED_HEADERS.issubset(headers):
                return True
        return False
    finally:
        wb.close()


def _detect_workshop_month(path_or_name: str) -> Optional[str]:
    """Infer YYYY-MM from names such as 2026年6月... or 全公司6月..."""
    name = Path(path_or_name).name
    m = re.search(r"(\d{4}).*?(\d{1,2})月", name)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.search(r"(\d{1,2})月", name)
    if m:
        return f"{datetime.now().year:04d}-{int(m.group(1)):02d}"
    # 兜底：识别 "5.1-5.31..." 这类日期区间开头的文件名，月份取第一个数字
    m = re.match(r"(\d{1,2})\.\d{1,2}\s*[-–~]", name)
    if m:
        return f"{datetime.now().year:04d}-{int(m.group(1)):02d}"
    return None


def _workshop_output_name(path_or_name: str) -> str:
    month = _detect_workshop_month(path_or_name)
    if month:
        month_no = int(month.split("-")[1])
        return f"{month_no}月奢匠车间考勤.xlsx"
    return "奢匠车间考勤.xlsx"


def _generate_auto(input_path: Path, output_dir: Path) -> Tuple[str, list]:
    """Generate by detected input content: workshop xlsx-like logs or office xls logs."""
    output_dir.mkdir(parents=True, exist_ok=True)
    if _detect_workshop_input(input_path):
        month = _detect_workshop_month(str(input_path))
        out_file = output_dir / _workshop_output_name(str(input_path))
        report = transform_attendance(
            input_file=input_path,
            template_file=WORKSHOP_TEMPLATE,
            output_file=out_file,
            month=month,
            config_file=WORKSHOP_CONFIG if WORKSHOP_CONFIG.exists() else None,
        )
        if not month:
            # 文件名识别不到月份时，用打卡数据推断出的月份命名
            named_file = output_dir / f"{report.month}月奢匠车间考勤.xlsx"
            out_file.replace(named_file)
            out_file = named_file
        return "workshop", [str(out_file)]
    paths = generate(str(input_path), str(output_dir))
    return "office", paths


@app.post("/generate")
async def generate_attendance(
    file: UploadFile = File(..., description="上传的员工刷卡记录文件"),
    output_format: str = Form("zip", description="输出格式: zip 或 json"),
):
    """上传刷卡记录，自动识别办公室/车间并生成考勤 xlsx"""
    t0 = time.time()

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / file.filename
        output_dir = Path(tmpdir) / "output"

        with open(input_path, "wb") as f:
            f.write(await file.read())

        try:
            mode, paths = _generate_auto(input_path, output_dir)
        except Exception as e:
            logger.exception("生成失败")
            raise HTTPException(500, detail=f"生成失败: {str(e)}")

        out_files = [Path(p) for p in paths] if paths else sorted(output_dir.glob("*.xlsx"))
        if not out_files:
            raise HTTPException(500, detail="未生成任何输出文件")

        cost = time.time() - t0

        if output_format == "json":
            return {
                "success": True,
                "mode": mode,
                "cost_seconds": round(cost, 3),
                "count": len(out_files),
                "files": [f.name for f in out_files],
            }

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in out_files:
                zf.write(f, f.name)
        zip_buffer.seek(0)

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=\"{Path(file.filename).stem}_output.zip\""
            },
        )


class ProcessRequest(BaseModel):
    file_content: str
    filename: str = "input.xls"


@app.post("/process")
async def process_attendance(req: ProcessRequest):
    """兼容 feishu-ws-gateway 的 /process 接口：base64 文件进，base64 文件出"""
    t0 = time.time()

    tmpdir = tempfile.mkdtemp()
    try:
        input_path = Path(tmpdir) / req.filename
        input_path.write_bytes(base64.b64decode(req.file_content))
        output_dir = Path(tmpdir) / "output"

        mode, paths = _generate_auto(input_path, output_dir)

        out_files = [Path(p) for p in paths] if paths else sorted(output_dir.glob("*.xlsx"))
        if not out_files:
            raise HTTPException(500, detail="未生成任何输出文件")

        output_files = []
        for f in out_files:
            output_files.append({
                "filename": f.name,
                "file_content": base64.b64encode(f.read_bytes()).decode("utf-8"),
            })

        cost = time.time() - t0
        return {
            "success": True,
            "mode": mode,
            "cost_seconds": round(cost, 3),
            "count": len(output_files),
            "output_files": output_files,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("处理失败")
        raise HTTPException(500, detail=f"生成失败: {str(e)}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@app.post("/generate-local")
async def generate_attendance_local(
    input_path: str = Form(..., description="本地考勤文件绝对路径"),
    output_dir: Optional[str] = Form(None, description="输出总目录，默认 data/output"),
):
    """通过本地文件路径生成考勤汇总（服务与文件在同一机器时使用）"""
    t0 = time.time()

    if not os.path.exists(input_path):
        raise HTTPException(400, detail=f"文件不存在: {input_path}")

    if output_dir is None:
        output_dir = str(WORK_DIR.parent / "data" / "output")

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    max_num = 0
    for f in output_path.iterdir():
        if f.is_dir():
            m = re.search(r'-(\d+)$', f.name)
            if m:
                max_num = max(max_num, int(m.group(1)))
    next_num = max_num + 1
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    task_dir = output_path / f"{timestamp}-{next_num}"
    task_dir.mkdir(parents=True, exist_ok=True)

    try:
        mode, paths = _generate_auto(Path(input_path), task_dir)
    except Exception as e:
        logger.exception("生成失败")
        raise HTTPException(500, detail=f"生成失败: {str(e)}")

    out_files = [Path(p) for p in paths] if paths else sorted(task_dir.glob("**/*.xlsx"))
    cost = time.time() - t0

    return {
        "success": True,
        "mode": mode,
        "cost_seconds": round(cost, 3),
        "count": len(out_files),
        "output_dir": str(task_dir),
        "files": [str(f.relative_to(task_dir)) for f in out_files],
    }


@app.get("/health")
async def health():
    return {"status": "ok", "service": "attendance-summary", "version": app.version}


def run(port: int = 8009):
    logger.info("[attendance-summary] HTTP 服务启动于 http://0.0.0.0:%d", port)
    uvicorn.run(app, host="0.0.0.0", port=port)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8009)
    args = parser.parse_args()
    run(args.port)
