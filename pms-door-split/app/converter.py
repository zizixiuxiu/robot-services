from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence


# PMS 优化门扇清单必须包含的列（其余列原样透传）
REQUIRED_SOURCE_HEADERS = ["板件名称", "数量", "材料描述", "厚度", "工艺路线"]

# 展示/剔除匹配时使用的订单号列（PMS 清单里物料编码是原订单号）
ORDER_HEADER_CANDIDATES = ("物料编码", "订单号")

THICKNESS_RE = re.compile(
    r"(\d+(?:\.\d+)?(?:\s*\+\s*\d+(?:\.\d+)?)*)\s*mm",
    re.IGNORECASE,
)

# 剔除门型（与 door-skin-converter 保持一致；清单更新需两边同步）
EXCLUDED_DOOR_MODELS = (
    "2091", "7066",
    "N1049", "N1050", "N1051", "N1052", "N1053",
    "N1054", "N1055", "N1056", "N1057", "N1058", "N1059",
    "YM-062", "YM-081", "XP-62",
    "YM-114", "YM-105", "YM-106", "YM-115",
    "YSM-079", "YSM-080", "YSM-081",
    "YM-082",
)

EXCLUDED_NAME_KEYWORDS = ("铝框门",)

DOOR_MODEL_RE = re.compile(r"[A-Z]{1,4}[^A-Z0-9]?\d{2,4}|\d{4}")

# 已是皮行的材料描述（如 4mm素板 / 8mm多层加密）：原样保留，不重复拆分/翻倍
ALREADY_SKIN_RE = re.compile(r"^\d+(?:\.\d+)?mm(?:素板|多层加密|黑[炭碳]晶)$")

# 板材识别：关键词 + 模糊匹配（与 door-skin-converter 一致）
#   多层.?加密 / 加密.?多层 → 统一输出“多层加密”
#   黑[炭碳]晶 → 按源表实际写法输出
BOARD_MATERIAL_PATTERNS = (
    (re.compile(r"多层.?加密|加密.?多层"), "多层加密"),
    (re.compile(r"黑[炭碳]晶"), None),
)

# 已见过的工艺描述（归一化：去厚度、忽略大小写/全半角/空格）
KNOWN_PROCESS_DESCRIPTIONS = frozenset({
    "美心YSM",
    "复合贴YSM皮",
    "复合贴YSM皮多层加密",
    "美心YSM装门门框",
    "美心YSM装板",
    "PVC系列",
    "PP系列",
    "复合贴YSM皮装板",
    "复合贴YSM皮工艺",
    "复合贴YSM皮工艺多层加密",
    "复合贴PM皮",
    "复合贴PM皮外凸",
    "黑碳晶贴YSM皮",
    "黑炭晶贴YSM皮",
    "多层加密",
    "黑碳晶",
    "黑炭晶",
    "PVC系列多层板加密",
})


class ConversionError(ValueError):
    """输入表格不符合转换要求。"""


@dataclass(frozen=True)
class ConversionStats:
    source_rows: int
    output_rows: int
    input_quantity_sum: float
    quantity_sum: float
    split_source_rows: int
    normal_output_rows: int
    split_output_rows: int
    blank_thickness_orders: tuple[str, ...]
    split_details: tuple[str, ...] = ()
    excluded_details: tuple[str, ...] = ()
    material_mismatch_details: tuple[str, ...] = ()
    new_process_details: tuple[str, ...] = ()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_number(value: Any) -> int | float | str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = clean_text(value)
        try:
            number = float(text)
        except ValueError:
            return text
    return int(number) if number.is_integer() else number


def _normalize_model(text: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", unicodedata.normalize("NFKC", str(text)).upper())


def extract_door_model(workpiece: Any) -> str:
    """从板件名称中提取门型编号并规范化（如 'N1046(N1046)单扇' -> 'N1046'）。"""
    text = unicodedata.normalize("NFKC", clean_text(workpiece)).upper()
    match = DOOR_MODEL_RE.search(text)
    return _normalize_model(match.group(0)) if match else ""


def extract_board_material(process_text: Any) -> str:
    """从工艺路线文本识别板材类型，默认素板。"""
    text = clean_text(process_text)
    for pattern, canonical in BOARD_MATERIAL_PATTERNS:
        match = pattern.search(text)
        if match:
            return canonical or match.group(0)
    return "素板"


def normalize_process_description(process_text: Any) -> str:
    """工艺描述归一化：去掉厚度（如 8MM / 5+8MM），忽略大小写、全半角和空格。"""
    text = unicodedata.normalize("NFKC", clean_text(process_text)).upper()
    text = re.sub(r"\d+(?:\.\d+)?(?:\s*\+\s*\d+(?:\.\d+)?)*\s*MM", "", text)
    return re.sub(r"\s+", "", text)


def extract_thicknesses(process_text: Any) -> list[str]:
    match = THICKNESS_RE.search(clean_text(process_text))
    if not match:
        return []
    return [normalize_decimal_text(part) for part in match.group(1).split("+")]


def normalize_decimal_text(value: Any) -> str:
    text = clean_text(value)
    try:
        number = float(text)
    except ValueError:
        return text
    return str(int(number)) if number.is_integer() else format(number, "g")


def multiply_quantity(value: Any) -> int | float:
    try:
        number = float(clean_text(value))
    except ValueError as exc:
        raise ConversionError(f"数量不是有效数字：{value!r}") from exc
    result = number * 2
    return int(result) if result.is_integer() else result


def transform_rows(
    headers: Sequence[Any], rows: Iterable[Sequence[Any]]
) -> tuple[list[list[Any]], ConversionStats]:
    """PMS 优化门扇清单 → 皮行清单。

    规则（与 door-skin-converter 一致）：
    - 工艺路线含 X+X 厚度 → 拆成多行，每行一个厚度
    - 数量一律 ×2
    - 材料描述 = 厚度mm + 板材（多层加密/黑碳晶模糊识别，否则素板）；无厚度时保留原材料描述
    - 命中剔除门型/关键词的行不进入输出
    - 其他列原样透传
    """
    normalized_headers = [clean_text(value) for value in headers]
    if len(set(normalized_headers)) != len(normalized_headers):
        raise ConversionError("源表表头存在重复字段，无法确定列映射")

    missing = [name for name in REQUIRED_SOURCE_HEADERS if name not in normalized_headers]
    if missing:
        raise ConversionError(f"源表缺少字段：{', '.join(missing)}")

    index_of = {name: normalized_headers.index(name) for name in normalized_headers}
    order_header = next((h for h in ORDER_HEADER_CANDIDATES if h in index_of), None)

    excluded_normalized = {_normalize_model(m) for m in EXCLUDED_DOOR_MODELS}

    output_rows: list[list[Any]] = []
    source_row_count = 0
    split_source_rows = 0
    input_quantity_sum = 0.0
    blank_thickness_orders: list[str] = []
    split_details: list[str] = []
    excluded_details: list[str] = []
    material_mismatch_details: list[str] = []
    new_process_details: list[str] = []

    for source_row in rows:
        padded = list(source_row) + [""] * max(0, len(normalized_headers) - len(source_row))
        if all(clean_text(value) == "" for value in padded):
            continue

        workpiece = clean_text(padded[index_of["板件名称"]])
        process_text = clean_text(padded[index_of["工艺路线"]])
        order_number = clean_text(padded[index_of[order_header]]) if order_header else ""

        # 备注行/说明行：没有板件名称或工艺路线的直接跳过
        if not workpiece or not process_text:
            continue

        if any(keyword in workpiece for keyword in EXCLUDED_NAME_KEYWORDS):
            excluded_details.append(f"订单 {order_number}：{workpiece}（关键词剔除）")
            continue
        door_model = extract_door_model(workpiece)
        if door_model in excluded_normalized:
            excluded_details.append(f"订单 {order_number}：{workpiece}")
            continue

        source_row_count += 1
        try:
            raw_quantity = float(clean_text(padded[index_of["数量"]]))
        except ValueError as exc:
            raise ConversionError(f"订单 {order_number or '(空)'} 的数量不是有效数字：{padded[index_of['数量']]!r}") from exc
        input_quantity_sum += raw_quantity

        original_desc = clean_text(padded[index_of["材料描述"]])
        original_thickness = padded[index_of["厚度"]]

        # 已是皮行（材料描述已是 Nmm板材）：原样透传，不重复拆分/翻倍
        if ALREADY_SKIN_RE.match(original_desc):
            output_rows.append([clean_number(value) for value in padded[: len(normalized_headers)]])
            continue

        quantity = multiply_quantity(padded[index_of["数量"]])

        thicknesses = extract_thicknesses(process_text)
        if len(thicknesses) > 1:
            split_source_rows += 1
            raw_match = ""
            m = THICKNESS_RE.search(process_text)
            if m:
                raw_match = m.group(0)
            split_details.append(
                f"订单 {order_number}：{raw_match} → "
                f"{', '.join(f'{t}mm' for t in thicknesses)}"
            )
        if not thicknesses:
            blank_thickness_orders.append(f"订单 {order_number}：{workpiece}")
            thicknesses = [""]

        board_material = extract_board_material(process_text)
        if normalize_process_description(process_text) not in KNOWN_PROCESS_DESCRIPTIONS:
            new_process_details.append(
                f"订单 {order_number}：{workpiece}（工艺「{process_text}」未见过，请检查）"
            )

        for final_thickness in thicknesses:
            material_description = (
                f"{final_thickness}mm{board_material}" if final_thickness else original_desc
            )
            # 一致性校验：工艺路线声明的板材/厚度必须体现在材料描述里
            mismatches = []
            if board_material != "素板" and board_material not in material_description:
                mismatches.append(f"工艺声明「{board_material}」但材料描述未体现")
            if final_thickness and f"{final_thickness}mm" not in material_description:
                mismatches.append(f"工艺厚度 {final_thickness}mm 与材料描述不符")
            for pattern, _ in BOARD_MATERIAL_PATTERNS:
                hit = pattern.search(material_description)
                if hit and hit.group(0) != board_material:
                    mismatches.append(f"材料描述含「{hit.group(0)}」但工艺未声明")
            if mismatches:
                material_mismatch_details.append(
                    f"订单 {order_number}：{workpiece}（工艺「{process_text}」↔ 材料描述「{material_description}」，{'；'.join(mismatches)}）"
                )

            new_row = [clean_number(value) for value in padded[: len(normalized_headers)]]
            new_row[index_of["数量"]] = quantity
            new_row[index_of["材料描述"]] = material_description
            new_row[index_of["厚度"]] = clean_number(final_thickness) if final_thickness else clean_number(original_thickness)
            output_rows.append(new_row)

    normal_source_rows = source_row_count - split_source_rows
    normal_output_rows = normal_source_rows
    split_output_rows = len(output_rows) - normal_output_rows

    stats = ConversionStats(
        source_rows=source_row_count,
        output_rows=len(output_rows),
        input_quantity_sum=input_quantity_sum,
        quantity_sum=sum(float(row[index_of["数量"]]) for row in output_rows),
        split_source_rows=split_source_rows,
        normal_output_rows=normal_output_rows,
        split_output_rows=split_output_rows,
        blank_thickness_orders=tuple(blank_thickness_orders),
        split_details=tuple(split_details),
        excluded_details=tuple(excluded_details),
        material_mismatch_details=tuple(material_mismatch_details),
        new_process_details=tuple(new_process_details),
    )
    return output_rows, stats


def read_excel(content: bytes, filename: str) -> tuple[list[Any], list[list[Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        for encoding in ("gb18030", "utf-8-sig"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ConversionError("无法读取 .csv 文件：编码不是 GB18030 或 UTF-8")
        values = [row for row in csv.reader(io.StringIO(text))]
    elif suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("缺少 xlrd 依赖，无法读取 .xls") from exc
        try:
            workbook = xlrd.open_workbook(file_contents=content)
            sheet = workbook.sheet_by_index(0)
            values = [sheet.row_values(index) for index in range(sheet.nrows)]
        except Exception as exc:
            raise ConversionError(f"无法读取 .xls 文件：{exc}") from exc
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法读取 .xlsx") from exc
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.worksheets[0]
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
        except Exception as exc:
            raise ConversionError(f"无法读取 .xlsx 文件：{exc}") from exc
    else:
        raise ConversionError("只支持 .xls、.xlsx 或 .csv 文件")

    first_nonempty = next(
        (index for index, row in enumerate(values) if any(clean_text(value) for value in row)),
        None,
    )
    if first_nonempty is None:
        raise ConversionError("源表为空")
    return values[first_nonempty], values[first_nonempty + 1 :]


def write_csv(headers: Sequence[Any], rows: Sequence[Sequence[Any]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow([clean_text(h) for h in headers])
    writer.writerows(rows)
    return buffer.getvalue().encode("gb18030")


def convert_excel_to_csv(
    content: bytes, filename: str
) -> tuple[bytes, ConversionStats]:
    headers, source_rows = read_excel(content, filename)
    output_rows, stats = transform_rows(headers, source_rows)
    return write_csv(headers, output_rows), stats
