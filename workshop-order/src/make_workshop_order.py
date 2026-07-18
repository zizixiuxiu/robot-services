#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Convert an original quotation workbook into a workshop-order workbook.

Rules implemented:
1. Tiepi orders use the full workshop conversion rules.
2. Hunyou orders only clear wood-box packaging and color-adjustment fee rows.
3. Legacy .xls files are converted through LibreOffice before processing so formatting is preserved.

Usage:
    python make_workshop_order.py input.xlsx
    python make_workshop_order.py input.xlsx output.xlsx
"""

from __future__ import annotations

import argparse
import ast
import operator
import os
import re
import subprocess
import tempfile
import threading
import time
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange


DEFAULT_DISCOUNT = 0.85
BUSINESS_SCAN_MAX_COL = 32
LONG_NOTE_MIN_LENGTH = 30
DEFAULT_COLUMN_WIDTH = 8.43
SOFFICE_PROFILE_DIR = Path(
    os.getenv("WORKSHOP_SOFFICE_PROFILE_DIR", str(Path(tempfile.gettempdir()) / "workshop-order-libreoffice"))
)
SOFFICE_TIMEOUT_SECONDS = int(os.getenv("WORKSHOP_SOFFICE_TIMEOUT_SECONDS", "180"))
_SOFFICE_LOCK = threading.RLock()
_SOFFICE_WARMED = False

KW_PRODUCT = "\u4ea7\u54c1"
KW_UNIT_PRICE = "\u5355\u4ef7"
KW_YUAN = "\u5143"
KW_SUBTOTAL = "\u5c0f\u8ba1"
KW_WORKPOINT = "\u5de5\u5206"
KW_WOOD_BOX = "\u6728\u7bb1\u5305\u88c5"
KW_NON_STANDARD_COLOR_FEE = "\u975e\u6807\u8272\u8d39"
KW_TINTING_FEE = "\u8c03\u8272\u8d39"
KW_HUNYOU = "\u6df7\u6cb9"
KW_COLOR = "\u989c\u8272"
KW_TOTAL = "\u5408\u8ba1"
KW_BOARD = "\u677f\u6750"
KW_HARDWARE = "\u4e94\u91d1"
KW_PAYMENT_TOTAL = "\u5408\u8ba1\u91d1\u989d"
KW_NEED_PAY = "\u9700\u652f\u4ed8"
KW_NEED_PAY_SHORT = "\u9700\u4ed8"
KW_DISCOUNTED = "\u6298\u540e"
KW_ORIGINAL_ORDER = "\u539f\u5355\u53f7"
KW_COLOR_DIFF = "\u8272\u5dee"
KW_AFTER_SALES = "\u552e\u540e"
KW_AGREE_BOARD = "\u540c\u610f\u677f\u6750"
KW_CRAFT = "\u5de5\u827a"

DROP_NOTE_KEYWORDS = (
    KW_ORIGINAL_ORDER,
)
PROTECTED_NOTE_KEYWORDS = (
    KW_AFTER_SALES,
    KW_CRAFT,
    "\u4ee5\u5b9e\u7269\u4e3a\u51c6",
)

PAGE_TEXT_RE = re.compile(r"\u6b64\u5355\u5171\s*(\d+)\s*\u9875")
PAGE_COUNT_RE = re.compile(r"\u5171\u8ba1\s*(\d+)\s*\u9875")
ORDER_NUMBER_RE = re.compile(r"^([A-Za-z]\d{4}(?:-\d+)+)")
FULL_ORDER_NUMBER_RE = re.compile(r"^[A-Za-z]\d{4}(?:-\d+)+$")

# Payment-related keywords that should be stripped from bottom notes.
PAYMENT_AMOUNT_KEYWORDS = (
    "\u5408\u8ba1\u91d1\u989d",
    "85\u6298\u4f18\u60e0",
    "\u9700\u652f\u4ed8",
    "\u9700\u4ed8",
    "\u5e94\u652f\u4ed8",
    "\u5e94\u4ed8",
    "\u4f18\u60e0\u91d1\u989d",
    "\u6298\u540e\u91d1\u989d",
)

PAYMENT_AMOUNT_PATTERNS = (
    re.compile(r"\u6b64\u5355[^，。,；;!！]*(?:\u91d1\u989d|\u91d1\u603b)"),
    re.compile(r"\d+(?:\.\d+)?\s*\u5143"),
    re.compile(r"\d+(?:\.\d+)?\s*\u6298"),
    re.compile(r"\d+(?:\.\d+)?\s*\u6298\s*\u4f18\u60e0"),
    re.compile(r"\u4f18\u60e0[^，。,；;!！]*\d+(?:\.\d+)?\s*\u5143"),
    re.compile(r"\u6298\u540e[^，。,；;!！]*\d+(?:\.\d+)?\s*\u5143"),
    re.compile(r"\u5408\u8ba1\u91d1\u989d[^，。,；;!！]*\d+(?:\.\d+)?\s*\u5143"),
    re.compile(r"\u9700\u652f?\u4ed8[^，。,；;!！]*\d+(?:\.\d+)?\s*\u5143"),
)

PAYMENT_AMOUNT_EMPTY_REMAINDERS = {
    "\u94f6\u955c",
    "\u6837\u54c1\u6253",
    "\u6b64\u5355\u91d1\u989d\u603b\u540c\u610f",
    "\u6b64\u5355\u91d1\u603b\u540c\u610f",
}

PAYMENT_NOTE_TRIM_CHARS = ",\uff0c\uff1b;:\uff1a "
PAYMENT_SEGMENT_SEPARATOR_RE = re.compile(r"(?:[\r\n]+|[ \t\u3000]{4,})")
DROP_NOTE_SEGMENT_KEYWORDS: tuple[str, ...] = ()


@dataclass(frozen=True)
class TextCleanupResult:
    value: str | None
    changed: bool


@dataclass
class TransformStats:
    order_type: str
    discounted_prices: int = 0
    cleared_wood_boxes: int = 0
    cleared_non_standard_color_fees: int = 0
    updated_page_notes: int = 0
    summary_constants_removed: int = 0
    cleared_bottom_notes: int = 0
    filled_workpoints: int = 0
    skipped_prices: int = 0
    deleted_blank_rows: int = 0
    moved_hardware_sheets: int = 0
    updated_order_numbers: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, int | str | list[str]]:
        return {
            "order_type": self.order_type,
            "discounted_prices": self.discounted_prices,
            "cleared_wood_boxes": self.cleared_wood_boxes,
            "cleared_non_standard_color_fees": self.cleared_non_standard_color_fees,
            "updated_page_notes": self.updated_page_notes,
            "summary_constants_removed": self.summary_constants_removed,
            "cleared_bottom_notes": self.cleared_bottom_notes,
            "filled_workpoints": self.filled_workpoints,
            "skipped_prices": self.skipped_prices,
            "deleted_blank_rows": self.deleted_blank_rows,
            "moved_hardware_sheets": self.moved_hardware_sheets,
            "updated_order_numbers": self.updated_order_numbers,
            "warnings": self.warnings,
        }


@dataclass(frozen=True)
class OrderContext:
    input_path: Path
    discount: float
    requested_order_type: str
    resolved_order_type: str
    page_count: str | None
    inferred_order_number: str | None


@dataclass(frozen=True)
class DetailSheetContext:
    wb: Any
    ws: Any
    cached_ws: Any | None
    header_row: int
    product_col: int
    price_col: int
    subtotal_col: int
    workpoint_col: int | None
    color_col: int | None


def payment_amount_info_index(value: str) -> int | None:
    indexes = [idx for kw in PAYMENT_AMOUNT_KEYWORDS if (idx := value.find(kw)) != -1]
    indexes.extend(match.start() for pattern in PAYMENT_AMOUNT_PATTERNS if (match := pattern.search(value)))
    return min(indexes) if indexes else None


def clean_payment_segment(value: str) -> str | None:
    min_idx = payment_amount_info_index(value)
    if min_idx is None:
        return value

    # Keep genuine page/after-sales text, but remove trailing payment details.
    prefix = value[:min_idx].rstrip(PAYMENT_NOTE_TRIM_CHARS)
    keep_prefix = (
        any(keyword in value[:min_idx] for keyword in PROTECTED_NOTE_KEYWORDS)
        or PAGE_TEXT_RE.search(prefix)
        or PAGE_COUNT_RE.search(prefix)
    )
    if keep_prefix and prefix and prefix not in PAYMENT_AMOUNT_EMPTY_REMAINDERS:
        # Discount approval text can be glued to an after-sales sentence, e.g.
        # "...不进入公司售后！此单某总同意板材95折优惠...". Keep only the
        # completed after-sales/page sentence before that price fragment.
        delimiter_positions = [prefix.rfind(ch) for ch in "\u3002\uff01!?\uff1f\uff1b;"]
        last_delimiter = max(delimiter_positions)
        if last_delimiter != -1:
            completed_prefix = prefix[: last_delimiter + 1].rstrip(PAYMENT_NOTE_TRIM_CHARS)
            dangling_fragment = prefix[last_delimiter + 1 :].strip()
            dangling_has_page_count = PAGE_TEXT_RE.search(dangling_fragment) or PAGE_COUNT_RE.search(dangling_fragment)
            if completed_prefix and dangling_fragment and KW_AFTER_SALES not in dangling_fragment and not dangling_has_page_count:
                return completed_prefix
        return prefix

    # If the segment is only meaningful because of the discount/amount, drop it entirely.
    return None


def clean_payment_amount_text(value: str) -> TextCleanupResult:
    """Remove payment/discount amounts from a note."""
    min_idx = payment_amount_info_index(value)
    if min_idx is None:
        return TextCleanupResult(value, False)

    segments = [segment for segment in PAYMENT_SEGMENT_SEPARATOR_RE.split(value) if segment.strip()]
    if len(segments) > 1:
        cleaned_segments = []
        changed = False
        for segment in segments:
            cleaned_segment = clean_payment_segment(segment.strip())
            if cleaned_segment is None:
                changed = True
                continue
            changed = changed or cleaned_segment != segment.strip()
            cleaned_segments.append(cleaned_segment)

        cleaned = "\n".join(cleaned_segments).strip()
        if not cleaned:
            return TextCleanupResult(None, True)
        return TextCleanupResult(cleaned, changed or cleaned != value)

    cleaned = clean_payment_segment(value)
    if not cleaned or cleaned in PAYMENT_AMOUNT_EMPTY_REMAINDERS:
        return TextCleanupResult(None, True)
    return TextCleanupResult(cleaned, cleaned != value)


def remove_payment_amount_info(value: str) -> str:
    """Strip payment amount suffixes from bottom notes."""
    result = clean_payment_amount_text(value)
    return result.value or ""


def should_drop_note_segment(segment: str) -> bool:
    if not any(keyword in segment for keyword in DROP_NOTE_SEGMENT_KEYWORDS):
        return False
    return "\u6b64\u5355" in segment or len(segment.strip()) >= LONG_NOTE_MIN_LENGTH


def clean_non_workshop_note_text(value: str) -> TextCleanupResult:
    """Drop after-sales/process note segments from workshop output notes."""
    segments = [segment for segment in PAYMENT_SEGMENT_SEPARATOR_RE.split(value) if segment.strip()]
    if not segments:
        return TextCleanupResult(value, False)

    cleaned_segments = []
    changed = False
    for segment in segments:
        stripped = segment.strip()
        if should_drop_note_segment(stripped):
            changed = True
            continue
        cleaned_segments.append(stripped)

    if not changed:
        return TextCleanupResult(value, False)
    cleaned = "\n".join(cleaned_segments).strip()
    return TextCleanupResult(cleaned or None, True)


ORDER_TYPES = {"auto", "tiepi", "hunyou"}
COLOR_FEE_KEYWORDS = (
    KW_NON_STANDARD_COLOR_FEE,
    KW_TINTING_FEE,
)


def text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def safe_eval_arithmetic(expr: str) -> float | None:
    """Evaluate simple numeric formulas such as '=794+120+150+80'."""
    if not expr or not expr.startswith("="):
        return None

    source = expr[1:].strip()
    if not re.fullmatch(r"[0-9+\-*/().\s]+", source):
        return None

    operators = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
        ast.USub: operator.neg,
        ast.UAdd: operator.pos,
    }

    def visit(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return visit(node.body)
        if isinstance(node, ast.Constant) and is_number(node.value):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and type(node.op) in operators:
            return operators[type(node.op)](visit(node.operand))
        if isinstance(node, ast.BinOp) and type(node.op) in operators:
            return operators[type(node.op)](visit(node.left), visit(node.right))
        raise ValueError("unsupported formula")

    try:
        return visit(ast.parse(source, mode="eval"))
    except Exception:
        return None


def find_header_columns(ws: Any) -> tuple[int, int, int, int, int | None] | None:
    """Return header_row, product_col, price_col, subtotal_col, workpoint_col."""
    max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
    for row in ws.iter_rows(max_col=max_col):
        product_col = price_col = subtotal_col = workpoint_col = None
        for cell in row:
            value = text(cell.value).replace("\n", "")
            # Accept both "单价" / "单价（元）" and "小计" / "小计（元）" forms.
            if KW_PRODUCT in value:
                product_col = cell.column
            elif KW_UNIT_PRICE in value:
                price_col = cell.column
            elif KW_SUBTOTAL in value:
                subtotal_col = cell.column
            elif KW_WORKPOINT in value:
                workpoint_col = cell.column
        if product_col and price_col and subtotal_col:
            return row[0].row, product_col, price_col, subtotal_col, workpoint_col
    return None


def get_price_value(formula_cell: Any, cached_cell: Any | None = None) -> float | None:
    raw = formula_cell.value
    if is_number(raw):
        return float(raw)
    if isinstance(raw, str):
        price = safe_eval_arithmetic(raw)
        if price is not None:
            return price
    if cached_cell is not None and is_number(cached_cell.value):
        return float(cached_cell.value)
    return None


def remove_numeric_formula_terms(formula: str) -> str:
    """Remove standalone numeric additions from a board-summary formula, e.g. '+160'."""
    if not formula or not formula.startswith("="):
        return formula
    result = re.sub(r"\+\s*\d+(?:\.\d+)?(?=$|[+\-])", "", formula)
    result = re.sub(r"-\s*\d+(?:\.\d+)?(?=$|[+\-])", "", result)
    return result


def is_color_fee_product(product: str) -> bool:
    return any(keyword in product for keyword in COLOR_FEE_KEYWORDS)


def is_hunyou_color(color: str) -> bool:
    return re.match(r"^[\s（(【\[]*\s*混油", color) is not None


def detect_order_type(wb: Any, requested: str) -> str:
    if requested != "auto":
        return requested

    for ws in wb.worksheets:
        # Only treat sheets with a full production-detail header as production
        # lists. BOM/summary sheets may also contain "产品名称" and "颜色"
        # columns but lack a "单价（元）" column, so they must not influence
        # order-type detection.
        headers = find_header_columns(ws)
        if not headers:
            continue
        header_row, product_col, price_col, subtotal_col, workpoint_col = headers

        color_col = None
        for cell in ws[header_row]:
            value = text(cell.value).replace("\n", "")
            if KW_COLOR in value:
                color_col = cell.column
                break
        if not color_col:
            continue

        current_product = ""
        current_color = ""
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_product = text(ws.cell(row_idx, product_col).value) if product_col else ""
            if raw_product:
                current_product = raw_product
            product = current_product

            if KW_WOOD_BOX in product or product == KW_TOTAL:
                break

            raw_color = text(ws.cell(row_idx, color_col).value)
            if raw_color:
                current_color = raw_color
            color = current_color

            if is_hunyou_color(color):
                return "hunyou"
    return "tiepi"


def find_color_column(ws: Any) -> tuple[int, int, int | None] | None:
    """Return header_row, color_col, product_col for production detail sheets."""
    max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
    for row in ws.iter_rows(max_col=max_col):
        color_col = None
        product_col = None
        for cell in row:
            value = text(cell.value).replace("\n", "")
            if KW_COLOR in value:
                color_col = cell.column
            elif KW_PRODUCT in value:
                product_col = cell.column
        if color_col and product_col:
            return row[0].row, color_col, product_col
    return None


def extract_page_count(wb: Any) -> str | None:
    for ws in wb.worksheets:
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for row in ws.iter_rows(max_col=max_col):
            for cell in row:
                value = text(cell.value)
                match = PAGE_TEXT_RE.search(value) or PAGE_COUNT_RE.search(value)
                if match:
                    return match.group(1)
    return None


def infer_order_number(input_path: Path) -> str | None:
    match = ORDER_NUMBER_RE.match(input_path.stem)
    return match.group(1) if match else None


def is_full_order_number(value: Any) -> bool:
    return isinstance(value, str) and FULL_ORDER_NUMBER_RE.fullmatch(value.strip()) is not None


def find_existing_order_number(wb: Any) -> str | None:
    """Return a concrete order number already present in the workbook header."""
    for ws in wb.worksheets:
        if not str(ws.title).isdigit():
            continue
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for row in ws.iter_rows(max_col=max_col):
            for cell in row:
                if text(cell.value) != "\u8ba2\u5355\u7f16\u53f7":
                    continue
                for offset in (1, 2):
                    value = ws.cell(cell.row, cell.column + offset).value
                    if is_full_order_number(value):
                        return str(value).strip()
    return None


def set_plain_cell(ws: Any, row: int, col: int, value: Any) -> bool:
    if row < 1 or col < 1 or row > ws.max_row or col > ws.max_column:
        return False
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell) or cell.value == value:
        return False
    cell.value = value
    return True


def should_update_suffix_cell(ws: Any, row: int, col: int) -> bool:
    if row < 1 or col < 1 or row > ws.max_row or col > ws.max_column:
        return False
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    return cell.value not in (None, "")


def normalize_order_numbers(wb: Any, order_number: str | None) -> int:
    if not order_number:
        return 0

    changed = 0
    for ws in wb.worksheets:
        if not str(ws.title).isdigit():
            continue
        suffix = -int(ws.title)
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for row in ws.iter_rows(max_col=max_col):
            for cell in row:
                value = text(cell.value)
                if value == "\u8ba2\u5355\u7f16\u53f7":
                    changed += int(set_plain_cell(ws, cell.row, cell.column + 1, order_number))
                    if should_update_suffix_cell(ws, cell.row, cell.column + 2):
                        changed += int(set_plain_cell(ws, cell.row, cell.column + 2, suffix))
                elif value == "\u751f\u4ea7\u5355\u53f7" and cell.column >= 14:
                    changed += int(set_plain_cell(ws, cell.row, cell.column + 2, order_number))
                    if should_update_suffix_cell(ws, cell.row, cell.column + 4):
                        changed += int(set_plain_cell(ws, cell.row, cell.column + 4, suffix))
    return changed


def translate_formula(formula: str, origin: str, target: str) -> str | None:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return None


CELL_REF_RE = re.compile(
    r"(?P<sheet>(?:'(?P<quoted_sheet>(?:[^']|'')+)'|(?P<plain_sheet>[^'!:+\-*/^&=<>(),\s]+))!)?"
    r"(?P<col_abs>\$?)(?P<col>[A-Z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)"
)


def formula_sheet_name(match: re.Match[str]) -> str | None:
    if match.group("quoted_sheet") is not None:
        return match.group("quoted_sheet").replace("''", "'")
    if match.group("plain_sheet") is not None:
        return match.group("plain_sheet")
    return None


def adjust_formula_after_row_delete(
    formula: str,
    formula_sheet: str,
    deleted_sheet: str,
    deleted_row: int,
    amount: int = 1,
) -> str:
    """Update A1-style row references after deleting rows from one sheet."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def replace(match: re.Match[str]) -> str:
        ref_sheet = formula_sheet_name(match) or formula_sheet
        if ref_sheet != deleted_sheet:
            return match.group(0)

        row = int(match.group("row"))
        if row > deleted_row:
            row -= amount
        elif row == deleted_row and match.start() > 0 and formula[match.start() - 1] == ":":
            row = max(1, row - amount)
        else:
            return match.group(0)

        return (
            f"{match.group('sheet') or ''}"
            f"{match.group('col_abs')}{match.group('col')}"
            f"{match.group('row_abs')}{row}"
        )

    return CELL_REF_RE.sub(replace, formula)


def adjust_workbook_formulas_after_row_delete(wb: Any, deleted_ws: Any, deleted_row: int, amount: int = 1) -> int:
    changed = 0
    for ws in wb.worksheets:
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for row in ws.iter_rows(max_col=max_col):
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                adjusted = adjust_formula_after_row_delete(value, ws.title, deleted_ws.title, deleted_row, amount)
                if adjusted != value:
                    cell.value = adjusted
                    changed += 1
    return changed


def prepare_merged_cells_for_row_delete(ws: Any, deleted_row: int, amount: int = 1) -> list[str]:
    shifted_ranges: list[str] = []
    for merged_range in list(ws.merged_cells.ranges):
        cell_range = CellRange(str(merged_range))
        if cell_range.min_row <= deleted_row <= cell_range.max_row:
            ws.unmerge_cells(str(merged_range))
        elif cell_range.min_row > deleted_row:
            ws.unmerge_cells(str(merged_range))
            cell_range.shift(row_shift=-amount)
            shifted_ranges.append(str(cell_range))
    return shifted_ranges


def fill_workpoint_formulas(
    ws: Any,
    header_row: int,
    product_col: int,
    subtotal_col: int,
    workpoint_col: int | None,
) -> int:
    if not workpoint_col:
        return 0

    filled = 0
    data_start = header_row + 1
    last_product_row = header_row
    known_formula_cell = None
    workpoint_letter = get_column_letter(workpoint_col)
    subtotal_letter = get_column_letter(subtotal_col)
    current_product = ""

    def normalized_formula(value: Any) -> str:
        if not isinstance(value, str):
            return ""
        return value.replace(" ", "").replace("=+", "=").upper()

    def find_workpoint_style_template() -> Any | None:
        scan_product = ""
        for scan_row in range(data_start, ws.max_row + 1):
            max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
            row_labels = [text(ws.cell(scan_row, col).value) for col in range(1, max_col + 1)]
            if any(KW_WOOD_BOX in value or value == KW_TOTAL for value in row_labels):
                break

            raw_product = text(ws.cell(scan_row, product_col).value)
            subtotal_value = ws.cell(scan_row, subtotal_col).value
            if raw_product:
                scan_product = raw_product
            elif scan_product and subtotal_value not in (None, ""):
                pass
            else:
                scan_product = ""
                continue
            if subtotal_value in (None, ""):
                continue

            candidate = ws.cell(scan_row, workpoint_col)
            expected = f"={subtotal_letter}{scan_row}/100"
            if isinstance(candidate.value, str) and normalized_formula(candidate.value) == normalized_formula(expected):
                return candidate
        return None

    style_template_cell = find_workpoint_style_template()

    def apply_workpoint_style(cell: Any) -> None:
        if style_template_cell is None or cell.coordinate == style_template_cell.coordinate:
            return
        cell._style = copy(style_template_cell._style)

    for row_idx in range(data_start, ws.max_row + 1):
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        row_labels = [text(ws.cell(row_idx, col).value) for col in range(1, max_col + 1)]
        if any(KW_WOOD_BOX in value or value == KW_TOTAL for value in row_labels):
            break

        raw_product = text(ws.cell(row_idx, product_col).value)
        subtotal_value = ws.cell(row_idx, subtotal_col).value
        if raw_product:
            current_product = raw_product
        elif current_product and subtotal_value not in (None, ""):
            # continuation of a vertically merged product cell
            pass
        else:
            current_product = ""
            continue
        if subtotal_value in (None, ""):
            continue

        product = current_product
        last_product_row = row_idx
        workpoint_cell = ws.cell(row_idx, workpoint_col)
        expected_formula = f"={subtotal_letter}{row_idx}/100"
        if isinstance(workpoint_cell.value, str) and workpoint_cell.value.startswith("="):
            if normalized_formula(workpoint_cell.value) != normalized_formula(expected_formula):
                workpoint_cell.value = expected_formula
                apply_workpoint_style(workpoint_cell)
                filled += 1
            known_formula_cell = workpoint_cell
            continue
        if workpoint_cell.value not in (None, ""):
            continue

        formula = None
        if known_formula_cell is not None:
            formula = translate_formula(
                known_formula_cell.value,
                known_formula_cell.coordinate,
                workpoint_cell.coordinate,
            )
        if not formula:
            formula = f"={subtotal_letter}{row_idx}/100"

        workpoint_cell.value = formula
        apply_workpoint_style(workpoint_cell)
        filled += 1

    if last_product_row <= header_row:
        return filled

    for row_idx in range(last_product_row + 1, ws.max_row + 1):
        labels = [text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, product_col + 1) + 1)]
        if any(value == KW_TOTAL for value in labels):
            total_cell = ws.cell(row_idx, workpoint_col)
            if total_cell.value in (None, ""):
                total_cell.value = f"=SUM({workpoint_letter}{data_start}:{workpoint_letter}{last_product_row})"
                filled += 1
            break

    return filled


def delete_blank_rows_between_data_and_summary(wb: Any, ws: Any, header_row: int) -> int:
    """Delete blank rows between product data and summary/wood-box rows.

    Merged cells that span the blank row are unmerged first. Formula references
    across the workbook are then shifted with Excel-like row-deletion semantics.
    """
    boundary_row = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for col_idx in range(1, max_col + 1):
            val = text(ws.cell(row_idx, col_idx).value)
            if val in (KW_TOTAL, KW_WOOD_BOX):
                boundary_row = row_idx
                break
        if boundary_row:
            break

    if boundary_row is None or boundary_row <= header_row + 1:
        return 0

    deleted = 0
    for row_idx in range(boundary_row - 1, header_row, -1):
        is_blank = True
        max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value not in (None, ""):
                is_blank = False
                break
        if not is_blank:
            continue

        shifted_merged_ranges = prepare_merged_cells_for_row_delete(ws, row_idx)

        ws.delete_rows(row_idx)
        for merged_range in shifted_merged_ranges:
            ws.merge_cells(merged_range)
        adjust_workbook_formulas_after_row_delete(wb, ws, row_idx)

        # Keep sequence numbers in column A continuous after the row deletion.
        for r in range(row_idx, ws.max_row + 1):
            seq_cell = ws.cell(r, 1)
            if isinstance(seq_cell.value, int):
                seq_cell.value -= 1

        deleted += 1

    return deleted


def is_hardware_sheet_name(sheet_name: str) -> bool:
    name = str(sheet_name).strip()
    return name.startswith(KW_HARDWARE) or name.endswith(KW_HARDWARE) or KW_HARDWARE in name


def move_hardware_sheets_to_end(wb: Any) -> int:
    moved = 0
    for ws in list(wb.worksheets):
        if is_hardware_sheet_name(ws.title):
            wb._sheets.remove(ws)
            wb._sheets.append(ws)
            moved += 1
    return moved


def build_order_context(input_path: Path, discount: float, requested_order_type: str, wb: Any) -> OrderContext:
    resolved_order_type = detect_order_type(wb, requested_order_type)
    return OrderContext(
        input_path=input_path,
        discount=discount,
        requested_order_type=requested_order_type,
        resolved_order_type=resolved_order_type,
        page_count=extract_page_count(wb),
        inferred_order_number=find_existing_order_number(wb) or infer_order_number(input_path),
    )


def build_detail_sheet_context(wb: Any, cached_wb: Any | None, ws: Any) -> DetailSheetContext | None:
    headers = find_header_columns(ws)
    if not headers:
        return None

    header_row, product_col, price_col, subtotal_col, workpoint_col = headers
    color_header = find_color_column(ws)
    color_col = color_header[1] if color_header else None
    return DetailSheetContext(
        wb=wb,
        ws=ws,
        cached_ws=cached_wb[ws.title] if cached_wb is not None else None,
        header_row=header_row,
        product_col=product_col,
        price_col=price_col,
        subtotal_col=subtotal_col,
        workpoint_col=workpoint_col,
        color_col=color_col,
    )


def workbook_needs_cached_prices(wb: Any) -> bool:
    for ws in wb.worksheets:
        headers = find_header_columns(ws)
        if not headers:
            continue

        header_row, product_col, price_col, subtotal_col, _workpoint_col = headers
        current_product = ""
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_product = text(ws.cell(row_idx, product_col).value)
            price_cell = ws.cell(row_idx, price_col)
            subtotal_cell = ws.cell(row_idx, subtotal_col)

            if raw_product:
                current_product = raw_product
            elif current_product and (price_cell.value not in (None, "") or subtotal_cell.value not in (None, "")):
                pass
            else:
                current_product = ""

            product = current_product
            if (
                is_color_fee_product(product)
                or KW_WOOD_BOX in product
                or not product
                or KW_TOTAL in product
                or isinstance(price_cell, MergedCell)
            ):
                continue

            raw_price = price_cell.value
            if isinstance(raw_price, str) and raw_price.startswith("=") and safe_eval_arithmetic(raw_price) is None:
                return True

    return False


def clear_color_fee_row(ctx: DetailSheetContext, row_idx: int) -> bool:
    last_clear_col = max(
        ctx.subtotal_col,
        ctx.workpoint_col or ctx.subtotal_col,
        min(ctx.ws.max_column, 18),
    )
    changed = False
    for col_idx in range(1, last_clear_col + 1):
        row_cell = ctx.ws.cell(row_idx, col_idx)
        if not isinstance(row_cell, MergedCell) and row_cell.value not in (None, ""):
            row_cell.value = None
            changed = True
    return changed


def clear_wood_box_row(ctx: DetailSheetContext, row_idx: int) -> bool:
    changed = False
    for col_idx in (ctx.subtotal_col, ctx.workpoint_col):
        if not col_idx:
            continue
        row_cell = ctx.ws.cell(row_idx, col_idx)
        if not isinstance(row_cell, MergedCell) and row_cell.value not in (None, ""):
            row_cell.value = None
            changed = True
    return changed


def process_detail_rows(ctx: DetailSheetContext, order_context: OrderContext, stats: TransformStats) -> None:
    current_product = ""
    for row_idx in range(ctx.header_row + 1, ctx.ws.max_row + 1):
        max_col = min(ctx.ws.max_column, BUSINESS_SCAN_MAX_COL)
        row_values = [text(ctx.ws.cell(row_idx, col_idx).value) for col_idx in range(1, max_col + 1)]
        if any(KW_WOOD_BOX in value for value in row_values):
            if clear_wood_box_row(ctx, row_idx):
                stats.cleared_wood_boxes += 1
            current_product = ""
            continue

        raw_product = text(ctx.ws.cell(row_idx, ctx.product_col).value)
        price_cell = ctx.ws.cell(row_idx, ctx.price_col)
        subtotal_cell = ctx.ws.cell(row_idx, ctx.subtotal_col)

        if raw_product:
            current_product = raw_product
        elif current_product and (price_cell.value not in (None, "") or subtotal_cell.value not in (None, "")):
            pass
        else:
            current_product = ""

        product = current_product
        if is_color_fee_product(product):
            if clear_color_fee_row(ctx, row_idx):
                stats.cleared_non_standard_color_fees += 1
            continue

        if KW_WOOD_BOX in product:
            if clear_wood_box_row(ctx, row_idx):
                stats.cleared_wood_boxes += 1
            continue

        if not product or KW_TOTAL in product or isinstance(price_cell, MergedCell):
            continue

        # 混油订单不清除行内容，但不对价格进行折扣
        if order_context.resolved_order_type == "hunyou":
            continue

        cached_price_cell = ctx.cached_ws.cell(row_idx, ctx.price_col) if ctx.cached_ws is not None else None
        price = get_price_value(price_cell, cached_price_cell)
        if price is None:
            if price_cell.value not in (None, ""):
                stats.skipped_prices += 1
            continue

        price_cell.value = price * order_context.discount
        stats.discounted_prices += 1


def cleanup_sheet_notes(ws: Any, order_context: OrderContext, stats: TransformStats) -> None:
    max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
    for row in ws.iter_rows(max_col=max_col):
        row_values = [text(cell.value) for cell in row]
        row_has_board_summary = any(KW_BOARD in value for value in row_values)
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue

            if value.startswith("=") and row_has_board_summary:
                new_formula = remove_numeric_formula_terms(value)
                if new_formula != value:
                    cell.value = new_formula
                    stats.summary_constants_removed += 1
                continue

            note_cleanup = clean_non_workshop_note_text(value)
            if note_cleanup.changed:
                cell.value = note_cleanup.value
                stats.updated_page_notes += 1
                if cell.value is None:
                    continue
                value = cell.value

            cleanup = clean_payment_amount_text(value)
            if cleanup.changed:
                cell.value = cleanup.value
                stats.updated_page_notes += 1


def display_text_units(value: str) -> int:
    return sum(2 if ord(ch) > 127 else 1 for ch in value)


def is_long_note_text(value: str) -> bool:
    if len(value.strip()) < LONG_NOTE_MIN_LENGTH:
        return False
    return KW_AFTER_SALES in value or KW_AGREE_BOARD in value or "\u56fe\u6587\u8bf4\u660e" in value


def normalize_note_display_text(value: str) -> str:
    """Use real line breaks instead of long space runs inside merged note cells."""
    return re.sub(r"[ \t\u3000]{4,}", "\n", value.strip())


def merged_range_for_cell(ws: Any, row_idx: int, col_idx: int) -> CellRange | None:
    coordinate = ws.cell(row_idx, col_idx).coordinate
    for merged_range in ws.merged_cells.ranges:
        if coordinate in merged_range:
            return CellRange(str(merged_range))
    return None


def cell_display_width(ws: Any, row_idx: int, col_idx: int) -> float:
    merged_range = merged_range_for_cell(ws, row_idx, col_idx)
    min_col = merged_range.min_col if merged_range else col_idx
    max_col = merged_range.max_col if merged_range else col_idx
    width = 0.0
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        width += ws.column_dimensions[letter].width or DEFAULT_COLUMN_WIDTH
    return max(width, DEFAULT_COLUMN_WIDTH)


def estimated_wrapped_lines(value: str, width: float) -> int:
    capacity = max(12, int(width * 1.15))
    lines = 0
    for part in value.splitlines() or [""]:
        units = max(1, display_text_units(part))
        lines += max(1, (units + capacity - 1) // capacity)
    return lines


def fit_long_note_rows(ws: Any) -> int:
    adjusted = 0
    max_col = min(ws.max_column, BUSINESS_SCAN_MAX_COL)
    for row in ws.iter_rows(max_col=max_col):
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not is_long_note_text(value):
                continue

            normalized = normalize_note_display_text(value)
            if normalized != value:
                cell.value = normalized
                value = normalized

            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = alignment.vertical or "center"
            cell.alignment = alignment

            width = cell_display_width(ws, cell.row, cell.column)
            lines = estimated_wrapped_lines(value, width)
            desired_height = min(180, max(42, lines * 24 + 18))

            merged_range = merged_range_for_cell(ws, cell.row, cell.column)
            if merged_range is not None and merged_range.max_row > cell.row:
                # 纵向跨多行的合并单元格：按整个合并区域的合计行高判断，
                # 不够时把差值补到首行，避免把全部高度堆到一行
                spanned_rows = range(cell.row, merged_range.max_row + 1)
                total_height = sum(
                    ws.row_dimensions[r].height or 15 for r in spanned_rows
                )
                if total_height < desired_height:
                    current_height = ws.row_dimensions[cell.row].height or 15
                    ws.row_dimensions[cell.row].height = (
                        current_height + desired_height - total_height
                    )
                    adjusted += 1
                continue

            current_height = ws.row_dimensions[cell.row].height or 15
            if current_height < desired_height:
                ws.row_dimensions[cell.row].height = desired_height
                adjusted += 1
    return adjusted


def process_detail_sheet(ctx: DetailSheetContext, order_context: OrderContext, stats: TransformStats) -> None:
    # 工分列处理：所有订单类型（含混油）均需填充工分公式
    if ctx.workpoint_col is None:
        stats.warnings.append(f"工作表 '{ctx.ws.title}' 缺少工分列，未补填工分")
    else:
        filled = fill_workpoint_formulas(
            ctx.ws,
            ctx.header_row,
            ctx.product_col,
            ctx.subtotal_col,
            ctx.workpoint_col,
        )
        stats.filled_workpoints += filled
        if filled > 0:
            stats.warnings.append(f"工作表 '{ctx.ws.title}' 补填了 {filled} 个工分公式")

    process_detail_rows(ctx, order_context, stats)
    cleanup_sheet_notes(ctx.ws, order_context, stats)
    stats.deleted_blank_rows += delete_blank_rows_between_data_and_summary(ctx.wb, ctx.ws, ctx.header_row)
    fit_long_note_rows(ctx.ws)


def enable_excel_recalculation(wb: Any) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass


def soffice_profile_uri(profile_dir: Path = SOFFICE_PROFILE_DIR) -> str:
    profile_dir.mkdir(parents=True, exist_ok=True)
    return profile_dir.resolve().as_uri()


def soffice_base_command(profile_dir: Path = SOFFICE_PROFILE_DIR) -> list[str]:
    return [
        "soffice",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nofirststartwizard",
        "--norestore",
        "--nolockcheck",
        f"-env:UserInstallation={soffice_profile_uri(profile_dir)}",
    ]


def warm_xls_converter(timeout: int = 30) -> float:
    """Initialize LibreOffice's profile so the first real .xls conversion is not cold."""
    global _SOFFICE_WARMED
    if _SOFFICE_WARMED:
        return 0.0

    with _SOFFICE_LOCK:
        if _SOFFICE_WARMED:
            return 0.0

        start = time.perf_counter()
        result = subprocess.run(
            soffice_base_command() + ["--terminate_after_init"],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        elapsed = time.perf_counter() - start
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(f"LibreOffice profile warm-up failed: {detail}")
        _SOFFICE_WARMED = True
        return elapsed


def transform(input_path: Path, output_path: Path, discount: float, order_type: str = "auto") -> dict[str, int | str]:
    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    workbook_path = input_path
    if input_path.suffix.lower() == ".xls":
        temp_dir = tempfile.TemporaryDirectory(prefix="workshop_order_")
        workbook_path = Path(temp_dir.name) / "converted.xlsx"
        convert_xls_to_xlsx(input_path, workbook_path)

    wb = load_workbook(workbook_path)
    order_context = build_order_context(input_path, discount, order_type, wb)
    cached_wb = (
        load_workbook(workbook_path, data_only=True)
        if order_context.resolved_order_type == "tiepi" and workbook_needs_cached_prices(wb)
        else None
    )
    stats = TransformStats(order_type=order_context.resolved_order_type)

    for ws in wb.worksheets:
        detail_context = build_detail_sheet_context(wb, cached_wb, ws)
        if detail_context:
            process_detail_sheet(detail_context, order_context, stats)

    enable_excel_recalculation(wb)
    stats.moved_hardware_sheets = move_hardware_sheets_to_end(wb)
    stats.updated_order_numbers = normalize_order_numbers(wb, order_context.inferred_order_number)

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return stats.to_dict()
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()


def convert_xls_to_xlsx(input_path: Path, output_path: Path) -> None:
    """Convert legacy .xls to .xlsx while preserving workbook formatting.

    Docker/Linux cannot use Excel COM. LibreOffice headless keeps styles,
    merged cells, column widths, images and sheet layout far better than a
    pandas data-frame rewrite. If LibreOffice is unavailable or fails, raise a
    clear error instead of silently producing a data-only workbook.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="xls_convert_") as tmp:
        tmpdir = Path(tmp)
        cmd = soffice_base_command() + [
            "--convert-to",
            "xlsx",
            "--outdir",
            str(tmpdir),
            str(input_path),
        ]

        with _SOFFICE_LOCK:
            warm_xls_converter()
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=SOFFICE_TIMEOUT_SECONDS)
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(f"LibreOffice failed to convert .xls to .xlsx: {detail}")

        converted = tmpdir / f"{input_path.stem}.xlsx"
        if not converted.exists():
            candidates = list(tmpdir.glob("*.xlsx"))
            if len(candidates) == 1:
                converted = candidates[0]
            else:
                detail = (result.stderr or result.stdout or "").strip()
                raise RuntimeError(f"LibreOffice did not create an .xlsx file: {detail}")

        output_path.write_bytes(converted.read_bytes())


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}\u4e0b\u8f66\u95f4.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert original order workbook to workshop order workbook.")
    parser.add_argument("input", type=Path, help="Original .xlsx workbook")
    parser.add_argument("output", type=Path, nargs="?", help="Output .xlsx workbook")
    parser.add_argument("--discount", type=float, default=DEFAULT_DISCOUNT, help="Unit-price multiplier, default: 0.85")
    parser.add_argument(
        "--order-type",
        choices=sorted(ORDER_TYPES),
        default="auto",
        help="Order rule set: auto detects hunyou only when a color-cell value starts with '混油'; tiepi uses full rules; hunyou only clears packaging and color fees.",
    )
    args = parser.parse_args()

    input_path = args.input.expanduser().resolve()
    output_path = (args.output or default_output_path(input_path)).expanduser().resolve()

    if input_path == output_path:
        raise SystemExit("Output path must be different from input path.")
    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    stats = transform(input_path, output_path, args.discount, args.order_type)
    print(f"Saved: {output_path}")
    print(
        "Stats: "
        f"order_type={stats['order_type']}, "
        f"discounted_prices={stats['discounted_prices']}, "
        f"cleared_wood_boxes={stats['cleared_wood_boxes']}, "
        f"cleared_non_standard_color_fees={stats['cleared_non_standard_color_fees']}, "
        f"summary_constants_removed={stats['summary_constants_removed']}, "
        f"updated_page_notes={stats['updated_page_notes']}, "
        f"cleared_bottom_notes={stats['cleared_bottom_notes']}, "
        f"filled_workpoints={stats['filled_workpoints']}, "
        f"skipped_prices={stats['skipped_prices']}, "
        f"deleted_blank_rows={stats['deleted_blank_rows']}, "
        f"moved_hardware_sheets={stats['moved_hardware_sheets']}, "
        f"updated_order_numbers={stats['updated_order_numbers']}"
    )


if __name__ == "__main__":
    main()
