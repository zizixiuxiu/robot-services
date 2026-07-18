import sys
import tempfile
import unittest
from base64 import b64decode
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.insert(0, str(SRC_DIR))

from make_workshop_order import (  # noqa: E402
    clean_payment_amount_text,
    find_existing_order_number,
    fit_long_note_rows,
    infer_order_number,
    is_hardware_sheet_name,
    is_hunyou_color,
    remove_payment_amount_info,
    transform,
)


PNG_1X1 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMB"
    "/6X7x2sAAAAASUVORK5CYII="
)


class WorkshopOrderRuleTests(unittest.TestCase):
    def test_hunyou_color_allows_leading_brackets(self) -> None:
        self.assertTrue(is_hunyou_color("（混油）Y05-S月影灰"))
        self.assertTrue(is_hunyou_color("(混油) Y05-S月影灰"))
        self.assertFalse(is_hunyou_color("天然木皮：白橡"))

    def test_note_cleanup_keeps_after_sales_sentence_and_strips_price(self) -> None:
        source = "玻璃运输易损坏，不进入公司售后！附色卡有色差，具体以出厂实物为准，不进入公司售后！92折优惠：972元"
        self.assertEqual(
            remove_payment_amount_info(source),
            "玻璃运输易损坏，不进入公司售后！附色卡有色差，具体以出厂实物为准，不进入公司售后！",
        )

    def test_payment_amount_cleanup_strips_discount_segment_after_after_sales(self) -> None:
        source = (
            "此单黄杨原木修红樱桃颜色纹理不同修色有色差不进入公司售后！工艺不同，以实物为准,不进入公司售后！"
            "               此单王董同意板材95折优惠，折后31197元，优惠1642元！"
        )
        self.assertEqual(
            remove_payment_amount_info(source),
            "此单黄杨原木修红樱桃颜色纹理不同修色有色差不进入公司售后！工艺不同，以实物为准,不进入公司售后！",
        )

    def test_payment_amount_cleanup_keeps_after_sales_before_discount_approval(self) -> None:
        prefix = "此单工艺不同，以实物为准，不进入公司售后！"
        examples = (
            prefix + "此单李总同意板材95折优惠，折后31197元，优惠1642元！",
            prefix + "此单张总确认8折，优惠100元！",
            prefix + "此单赵董同意折后31197元！",
        )
        for source in examples:
            self.assertEqual(remove_payment_amount_info(source), prefix)

    def test_payment_amount_cleanup_clears_price_only_fragments(self) -> None:
        self.assertIsNone(clean_payment_amount_text("银镜320元/平").value)
        self.assertIsNone(clean_payment_amount_text("样品打8折、折后：119810元，优惠：29953元").value)

    def test_payment_amount_cleanup_keeps_page_count(self) -> None:
        self.assertEqual(remove_payment_amount_info("此单共2页，92折优惠：972元"), "此单共2页")

    def test_infer_order_number_from_filename(self) -> None:
        path = Path("S2606-4106阮国政（西安越秀铁建樽越）.xlsx")
        self.assertEqual(infer_order_number(path), "S2606-4106")

    def test_existing_header_order_number_wins_over_filename(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "1"
        ws["P2"] = "订单编号"
        ws["Q2"] = "S2606-6111-2"
        ws["R2"] = -1

        self.assertEqual(find_existing_order_number(wb), "S2606-6111-2")

    def test_hardware_sheet_name(self) -> None:
        self.assertTrue(is_hardware_sheet_name("五金-1"))
        self.assertTrue(is_hardware_sheet_name("订单五金汇总"))
        self.assertFalse(is_hardware_sheet_name("1"))

    def test_transform_preserves_embedded_images(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            image_path = tmpdir / "image.png"
            image_path.write_bytes(b64decode(PNG_1X1))

            source_path = tmpdir / "S2606-0001-image.xlsx"
            output_path = tmpdir / "S2606-0001-image下车间.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "1"
            ws["A1"] = "测试"
            ws.add_image(Image(str(image_path)), "S8")
            wb.save(source_path)

            transform(source_path, output_path, 0.85, "auto")

            with ZipFile(output_path) as zf:
                media = [name for name in zf.namelist() if name.startswith("xl/media/") and not name.endswith("/")]
            self.assertEqual(len(media), 1)

    def test_transform_preserves_existing_header_order_number(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            source_path = tmpdir / "S2606-6111-sample.xlsx"
            output_path = tmpdir / "S2606-6111-sample下车间.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "1"
            ws["P2"] = "订单编号"
            ws["Q2"] = "S2606-6111-2"
            ws["R2"] = -1
            wb.save(source_path)

            transform(source_path, output_path, 0.85, "auto")

            out = load_workbook(output_path)
            self.assertEqual(out["1"]["Q2"].value, "S2606-6111-2")
            self.assertEqual(out["1"]["R2"].value, -1)

    def test_transform_does_not_create_missing_order_suffix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            source_path = tmpdir / "S2607-7506-sample.xlsx"
            output_path = tmpdir / "S2607-7506-sample下车间.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "1"
            ws["P2"] = "订单编号"
            ws["Q2"] = "S2607-7506"
            ws["R2"] = None
            wb.save(source_path)

            transform(source_path, output_path, 0.85, "auto")

            out = load_workbook(output_path)
            self.assertEqual(out["1"]["Q2"].value, "S2607-7506")
            self.assertIsNone(out["1"]["R2"].value)

    def test_fit_long_note_rows_replaces_space_padding_with_line_break(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("D25:R25")
        ws["D25"] = "此单黄杨原木修红樱桃颜色纹理不同修色有色差不进入公司售后！        此单王董同意板材"
        ws.row_dimensions[25].height = 20

        fit_long_note_rows(ws)

        self.assertIn("\n", ws["D25"].value)
        self.assertTrue(ws["D25"].alignment.wrap_text)
        self.assertGreater(ws.row_dimensions[25].height, 20)

    def test_fit_long_note_rows_skips_tall_multi_row_merged_note(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A15:C18")
        ws["A15"] = "图文说明:请车间注意在外包装上写上，内有玻璃，勿压!!!!!"
        for row_idx in range(15, 19):
            ws.row_dimensions[row_idx].height = 25

        fit_long_note_rows(ws)

        # 合并区域 4 行合计 100pt 已足够，首行不应被加高
        self.assertEqual(ws.row_dimensions[15].height, 25)

    def test_fit_long_note_rows_tops_up_short_multi_row_merged_note(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A15:C16")
        ws["A15"] = "图文说明:" + "请车间注意在外包装上写上，内有玻璃，勿压!!!!!" * 5
        for row_idx in range(15, 17):
            ws.row_dimensions[row_idx].height = 25

        fit_long_note_rows(ws)

        # 合并区域合计高度不够时，差值补到首行，而不是整段高度
        self.assertGreater(ws.row_dimensions[15].height, 25)
        self.assertLessEqual(ws.row_dimensions[15].height + 25, 180)


if __name__ == "__main__":
    unittest.main()
