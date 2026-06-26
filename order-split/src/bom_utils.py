"""BOM MCP 服务公共工具函数"""
import os
import re
import shutil
import tempfile
import base64
from pathlib import Path
from contextlib import contextmanager


def detect_file_type(input_path: str) -> str:
    """根据 sheet 名判断文件类型：'order_split' | 'hardware' | 'unknown'"""
    try:
        import xlrd
        book = xlrd.open_workbook(input_path)
        names = book.sheet_names()
        if any(s in names for s in ['实木附件', '实木柜门']):
            return 'order_split'
        if any('五' in s for s in names):
            return 'hardware'
    except Exception:
        pass
    try:
        import openpyxl
        wb = openpyxl.load_workbook(input_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        if any(s in names for s in ['实木附件', '实木柜门', '柜体']):
            return 'order_split'
        if any('五' in s for s in names):
            return 'hardware'
    except Exception:
        pass
    return 'unknown'


def xlsx_to_xls(src: str, dst: str):
    """xlsx 转 xls（只复制值）"""
    import openpyxl, xlwt
    wb_in = openpyxl.load_workbook(src)
    wb_out = xlwt.Workbook()
    for sn in wb_in.sheetnames:
        ws_in = wb_in[sn]
        ws_out = wb_out.add_sheet(sn[:31])
        for ri, row in enumerate(ws_in.iter_rows()):
            for ci, cell in enumerate(row):
                if cell.value is not None:
                    ws_out.write(ri, ci, cell.value)
    wb_out.save(dst)


def get_clean_filename(filename: str) -> str:
    """去掉缓存前缀如 doc_xxx_ 和浏览器重复下载后缀如 (1)，返回原始业务文件名"""
    cleaned = re.sub(r'^doc_[a-f0-9]{12}_', '', filename)
    # 去掉浏览器重复下载后缀：(1), (2) 等，如 "S2604-4090 张姝媛(1).xls"
    cleaned = re.sub(r'\(\d+\)(?=\.|$)', '', cleaned)
    if cleaned.endswith('.xls') or cleaned.endswith('.xlsx'):
        cleaned = Path(cleaned).stem
    return cleaned


@contextmanager
def suppress_output():
    """屏蔽 stdout/stderr，防止 print 污染 JSON-RPC 流"""
    old_stdout = os.dup(1)
    old_stderr = os.dup(2)
    with open(os.devnull, 'w') as devnull:
        os.dup2(devnull.fileno(), 1)
        os.dup2(devnull.fileno(), 2)
    try:
        yield
    finally:
        os.dup2(old_stdout, 1)
        os.dup2(old_stderr, 2)
        os.close(old_stdout)
        os.close(old_stderr)


def detect_unusual_sheets(input_path: str) -> list[str]:
    """检测 Excel 中是否存在非标准 sheet，返回 unusual sheet 名称列表。

    标准 sheet 规则：
      - 纯数字名称（如 1、2、14）视为普通表
      - 包含"五"视为五金表
      - 以"外购表"开头视为外购表
      - "特殊产品表"和"汇总"视为已知表
      - 隐藏 sheet 跳过
    """
    names = []
    hidden = set()

    # 优先尝试 xlrd（.xls）
    try:
        import xlrd
        book = xlrd.open_workbook(input_path, formatting_info=True)
        names = book.sheet_names()
        if hasattr(book, 'sheet_visibility'):
            for i, vis in enumerate(book.sheet_visibility):
                if vis != 0:
                    hidden.add(names[i])
    except Exception:
        pass

    # xlrd 失败再尝试 openpyxl（.xlsx）
    if not names:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path, read_only=True)
            names = wb.sheetnames
            for name in names:
                ws = wb[name]
                if hasattr(ws, 'sheet_state') and ws.sheet_state != 'visible':
                    hidden.add(name)
            wb.close()
        except Exception:
            pass

    unusual = []
    for name in names:
        if name in hidden:
            continue
        if name.isdigit():
            continue
        if '五' in name:
            continue
        if name.startswith('外购表'):
            continue
        if name in ('特殊产品表', '汇总', '实木'):
            continue
        unusual.append(name)

    return unusual
