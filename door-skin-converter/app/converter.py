from __future__ import annotations

import csv
import io
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence


OUTPUT_HEADERS = [
    "订单号",
    "板件名称",
    "加工长度",
    "加工宽度",
    "数量",
    "材料描述",
    "纹理",
    "完工长度",
    "完工宽度",
    "打孔工艺",
    "特殊工艺",
    "正面条码",
    "开槽工艺",
    "客户",
    "柜体名称",
    "分柜号",
    "品牌",
    "厚度",
    "面积",
    "批次号",
    "工艺路线",
    "材料描述2",
    "物料编码",
    "异型",
    "类型",
    "封边类型",
]

REQUIRED_SOURCE_HEADERS = [
    "材料",
    "开料长",
    "开料宽",
    "数量",
    "厚度",
    "工件名称",
    "颜色",
    "订单编号",
    "客户地址",
    "工艺",
    "特殊要求",
    "品牌",
]

THICKNESS_RE = re.compile(
    r"(\d+(?:\.\d+)?(?:\s*\+\s*\d+(?:\.\d+)?)*)\s*mm",
    re.IGNORECASE,
)

# 《PVC门型分类表》「门型汇总」sheet 中标黄“剔除”的门型编号，
# 这些门型的订单行不进入生成文件。分类表更新后需同步修改此处并重新发版。
EXCLUDED_DOOR_MODELS = (
    "2091", "7066",                                    # 玻璃门（美心）
    "N1049", "N1050", "N1051", "N1052", "N1053",
    "N1054", "N1055", "N1056", "N1057", "N1058", "N1059",  # 装板门+线条
    "YM-062", "YM-081", "XP-62",                       # 装板门（固豪）
    "YM-114", "YM-105", "YM-106", "YM-115",            # 玻璃门（固豪）
    "YSM-079", "YSM-080", "YSM-081",
    "YM-082",
)

# 工件名称中包含以下关键词的订单行同样剔除（不进入生成文件）
EXCLUDED_NAME_KEYWORDS = ("铝框门",)

# 运行时剔除清单：群里 @机器人 加的型号/关键词持久化在这个 JSON，
# 与上面的内置清单合并生效，容器重建不丢失（data 目录已挂载到宿主机）。
RUNTIME_EXCLUSIONS_FILE = Path(
    os.getenv("RUNTIME_EXCLUSIONS_FILE", "/app/data/runtime_exclusions.json")
)


def load_runtime_exclusions() -> dict:
    """读取运行时剔除清单（不存在/损坏时返回空清单）。"""
    try:
        data = json.loads(RUNTIME_EXCLUSIONS_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {"models": [], "keywords": []}
    return {
        "models": [str(m) for m in data.get("models", []) if str(m).strip()],
        "keywords": [str(k) for k in data.get("keywords", []) if str(k).strip()],
    }


def save_runtime_exclusions(exclusions: dict) -> None:
    RUNTIME_EXCLUSIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    RUNTIME_EXCLUSIONS_FILE.write_text(
        json.dumps(exclusions, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def get_excluded_models() -> list[str]:
    """内置型号 + 运行时型号（去重，保序）。"""
    merged = list(EXCLUDED_DOOR_MODELS)
    for model in load_runtime_exclusions()["models"]:
        if _normalize_model(model) not in {_normalize_model(m) for m in merged}:
            merged.append(model)
    return merged


def get_excluded_keywords() -> list[str]:
    """内置关键词 + 运行时关键词（去重，保序）。"""
    merged = list(EXCLUDED_NAME_KEYWORDS)
    for keyword in load_runtime_exclusions()["keywords"]:
        if keyword not in merged:
            merged.append(keyword)
    return merged


# 比较时只看字母+数字内容：忽略大小写、全半角、空格和“-”等分隔符
#（源表可能写作 YSM079 / ysm-079 / ＹＳＭ－０７９，均视为同一型号）
def _normalize_model(text: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", unicodedata.normalize("NFKC", str(text)).upper())


DOOR_MODEL_RE = re.compile(r"[A-Z]{1,4}[^A-Z0-9]?\d{2,4}|\d{4}")

# 工艺列命中这些模式时，材料描述里的“素板”换成对应板材名（如 8mm素板 → 8mm多层加密）
# 关键词 + 模糊匹配：
#   多层.?加密 / 加密.?多层 —— 多层加密 / 多层板加密 / 加密多层 等变体都认，输出统一为“多层加密”
#   黑[炭碳]晶 —— 黑炭晶 / 黑碳晶 都认，输出名为 None 时按源表实际写法输出
BOARD_MATERIAL_PATTERNS = (
    (re.compile(r"多层.?加密|加密.?多层"), "多层加密"),
    (re.compile(r"黑[炭碳]晶"), None),
)


def extract_board_material(process_text: Any) -> str:
    """从工艺文本识别板材类型，默认素板。"""
    text = clean_text(process_text)
    for pattern, canonical in BOARD_MATERIAL_PATTERNS:
        match = pattern.search(text)
        if match:
            return canonical or match.group(0)
    return "素板"

# 已见过的工艺描述（归一化：去厚度、忽略大小写/全半角/空格）。
# 不在此表的工艺会在飞书卡片里提醒人工检查，材料描述仍按现有规则（素板/板材词）生成。
KNOWN_PROCESS_DESCRIPTIONS = frozenset({
    "美心YSM",
    "复合贴YSM皮",
    "复合贴YSM皮多层加密",
    "美心YSM装门门框",
    "美心YSM装板",
    "PVC系列",
    "PP系列",
    "复合贴YSM皮装板",
    "复合贴YSM皮工艺",
    "复合贴YSM皮工艺多层加密",
    "复合贴PM皮",
    "复合贴PM皮外凸",
    "黑碳晶贴YSM皮",
    "黑炭晶贴YSM皮",
    "多层加密",
    "黑碳晶",
    "黑炭晶",
    "PVC系列多层板加密",
})


def normalize_process_description(process_text: Any) -> str:
    """工艺描述归一化：去掉厚度（如 8MM / 5+8MM），忽略大小写、全半角和空格。"""
    text = unicodedata.normalize("NFKC", clean_text(process_text)).upper()
    text = re.sub(r"\d+(?:\.\d+)?(?:\s*\+\s*\d+(?:\.\d+)?)*\s*MM", "", text)
    return re.sub(r"\s+", "", text)


class ConversionError(ValueError):
    """输入表格不符合转换要求。"""


@dataclass(frozen=True)
class ConversionStats:
    source_rows: int
    output_rows: int
    input_quantity_sum: float
    quantity_sum: float
    split_source_rows: int
    normal_output_rows: int
    split_output_rows: int
    blank_thickness_orders: tuple[str, ...]
    split_details: tuple[str, ...] = ()
    excluded_details: tuple[str, ...] = ()
    material_mismatch_details: tuple[str, ...] = ()
    new_process_details: tuple[str, ...] = ()


SplitDetail = dict[str, Any]
def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_number(value: Any) -> int | float | str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = clean_text(value)
        try:
            number = float(text)
        except ValueError:
            return text
    return int(number) if number.is_integer() else number


def extract_door_model(workpiece: Any) -> str:
    """从工件名称中提取门型编号并规范化（如 'N1046(N1046)单扇' -> 'N1046'，
    'ｙｓｍ－０７９单扇' -> 'YSM079'），只保留字母和数字便于比较。"""
    text = unicodedata.normalize("NFKC", clean_text(workpiece)).upper()
    match = DOOR_MODEL_RE.search(text)
    return _normalize_model(match.group(0)) if match else ""


def extract_thicknesses(process_text: Any) -> list[str]:
    match = THICKNESS_RE.search(clean_text(process_text))
    if not match:
        return []
    return [normalize_decimal_text(part) for part in match.group(1).split("+")]


def normalize_decimal_text(value: Any) -> str:
    text = clean_text(value)
    try:
        number = float(text)
    except ValueError:
        return text
    return str(int(number)) if number.is_integer() else format(number, "g")


def multiply_quantity(value: Any) -> int | float:
    try:
        number = float(clean_text(value))
    except ValueError as exc:
        raise ConversionError(f"数量不是有效数字：{value!r}") from exc
    result = number * 2
    return int(result) if result.is_integer() else result


def build_special_process(row: dict[str, Any]) -> str:
    material = clean_text(row["材料"])
    door_thickness = re.sub(
        r"\s*mm\s*$", "", clean_text(row["厚度"]), flags=re.IGNORECASE
    )
    requirement = clean_text(row["特殊要求"])
    thickness_text = f"厚度{door_thickness}mm" if door_thickness else "厚度"
    return f"{material}{thickness_text}.{requirement}"


def transform_rows(
    headers: Sequence[Any], rows: Iterable[Sequence[Any]]
) -> tuple[list[list[Any]], ConversionStats]:
    normalized_headers = [clean_text(value) for value in headers]
    if len(set(normalized_headers)) != len(normalized_headers):
        raise ConversionError("源表表头存在重复字段，无法确定列映射")

    missing = [name for name in REQUIRED_SOURCE_HEADERS if name not in normalized_headers]
    if missing:
        raise ConversionError(f"源表缺少字段：{', '.join(missing)}")

    source_indexes = {name: normalized_headers.index(name) for name in REQUIRED_SOURCE_HEADERS}
    output_rows: list[list[Any]] = []
    source_row_count = 0
    split_source_rows = 0
    input_quantity_sum = 0.0
    blank_thickness_orders: list[str] = []
    split_details: list[str] = []
    excluded_details: list[str] = []
    material_mismatch_details: list[str] = []
    new_process_details: list[str] = []
    excluded_keywords = get_excluded_keywords()
    excluded_normalized = {_normalize_model(m) for m in get_excluded_models()}

    for source_row in rows:
        padded = list(source_row) + [""] * max(0, len(normalized_headers) - len(source_row))
        if all(clean_text(value) == "" for value in padded):
            continue

        row = {name: padded[index] for name, index in source_indexes.items()}
        order_number = clean_text(row["订单编号"])
        workpiece = clean_text(row["工件名称"])
        if any(keyword in workpiece for keyword in excluded_keywords):
            excluded_details.append(f"订单 {order_number}：{workpiece}（关键词剔除）")
            continue
        door_model = extract_door_model(workpiece)
        if door_model in excluded_normalized:
            excluded_details.append(f"订单 {order_number}：{workpiece}")
            continue

        source_row_count += 1
        try:
            raw_quantity = float(clean_text(row["数量"]))
        except ValueError as exc:
            raise ConversionError(f"订单 {order_number or '(空)'} 的数量不是有效数字：{row['数量']!r}") from exc
        input_quantity_sum += raw_quantity
        try:
            quantity = multiply_quantity(row["数量"])
        except ConversionError as exc:
            raise ConversionError(f"订单 {order_number or '(空)'} 的{exc}") from exc

        thicknesses = extract_thicknesses(row["工艺"])
        if len(thicknesses) > 1:
            split_source_rows += 1
            raw_match = ""
            m = THICKNESS_RE.search(clean_text(row["工艺"]))
            if m:
                raw_match = m.group(0)
            split_details.append(
                f"订单 {order_number}：{raw_match or clean_text(row['厚度'])} → "
                f"{', '.join(f'{t}mm' for t in thicknesses)}"
            )
        if not thicknesses:
            blank_thickness_orders.append(
                f"订单 {order_number}：{clean_text(row['工件名称'])}"
            )
            thicknesses = [""]

        color = clean_text(row["颜色"])
        process_text = clean_text(row["工艺"])
        if process_text and normalize_process_description(process_text) not in KNOWN_PROCESS_DESCRIPTIONS:
            new_process_details.append(
                f"订单 {order_number}：{workpiece}（工艺「{process_text}」未见过，请检查）"
            )
        board_material = extract_board_material(row["工艺"])
        for final_thickness in thicknesses:
            material_description_1 = (
                f"{final_thickness}mm{board_material}" if final_thickness else color
            )
            # 一致性校验：工艺列声明的板材/厚度必须体现在材料描述里
            mismatches = []
            if board_material != "素板" and board_material not in material_description_1:
                mismatches.append(f"工艺声明「{board_material}」但材料描述未体现")
            if final_thickness and f"{final_thickness}mm" not in material_description_1:
                mismatches.append(f"工艺厚度 {final_thickness}mm 与材料描述不符")
            for pattern, _ in BOARD_MATERIAL_PATTERNS:
                hit = pattern.search(material_description_1)
                if hit and hit.group(0) != board_material:
                    mismatches.append(f"材料描述含「{hit.group(0)}」但工艺未声明")
            if mismatches:
                material_mismatch_details.append(
                    f"订单 {order_number}：{workpiece}（工艺「{process_text}」↔ 材料描述「{material_description_1}」，{'；'.join(mismatches)}）"
                )
            material_description_2 = (
                f"{final_thickness}mm{color}" if final_thickness else color
            )
            output_rows.append(
                [
                    order_number,
                    clean_text(row["工件名称"]),
                    clean_number(row["开料长"]),
                    clean_number(row["开料宽"]),
                    quantity,
                    material_description_1,
                    1,
                    "",
                    "",
                    clean_text(row["工艺"]),  # 打孔工艺列 ← 源表“工艺”列
                    build_special_process(row),
                    "",
                    "",
                    clean_text(row["客户地址"]),
                    "",
                    "",
                    clean_text(row["品牌"]),
                    clean_number(final_thickness),
                    "",
                    "",
                    "",
                    material_description_2,
                    "",
                    "",
                    "",
                    "",
                ]
            )

    normal_source_rows = source_row_count - split_source_rows
    normal_output_rows = normal_source_rows
    split_output_rows = len(output_rows) - normal_output_rows

    stats = ConversionStats(
        source_rows=source_row_count,
        output_rows=len(output_rows),
        input_quantity_sum=input_quantity_sum,
        quantity_sum=sum(float(row[4]) for row in output_rows),
        split_source_rows=split_source_rows,
        normal_output_rows=normal_output_rows,
        split_output_rows=split_output_rows,
        blank_thickness_orders=tuple(blank_thickness_orders),
        split_details=tuple(split_details),
        excluded_details=tuple(excluded_details),
        material_mismatch_details=tuple(material_mismatch_details),
        new_process_details=tuple(new_process_details),
    )
    return output_rows, stats


def read_excel(content: bytes, filename: str) -> tuple[list[Any], list[list[Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("缺少 xlrd 依赖，无法读取 .xls") from exc
        try:
            workbook = xlrd.open_workbook(file_contents=content)
            sheet = workbook.sheet_by_index(0)
            values = [sheet.row_values(index) for index in range(sheet.nrows)]
        except Exception as exc:
            raise ConversionError(f"无法读取 .xls 文件：{exc}") from exc
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法读取 .xlsx") from exc
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.worksheets[0]
            values = [list(row) for row in sheet.iter_rows(values_only=True)]
        except Exception as exc:
            raise ConversionError(f"无法读取 .xlsx 文件：{exc}") from exc
    else:
        raise ConversionError("只支持 .xls 或 .xlsx 文件")

    first_nonempty = next(
        (index for index, row in enumerate(values) if any(clean_text(value) for value in row)),
        None,
    )
    if first_nonempty is None:
        raise ConversionError("源表为空")
    return values[first_nonempty], values[first_nonempty + 1 :]


def write_csv(rows: Sequence[Sequence[Any]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(OUTPUT_HEADERS)
    writer.writerows(rows)
    return buffer.getvalue().encode("gb18030")


def convert_excel_to_csv(
    content: bytes, filename: str
) -> tuple[bytes, ConversionStats]:
    headers, source_rows = read_excel(content, filename)
    output_rows, stats = transform_rows(headers, source_rows)
    return write_csv(output_rows), stats
