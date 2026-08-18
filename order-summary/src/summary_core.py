#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PVC 订单汇总核心逻辑：
- 读取订单明细 .xls/.xlsx（第 1 行为标题，第 2 行为表头）
- 按「回传单号 + 类型」分组
- 数量（明细表第 70 列，即 宽/厚、厚 后的“数量”列）求和；主数量为 0 时回退“单线条”列
- 产品列去重后用 '/' 连接；生产单号同样合并，作为「订单明细」表 E 列隐藏保留
- 隐藏工作表「系统导入数据」（按回传单号一单一行，仅含门扇/门套窗套计划数
  至少一个大于 0 的单号）：生产单号列表、终端客户、门扇计划数（全行累计 单扇+门扇）、
  门套窗套计划数（全行累计 门套+窗套，含类型=门行自带的门套）、门扇/门套窗套产品摘要、
  财务下单日期（取“办公室”列）、数据生成时间
- 输出带格式的 .xlsx（可见工作表“订单明细”，表头加粗、冻结首行、边框、自动筛选，
  按回传单号交替灰白底色，E 列隐藏；“系统导入数据”工作表整体隐藏）
"""
from pathlib import Path
from datetime import datetime, timezone, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 数据生成时间固定用东八区（容器内默认 UTC，需显式 +8）
CN_TZ = timezone(timedelta(hours=8))

# 数量列在原始明细表中的列索引（0 基）：门洞尺寸、宽/厚、厚 之后的“数量”
QTY_COL_INDEX = 69

# 回退数量列：类型为「线条」的行不填主数量列，数量记在“单线条”列
QTY_FALLBACK_COL_INDEX = 60

# 隐藏统计列取数位置（0 基）：门扇 / 门套 / 单扇 / 窗套
DOOR_COL_INDEX = 53      # 门扇
TAOK_COL_INDEX = 54      # 门套
SINGLE_DOOR_COL_INDEX = 55  # 单扇
WINDOW_TAO_COL_INDEX = 56   # 窗套

REQUIRED_COLUMNS = ["回传单号", "类型", "产品", "生产单号"]

# 隐藏工作表「系统导入数据」列（按回传单号一单一行）
IMPORT_SHEET_NAME = "系统导入数据"
IMPORT_STAT_COLUMNS = [
    "生产单号列表", "终端客户", "门扇计划数", "门套窗套计划数",
    "门扇产品摘要", "门套窗套产品摘要", "财务下单日期", "数据生成时间",
]

GRAY_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")


def _num(series) -> "pd.Series":
    return pd.to_numeric(series, errors="coerce").fillna(0)


def _unique_join(series, sep: str) -> str:
    """去重后连接（忽略空值，保持出现顺序）"""
    vals = []
    for v in series:
        if pd.isna(v):
            continue
        v = str(v).strip()
        if v and v not in vals:
            vals.append(v)
    return sep.join(vals)


def _build_order_stats(df: "pd.DataFrame") -> dict:
    """按回传单号统计车间扫码系统需要的数据。

    返回 {回传单号: {...}}，仅包含门扇计划数或门套窗套计划数至少一个大于 0 的单号。
    """
    door = _num(df[df.columns[DOOR_COL_INDEX]])          # 门扇
    tao = _num(df[df.columns[TAOK_COL_INDEX]])           # 门套
    single = _num(df[df.columns[SINGLE_DOOR_COL_INDEX]])  # 单扇
    window = _num(df[df.columns[WINDOW_TAO_COL_INDEX]])   # 窗套

    door_qty = single + door          # 每行门扇计划数量 = 单扇 + 门扇
    taoc_qty = tao + window           # 每行门套窗套数量 = 门套 + 窗套

    order_date_raw = df["办公室"] if "办公室" in df.columns else pd.Series([None] * len(df))

    stats = {}
    for order_no, idx in df.groupby("回传单号", sort=True, dropna=True).groups.items():
        idx = list(idx)
        # 门扇计划数 = 全部行的 单扇+门扇（门扇只出现在类型=门的行）
        door_plan = float(door_qty[idx].sum())
        # 门套窗套计划数 = 全部行的 门套+窗套（类型=门的行通常也带门套，不能漏）
        taoc_plan = float(taoc_qty[idx].sum())

        if door_plan <= 0 and taoc_plan <= 0:
            continue  # 两者都为 0 的订单不写入

        door_products = _unique_join(df.loc[idx, "产品"][door_qty[idx] > 0], " / ")
        taoc_products = _unique_join(df.loc[idx, "产品"][taoc_qty[idx] > 0], " / ")

        order_date = ""
        for v in order_date_raw[idx]:
            if pd.notna(v) and str(v).strip():
                try:
                    order_date = pd.to_datetime(v).strftime("%Y-%m-%d")
                except Exception:
                    order_date = str(v).strip()
                break

        stats[order_no] = {
            "生产单号列表": _unique_join(df.loc[idx, "生产单号"], "；"),
            "终端客户": _unique_join(df.loc[idx, "终端客户"], "；") if "终端客户" in df.columns else "",
            "门扇计划数": int(door_plan),
            "门套窗套计划数": int(taoc_plan),
            "门扇产品摘要": door_products,
            "门套窗套产品摘要": taoc_products,
            "财务下单日期": order_date,
        }
    return stats


def _merge_values(series) -> str:
    """组内去重后用 '/' 连接（忽略空值，保持出现顺序）"""
    vals = []
    for v in series:
        if pd.isna(v):
            continue
        v = str(v).strip()
        if v and v not in vals:
            vals.append(v)
    return "/".join(vals)


def _apply_format(xlsx_path: Path) -> None:
    """格式优化：
    - 「订单明细」表：表头加粗居中、冻结首行、细边框、列宽、自动筛选；
      数据区按回传单号（A 列）交替填充灰/白底色；E 列（生产单号）隐藏
    - 「系统导入数据」表：表头加粗、列宽，整个工作表隐藏"""
    wb = load_workbook(xlsx_path)
    ws = wb["订单明细"]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    gray = False
    prev_key = object()
    for row in ws.iter_rows(min_row=2):
        key = row[0].value  # A 列：回传单号
        if key != prev_key:
            gray = not gray
            prev_key = key
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = border
            if gray:
                cell.fill = GRAY_FILL
            if cell.column_letter == "D":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, w in {"A": 14, "B": 10, "C": 8, "D": 50, "E": 22}.items():
        ws.column_dimensions[col].width = w

    # E 列「生产单号」隐藏保留，需要时可取消隐藏查看
    ws.column_dimensions["E"].hidden = True

    ws.auto_filter.ref = ws.dimensions

    # 「系统导入数据」工作表：简单表头格式 + 列宽，整体隐藏
    if IMPORT_SHEET_NAME in wb.sheetnames:
        imp = wb[IMPORT_SHEET_NAME]
        for cell in imp[1]:
            cell.font = Font(name="微软雅黑", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        imp.row_dimensions[1].height = 22
        for row in imp.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col, w in {"A": 14, "B": 24, "C": 20, "D": 12, "E": 14,
                       "F": 30, "G": 30, "H": 14, "I": 20}.items():
            imp.column_dimensions[col].width = w
        imp.sheet_state = "hidden"

    wb.save(xlsx_path)


def summarize_order(input_path: str, output_dir: str) -> dict:
    """处理单个订单明细文件，返回 {output_path, groups, quantity_total}"""
    df = pd.read_excel(input_path, header=1)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"文件缺少必需列: {', '.join(missing)}")
    if len(df.columns) <= QTY_COL_INDEX:
        raise ValueError(f"文件列数不足（{len(df.columns)} 列），无法定位数量列（第 {QTY_COL_INDEX + 1} 列）")

    qty_col = df.columns[QTY_COL_INDEX]

    sub = df[REQUIRED_COLUMNS].copy()
    qty = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    # 主数量列为 0 的行（如「线条」类型）回退取“单线条”列的数量
    if len(df.columns) > QTY_FALLBACK_COL_INDEX:
        fallback = pd.to_numeric(df[df.columns[QTY_FALLBACK_COL_INDEX]], errors="coerce").fillna(0)
        qty = qty.where(qty != 0, fallback)
    sub["数量"] = qty

    out = (
        sub.groupby(["回传单号", "类型"], sort=True, dropna=False)
        .agg({"数量": "sum", "产品": _merge_values, "生产单号": _merge_values})
        .reset_index()
    )
    out["数量"] = out["数量"].astype(int)
    out = out[["回传单号", "类型", "数量", "产品", "生产单号"]]

    # 隐藏工作表「系统导入数据」：按回传单号一单一行
    stats = _build_order_stats(df)
    gen_time = datetime.now(CN_TZ).strftime("%Y-%m-%d %H:%M:%S")
    import_rows = []
    for order_no, vals in stats.items():
        row = {"回传单号": order_no}
        row.update(vals)
        row["数据生成时间"] = gen_time
        import_rows.append(row)
    import_df = pd.DataFrame(import_rows, columns=["回传单号"] + IMPORT_STAT_COLUMNS)

    stem = Path(input_path).stem
    output_path = Path(output_dir) / f"{stem}_按回传单号汇总.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="订单明细")
        import_df.to_excel(writer, index=False, sheet_name=IMPORT_SHEET_NAME)
    _apply_format(output_path)

    return {
        "output_path": str(output_path),
        "groups": len(out),
        "quantity_total": int(out["数量"].sum()),
    }
