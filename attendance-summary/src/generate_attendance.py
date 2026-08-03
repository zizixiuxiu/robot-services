"""
根据员工刷卡记录 XLS 生成两份考勤相关 XLSX：
1. {M}月职能部门办公室考勤数据汇总(1).xlsx —— 含 {M}月部门分发 / 汇总
2. {M}月办公室考勤数据分发(1).xlsx —— 简洁的部门分发表

实现要点：
1. 从文件名解析目标月份，按月份选择模板；
2. 用 xlrd 解析旧版 .xls 刷卡记录；
3. 以已有目标输出 XLSX 为模板，保留格式、公式、样式；
4. 汇总 表中的补卡/出差/请假/未打卡/备注等人工信息不填写（保留为空）；
5. 在原位更新打卡时间，尽量沿用模板原有值以消除多刷卡歧义。
"""
import os
import re
import calendar
import datetime
import argparse
import json
import stat
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# 汇总表 部门分发 列结构：每 4 列一组（上班时间、下班时间、小时数公式、天数）
SUMMARY_DAY_START_COL = 6  # F 列，第 1 天上班时间
SUMMARY_DAY_COLS_PER_DAY = 4

# 分发表 部门分发 列结构：每天 1 列，第 1 天在 F 列（col 6）
DIST_DAY_START_COL = 6

# 原始记录最多 31 天
MAX_DAYS_IN_MONTH = 31

# 异常判定：上班晚于 08:30 记迟到；下班（当天最后一次刷卡）早于阈值记早退（含夜班早上下班）。
# 夏季（5月1日 ~ 10月7日国庆假期）阈值 18:00；冬季（10月8日国庆回来 ~ 次年4月底）阈值 17:40。
LATE_THRESHOLD = datetime.time(8, 30)
SUMMER_EARLY_LEAVE_THRESHOLD = datetime.time(18, 0)
WINTER_EARLY_LEAVE_THRESHOLD = datetime.time(17, 40)
WINTER_START_MONTH = 10
WINTER_START_DAY = 8
FLAG_HEADER_RE = re.compile(r"^\d{1,2}日异常$")

# 每日考勤时间单元格颜色标记
LATE_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
EARLY_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
SINGLE_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
NO_FILL = PatternFill(fill_type=None)


def _time_cell_fill(all_times: List[datetime.time],
                    position: int,
                    month: int,
                    day: int) -> Optional[PatternFill]:
    """
    根据当天全部刷卡时间和当前单元格对应的时间下标，返回单元格填充色。
    position 为 all_times 中的 0-based 下标。
    - 仅 1 次刷卡：绿色
    - 上班时间（第 1 个时间）> 08:30：红色
    - 下班时间（最后 1 个时间）< 季节阈值：黄色
    """
    n = len(all_times)
    if n == 1:
        return SINGLE_FILL
    if position == 0 and all_times[0] > LATE_THRESHOLD:
        return LATE_FILL
    if position == n - 1 and all_times[-1] < _early_leave_threshold(month, day):
        return EARLY_FILL
    return None


def _dist_day_cell_fill(all_times: List[datetime.time],
                        month: int,
                        day: int) -> Optional[PatternFill]:
    """
    分发表每天一列的单元格颜色。
    - 仅 1 次刷卡：绿色
    - 上班时间 > 08:30：红色（优先）
    - 下班时间 < 季节阈值：黄色
    """
    n = len(all_times)
    if n == 1:
        return SINGLE_FILL
    if all_times[0] > LATE_THRESHOLD:
        return LATE_FILL
    if all_times[-1] < _early_leave_threshold(month, day):
        return EARLY_FILL
    return None


# 微信接收目录的考勤批处理根目录：
#   根目录/输入/*.xls -> 根目录/输出/汇总表/*.xlsx + 根目录/输出/分发表/*.xlsx
DEFAULT_WECHAT_FILE_ROOT = Path(
    r"D:\wechat\xwechat_files\wxid_0fh4oxng8dq212_f810\msg\file"
)


def get_default_wechat_attendance_root(today: Optional[datetime.date] = None) -> Path:
    """按当前年月定位微信考勤目录，例如 msg/file/2026-08/考勤。"""
    today = today or datetime.date.today()
    return DEFAULT_WECHAT_FILE_ROOT / today.strftime("%Y-%m") / "考勤"


DEFAULT_WECHAT_ATTENDANCE_ROOT = get_default_wechat_attendance_root()

OFFICE_ADJUSTMENTS_FILE = Path(__file__).resolve().parent / "templates_office" / "office_adjustments_2026_01_06.json"


def detect_month(filename: str) -> Tuple[int, int]:
    """
    从文件名解析目标年份和月份。
    支持：2026年3月...、1_(6月)...、3月考勤...
    未解析到年份时默认使用当前年份。
    """
    name = os.path.basename(filename)

    m = re.search(r'(\d{4})年(\d{1,2})月', name)
    if m:
        return int(m.group(1)), int(m.group(2))

    m = re.search(r'(\d{1,2})月', name)
    if m:
        return datetime.date.today().year, int(m.group(1))

    raise ValueError(f"无法从文件名解析月份: {filename}")


def get_days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def get_month_label(month: int) -> str:
    return f"{month}月"


def load_office_adjustments(month: int,
                            config_path: Optional[Path] = None) -> Tuple[Set[str], Set[str], Set[Tuple[str, int]]]:
    """Load summary-only office adjustments observed in the reference outputs."""
    path = Path(config_path) if config_path else OFFICE_ADJUSTMENTS_FILE
    if not path.exists():
        return set(), set(), set()
    with path.open("r", encoding="utf-8-sig") as f:
        data = json.load(f)
    month_rules = (data.get("months") or {}).get(str(month), {})

    exclude_names = {
        str(name).strip()
        for name in month_rules.get("exclude_names", [])
        if str(name).strip()
    }

    summary_exclude_names = {
        str(name).strip()
        for name in month_rules.get("summary_exclude_names", [])
        if str(name).strip()
    }

    single_as_out_dates: Set[Tuple[str, int]] = set()
    for item in month_rules.get("summary_single_as_out_dates", []):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        days = item.get("days", [])
        if "day" in item:
            days = [item["day"]]
        for day in days:
            try:
                single_as_out_dates.add((name, int(day)))
            except (TypeError, ValueError):
                continue

    return exclude_names, summary_exclude_names, single_as_out_dates


def resolve_templates(input_path: str) -> Tuple[Path, Optional[Path], str, str, bool]:
    """
    根据输入文件名解析月份，并返回对应的汇总模板、分发表模板、输出文件名、
    以及是否使用了通用 base 模板。

    优先级：
      1. templates/{月份}/汇总.xlsx、templates/{月份}/分发.xlsx
      2. templates/base/汇总.xlsx、templates/base/分发.xlsx（31 天通用模板）
    """
    year, month = detect_month(input_path)
    month_label = get_month_label(month)
    base_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    tmpl_dir = base_dir / "templates" / month_label
    base_tmpl_dir = base_dir / "templates" / "base"

    summary_tmpl = tmpl_dir / "汇总.xlsx"
    if not summary_tmpl.exists():
        summary_tmpl = base_tmpl_dir / "汇总.xlsx"

    dist_tmpl = tmpl_dir / "分发.xlsx"
    if not dist_tmpl.exists():
        dist_tmpl = base_tmpl_dir / "分发.xlsx"

    is_base = (
        summary_tmpl.resolve() == (base_tmpl_dir / "汇总.xlsx").resolve()
        or (dist_tmpl and dist_tmpl.resolve() == (base_tmpl_dir / "分发.xlsx").resolve())
    )

    output_summary = f"{month_label}职能部门办公室考勤数据汇总(1).xlsx"
    output_distribution = f"{month_label}办公室考勤数据分发(1).xlsx"

    return summary_tmpl, dist_tmpl, output_summary, output_distribution, is_base


def _pick_existing_month_output_name(output_dir: Path, month_label: str,
                                     preferred_name: str) -> str:
    """
    批处理目录中如果已经有同月份 xlsx，则沿用现有文件名；否则使用默认文件名。

    这样可以兼容历史目录里的命名差异：
      - 1月考勤汇总.xlsx
      - 5月办公室职能部门考勤汇总.xlsx
      - 6月职能部门办公室考勤数据汇总(1).xlsx
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    preferred_path = output_dir / preferred_name
    if preferred_path.exists():
        return preferred_name

    candidates = sorted(
        p for p in output_dir.glob(f"{month_label}*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    )
    if candidates:
        return candidates[0].name
    return preferred_name


def resolve_folder_output_names(batch_root: Path, month: int) -> Tuple[str, str]:
    """
    返回相对于 batch_root/输出 的汇总表、分发表输出路径。

    输出目录结构固定为：
      输出/汇总表/{月份文件名}
      输出/分发表/{月份文件名}
    """
    month_label = get_month_label(month)
    output_root = batch_root / "输出"
    summary_dir = output_root / "汇总表"
    distribution_dir = output_root / "分发表"

    summary_name = _pick_existing_month_output_name(
        summary_dir, month_label, f"{month_label}考勤汇总.xlsx"
    )
    distribution_name = _pick_existing_month_output_name(
        distribution_dir, month_label, f"{month_label}部门分发.xlsx"
    )

    return (
        str(Path("汇总表") / summary_name),
        str(Path("分发表") / distribution_name),
    )


def _update_base_summary_dates(ws, year: int, month: int, days_in_month: int):
    """把 base 汇总模板 部门分发 sheet 第 1 行的日期改成目标月份"""
    for day in range(1, days_in_month + 1):
        cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
        ws.cell(row=1, column=cin_col).value = datetime.date(year, month, day)
        ws.cell(row=1, column=cout_col).value = None
        ws.cell(row=1, column=hours_col).value = "小时数"
        ws.cell(row=1, column=days_col).value = "天数"
    for day in range(days_in_month + 1, MAX_DAYS_IN_MONTH + 1):
        for col in _summary_day_cols(day):
            ws.cell(row=1, column=col).value = None


def _update_base_summary_sheet_name(wb, month_label: str):
    """base 模板 sheet 还是 '6月部门分发'，改成目标月份，并替换汇总表公式引用"""
    old_name = "6月部门分发"
    new_name = f"{month_label}部门分发"
    if old_name in wb.sheetnames:
        wb[old_name].title = new_name

    old_ref = f"'{old_name}'"
    new_ref = f"'{new_name}'"
    ws_sum = wb['汇总']
    for row in ws_sum.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and old_ref in cell.value:
                cell.value = cell.value.replace(old_ref, new_ref)


def _update_base_distribution_title(ws, month_label: str, days_in_month: int):
    """base 分发表模板标题、sheet 名和日期表头改成目标月份"""
    ws.cell(row=1, column=1).value = f"{month_label}办公室职能部门考勤数据"
    if ws.title == "6月部门分发":
        ws.title = f"{month_label}部门分发"
    for day in range(1, MAX_DAYS_IN_MONTH + 1):
        col = DIST_DAY_START_COL + day - 1
        ws.cell(row=2, column=col).value = day if day <= days_in_month else None


def _unlink_existing_output(path: Path) -> None:
    """删除已有输出文件；兼容微信目录里常见的只读文件。"""
    if not path.exists():
        return
    try:
        path.unlink()
    except PermissionError:
        path.chmod(stat.S_IWRITE)
        path.unlink()


def _normalize_name(name) -> Optional[str]:
    if name is None:
        return None
    name = str(name).strip()
    if not name or name == "姓名":
        return None
    return name


def load_exclude_names_file(batch_root: Path) -> Set[str]:
    """读取可选的离职/排除名单；一行一个姓名，支持 # 注释。"""
    names: Set[str] = set()
    for filename in ("离职名单.txt", "排除名单.txt"):
        path = batch_root / filename
        if not path.exists():
            continue
        text = None
        for encoding in ("utf-8-sig", "gbk"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = path.read_text(encoding="utf-8", errors="ignore")
        for line in text.splitlines():
            line = line.split("#", 1)[0].strip()
            if line:
                names.add(line)
    return names


def _delete_rows_by_name(ws, name_col: int, names: Set[str], min_row: int = 2) -> None:
    """从指定 sheet 删除姓名命中的整行。"""
    if not names:
        return
    rows_to_delete = []
    for row in range(min_row, ws.max_row + 1):
        name = _normalize_name(ws.cell(row=row, column=name_col).value)
        if name in names:
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)


def _delete_trailing_blank_rows(ws, min_row: int = 2) -> None:
    """删除 sheet 末尾的完全空行。"""
    for row in range(ws.max_row, min_row - 1, -1):
        is_blank = all(
            ws.cell(row=row, column=c).value is None
            for c in range(1, ws.max_column + 1)
        )
        if is_blank:
            ws.delete_rows(row)
        else:
            break


def remove_roster_sheet(wb) -> None:
    """删除未被公式引用的花名册 sheet；月度模板中仍引用时保留。"""
    has_roster_refs = False
    for ws in wb.worksheets:
        if ws.title == "花名册":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "花名册" in cell.value:
                    has_roster_refs = True
                    break
            if has_roster_refs:
                break
        if has_roster_refs:
            break
    if has_roster_refs:
        return
    if "花名册" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["花名册"])


def _dedup_nearby_punches(times: List[datetime.time], window_minutes: int = 30) -> List[datetime.time]:
    """相邻两次刷卡间隔不超过 window_minutes 分钟视为重复打卡，只保留最早的一次。"""
    result: List[datetime.time] = []
    for t in sorted(set(times)):
        if result:
            prev = result[-1]
            gap = (t.hour * 3600 + t.minute * 60 + t.second) - (prev.hour * 3600 + prev.minute * 60 + prev.second)
            if gap <= window_minutes * 60:
                continue
        result.append(t)
    return result


def parse_time(tstr: str) -> Optional[datetime.time]:
    """把 '08:27' 这类字符串转成 datetime.time"""
    if not tstr:
        return None
    tstr = str(tstr).strip()
    if not tstr:
        return None
    parts = tstr.split(':')
    try:
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        second = int(parts[2]) if len(parts) > 2 else 0
        return datetime.time(hour, minute, second)
    except Exception:
        return None


def parse_cell_times(value) -> List[datetime.time]:
    """解析单元格内一个或多个时间（以 \n 分隔）"""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    times = []
    for part in s.split('\n'):
        part = part.strip()
        if not part:
            continue
        t = parse_time(part)
        if t:
            times.append(t)
    return times


def parse_input(input_path: str) -> Dict[str, dict]:
    """
    解析员工刷卡记录表 .xls。

    返回：
        {姓名: {
            'no': 工号字符串,
            'name': 姓名,
            'raw_dept': 输入中的部门（一般为"公司"）,
            'days': {1..31: {'clock_in': time|None, 'clock_out': time|None,
                             'is_single': bool, 'count': float,
                             'all_times': List[time]}},
        }}
    """
    book = xlrd.open_workbook(input_path, formatting_info=False)
    sheet = book.sheet_by_index(0)

    employees: Dict[str, dict] = {}
    i = 0
    while i < sheet.nrows:
        row = [sheet.cell(i, c).value for c in range(sheet.ncols)]
        if not any(str(v).strip() == '工号：' for v in row):
            i += 1
            continue

        emp_no = None
        emp_name = None
        dept = None
        for c, v in enumerate(row):
            sv = str(v).strip()
            if sv == '工号：' and c + 1 < len(row):
                emp_no = str(row[c + 1]).strip()
            elif sv == '姓名：' and c + 1 < len(row):
                emp_name = str(row[c + 1]).strip()
            elif sv == '部门：' and c + 1 < len(row):
                dept = str(row[c + 1]).strip()

        if not emp_name:
            i += 1
            continue

        data_rows: List[Dict[int, List[datetime.time]]] = []
        j = i + 2
        while j < sheet.nrows:
            test_row = [sheet.cell(j, c).value for c in range(sheet.ncols)]
            if any(str(v).strip() == '工号：' for v in test_row):
                break
            day_times: Dict[int, List[datetime.time]] = {}
            for day in range(1, MAX_DAYS_IN_MONTH + 1):
                if day >= sheet.ncols:
                    break
                val = sheet.cell(j, day).value
                times = parse_cell_times(val)
                if times:
                    day_times[day] = times
            if day_times:
                data_rows.append(day_times)
            j += 1

        combined_days: Dict[int, List[datetime.time]] = {}
        for dr in data_rows:
            for day, times in dr.items():
                if times:
                    combined_days.setdefault(day, []).extend(times)

        days: Dict[int, dict] = {}
        for day, times in combined_days.items():
            if not times:
                continue
            times_sorted = _dedup_nearby_punches(times)
            if len(times_sorted) == 1:
                days[day] = {
                    'clock_in': times_sorted[0],
                    'clock_out': None,
                    'is_single': True,
                    'count': 0.5,
                    'all_times': times_sorted,
                }
            else:
                days[day] = {
                    'clock_in': times_sorted[0],
                    'clock_out': times_sorted[-1],
                    'is_single': False,
                    'count': 1.0,
                    'all_times': times_sorted,
                }

        employees[emp_name] = {
            'no': emp_no,
            'name': emp_name,
            'raw_dept': dept,
            'days': days,
        }
        i = j

    return employees


def _summary_day_cols(day: int) -> Tuple[int, int, int, int]:
    """汇总表：返回某一天的（上班时间列，下班时间列，小时数公式列，天数列），1-based"""
    base = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (day - 1)
    return base, base + 1, base + 2, base + 3


def _time_to_cell_value(t: Optional[datetime.time], template_value) -> object:
    """保持与模板单元格相同的存储类型"""
    if t is None:
        return None
    if isinstance(template_value, str):
        return f"{t.hour:02d}:{t.minute:02d}"
    return t


def _cell_to_time(value) -> Optional[datetime.time]:
    """把模板单元格中的时间值（time 或 'HH:MM' 字符串）转成 datetime.time"""
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value
    return parse_time(str(value))


NOON = datetime.time(12, 0)


def _times_span_noon(times: List[datetime.time]) -> bool:
    """判断一组时间是否跨中午（存在早于 12:00 和晚于等于 12:00 的时间）"""
    if len(times) < 2:
        return False
    has_before = any(t < NOON for t in times)
    has_after = any(t >= NOON for t in times)
    return has_before and has_after


def _time_to_hours(t: datetime.time) -> float:
    """把 datetime.time 转成当天小时数（浮点数）"""
    return t.hour + t.minute / 60 + t.second / 3600


def _compute_summary_row_days(all_times: List[datetime.time],
                              is_extra: bool = False) -> float:
    """
    计算汇总表 部门分发 单行的天数。
    规则与 3 月参考表一致：
      - 续行固定 0.5 天；
      - 仅 1 个时间：0.5 天；
      - 2 个时间：上班时间 <= 10:00 且下班时间 >= 14:00 为 1 天，否则 0.5 天；
      - 3 个及以上：主行 0.5 天。
    """
    if is_extra:
        return 0.5
    n = len(all_times)
    if n == 1:
        return 0.5
    if n == 2:
        if all_times[0] <= datetime.time(10, 0) and all_times[1] >= datetime.time(14, 0):
            return 1.0
        return 0.5
    return 0.5


def _choose_time(all_times: List[datetime.time],
                 template_time: Optional[datetime.time]) -> datetime.time:
    """
    优先沿用模板原值；若模板无值，则取最早/最晚。
    这里用于汇总表：上班时间取最早，下班时间取最晚。
    """
    if template_time is not None:
        return template_time
    return sorted(all_times)[0 if template_time is None else -1]


def _copy_row_style(ws, src_row: int, dst_row: int):
    """把 src_row 的样式复制到 dst_row（包括行高和每个单元格样式）"""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def build_summary_distribution_sheet(ws, employees: Dict[str, dict],
                                     days_in_month: int,
                                     month: int) -> None:
    """
    清空模板原有数据行，按输入员工顺序重建汇总表的 部门分发 sheet。
    每个员工一行主数据 + 有需要时的续行，保留模板第 2 行样式。
    """
    style_source_row = 2
    # 清空原有数据行
    for r in range(ws.max_row, style_source_row, -1):
        ws.delete_rows(r)
    for c in range(1, ws.max_column + 1):
        ws.cell(row=style_source_row, column=c).value = None

    current_row = style_source_row
    for name in employees.keys():
        emp = employees[name]
        days = emp['days']
        need_extra = any(len(set(d['all_times'])) >= 3 for d in days.values())

        # 主行：最早 + 最晚
        _copy_row_style(ws, style_source_row, current_row)
        ws.cell(row=current_row, column=1).value = '五楼'
        ws.cell(row=current_row, column=2).value = name
        ws.cell(row=current_row, column=3).value = emp.get('no')
        ws.cell(row=current_row, column=4).value = None
        ws.cell(row=current_row, column=5).value = None

        for day in range(1, days_in_month + 1):
            cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
            day_data = days.get(day)
            template_cin_val = ws.cell(row=style_source_row, column=cin_col).value
            template_cout_val = ws.cell(row=style_source_row, column=cout_col).value
            cin_letter = get_column_letter(cin_col)
            cout_letter = get_column_letter(cout_col)

            if not day_data:
                ws.cell(row=current_row, column=cin_col).value = None
                ws.cell(row=current_row, column=cout_col).value = None
                ws.cell(row=current_row, column=hours_col).value = None
                ws.cell(row=current_row, column=days_col).value = None
                ws.cell(row=current_row, column=cin_col).fill = NO_FILL
                ws.cell(row=current_row, column=cout_col).fill = NO_FILL
                continue

            all_times = sorted(set(day_data['all_times']))
            n_times = len(all_times)

            if n_times == 1:
                ws.cell(row=current_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=current_row, column=cout_col).value = None
                ws.cell(row=current_row, column=hours_col).value = None
                ws.cell(row=current_row, column=days_col).value = 0.5
                fill = _time_cell_fill(all_times, 0, month, day)
                ws.cell(row=current_row, column=cin_col).fill = fill if fill else NO_FILL
                ws.cell(row=current_row, column=cout_col).fill = NO_FILL
            elif n_times == 2:
                ws.cell(row=current_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=current_row, column=cout_col).value = _time_to_cell_value(
                    all_times[1], template_cout_val
                )
                ws.cell(row=current_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{current_row}-{cin_letter}{current_row},1)*24"
                )
                ws.cell(row=current_row, column=days_col).value = _compute_summary_row_days(
                    all_times
                )
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, 1, month, day)
                ws.cell(row=current_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=current_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
            else:
                # 3 个及以上：主行取前两个时间
                ws.cell(row=current_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=current_row, column=cout_col).value = _time_to_cell_value(
                    all_times[1], template_cout_val
                )
                ws.cell(row=current_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{current_row}-{cin_letter}{current_row},1)*24"
                )
                ws.cell(row=current_row, column=days_col).value = 0.5
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, 1, month, day)
                ws.cell(row=current_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=current_row, column=cout_col).fill = fill_out if fill_out else NO_FILL

        current_row += 1

        if need_extra:
            # 续行：剩余时间
            _copy_row_style(ws, style_source_row, current_row)
            ws.cell(row=current_row, column=1).value = '五楼'
            ws.cell(row=current_row, column=2).value = name
            ws.cell(row=current_row, column=3).value = emp.get('no')
            ws.cell(row=current_row, column=4).value = None
            ws.cell(row=current_row, column=5).value = None
            for day in range(1, days_in_month + 1):
                cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
                cin_letter = get_column_letter(cin_col)
                cout_letter = get_column_letter(cout_col)
                ws.cell(row=current_row, column=cin_col).value = None
                ws.cell(row=current_row, column=cout_col).value = None
                ws.cell(row=current_row, column=hours_col).value = None
                ws.cell(row=current_row, column=days_col).value = None
                ws.cell(row=current_row, column=cin_col).fill = NO_FILL
                ws.cell(row=current_row, column=cout_col).fill = NO_FILL
                day_data = days.get(day)
                if day_data and len(set(day_data['all_times'])) >= 3:
                    extra_times = sorted(set(day_data['all_times']))
                    template_cin_val = ws.cell(row=style_source_row, column=cin_col).value
                    template_cout_val = ws.cell(row=style_source_row, column=cout_col).value
                    if len(extra_times) == 3:
                        ws.cell(row=current_row, column=cin_col).value = _time_to_cell_value(
                            extra_times[2], template_cin_val
                        )
                        ws.cell(row=current_row, column=hours_col).value = None
                        fill = _time_cell_fill(extra_times, 2, month, day)
                        ws.cell(row=current_row, column=cin_col).fill = fill if fill else NO_FILL
                        ws.cell(row=current_row, column=cout_col).fill = NO_FILL
                    elif len(extra_times) >= 4:
                        ws.cell(row=current_row, column=cin_col).value = _time_to_cell_value(
                            extra_times[2], template_cin_val
                        )
                        ws.cell(row=current_row, column=cout_col).value = _time_to_cell_value(
                            extra_times[3], template_cout_val
                        )
                        ws.cell(row=current_row, column=hours_col).value = (
                            f"=MOD({cout_letter}{current_row}-{cin_letter}{current_row},1)*24"
                        )
                        fill_in = _time_cell_fill(extra_times, 2, month, day)
                        fill_out = _time_cell_fill(extra_times, 3, month, day)
                        ws.cell(row=current_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                        ws.cell(row=current_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
                    ws.cell(row=current_row, column=days_col).value = 0.5
            current_row += 1

    # 在末尾增加两列合计，供汇总表 SUMIF 引用
    add_summary_distribution_totals(ws, days_in_month, start_row=style_source_row)


def clear_summary_manual_fields(ws) -> None:
    """清空汇总表中人工填写的补卡/出差/请假/未打卡/备注信息（S-W 列）"""
    # S=19 补卡时间, T=20 出差时间, U=21 请假时间, V=22 未打卡时间, W=23 备注
    manual_cols = [19, 20, 21, 22, 23]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        for c in manual_cols:
            ws.cell(row=r, column=c).value = None


def build_summary_sheet(ws, employees: Dict[str, dict], month_label: str,
                        days_in_month: int) -> None:
    """
    清空模板原有数据行，按输入员工顺序重建汇总 sheet。
    保留模板第 2 行样式，B-R 列留空给人事填写，公式引用对应月份部门分发。
    """
    style_source_row = 2
    for r in range(ws.max_row, style_source_row, -1):
        ws.delete_rows(r)
    for c in range(1, ws.max_column + 1):
        ws.cell(row=style_source_row, column=c).value = None

    # 汇总表公式引用 部门分发 最后两列的合计
    sheet_ref = f"'{month_label}部门分发'"
    count_end_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (days_in_month - 1) + 3
    hours_total_letter = get_column_letter(count_end_col + 1)
    count_total_letter = get_column_letter(count_end_col + 2)

    current_row = style_source_row
    for name in employees.keys():
        _copy_row_style(ws, style_source_row, current_row)
        ws.cell(row=current_row, column=1).value = name
        for c in range(2, 19):
            ws.cell(row=current_row, column=c).value = None
        ws.cell(row=current_row, column=8).value = (
            f"=SUMIF({sheet_ref}!B:B,A{current_row},{sheet_ref}!{hours_total_letter}:{hours_total_letter})"
        )
        ws.cell(row=current_row, column=9).value = f"=H{current_row}*60"
        ws.cell(row=current_row, column=10).value = (
            f"=SUMIF({sheet_ref}!B:B,A{current_row},{sheet_ref}!{count_total_letter}:{count_total_letter})"
        )
        ws.cell(row=current_row, column=18).value = (
            f"=J{current_row}+K{current_row}+L{current_row}+M{current_row}+"
            f"N{current_row}+O{current_row}+P{current_row}+Q{current_row}"
        )
        current_row += 1

    clear_summary_manual_fields(ws)


def update_summary_distribution_sheet(ws, employees: Dict[str, dict],
                                      days_in_month: int,
                                      month: int,
                                      split_3plus_days: bool = False,
                                      single_as_out_dates: Optional[Set[Tuple[str, int]]] = None) -> None:
    """
    月份专用模板：保守更新汇总表的 部门分发 sheet（保留模板原有员工信息和歧义处理）。
    split_3plus_days=True 时，3 次及以上拆成两行，每行 0.5 天（匹配 3 月参考表规则）。
    """
    # 收集每个员工在模板中的行号（主行 + 续行）
    emp_rows: Dict[str, List[int]] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if name:
            emp_rows.setdefault(name, []).append(r)

    for name, emp in employees.items():
        rows = emp_rows.get(name, [])
        if not rows:
            continue
        main_row = rows[0]
        extra_row = rows[1] if len(rows) > 1 else None
        days = emp.get('days') or {}

        for day in range(1, days_in_month + 1):
            cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
            day_data = days.get(day)
            template_cin = ws.cell(row=main_row, column=cin_col).value
            template_cout = ws.cell(row=main_row, column=cout_col).value

            if not day_data:
                ws.cell(row=main_row, column=cin_col).value = None
                ws.cell(row=main_row, column=cout_col).value = None
                ws.cell(row=main_row, column=hours_col).value = None
                ws.cell(row=main_row, column=days_col).value = None
                ws.cell(row=main_row, column=cin_col).fill = NO_FILL
                ws.cell(row=main_row, column=cout_col).fill = NO_FILL
                if extra_row:
                    ws.cell(row=extra_row, column=cin_col).value = None
                    ws.cell(row=extra_row, column=cout_col).value = None
                    ws.cell(row=extra_row, column=hours_col).value = None
                    ws.cell(row=extra_row, column=days_col).value = None
                    ws.cell(row=extra_row, column=cin_col).fill = NO_FILL
                    ws.cell(row=extra_row, column=cout_col).fill = NO_FILL
                continue

            all_times = sorted(set(day_data.get('all_times', [])))
            n_times = len(all_times)
            cin_letter = get_column_letter(cin_col)
            cout_letter = get_column_letter(cout_col)

            if n_times == 1:
                single_time = all_times[0]
                if single_as_out_dates and (name, day) in single_as_out_dates:
                    ws.cell(row=main_row, column=cin_col).value = None
                    ws.cell(row=main_row, column=cout_col).value = _time_to_cell_value(
                        single_time, template_cout
                    )
                    ws.cell(row=main_row, column=cout_col).fill = SINGLE_FILL
                    ws.cell(row=main_row, column=cin_col).fill = NO_FILL
                elif template_cin is not None:
                    ws.cell(row=main_row, column=cin_col).value = _time_to_cell_value(
                        single_time, template_cin
                    )
                    ws.cell(row=main_row, column=cout_col).value = None
                    ws.cell(row=main_row, column=cin_col).fill = SINGLE_FILL
                    ws.cell(row=main_row, column=cout_col).fill = NO_FILL
                else:
                    ws.cell(row=main_row, column=cin_col).value = None
                    ws.cell(row=main_row, column=cout_col).value = _time_to_cell_value(
                        single_time, template_cout
                    )
                    ws.cell(row=main_row, column=cout_col).fill = SINGLE_FILL
                    ws.cell(row=main_row, column=cin_col).fill = NO_FILL
                ws.cell(row=main_row, column=hours_col).value = None
                ws.cell(row=main_row, column=days_col).value = 0.5
                if extra_row:
                    ws.cell(row=extra_row, column=cin_col).value = None
                    ws.cell(row=extra_row, column=cout_col).value = None
                    ws.cell(row=extra_row, column=hours_col).value = None
                    ws.cell(row=extra_row, column=days_col).value = None
                    ws.cell(row=extra_row, column=cin_col).fill = NO_FILL
                    ws.cell(row=extra_row, column=cout_col).fill = NO_FILL
                continue

            if split_3plus_days and n_times >= 3:
                # 主行：前两个时间，0.5 天
                ws.cell(row=main_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin
                )
                ws.cell(row=main_row, column=cout_col).value = _time_to_cell_value(
                    all_times[1], template_cout
                )
                ws.cell(row=main_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{main_row}-{cin_letter}{main_row},1)*24"
                )
                ws.cell(row=main_row, column=days_col).value = 0.5
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, 1, month, day)
                ws.cell(row=main_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=main_row, column=cout_col).fill = fill_out if fill_out else NO_FILL

                # 续行：剩余时间，0.5 天
                if extra_row:
                    if n_times == 3:
                        ws.cell(row=extra_row, column=cin_col).value = _time_to_cell_value(
                            all_times[2], template_cin
                        )
                        ws.cell(row=extra_row, column=cout_col).value = None
                        ws.cell(row=extra_row, column=hours_col).value = None
                        fill = _time_cell_fill(all_times, 2, month, day)
                        ws.cell(row=extra_row, column=cin_col).fill = fill if fill else NO_FILL
                        ws.cell(row=extra_row, column=cout_col).fill = NO_FILL
                    else:
                        ws.cell(row=extra_row, column=cin_col).value = _time_to_cell_value(
                            all_times[2], template_cin
                        )
                        ws.cell(row=extra_row, column=cout_col).value = _time_to_cell_value(
                            all_times[3], template_cout
                        )
                        ws.cell(row=extra_row, column=hours_col).value = (
                            f"=MOD({cout_letter}{extra_row}-{cin_letter}{extra_row},1)*24"
                        )
                        fill_in = _time_cell_fill(all_times, 2, month, day)
                        fill_out = _time_cell_fill(all_times, 3, month, day)
                        ws.cell(row=extra_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                        ws.cell(row=extra_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
                    ws.cell(row=extra_row, column=days_col).value = 0.5
                continue

            # 默认规则：最早 + 最晚，优先沿用模板原值消除多刷卡歧义
            template_cin_time = _cell_to_time(template_cin)
            template_cout_time = _cell_to_time(template_cout)
            clock_in = template_cin_time if template_cin_time is not None else all_times[0]
            clock_out = template_cout_time if template_cout_time is not None else all_times[-1]

            ws.cell(row=main_row, column=cin_col).value = _time_to_cell_value(
                clock_in, template_cin
            )
            ws.cell(row=main_row, column=cout_col).value = _time_to_cell_value(
                clock_out, template_cout
            )
            ws.cell(row=main_row, column=hours_col).value = (
                f"=MOD({cout_letter}{main_row}-{cin_letter}{main_row},1)*24"
            )
            if split_3plus_days:
                ws.cell(row=main_row, column=days_col).value = _compute_summary_row_days(
                    [clock_in, clock_out]
                )
            # 非拆分月份沿用模板天数，不覆盖
            fill_in = _time_cell_fill(all_times, 0, month, day)
            fill_out = _time_cell_fill(all_times, n_times - 1, month, day)
            ws.cell(row=main_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
            ws.cell(row=main_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
            if extra_row:
                ws.cell(row=extra_row, column=cin_col).value = None
                ws.cell(row=extra_row, column=cout_col).value = None
                ws.cell(row=extra_row, column=hours_col).value = None
                ws.cell(row=extra_row, column=days_col).value = None
                ws.cell(row=extra_row, column=cin_col).fill = NO_FILL
                ws.cell(row=extra_row, column=cout_col).fill = NO_FILL


def append_summary_rows(ws, employees: Dict[str, dict],
                        month_label: str, days_in_month: int) -> None:
    """月份专用模板：把输入中有、但模板 汇总 sheet 里没有的员工追加到末尾"""
    existing_names = set()
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name:
            existing_names.add(name)
            last_data_row = r

    to_append = [name for name in employees.keys() if name not in existing_names]
    if not to_append:
        return

    sheet_ref = f"'{month_label}部门分发'"
    hours_start_col = SUMMARY_DAY_START_COL + 2
    count_start_col = SUMMARY_DAY_START_COL + 3
    count_end_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (days_in_month - 1) + 3
    hours_start_letter = get_column_letter(hours_start_col)
    count_start_letter = get_column_letter(count_start_col)
    count_end_letter = get_column_letter(count_end_col)

    next_row = last_data_row + 1
    for name in to_append:
        _copy_row_style(ws, last_data_row, next_row)
        ws.cell(row=next_row, column=1).value = name
        for c in range(2, 19):
            ws.cell(row=next_row, column=c).value = None
        ws.cell(row=next_row, column=8).value = (
            f"=SUMPRODUCT(({sheet_ref}!$B:$B=A{next_row})*"
            f"(MOD(COLUMN({sheet_ref}!${hours_start_letter}:${count_end_letter})-"
            f"COLUMN({sheet_ref}!${hours_start_letter}:${hours_start_letter}),4)=0)*"
            f"({sheet_ref}!${hours_start_letter}:${count_end_letter}))"
        )
        ws.cell(row=next_row, column=9).value = f"=H{next_row}*60"
        ws.cell(row=next_row, column=10).value = (
            f"=SUMPRODUCT(({sheet_ref}!$B:$B=A{next_row})*"
            f"(MOD(COLUMN({sheet_ref}!${count_start_letter}:${count_end_letter})-"
            f"COLUMN({sheet_ref}!${count_start_letter}:${count_start_letter}),4)=0)*"
            f"({sheet_ref}!${count_start_letter}:${count_end_letter}))"
        )
        ws.cell(row=next_row, column=18).value = (
            f"=J{next_row}+K{next_row}+L{next_row}+M{next_row}+"
            f"N{next_row}+O{next_row}+P{next_row}+Q{next_row}"
        )
        last_data_row = next_row
        next_row += 1


def _build_summary_totals_formula(start_letter: str, end_letter: str, row: int) -> str:
    """生成对 部门分发 行内多列求和的 SUM 公式（列不连续，每 4 列取 1）"""
    # 把列字母转成数字，再每隔 4 列取一个
    start_num = column_index_from_string(start_letter)
    end_num = column_index_from_string(end_letter)
    cols = [get_column_letter(c) for c in range(start_num, end_num + 1, 4)]
    terms = [f"{col}{row}" for col in cols]
    return "=SUM(" + ",".join(terms) + ")"


def add_summary_distribution_totals(ws, days_in_month: int,
                                    start_row: int = 2) -> Tuple[str, str]:
    """
    在 部门分发 sheet 最后两列增加"工时合计"、"天数合计"，返回对应的列字母。
    """
    count_end_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (days_in_month - 1) + 3
    hours_total_col = count_end_col + 1
    count_total_col = count_end_col + 2
    hours_total_letter = get_column_letter(hours_total_col)
    count_total_letter = get_column_letter(count_total_col)

    hours_start = get_column_letter(SUMMARY_DAY_START_COL + 2)
    hours_end = get_column_letter(SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (days_in_month - 1) + 2)
    count_start = get_column_letter(SUMMARY_DAY_START_COL + 3)
    count_end = get_column_letter(count_end_col)

    for r in range(start_row, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is None:
            continue
        ws.cell(row=r, column=hours_total_col).value = _build_summary_totals_formula(
            hours_start, hours_end, r
        )
        ws.cell(row=r, column=count_total_col).value = _build_summary_totals_formula(
            count_start, count_end, r
        )

    return hours_total_letter, count_total_letter


def update_summary_formulas(ws, month_label: str, hours_total_letter: str,
                            count_total_letter: str) -> None:
    """更新汇总表 H/I/J/R 列公式，引用 部门分发 的合计列"""
    sheet_ref = f"'{month_label}部门分发'"
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        ws.cell(row=r, column=8).value = (
            f"=SUMIF({sheet_ref}!B:B,A{r},{sheet_ref}!{hours_total_letter}:{hours_total_letter})"
        )
        ws.cell(row=r, column=9).value = f"=H{r}*60"
        ws.cell(row=r, column=10).value = (
            f"=SUMIF({sheet_ref}!B:B,A{r},{sheet_ref}!{count_total_letter}:{count_total_letter})"
        )
        ws.cell(row=r, column=18).value = (
            f"=J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}+Q{r}"
        )


def append_summary_distribution_rows(ws, employees: Dict[str, dict],
                                     days_in_month: int,
                                     month: int,
                                     split_3plus_days: bool = False,
                                     single_as_out_dates: Optional[Set[Tuple[str, int]]] = None) -> None:
    """
    月份专用模板：把输入中有、但模板里没有的员工追加到汇总表 部门分发 sheet 末尾。
    split_3plus_days=True 时，3 次及以上拆成两行，每行 0.5 天（匹配 3 月参考表规则）。
    """
    existing_names = set()
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if name:
            existing_names.add(name)
            last_data_row = r

    to_append = [name for name in employees.keys() if name not in existing_names]
    if not to_append:
        return

    next_row = last_data_row + 1
    for name in to_append:
        emp = employees[name]
        days = emp['days']
        need_extra = split_3plus_days and any(len(set(d['all_times'])) >= 3 for d in days.values())

        # 主行
        _copy_row_style(ws, last_data_row, next_row)
        ws.cell(row=next_row, column=1).value = '五楼'
        ws.cell(row=next_row, column=2).value = name
        ws.cell(row=next_row, column=3).value = emp.get('no')
        ws.cell(row=next_row, column=4).value = None
        ws.cell(row=next_row, column=5).value = None

        for day in range(1, days_in_month + 1):
            cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
            day_data = days.get(day)
            template_cin_val = ws.cell(row=last_data_row, column=cin_col).value
            template_cout_val = ws.cell(row=last_data_row, column=cout_col).value
            cin_letter = get_column_letter(cin_col)
            cout_letter = get_column_letter(cout_col)

            if not day_data:
                ws.cell(row=next_row, column=cin_col).value = None
                ws.cell(row=next_row, column=cout_col).value = None
                ws.cell(row=next_row, column=hours_col).value = None
                ws.cell(row=next_row, column=days_col).value = None
                ws.cell(row=next_row, column=cin_col).fill = NO_FILL
                ws.cell(row=next_row, column=cout_col).fill = NO_FILL
                continue

            all_times = sorted(set(day_data['all_times']))
            n_times = len(all_times)

            if n_times == 1:
                if single_as_out_dates and (name, day) in single_as_out_dates:
                    ws.cell(row=next_row, column=cin_col).value = None
                    ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                        all_times[0], template_cout_val
                    )
                    ws.cell(row=next_row, column=cout_col).fill = SINGLE_FILL
                    ws.cell(row=next_row, column=cin_col).fill = NO_FILL
                elif template_cin_val is not None:
                    ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                        all_times[0], template_cin_val
                    )
                    ws.cell(row=next_row, column=cout_col).value = None
                    ws.cell(row=next_row, column=cin_col).fill = SINGLE_FILL
                    ws.cell(row=next_row, column=cout_col).fill = NO_FILL
                else:
                    ws.cell(row=next_row, column=cin_col).value = None
                    ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                        all_times[0], template_cout_val
                    )
                    ws.cell(row=next_row, column=cout_col).fill = SINGLE_FILL
                    ws.cell(row=next_row, column=cin_col).fill = NO_FILL
                ws.cell(row=next_row, column=hours_col).value = None
                ws.cell(row=next_row, column=days_col).value = 0.5
            elif n_times == 2:
                ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                    all_times[1], template_cout_val
                )
                ws.cell(row=next_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{next_row}-{cin_letter}{next_row},1)*24"
                )
                ws.cell(row=next_row, column=days_col).value = _compute_summary_row_days(
                    all_times
                )
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, 1, month, day)
                ws.cell(row=next_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=next_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
            elif split_3plus_days and n_times >= 3:
                # 主行：前两个时间，0.5 天
                ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                    all_times[1], template_cout_val
                )
                ws.cell(row=next_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{next_row}-{cin_letter}{next_row},1)*24"
                )
                ws.cell(row=next_row, column=days_col).value = 0.5
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, 1, month, day)
                ws.cell(row=next_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=next_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
            else:
                ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                    all_times[0], template_cin_val
                )
                ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                    all_times[-1], template_cout_val
                )
                ws.cell(row=next_row, column=hours_col).value = (
                    f"=MOD({cout_letter}{next_row}-{cin_letter}{next_row},1)*24"
                )
                ws.cell(row=next_row, column=days_col).value = 1.0
                fill_in = _time_cell_fill(all_times, 0, month, day)
                fill_out = _time_cell_fill(all_times, n_times - 1, month, day)
                ws.cell(row=next_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                ws.cell(row=next_row, column=cout_col).fill = fill_out if fill_out else NO_FILL

        last_data_row = next_row
        next_row += 1

        if need_extra:
            _copy_row_style(ws, last_data_row - 1, next_row)
            ws.cell(row=next_row, column=1).value = '五楼'
            ws.cell(row=next_row, column=2).value = name
            ws.cell(row=next_row, column=3).value = emp.get('no')
            ws.cell(row=next_row, column=4).value = None
            ws.cell(row=next_row, column=5).value = None
            for day in range(1, days_in_month + 1):
                cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
                day_data = days.get(day)
                ws.cell(row=next_row, column=cin_col).value = None
                ws.cell(row=next_row, column=cout_col).value = None
                ws.cell(row=next_row, column=hours_col).value = None
                ws.cell(row=next_row, column=days_col).value = None
                ws.cell(row=next_row, column=cin_col).fill = NO_FILL
                ws.cell(row=next_row, column=cout_col).fill = NO_FILL
                if day_data and len(set(day_data['all_times'])) >= 3:
                    extra_times = sorted(set(day_data['all_times']))
                    template_cin_val = ws.cell(row=last_data_row, column=cin_col).value
                    template_cout_val = ws.cell(row=last_data_row, column=cout_col).value
                    cin_letter = get_column_letter(cin_col)
                    cout_letter = get_column_letter(cout_col)
                    if split_3plus_days:
                        if len(extra_times) == 3:
                            ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                                extra_times[2], template_cin_val
                            )
                            ws.cell(row=next_row, column=cout_col).value = None
                            ws.cell(row=next_row, column=hours_col).value = None
                            fill = _time_cell_fill(extra_times, 2, month, day)
                            ws.cell(row=next_row, column=cin_col).fill = fill if fill else NO_FILL
                            ws.cell(row=next_row, column=cout_col).fill = NO_FILL
                        elif len(extra_times) >= 4:
                            ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                                extra_times[2], template_cin_val
                            )
                            ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                                extra_times[3], template_cout_val
                            )
                            ws.cell(row=next_row, column=hours_col).value = (
                                f"=MOD({cout_letter}{next_row}-{cin_letter}{next_row},1)*24"
                            )
                            fill_in = _time_cell_fill(extra_times, 2, month, day)
                            fill_out = _time_cell_fill(extra_times, 3, month, day)
                            ws.cell(row=next_row, column=cin_col).fill = fill_in if fill_in else NO_FILL
                            ws.cell(row=next_row, column=cout_col).fill = fill_out if fill_out else NO_FILL
                        ws.cell(row=next_row, column=days_col).value = 0.5
                    else:
                        if len(extra_times) == 3:
                            ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                                extra_times[2], template_cin_val
                            )
                        elif len(extra_times) >= 4:
                            ws.cell(row=next_row, column=cin_col).value = _time_to_cell_value(
                                extra_times[2], template_cin_val
                            )
                            ws.cell(row=next_row, column=cout_col).value = _time_to_cell_value(
                                extra_times[3], template_cout_val
                            )
                        ws.cell(row=next_row, column=hours_col).value = (
                            f"=MOD({cout_letter}{next_row}-{cin_letter}{next_row},1)*24"
                        )
                        ws.cell(row=next_row, column=days_col).value = 1
            last_data_row = next_row
            next_row += 1


def update_distribution_sheet(ws, employees: Dict[str, dict],
                              days_in_month: int,
                              month: int) -> None:
    """月份专用模板：保守更新分发表（保留模板原有员工信息和歧义处理）"""
    seen_first_row: Dict[str, bool] = {}

    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or name == '姓名':
            continue
        emp = employees.get(name)
        days = emp['days'] if emp else {}

        is_first_row = not seen_first_row.get(name, False)
        seen_first_row[name] = True

        for day in range(1, days_in_month + 1):
            col = DIST_DAY_START_COL + (day - 1)
            template_value = ws.cell(row=r, column=col).value
            day_data = days.get(day)

            if is_first_row:
                if not day_data:
                    if template_value is not None:
                        ws.cell(row=r, column=col).value = None
                    continue
                input_times = day_data['all_times']
                all_times = sorted(set(input_times))

                new_value = _build_dist_main_cell_value(input_times)
                if new_value != template_value:
                    ws.cell(row=r, column=col).value = new_value
                fill = _dist_day_cell_fill(all_times, month, day)
                ws.cell(row=r, column=col).fill = fill if fill else NO_FILL
            else:
                if not day_data:
                    if template_value is not None:
                        ws.cell(row=r, column=col).value = None
                    continue
                if len(set(day_data['all_times'])) >= 3:
                    new_value = _build_dist_extra_cell_value(day_data['all_times'])
                    if new_value != template_value:
                        ws.cell(row=r, column=col).value = new_value
                else:
                    if template_value is not None:
                        ws.cell(row=r, column=col).value = None


def append_distribution_rows(ws, employees: Dict[str, dict],
                             days_in_month: int,
                             month: int) -> None:
    """月份专用模板：把输入中有、但模板里没有的员工追加到分发表末尾"""
    existing_names = set()
    last_data_row = 2
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if name and name != '姓名':
            existing_names.add(name)
            last_data_row = r

    to_append = [name for name in employees.keys() if name not in existing_names]
    if not to_append:
        return

    next_row = last_data_row + 1
    for name in to_append:
        emp = employees[name]
        days = emp['days']
        need_extra = any(len(set(d['all_times'])) >= 3 for d in days.values())

        # 主行
        _copy_row_style(ws, last_data_row, next_row)
        ws.cell(row=next_row, column=1).value = '五楼'
        ws.cell(row=next_row, column=2).value = name
        ws.cell(row=next_row, column=3).value = emp.get('no')
        ws.cell(row=next_row, column=4).value = None
        ws.cell(row=next_row, column=5).value = None
        for day in range(1, days_in_month + 1):
            col = DIST_DAY_START_COL + (day - 1)
            day_data = days.get(day)
            cell = ws.cell(row=next_row, column=col)
            if day_data:
                all_times = sorted(set(day_data['all_times']))
                cell.value = _build_dist_main_cell_value(day_data['all_times'])
                fill = _dist_day_cell_fill(all_times, month, day)
                cell.fill = fill if fill else NO_FILL
            else:
                cell.fill = NO_FILL
        last_data_row = next_row
        next_row += 1

        if need_extra:
            _copy_row_style(ws, last_data_row, next_row)
            ws.cell(row=next_row, column=1).value = '五楼'
            ws.cell(row=next_row, column=2).value = name
            ws.cell(row=next_row, column=3).value = emp.get('no')
            ws.cell(row=next_row, column=4).value = None
            ws.cell(row=next_row, column=5).value = None
            for day in range(1, days_in_month + 1):
                col = DIST_DAY_START_COL + (day - 1)
                day_data = days.get(day)
                if day_data and len(set(day_data['all_times'])) >= 3:
                    ws.cell(row=next_row, column=col).value = _build_dist_extra_cell_value(day_data['all_times'])
            last_data_row = next_row
            next_row += 1


def _format_dist_time(t: datetime.time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def _build_dist_main_cell_value(input_times: List[datetime.time]) -> Optional[str]:
    """主行：1 个时间或最早两个时间"""
    if not input_times:
        return None
    times = sorted(set(input_times))
    if len(times) == 1:
        return f"{_format_dist_time(times[0])}\n     "
    return f"{_format_dist_time(times[0])}\n{_format_dist_time(times[1])}"


def _build_dist_extra_cell_value(input_times: List[datetime.time]) -> Optional[str]:
    """续行：取剩余时间；3 个时间取第 3 个，4 个及以上取第 3、4 个"""
    if not input_times:
        return None
    times = sorted(set(input_times))
    if len(times) == 3:
        return f"{_format_dist_time(times[2])}\n     "
    if len(times) >= 4:
        return f"{_format_dist_time(times[2])}\n{_format_dist_time(times[3])}"
    return None


def _early_leave_threshold(month: int, day: int) -> datetime.time:
    """夏季（5/1 ~ 10/7 国庆假期）18:00；冬季（10/8 国庆回来 ~ 次年 4 月底）17:40。"""
    is_winter = (
        month < 5
        or month > WINTER_START_MONTH
        or (month == WINTER_START_MONTH and day >= WINTER_START_DAY)
    )
    return WINTER_EARLY_LEAVE_THRESHOLD if is_winter else SUMMER_EARLY_LEAVE_THRESHOLD


def _attendance_flag(day_data: Optional[dict], month: int, day: int) -> Optional[str]:
    """返回某天迟到/早退异常文本；无异常返回 None。"""
    if not day_data:
        return None
    times = sorted(set(day_data.get('all_times', [])))
    if not times:
        return None
    if len(times) == 1:
        return "只打一次卡"

    flags = []
    clock_in = times[0]
    clock_out = times[-1]

    if clock_in > LATE_THRESHOLD:
        flags.append("迟到")
    if clock_out < _early_leave_threshold(month, day):
        flags.append("早退")

    if not flags:
        return None
    return "/".join(flags)


def _remove_existing_attendance_flag_columns(ws, header_row: int) -> None:
    """删除之前生成过的 N日异常 列，避免重复运行时不断追加。"""
    cols = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and FLAG_HEADER_RE.match(value.strip()):
            cols.append(col)
    for col in reversed(cols):
        ws.delete_cols(col)


def _copy_column_style(ws, src_col: int, dst_col: int, start_row: int = 1) -> None:
    """复制一列的宽度和已有单元格样式到新列。"""
    src_letter = get_column_letter(src_col)
    dst_letter = get_column_letter(dst_col)
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    for row in range(start_row, ws.max_row + 1):
        src = ws.cell(row=row, column=src_col)
        dst = ws.cell(row=row, column=dst_col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def add_attendance_flag_columns(ws, employees: Dict[str, dict], month: int,
                                days_in_month: int, header_row: int,
                                data_start_row: int, name_col: int = 2) -> None:
    """
    在部门分发/分发表末尾追加每日异常列。

    同名续行只在第一行写异常，避免 3 次/4 次打卡拆行时重复显示。
    """
    _remove_existing_attendance_flag_columns(ws, header_row)
    start_col = ws.max_column + 1
    style_col = max(1, start_col - 1)

    for day in range(1, days_in_month + 1):
        col = start_col + day - 1
        _copy_column_style(ws, style_col, col, start_row=header_row)
        header = ws.cell(row=header_row, column=col)
        header.value = f"{day}日异常"
        header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 11

    seen_names: Set[str] = set()
    for row in range(data_start_row, ws.max_row + 1):
        name = _normalize_name(ws.cell(row=row, column=name_col).value)
        if not name:
            continue
        is_first_row = name not in seen_names
        seen_names.add(name)
        emp = employees.get(name)
        for day in range(1, days_in_month + 1):
            cell = ws.cell(row=row, column=start_col + day - 1)
            if not is_first_row:
                cell.value = None
                continue
            cell.value = _attendance_flag(emp.get('days', {}).get(day) if emp else None, month, day)


def apply_attendance_column_widths(ws, days_in_month: int, summary_layout: bool) -> None:
    """设置考勤日期/时间/异常列宽，避免 WPS 中显示 ##### 或文字截断。"""
    if summary_layout:
        for day in range(1, days_in_month + 1):
            cin_col, cout_col, hours_col, days_col = _summary_day_cols(day)
            ws.column_dimensions[get_column_letter(cin_col)].width = 10
            ws.column_dimensions[get_column_letter(cout_col)].width = 10
            ws.column_dimensions[get_column_letter(hours_col)].width = 9
            ws.column_dimensions[get_column_letter(days_col)].width = 7
            for col in (cin_col, cout_col, hours_col, days_col):
                ws.cell(row=1, column=col).alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
    else:
        for day in range(1, days_in_month + 1):
            col = DIST_DAY_START_COL + day - 1
            ws.column_dimensions[get_column_letter(col)].width = 10
            ws.cell(row=2, column=col).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True
            )


def _template_times_subset_of_input(template_value, input_times: List[datetime.time]) -> bool:
    """判断模板单元格里的所有时间是否都存在于输入时间中"""
    template_times = parse_cell_times(template_value)
    if not template_times:
        return True  # 模板为空，视为可更新
    input_set = set(input_times)
    return all(t in input_set for t in template_times)


def build_distribution_sheet(ws, employees: Dict[str, dict],
                             days_in_month: int,
                             month: int,
                             style_source_row: int = 3) -> None:
    """
    清空模板原有数据行，按输入员工顺序重建分发表。
    每个员工一行主数据 + 有需要时的续行，保留模板第 style_source_row 行样式。
    """
    # 清空原有数据行
    for r in range(ws.max_row, style_source_row, -1):
        ws.delete_rows(r)
    for c in range(1, ws.max_column + 1):
        ws.cell(row=style_source_row, column=c).value = None

    info_font = Font(name='微软雅黑', size=10)
    info_align = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='微软雅黑', size=9)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = style_source_row
    for name in employees.keys():
        emp = employees[name]
        days = emp['days']
        need_extra = any(len(set(d['all_times'])) >= 3 for d in days.values())

        # 主行（部门留空，由人事后续导入）
        _copy_row_style(ws, style_source_row, current_row)
        ws.cell(row=current_row, column=1).value = '五楼'
        ws.cell(row=current_row, column=2).value = name
        ws.cell(row=current_row, column=3).value = emp.get('no')
        ws.cell(row=current_row, column=4).value = None
        ws.cell(row=current_row, column=5).value = None
        for c in range(1, 6):
            cell = ws.cell(row=current_row, column=c)
            cell.font = info_font
            cell.alignment = info_align
        for day in range(1, days_in_month + 1):
            col = DIST_DAY_START_COL + (day - 1)
            cell = ws.cell(row=current_row, column=col)
            day_data = days.get(day)
            if day_data:
                all_times = sorted(set(day_data['all_times']))
                cell.value = _build_dist_main_cell_value(all_times)
                fill = _dist_day_cell_fill(all_times, month, day)
                cell.fill = fill if fill else NO_FILL
            else:
                cell.fill = NO_FILL
            cell.font = data_font
            cell.alignment = data_align
        current_row += 1

        if need_extra:
            _copy_row_style(ws, style_source_row, current_row)
            ws.cell(row=current_row, column=1).value = '五楼'
            ws.cell(row=current_row, column=2).value = name
            ws.cell(row=current_row, column=3).value = emp.get('no')
            ws.cell(row=current_row, column=4).value = None
            ws.cell(row=current_row, column=5).value = None
            for c in range(1, 6):
                cell = ws.cell(row=current_row, column=c)
                cell.font = info_font
                cell.alignment = info_align
            for day in range(1, days_in_month + 1):
                col = DIST_DAY_START_COL + (day - 1)
                cell = ws.cell(row=current_row, column=col)
                day_data = days.get(day)
                if day_data and len(set(day_data['all_times'])) >= 3:
                    cell.value = _build_dist_extra_cell_value(day_data['all_times'])
                cell.font = data_font
                cell.alignment = data_align
            current_row += 1


def _copy_cell_style(src, dst):
    """复制单元格样式（字体、填充、边框、对齐、数字格式）"""
    if src.has_style:
        dst.font = src.font.copy()
        dst.border = src.border.copy()
        dst.fill = src.fill.copy()
        dst.number_format = src.number_format
        dst.protection = src.protection.copy()
        dst.alignment = src.alignment.copy()


def _build_distribution_workbook(employees: Dict[str, dict], month_label: str,
                                 days_in_month: int,
                                 month: int) -> object:
    """
    当不存在月份分发表模板时，动态构建一个简洁的分发表。
    结构与 June 模板一致：打卡地点、姓名、工号、一级部门、二级部门 + 每天 1 列。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_label}部门分发"

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + days_in_month)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{month_label}办公室职能部门考勤数据"
    title_cell.font = Font(name='微软雅黑', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 表头
    headers = ['打卡地点', '姓名', '工号', '一级部门', '二级部门'] + [str(d) for d in range(1, days_in_month + 1)]
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    header_align = Alignment(horizontal='center', vertical='center')

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c)
        cell.value = h
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 22

    # 列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(DIST_DAY_START_COL + d - 1)].width = 11

    # 填充员工数据
    row = 3
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for name in sorted(employees.keys()):
        emp = employees[name]
        days = emp['days']

        # 判断是否需要续行（存在任意一天有 >=3 个时间）
        need_extra = any(len(set(d['all_times'])) >= 3 for d in days.values())

        # 主行（部门留空，后续由人事通过花名册导入）
        ws.cell(row=row, column=1).value = '五楼'
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = emp.get('no')
        ws.cell(row=row, column=4).value = None
        ws.cell(row=row, column=5).value = None
        for day in range(1, days_in_month + 1):
            cell = ws.cell(row=row, column=DIST_DAY_START_COL + day - 1)
            day_data = days.get(day)
            if day_data:
                all_times = sorted(set(day_data['all_times']))
                cell.value = _build_dist_main_cell_value(day_data['all_times'])
                fill = _dist_day_cell_fill(all_times, month, day)
                cell.fill = fill if fill else NO_FILL
            else:
                cell.fill = NO_FILL
            cell.alignment = data_align
            cell.border = border
        for c in range(1, 6):
            ws.cell(row=row, column=c).alignment = data_align
            ws.cell(row=row, column=c).border = border
        ws.row_dimensions[row].height = 22
        row += 1

        if need_extra:
            # 续行（部门留空）
            ws.cell(row=row, column=1).value = '五楼'
            ws.cell(row=row, column=2).value = name
            ws.cell(row=row, column=3).value = emp.get('no')
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=5).value = None
            for day in range(1, days_in_month + 1):
                cell = ws.cell(row=row, column=DIST_DAY_START_COL + day - 1)
                day_data = days.get(day)
                if day_data and len(set(day_data['all_times'])) >= 3:
                    cell.value = _build_dist_extra_cell_value(day_data['all_times'])
                cell.alignment = data_align
                cell.border = border
            for c in range(1, 6):
                ws.cell(row=row, column=c).alignment = data_align
                ws.cell(row=row, column=c).border = border
            ws.row_dimensions[row].height = 22
            row += 1

    return wb


def generate(input_path: str, output_dir: str,
             summary_template_path: Optional[str] = None,
             distribution_template_path: Optional[str] = None,
             output_summary_name: Optional[str] = None,
             output_distribution_name: Optional[str] = None,
             include_names: Optional[Set[str]] = None,
             exclude_names: Optional[Set[str]] = None,
             append_missing_employees: bool = True,
             preserve_template_roster: bool = False,
             drop_empty_employees: bool = True,
             summary_exclude_names: Optional[Set[str]] = None,
             summary_single_as_out_dates: Optional[Set[Tuple[str, int]]] = None) -> List[str]:
    """
    生成两份输出 XLSX 文件。

    Args:
        input_path: 输入 .xls 文件路径
        output_dir: 输出目录
        summary_template_path: 汇总表模板路径（可选，默认按月份自动选择）
        distribution_template_path: 分发表模板路径（可选，默认按月份自动选择）
        output_summary_name: 汇总表输出文件名/相对路径（可选）
        output_distribution_name: 分发表输出文件名/相对路径（可选）
        include_names: 只保留这些员工（可选，用于按当月名单过滤）
        exclude_names: 排除这些员工（可选）
        append_missing_employees: 是否追加输入有但模板没有的员工
        preserve_template_roster: 是否把传入模板视为本月名单模板
        drop_empty_employees: 是否过滤整月无打卡记录的员工

    Returns:
        生成的两个文件路径列表
    """
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    year, month = detect_month(str(input_path))
    month_label = get_month_label(month)
    days_in_month = get_days_in_month(year, month)
    default_excludes, default_summary_excludes, default_single_as_out = load_office_adjustments(month)
    if exclude_names:
        exclude_names = set(exclude_names) | default_excludes
    else:
        exclude_names = default_excludes
    if summary_exclude_names is None:
        summary_exclude_names = default_summary_excludes
    else:
        summary_exclude_names = set(summary_exclude_names)
    if summary_single_as_out_dates is None:
        summary_single_as_out_dates = default_single_as_out
    else:
        summary_single_as_out_dates = set(summary_single_as_out_dates)

    auto_summary_tmpl, auto_dist_tmpl, auto_summary_name, auto_distribution_name, is_base = resolve_templates(str(input_path))

    if output_summary_name is None:
        output_summary_name = auto_summary_name
    if output_distribution_name is None:
        output_distribution_name = auto_distribution_name

    if summary_template_path is None:
        summary_template_path = auto_summary_tmpl
    if distribution_template_path is None:
        distribution_template_path = auto_dist_tmpl
    if preserve_template_roster:
        is_base = False

    employees = parse_input(str(input_path))
    if include_names is not None:
        employees = {
            name: emp for name, emp in employees.items()
            if name in include_names
        }
    if exclude_names:
        employees = {
            name: emp for name, emp in employees.items()
            if name not in exclude_names
        }
    if drop_empty_employees:
        employees = {
            name: emp for name, emp in employees.items()
            if emp.get('days')
        }
    if include_names is not None:
        allowed_names = include_names - (exclude_names or set())
        for name in allowed_names:
            employees.setdefault(name, {
                'no': None,
                'name': name,
                'raw_dept': None,
                'days': {},
            })
    if not employees:
        raise ValueError("未能从输入文件中解析到员工刷卡数据")

    # 过滤整月无打卡记录的员工
    employees = {name: emp for name, emp in employees.items() if emp.get('days')}
    if not employees:
        raise ValueError("过滤后没有员工在本月有打卡数据")

    # 分发模板：基模板或 3 月新模板需要清空重建；汇总表保持保守更新以保留模板信息
    summary_distribution_employees = {
        name: emp for name, emp in employees.items()
        if name not in (summary_exclude_names or set())
    }

    rebuild_summary = is_base
    rebuild_dist = is_base or month_label == '3月'

    generated = []

    # 1. 汇总文件
    out_summary = output_dir / output_summary_name
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    wb_summary = load_workbook(summary_template_path)
    if is_base:
        _update_base_summary_sheet_name(wb_summary, month_label)
    dist_sheet_name = f"{month_label}部门分发"
    if dist_sheet_name not in wb_summary.sheetnames:
        dist_sheet_name = wb_summary.sheetnames[0]
    ws_dist = wb_summary[dist_sheet_name]
    if is_base:
        _update_base_summary_dates(ws_dist, year, month, days_in_month)
    ws_sum = wb_summary['汇总']
    if exclude_names:
        _delete_rows_by_name(ws_dist, 2, exclude_names)
        _delete_rows_by_name(ws_sum, 1, exclude_names)
    if summary_exclude_names:
        _delete_rows_by_name(ws_dist, 2, summary_exclude_names)

    # 删除模板中存在但本月无数据的员工行
    present_names = {_normalize_name(n) for n in employees.keys()}
    absent_in_summary_dist = set()
    for r in range(2, ws_dist.max_row + 1):
        name = _normalize_name(ws_dist.cell(row=r, column=2).value)
        if name and name not in present_names:
            absent_in_summary_dist.add(name)
    _delete_rows_by_name(ws_dist, 2, absent_in_summary_dist)

    absent_in_summary = set()
    for r in range(2, ws_sum.max_row + 1):
        name = _normalize_name(ws_sum.cell(row=r, column=1).value)
        if name and name not in present_names:
            absent_in_summary.add(name)
    _delete_rows_by_name(ws_sum, 1, absent_in_summary)

    if rebuild_summary:
        build_summary_distribution_sheet(ws_dist, summary_distribution_employees, days_in_month,
                                         month=month)
        build_summary_sheet(ws_sum, employees, month_label, days_in_month)
    else:
        split_3plus = month != 6
        update_summary_distribution_sheet(ws_dist, summary_distribution_employees, days_in_month,
                                          month=month,
                                          split_3plus_days=split_3plus,
                                          single_as_out_dates=summary_single_as_out_dates)
        if append_missing_employees:
            append_summary_distribution_rows(ws_dist, summary_distribution_employees, days_in_month,
                                             month=month,
                                             split_3plus_days=split_3plus,
                                             single_as_out_dates=summary_single_as_out_dates)
        # 3 月在保守更新后增加合计列并修正汇总公式
        if month_label == '3月' and is_base:
            hours_total_letter, count_total_letter = add_summary_distribution_totals(
                ws_dist, days_in_month, start_row=2
            )
        if append_missing_employees:
            append_summary_rows(ws_sum, employees, month_label, days_in_month)
        _delete_trailing_blank_rows(ws_sum)
        if month_label == '3月':
            # 月专用模板可能保留了硬编码行号的旧公式，删除无数据员工后重新写入 SUMIF 公式
            if not (is_base and hours_total_letter and count_total_letter):
                count_end_col = SUMMARY_DAY_START_COL + SUMMARY_DAY_COLS_PER_DAY * (days_in_month - 1) + 3
                hours_total_letter = get_column_letter(count_end_col + 1)
                count_total_letter = get_column_letter(count_end_col + 2)
            update_summary_formulas(ws_sum, month_label, hours_total_letter, count_total_letter)
    apply_attendance_column_widths(ws_dist, days_in_month, summary_layout=True)
    remove_roster_sheet(wb_summary)
    _unlink_existing_output(out_summary)
    wb_summary.save(out_summary)
    wb_summary.close()
    generated.append(str(out_summary))

    # 2. 分发文件
    out_distribution = output_dir / output_distribution_name
    out_distribution.parent.mkdir(parents=True, exist_ok=True)
    wb_distribution = load_workbook(distribution_template_path)
    if is_base:
        _update_base_distribution_title(wb_distribution.active, month_label, days_in_month)
    dist_sheet_name2 = f"{month_label}部门分发"
    if dist_sheet_name2 not in wb_distribution.sheetnames:
        dist_sheet_name2 = wb_distribution.sheetnames[0]
    ws_dist2 = wb_distribution[dist_sheet_name2]
    if exclude_names:
        _delete_rows_by_name(ws_dist2, 2, exclude_names)

    # 删除分发模板中存在但本月无数据的员工行
    absent_in_dist = set()
    for r in range(3, ws_dist2.max_row + 1):
        name = _normalize_name(ws_dist2.cell(row=r, column=2).value)
        if name and name not in present_names:
            absent_in_dist.add(name)
    _delete_rows_by_name(ws_dist2, 2, absent_in_dist, min_row=3)

    if rebuild_dist:
        # 月份专用分发模板的数据从第 2 行开始；基模板从第 3 行开始
        dist_style_source_row = 3 if is_base else 2
        build_distribution_sheet(ws_dist2, employees, days_in_month,
                                 month=month,
                                 style_source_row=dist_style_source_row)
    else:
        update_distribution_sheet(ws_dist2, employees, days_in_month,
                                  month=month)
        if append_missing_employees:
            append_distribution_rows(ws_dist2, employees, days_in_month,
                                     month=month)
    apply_attendance_column_widths(ws_dist2, days_in_month, summary_layout=False)
    remove_roster_sheet(wb_distribution)
    _unlink_existing_output(out_distribution)
    wb_distribution.save(out_distribution)
    wb_distribution.close()
    generated.append(str(out_distribution))

    return generated


def _sorted_input_files(input_dir: Path) -> List[Path]:
    """按年月和文件名稳定排序输入目录中的 .xls 原始记录。"""
    files = [
        p for p in input_dir.glob("*.xls")
        if p.is_file() and not p.name.startswith("~$")
    ]

    def sort_key(path: Path) -> Tuple[int, int, str]:
        year, month = detect_month(str(path))
        return year, month, path.name

    return sorted(files, key=sort_key)


def generate_attendance_folder(batch_root: str = str(DEFAULT_WECHAT_ATTENDANCE_ROOT),
                               month_filter: Optional[int] = None,
                               exclude_names: Optional[Set[str]] = None) -> List[dict]:
    """
    批量处理微信考勤目录。

    目录约定：
      {batch_root}/输入/*.xls
      {batch_root}/输出/汇总表/*.xlsx
      {batch_root}/输出/分发表/*.xlsx
    """
    root = Path(batch_root)
    input_dir = root / "输入"
    output_dir = root / "输出"

    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在: {input_dir}")

    input_files = _sorted_input_files(input_dir)
    if month_filter is not None:
        input_files = [
            p for p in input_files
            if detect_month(str(p))[1] == month_filter
        ]
    if not input_files:
        if month_filter is None:
            raise FileNotFoundError(f"输入目录中没有 .xls 文件: {input_dir}")
        raise FileNotFoundError(f"输入目录中没有 {month_filter}月 .xls 文件: {input_dir}")

    results = []
    combined_exclude_names = set(exclude_names or set())
    combined_exclude_names.update(load_exclude_names_file(root))
    for input_file in input_files:
        year, month = detect_month(str(input_file))
        summary_name, distribution_name = resolve_folder_output_names(root, month)
        paths = generate(
            str(input_file),
            str(output_dir),
            output_summary_name=summary_name,
            output_distribution_name=distribution_name,
            exclude_names=combined_exclude_names,
        )
        results.append({
            "year": year,
            "month": month,
            "input": str(input_file),
            "summary": paths[0],
            "distribution": paths[1],
        })

    return results


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成考勤汇总表和分发表")
    parser.add_argument("input_path", nargs="?", help="单个输入 .xls 文件；省略时默认批处理微信考勤目录")
    parser.add_argument("output_dir", nargs="?", default="output", help="单文件模式输出目录")
    parser.add_argument("summary_template", nargs="?", default=None, help="单文件模式汇总模板")
    parser.add_argument("distribution_template", nargs="?", default=None, help="单文件模式分发模板")
    parser.add_argument("--batch", action="store_true", help="批量处理 batch-root/输入 下所有 .xls")
    parser.add_argument(
        "--batch-root",
        default=str(DEFAULT_WECHAT_ATTENDANCE_ROOT),
        help="批处理根目录，默认指向微信考勤目录",
    )
    parser.add_argument("--month", type=int, default=None, help="批处理时只处理指定月份，例如 --month 7")
    parser.add_argument(
        "--exclude-name",
        action="append",
        default=[],
        help="从输出中排除指定姓名；可重复传入",
    )
    return parser


if __name__ == '__main__':
    args = _build_arg_parser().parse_args()

    if args.batch or not args.input_path:
        rows = generate_attendance_folder(
            args.batch_root,
            args.month,
            exclude_names=set(args.exclude_name) if args.exclude_name else None,
        )
        print("Batch generated:")
        for row in rows:
            print(f"  {row['month']}月:")
            print(f"    input: {row['input']}")
            print(f"    summary: {row['summary']}")
            print(f"    distribution: {row['distribution']}")
    else:
        paths = generate(
            input_path=args.input_path,
            output_dir=args.output_dir,
            summary_template_path=args.summary_template,
            distribution_template_path=args.distribution_template,
            exclude_names=set(args.exclude_name) if args.exclude_name else None,
        )
        print('Generated:')
        for p in paths:
            print(' ', p)
