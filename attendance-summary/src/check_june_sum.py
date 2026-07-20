import openpyxl
import sys
wb = openpyxl.load_workbook(sys.argv[1], data_only=False)
ws = wb['汇总']
for r in range(2, 6):
    print(f'row {r}: A={ws.cell(r,1).value} H={ws.cell(r,8).value} I={ws.cell(r,9).value} J={ws.cell(r,10).value} R={ws.cell(r,18).value}')
